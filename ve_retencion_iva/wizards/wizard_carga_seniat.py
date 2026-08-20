import base64
import io
import logging
import re
import unicodedata

from markupsafe import Markup
from odoo import models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _formatear_rif(rif):
    """Mismo criterio que ve_conecta_carga_ventas.py::_formatear_rif (ver
    ese archivo para el detalle) -- RIF sin guión (letra + 9 dígitos) se
    formatea a LETRA-12345678-9. Pedido explícito 2026-08-05."""
    limpio = (rif or '').upper().strip()
    if not limpio or '-' in limpio:
        return limpio
    m = re.match(r'^([VEJPG])(\d{9})$', limpio)
    if not m:
        return limpio
    letra, digitos = m.groups()
    return f'{letra}-{digitos[:8]}-{digitos[8]}'


def _norm_header(text):
    text = text.lower().strip()
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[°#.\-_/\\]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _parse_amount(val):
    if val is None:
        return 0.0
    s = re.sub(r'[^\d.,]', '', str(val).strip())
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date(val):
    if val is None:
        return False
    # openpyxl puede devolver un objeto date/datetime directamente
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    # Número serial de Excel (ej. 46118 → 2026-03-28)
    if isinstance(val, (int, float)) and 1000 < val < 100000:
        try:
            from datetime import date, timedelta
            return (date(1899, 12, 30) + timedelta(days=int(val))).strftime('%Y-%m-%d')
        except Exception:
            pass
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return False


def _normalizar_periodo(val):
    _MESES = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12',
        'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04', 'may': '05',
        'jun': '06', 'jul': '07', 'ago': '08', 'sep': '09', 'oct': '10',
        'nov': '11', 'dic': '12',
    }
    if not val:
        return ''
    texto = str(val).strip()
    m = re.match(r'^(\d{4})-(\d{1,2})$', texto)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r'^(\d{1,2})[/-](\d{4})$', texto)
    if m and 1 <= int(m.group(1)) <= 12:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    for nombre, num in _MESES.items():
        match = re.search(rf'\b{nombre}\b.*?(\d{{4}})', texto, re.I)
        if match:
            return f"{match.group(1)}-{num}"
        match = re.search(rf'(\d{{4}}).*?\b{nombre}\b', texto, re.I)
        if match:
            return f"{match.group(1)}-{num}"
    return ''


_MESES_NOMBRE = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12',
    'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04', 'may': '05',
    'jun': '06', 'jul': '07', 'ago': '08', 'sep': '09', 'oct': '10',
    'nov': '11', 'dic': '12',
}

_HEADER_MAP = {
    # RIF
    'rif': 'rif_agente', 'rif agente': 'rif_agente',
    'rif agente retencion': 'rif_agente',
    'rif del agente': 'rif_agente', 'n rif': 'rif_agente', 'nro rif': 'rif_agente',
    # Nombre agente
    'nombre': 'nombre_agente', 'nombre agente': 'nombre_agente',
    'agente retencion': 'nombre_agente',
    'razon social': 'nombre_agente', 'nombre del agente': 'nombre_agente',
    # Período
    'periodo': 'periodo', 'periodo fiscal': 'periodo', 'mes': 'periodo',
    # Fecha
    'fecha': 'fecha', 'fecha operacion': 'fecha', 'fecha emision': 'fecha',
    'fecha documento': 'fecha',
    # N° Comprobante
    'n comprobante': 'name', 'nro comprobante': 'name', 'comprobante': 'name',
    'n de comprobante': 'name', 'numero comprobante': 'name',
    # N° Control
    'n control': 'nro_control', 'nro control': 'nro_control',
    'control': 'nro_control', 'numero control': 'nro_control',
    # N° Documento / Factura
    'n factura': 'nro_documento', 'nro factura': 'nro_documento',
    'factura': 'nro_documento', 'n documento': 'nro_documento',
    'nro documento': 'nro_documento', 'numero factura': 'nro_documento',
    'numero documento': 'nro_documento',
    # Tipo documento
    'tipo': 'tipo_documento', 'tipo documento': 'tipo_documento',
    # Base imponible
    'base imponible': 'monto_base', 'base': 'monto_base', 'imponible': 'monto_base',
    # Alícuota
    'alicuota': 'alicuota', 'alicuota iva': 'alicuota', 'tasa': 'alicuota',
    # Monto retenido (BsS = bolívares soberanos, la columna vigente)
    'iva retenido': 'monto_retenido', 'monto retenido': 'monto_retenido',
    'monto retenido bss': 'monto_retenido', 'monto retenido bsf': '_ignorar',
    'retencion': 'monto_retenido', 'retenido': 'monto_retenido',
    'impuesto retenido': 'monto_retenido',
    # Monto documento
    'total': 'monto_documento', 'monto total': 'monto_documento',
    'monto documento': 'monto_documento', 'monto con iva': 'monto_documento',
    'monto documento bss': 'monto_documento', 'monto documento bsf': '_ignorar',
    # Exento
    'exento': 'monto_exento', 'monto exento': 'monto_exento',
    'monto exento bss': 'monto_exento', 'monto exento bsf': '_ignorar',
    # Doc afectado
    'doc afectado': 'doc_afectado', 'documento afectado': 'doc_afectado',
    'numero document afectado': 'doc_afectado', 'numero documento afectado': 'doc_afectado',
}

_AMOUNT_FIELDS = {'monto_base', 'monto_retenido', 'monto_documento',
                  'monto_exento', 'alicuota'}


def _encontrar_hoja(wb, mes_num):
    """Busca la pestaña cuyo nombre coincida con el mes (número o nombre).
    mes_num: string '05'. Devuelve la hoja o None."""
    mes_int = int(mes_num)
    for name in wb.sheetnames:
        n = name.strip()
        # Coincidencia numérica: "5" o "05"
        try:
            if int(n) == mes_int:
                return wb[name], name
        except ValueError:
            pass
        # Coincidencia por nombre: "mayo", "Mayo", "may", etc.
        if _MESES_NOMBRE.get(n.lower()) == mes_num:
            return wb[name], name
    return None, None


class WizardCargaSeniat(models.TransientModel):
    _name = 've.seniat.wizard.carga'
    _description = 'Carga Manual de Retenciones SENIAT desde XLSX'

    conciliacion_id = fields.Many2one(
        've.conciliacion.periodo',
        string='Período de Conciliación',
        required=True,
    )
    archivo_xlsx = fields.Binary(string='Archivo XLSX')
    archivo_nombre = fields.Char()

    preview = fields.Text(string='Vista Previa', readonly=True)
    resultado = fields.Text(string='Resultado', readonly=True)
    importado = fields.Boolean(readonly=True)

    # ------------------------------------------------------------------ #
    #  Lógica de parseo                                                    #
    # ------------------------------------------------------------------ #

    def _leer_xlsx(self):
        """Devuelve (col_map, vals_list, errores, hoja_nombre).
        Selecciona automáticamente la pestaña que coincide con el mes del período."""
        if not self.archivo_xlsx:
            raise UserError('Suba un archivo XLSX primero.')

        try:
            import openpyxl
        except ImportError:
            raise UserError('La librería openpyxl no está disponible en este servidor.')

        raw = base64.b64decode(self.archivo_xlsx)
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(f'No se pudo abrir el archivo XLSX: {e}')

        # Determinar mes del período
        periodo = self.conciliacion_id.periodo or ''   # yyyy-mm
        if len(periodo) < 7:
            raise UserError(
                f'El período "{periodo}" no tiene formato yyyy-mm válido.')
        mes_num = periodo[5:7]   # '05'

        hoja, hoja_nombre = _encontrar_hoja(wb, mes_num)
        if hoja is None:
            pestanas = ', '.join(wb.sheetnames)
            raise UserError(
                f'No se encontró una pestaña para el mes {int(mes_num):02d} '
                f'({periodo}).\n'
                f'Pestañas disponibles en el archivo: {pestanas}')

        filas = list(hoja.iter_rows(values_only=True))
        wb.close()
        # Mismo bug real que ve_conecta_carga_ventas.py::_leer_filas (ver ese
        # archivo para el detalle) -- read_only=True puede cortar la lectura
        # casi de inmediato si el XLSX trae mal su metadato <dimension>.
        # Reintentar sin read_only solo si la primera lectura salió
        # sospechosamente corta.
        if len(filas) <= 1:
            try:
                wb2 = openpyxl.load_workbook(
                    filename=io.BytesIO(raw), read_only=False, data_only=True)
                hoja2, _ = _encontrar_hoja(wb2, mes_num)
                if hoja2 is not None:
                    filas_reintento = list(hoja2.iter_rows(values_only=True))
                    if len(filas_reintento) > len(filas):
                        filas = filas_reintento
                wb2.close()
            except Exception:
                pass

        if not filas:
            raise UserError(f'La pestaña "{hoja_nombre}" está vacía.')

        # Fila de encabezados: primera fila no vacía
        encabezado = None
        data_start = 0
        for i, fila in enumerate(filas):
            if any(c is not None and str(c).strip() for c in fila):
                encabezado = fila
                data_start = i + 1
                break

        if encabezado is None:
            raise UserError(f'La pestaña "{hoja_nombre}" no tiene encabezados.')

        # Mapear encabezados
        col_map = {}
        for i, h in enumerate(encabezado):
            if h is None:
                continue
            norm = _norm_header(str(h))
            if norm in _HEADER_MAP:
                col_map[i] = _HEADER_MAP[norm]

        # Descartar columnas marcadas como ignorar (BsF, duplicadas, etc.)
        col_map = {i: f for i, f in col_map.items() if f != '_ignorar'}

        if not col_map:
            raise UserError(
                f'No se reconoció ninguna columna en la pestaña "{hoja_nombre}".\n'
                'Verifique que la primera fila tenga los encabezados correctos.')

        mapped = set(col_map.values())
        if 'rif_agente' not in mapped:
            raise UserError('Columna obligatoria faltante: RIF')
        if 'nro_control' not in mapped:
            raise UserError('Columna obligatoria faltante: N° Control')

        errores = []
        vals_list = []

        for row_num, row in enumerate(filas[data_start:], start=data_start + 1):
            if not any(c is not None and str(c).strip() for c in row):
                continue

            # conciliacion_id NO se fija acá a un valor fijo -- bug real
            # 2026-08-04 (Cementos): el wizard se abre desde UN período
            # puntual (ej. "2025-12 1Q") pero el XLSX trae el MES completo
            # (1Q y 2Q mezclados); forzar conciliacion_id acá metía TODAS
            # las filas en el período desde el que se abrió el wizard, sin
            # importar su fecha real. Se resuelve por fila en
            # action_importar() (mismo criterio que el controller del RPA:
            # _asegurar_periodo según la fecha real de cada retención).
            vals = {
                'periodo': self.conciliacion_id.periodo or '',
                'cargado_por_rpa': False,
                'estado': 'cargado',
            }

            for col_idx, field in col_map.items():
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if cell is None or str(cell).strip() == '':
                    continue
                if field in _AMOUNT_FIELDS:
                    vals[field] = _parse_amount(cell)
                elif field == 'fecha':
                    fecha = _parse_date(cell)
                    if fecha:
                        vals[field] = fecha
                    else:
                        errores.append(
                            f'Fila {row_num}: fecha inválida "{cell}" (use DD/MM/YYYY)')
                elif field == 'periodo':
                    vals[field] = _normalizar_periodo(cell) or str(cell)
                elif field == 'rif_agente':
                    vals[field] = _formatear_rif(str(cell).strip())
                else:
                    vals[field] = str(cell).strip()

            if not vals.get('rif_agente'):
                errores.append(f'Fila {row_num}: RIF vacío — omitida')
                continue
            if not vals.get('nro_control'):
                errores.append(f'Fila {row_num}: N° Control vacío — omitida')
                continue

            vals_list.append(vals)

        return col_map, vals_list, errores, hoja_nombre

    # ------------------------------------------------------------------ #
    #  Acciones                                                            #
    # ------------------------------------------------------------------ #

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_previsualizar(self):
        self.ensure_one()
        col_map, vals_list, errores, hoja_nombre = self._leer_xlsx()

        campos = ', '.join(sorted(set(col_map.values())))
        periodo = self.conciliacion_id.periodo or ''
        mes_num = int(periodo[5:7]) if len(periodo) >= 7 else '?'
        lineas = [
            f'Pestaña cargada    : {hoja_nombre} (mes {mes_num})',
            f'Columnas detectadas: {campos}',
            f'Filas a importar   : {len(vals_list)}',
        ]
        if errores:
            lineas += ['', 'Advertencias:']
            lineas += [f'  • {e}' for e in errores[:10]]
            if len(errores) > 10:
                lineas.append(f'  ... y {len(errores) - 10} advertencias más')

        if vals_list:
            lineas += ['', 'Primeras 5 filas:']
            for v in vals_list[:5]:
                lineas.append(
                    f"  {v.get('rif_agente', ''):>14}  "
                    f"ctrl={v.get('nro_control', ''):12}  "
                    f"ret={v.get('monto_retenido', 0):>12,.2f}"
                )

        self.write({'preview': '\n'.join(lineas)})
        return self._reload()

    def action_importar(self):
        self.ensure_one()
        _, vals_list, _, hoja_nombre = self._leer_xlsx()

        if not vals_list:
            raise UserError('No hay filas válidas para importar.')

        total_parseado = len(vals_list)
        company = self.conciliacion_id.company_id
        ConcModel = self.env['ve.conciliacion.periodo']

        # Resolver conciliacion_id/periodo_retencion POR FILA según su
        # propia fecha (no la del período desde el que se abrió el
        # wizard) -- mismo criterio que ya usa el controller del RPA
        # (_asegurar_periodo). Cache por periodo_retencion para no
        # buscar/crear el mismo período de nuevo en cada fila.
        _periodo_cache = {}
        periodos_tocados = ConcModel.browse()
        SeniatRet = self.env['ve.seniat.retencion']
        # De-duplicar POR FILA del propio XLSX (nro_control+nro_documento+
        # rif_agente) -- bug real 2026-08-04 (Cementos): a diferencia del
        # controller del RPA (que busca "ya_existe" antes de crear), este
        # wizard llamaba create() directo sobre TODO vals_list -- si el
        # archivo traía la misma retención 2 veces (confirmado: 3 pares
        # exactos, misma fecha+monto+control+rif), se creaban 2 filas
        # físicas en vez de 1. Se ignora silenciosamente la 2da aparición de
        # la misma clave (nro_control puede repetirse legítimamente entre
        # agentes distintos -- ej. "00" como relleno -- por eso la clave
        # incluye rif_agente, no solo nro_control).
        #
        # Segundo bug real encontrado 2026-08-05, ANTES de llegar a
        # producción (verificado por RPC en Cementos): "00" como relleno de
        # nro_control NO es un caso raro -- 34 retenciones reales, 5 RIF
        # distintos con 2 a 5 retenciones CADA UNO compartiendo
        # nro_control="00", cada una con su propio nro_documento real. Sin
        # nro_documento en la clave, ese mismo RIF con 2+ retenciones "00"
        # en el mismo archivo se habría de-duplicado por error, descartando
        # retenciones genuinamente distintas.
        vistos = set()
        duplicados_en_archivo = 0
        vals_list_dedup = []
        for vals in vals_list:
            fecha_ref = vals.get('fecha') or f'{self.conciliacion_id.periodo}-01'
            pr = SeniatRet._calc_pr(vals.get('periodo') or self.conciliacion_id.periodo, fecha_ref)
            clave = (vals.get('nro_control'), vals.get('nro_documento'), vals.get('rif_agente'), pr)
            if clave in vistos:
                duplicados_en_archivo += 1
                continue
            vistos.add(clave)
            if pr not in _periodo_cache:
                _periodo_cache[pr] = ConcModel._asegurar_periodo(company, fecha_ref)
            periodo_dest = _periodo_cache[pr]
            vals['conciliacion_id'] = periodo_dest.id
            vals['periodo_retencion'] = periodo_dest.periodo_retencion
            periodos_tocados |= periodo_dest
            vals_list_dedup.append(vals)
        vals_list = vals_list_dedup

        # Eliminar registros SENIAT existentes del MES completo (ambas
        # quincenas) antes de reimportar -- el XLSX trae el mes entero, no
        # solo la quincena desde la que se abrió el wizard (mismo bug de
        # arriba: antes solo se borraba conciliacion_id == período actual,
        # dejando basura mezclada con lo recién importado si ya existían
        # filas de la OTRA quincena del mismo mes).
        existentes = self.env['ve.seniat.retencion'].search([
            ('periodo', '=', self.conciliacion_id.periodo),
            ('company_id', '=', company.id),
        ])
        eliminados = len(existentes)
        if existentes:
            existentes.unlink()

        creados = self.env['ve.seniat.retencion'].create(vals_list)

        # Una retención SENIAT real para un RIF no marcado como Agente de
        # Retención es una contradicción entre el Libro de Ventas del
        # cliente y SENIAT que amerita revisión humana -- YA NO se marca
        # sola (pedido explícito 2026-08-18, ver res_partner.py::
        # _detectar_agentes_retencion_por_rif). Se reporta abajo en el
        # chatter para revisión manual.
        rifs_archivo = {v.get('rif_agente') for v in vals_list if v.get('rif_agente')}
        no_spe = self.env['res.partner']._detectar_agentes_retencion_por_rif(company, rifs_archivo)
        no_spe_txt = ''
        no_spe_lineas = False
        if no_spe:
            no_spe_lineas = '\n'.join(f'{p.vat} — {p.name}' for p in no_spe)
            listado = '<br/>'.join(
                f'&#8226; {p.vat} — {p.name}' for p in no_spe[:30])
            mas = f'<br/><i>... y {len(no_spe) - 30} más</i>' if len(no_spe) > 30 else ''
            no_spe_txt = (
                f'<br/><br/><b>&#9888; {len(no_spe)} RIF(s) con retención SENIAT '
                f'pero SIN marcar como Contribuyente Especial — revisar antes de '
                f'marcarlos a mano:</b><br/>{listado}{mas}'
            )

        from datetime import datetime
        now_fmt = datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S')
        nombre  = self.archivo_nombre or 'archivo.xlsx'

        reemplazo = f'<br/><i>({eliminados} registros anteriores reemplazados)</i>' if eliminados else ''
        dup_txt = f' ({duplicados_en_archivo} duplicados en el archivo, ignorados)' if duplicados_en_archivo else ''

        # El XLSX puede traer ambas quincenas del mes -- se avisa/actualiza
        # CADA período tocado (no solo el que abrió el wizard), con su
        # propio conteo, para que ningún período quede con estado_extraccion
        # desactualizado pese a haber recibido datos reales.
        for periodo_dest in periodos_tocados:
            n_periodo = len(creados.filtered(lambda r: r.conciliacion_id.id == periodo_dest.id))
            # Mismo formato que el interceptor de mensajes del RPA
            # (ve_conciliacion.py::message_post) -- "extraídas del mes" vs
            # "vinculadas a este período", para que no se lea como una
            # discrepancia cuando el archivo trae el mes completo.
            periodo_dest.write({
                'estado_extraccion': 'completada',
                'fecha_estado_extraccion': fields.Datetime.now(),
                'mensaje_estado_extraccion':
                    f'{total_parseado} extraídas del mes, {n_periodo} vinculadas a este período'
                    f'{dup_txt} (XLSX: {nombre})',
                'rifs_seniat_no_spe': no_spe_lineas,
                'partners_seniat_no_spe_ids': [(6, 0, no_spe.ids)],
            })
            periodo_dest.message_post(
                body=Markup(
                    f'<b>Retenciones SENIAT cargadas desde XLSX</b><br/>'
                    f'<b>Archivo:</b> {nombre}<br/>'
                    f'<b>Pestaña:</b> {hoja_nombre}<br/>'
                    f'<b>Fecha / Hora:</b> {now_fmt} UTC<br/>'
                    f'<b>Registros importados:</b> {n_periodo} (de {total_parseado} en todo el mes)'
                    f'{dup_txt}{reemplazo}<br/><br/>'
                    f'Use el botón <b>Conciliar</b> para cruzarlas contra '
                    f'las retenciones Odoo del período {periodo_dest.periodo_retencion}.'
                    f'{no_spe_txt}'
                ),
                message_type='comment',
                subtype_xmlid='mail.mt_note',
            )

        return {
            'type': 'ir.actions.act_window',
            'res_model': 've.conciliacion.periodo',
            'res_id': self.conciliacion_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
