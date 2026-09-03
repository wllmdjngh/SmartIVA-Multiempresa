import io
import logging
import re

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_TIPO_DOC = {
    'out_invoice': '01',
    'out_refund':  '03',
}


class WizardLibroVentas(models.TransientModel):
    _name = 've.wizard.libro.ventas'
    _description = 'Asistente Libro de Ventas'

    periodo = fields.Char(
        string='Período Fiscal',
        required=True,
        default=lambda self: fields.Date.today().strftime('%Y-%m'),
        help='Formato yyyy-mm. Ej: 2026-05',
    )
    quincena = fields.Selection([
        ('completo', 'Mes completo'),
        ('1Q',       '1ra quincena (1-15)'),
        ('2Q',       '2da quincena (16-fin)'),
    ], string='Quincena', default='completo', required=True)
    company_id = fields.Many2one(
        'res.company',
        string='Empresa',
        required=True,
        default=lambda self: self.env.company,
    )

    @api.constrains('periodo')
    def _check_periodo(self):
        for rec in self:
            if not re.match(r'^\d{4}-\d{2}$', rec.periodo):
                raise UserError(
                    'El período debe tener el formato yyyy-mm (ej: 2026-05).'
                )

    def _get_facturas(self):
        self.ensure_one()
        import calendar
        year, month = int(self.periodo[:4]), int(self.periodo[5:7])
        last_day = calendar.monthrange(year, month)[1]
        if self.quincena == '1Q':
            date_from = f'{year:04d}-{month:02d}-01'
            date_to   = f'{year:04d}-{month:02d}-15'
        elif self.quincena == '2Q':
            date_from = f'{year:04d}-{month:02d}-16'
            date_to   = f'{year:04d}-{month:02d}-{last_day:02d}'
        else:
            date_from = f'{year:04d}-{month:02d}-01'
            date_to   = f'{year:04d}-{month:02d}-{last_day:02d}'
        return self.env['account.move'].search([
            ('move_type', 'in', ('out_invoice', 'out_refund')),
            ('state', '=', 'posted'),
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('company_id', '=', self.company_id.id),
        ], order='invoice_date, name')

    def _compute_line_data(self, move, op_nro=0):
        """Devuelve dict con los campos del libro para una factura."""
        sign = -1 if move.move_type == 'out_refund' else 1

        base_16 = iva_16 = 0.0
        base_8 = iva_8 = 0.0
        base_exp = base_exenta = 0.0

        for line in move.invoice_line_ids:
            if line.display_type in ('line_section', 'line_note'):
                continue
            taxes = line.tax_ids
            if not taxes:
                base_exenta += line.price_subtotal
                continue
            for tax in taxes:
                if not hasattr(tax, 'amount'):
                    continue
                if tax.amount == 0:
                    base_exp += line.price_subtotal
                elif abs(tax.amount - 16.0) < 0.01:
                    base_16 += line.price_subtotal
                    iva_16 += line.price_subtotal * 0.16
                elif abs(tax.amount - 8.0) < 0.01:
                    base_8 += line.price_subtotal
                    iva_8 += line.price_subtotal * 0.08

        # Datos de nota de crédito/débito
        nro_nc = ''
        nro_nd = ''
        nro_fact_afectada = ''
        if move.move_type == 'out_refund':
            nro_nc = move.name
            if move.reversed_entry_id:
                nro_fact_afectada = move.reversed_entry_id.name
        elif move.move_type == 'out_invoice':
            debit_origin = getattr(move, 'debit_origin_id', False)
            if debit_origin:
                nro_nd = move.name
                nro_fact_afectada = debit_origin.name

        # Tipo de documento SENIAT: 01 Factura, 02 Nota de Débito, 03 Nota de
        # Crédito -- _TIPO_DOC.get(move.move_type) no distingue una ND, que
        # sigue teniendo move_type=='out_invoice' (solo cambia debit_origin_id,
        # ya calculado arriba en nro_nd). Bug real corregido 2026-09-03: antes
        # exportaba '01' para toda ND, código '02' nunca salía de este reporte.
        tipo_doc = '02' if nro_nd else _TIPO_DOC.get(move.move_type, '01')

        # Comprobante de retención vinculado
        wh = move.ve_wh_iva_ids[:1]
        nro_comp = ''
        monto_ret = 0.0
        estado_comp = ''
        if wh:
            nro_comp = wh.name or ''
            monto_ret = wh.monto_retenido
            estado_comp = dict(wh._fields['state'].selection).get(wh.state, '')

        return {
            'op_nro':            op_nro,
            'fecha':             move.invoice_date,
            'rif':               move.partner_id.vat or '',
            'cliente':           move.partner_id.name or '',
            'nro_comp':          nro_comp,
            'nro_factura':       move.name,
            'nro_control':       getattr(move, 'nro_control', '') or '',
            'nro_nd':            nro_nd,
            'nro_nc':            nro_nc,
            'tipo_doc':          tipo_doc,
            'nro_fact_afectada': nro_fact_afectada,
            'total_iva':         sign * move.amount_total,
            'ventas_no_grav':    sign * base_exenta,
            'base_16':           sign * base_16,
            'iva_16':            sign * iva_16,
            'base_8':            sign * base_8,
            'iva_8':             sign * iva_8,
            'monto_ret':         sign * monto_ret,
            'iva_percibido':     0.0,
            # Para acumulados en resumen
            'base_exp':          sign * base_exp,
            'estado_comp':       estado_comp,
        }

    def _get_summary(self):
        """Totales por categoría para la sección de resumen RLIVA Arts. 70-78."""
        self.ensure_one()
        facturas = self._get_facturas()
        s = {
            'no_gravadas':  0.0,
            'exportaciones': 0.0,
            'base_16':  0.0, 'iva_16':  0.0, 'ret_16':  0.0,
            'base_8':   0.0, 'iva_8':   0.0, 'ret_8':   0.0,
            'total_base': 0.0, 'total_iva': 0.0, 'total_ret': 0.0,
            'total_total': 0.0,
        }
        for move in facturas:
            d = self._compute_line_data(move)
            s['no_gravadas']   += d['ventas_no_grav']
            s['exportaciones'] += d['base_exp']
            s['base_16']       += d['base_16']
            s['iva_16']        += d['iva_16']
            s['base_8']        += d['base_8']
            s['iva_8']         += d['iva_8']
            s['total_total']   += d['total_iva']
            # Asignar retención a la alícuota mayoritaria
            if d['base_16'] >= d['base_8']:
                s['ret_16'] += d['monto_ret']
            else:
                s['ret_8'] += d['monto_ret']

        s['total_base'] = s['base_16'] + s['base_8']
        s['total_iva']  = s['iva_16'] + s['iva_8']
        s['total_ret']  = s['ret_16'] + s['ret_8']
        return s

    def action_imprimir_pdf(self):
        self.ensure_one()
        facturas = self._get_facturas()
        if not facturas:
            raise UserError(
                f'No se encontraron facturas de venta publicadas '
                f'para el período {self.periodo}.'
            )
        return self.env.ref(
            've_retencion_iva.action_report_libro_ventas'
        ).report_action(self)

    def action_exportar_excel(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise UserError(
                'La librería openpyxl no está instalada en este servidor.\n'
                'Use el botón "Imprimir PDF" o pida al administrador '
                'que instale openpyxl (pip install openpyxl).'
            )

        facturas = self._get_facturas()
        if not facturas:
            raise UserError(
                f'No se encontraron facturas para el período {self.periodo}.'
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        periodo_label = (f'{self.periodo} {self.quincena}'
                         if self.quincena != 'completo' else self.periodo)
        ws.title = f'Libro Ventas {periodo_label}'[:31]

        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        total_fill  = PatternFill('solid', fgColor='D6E4F0')
        total_font  = Font(bold=True, size=9)
        thin   = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            'Op. N°', 'Fecha', 'RIF', 'Nombre o Razón Social',
            'N° Comprobante', 'N° Factura', 'N° Control',
            'N° Nota Débito', 'N° Nota Crédito',
            'Tipo', 'N° Fact. Afectada',
            'Total c/IVA', 'V. Internas No Gravadas',
            'Base 16%', 'IVA 16%',
            'Base 8%', 'IVA 8%',
            'IVA Retenido', 'IVA Percibido',
        ]
        col_widths = [6, 12, 12, 35, 18, 18, 14, 14, 14, 6, 18,
                      14, 18, 16, 14, 16, 14, 16, 14]

        # Título
        ncols = len(headers)
        last_col_letter = openpyxl.utils.get_column_letter(ncols)
        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.value = (
            f'LIBRO DE VENTAS  –  {self.company_id.name}  –  '
            f'Período: {periodo_label}'
        )
        title_cell.font = Font(bold=True, size=11)
        title_cell.alignment = Alignment(horizontal='center')

        # Cabecera
        ws.row_dimensions[2].height = 32
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = w

        # Filas de datos
        num_fmt  = '#,##0.00'
        date_fmt = 'DD/MM/YYYY'
        money_cols = {12, 13, 14, 15, 16, 17, 18, 19}

        for row_idx, (op_nro, move) in enumerate(
            enumerate(facturas, 1), 3
        ):
            d = self._compute_line_data(move, op_nro)
            row_data = [
                d['op_nro'], d['fecha'], d['rif'], d['cliente'],
                d['nro_comp'], d['nro_factura'], d['nro_control'],
                d['nro_nd'], d['nro_nc'],
                d['tipo_doc'], d['nro_fact_afectada'],
                d['total_iva'], d['ventas_no_grav'],
                d['base_16'], d['iva_16'],
                d['base_8'], d['iva_8'],
                d['monto_ret'], d['iva_percibido'],
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.font = Font(size=9)
                if col == 2 and val:
                    cell.number_format = date_fmt
                elif col in money_cols:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='right')

        # Fila de totales de detalle
        total_row = len(facturas) + 3
        ws.cell(total_row, 1, 'TOTALES').font = total_font
        ws.merge_cells(
            start_row=total_row, start_column=1,
            end_row=total_row, end_column=11,
        )

        s = self._get_summary()
        total_map = {
            12: s['total_total'],
            13: s['no_gravadas'],
            14: s['base_16'],
            15: s['iva_16'],
            16: s['base_8'],
            17: s['iva_8'],
            18: s['total_ret'],
            19: 0.0,
        }
        for col, val in total_map.items():
            cell = ws.cell(total_row, col, val)
            cell.number_format = num_fmt
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal='right')
            cell.border = border

        # Sección resumen RLIVA (debajo de los totales)
        r = total_row + 2
        summary_fill = PatternFill('solid', fgColor='F0F4FF')
        summary_hdr_fill = PatternFill('solid', fgColor='2E75B6')

        ws.merge_cells(f'A{r}:{last_col_letter}{r}')
        hdr = ws[f'A{r}']
        hdr.value = 'RESUMEN RLIVA  — Arts. 70-78'
        hdr.fill = summary_hdr_fill
        hdr.font = Font(color='FFFFFF', bold=True, size=10)
        hdr.alignment = Alignment(horizontal='center')
        r += 1

        sub_headers = ['Categoría', 'Base Imponible', 'Débito Fiscal', 'IVA Ret. x Comprador']
        for ci, sh in enumerate(sub_headers):
            ws.cell(r, ci + 1, sh).font = Font(bold=True, size=9)
        r += 1

        summary_rows = [
            ('Total: Ventas Internas No Gravadas',          s['no_gravadas'],  0.0,           0.0),
            ('Total Ventas de Exportación',                  s['exportaciones'], 0.0,          0.0),
            ('Total Ventas Exportación — Alíc. General (16%)', 0.0,            0.0,           0.0),
            ('Total Ventas Exportación — Alíc. Reducida (8%)', 0.0,            0.0,           0.0),
            ('Total Ventas Exportación — Ambas Alícuotas',   0.0,              0.0,           0.0),
            ('Total Ventas Internas — Alíc. General (16%)',  s['base_16'],     s['iva_16'],   s['ret_16']),
            ('Total Ventas Internas — Alíc. Reducida (8%)',  s['base_8'],      s['iva_8'],    s['ret_8']),
        ]
        for label, base, iva, ret in summary_rows:
            ws.cell(r, 1, label).font = Font(size=9)
            for ci, val in enumerate([base, iva, ret], 2):
                cell = ws.cell(r, ci, val)
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='right')
                cell.fill = summary_fill
                cell.font = Font(size=9)
            r += 1

        r += 1
        ws.cell(r, 1, 'TOTALES GENERALES').font = Font(bold=True, size=10)
        for ci, val in enumerate([s['total_base'], s['total_iva'], s['total_ret']], 2):
            cell = ws.cell(r, ci, val)
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal='right')
            cell.fill = total_fill
            cell.font = total_font

        # Guardar y devolver como descarga
        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()

        import base64
        fname_key = (f'{self.periodo}_{self.quincena}'
                     if self.quincena != 'completo' else self.periodo)
        filename = f'libro_ventas_{fname_key}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(xlsx_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
