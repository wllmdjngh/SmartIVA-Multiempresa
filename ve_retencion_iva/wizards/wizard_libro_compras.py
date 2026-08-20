import io
import re
import logging
import calendar

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_TIPO_DOC = {
    'in_invoice': '01-reg',
    'in_refund':  '03',
}


class WizardLibroCompras(models.TransientModel):
    _name = 've.wizard.libro.compras'
    _description = 'Asistente Libro de Compras'

    periodo = fields.Char(
        string='Período Fiscal',
        required=True,
        default=lambda self: self._default_periodo(),
        help='Formato yyyy-mm. Ej: 2026-05',
    )

    @staticmethod
    def _default_periodo():
        today = fields.Date.today()
        if today.day <= 15:
            # Aún en la 1Q del mes actual → el período a procesar es el mes anterior
            if today.month == 1:
                return f'{today.year - 1:04d}-12'
            return f'{today.year:04d}-{today.month - 1:02d}'
        return today.strftime('%Y-%m')
    quincena = fields.Selection([
        ('completo', 'Mes completo'),
        ('1Q',       '1ra quincena (1-15)'),
        ('2Q',       '2da quincena (16-fin)'),
    ], string='Quincena', default='completo', required=True)
    company_id = fields.Many2one(
        'res.company',
        string='Empresa',
        required=True,
        default=lambda self: self.env.company,
    )

    @api.constrains('periodo')
    def _check_periodo(self):
        for rec in self:
            if not re.match(r'^\d{4}-\d{2}$', rec.periodo):
                raise UserError(
                    'El período debe tener el formato yyyy-mm (ej: 2026-05).'
                )

    def _get_fechas(self):
        self.ensure_one()
        year, month = int(self.periodo[:4]), int(self.periodo[5:7])
        last_day = calendar.monthrange(year, month)[1]
        if self.quincena == '1Q':
            return f'{year:04d}-{month:02d}-01', f'{year:04d}-{month:02d}-15'
        if self.quincena == '2Q':
            return f'{year:04d}-{month:02d}-16', f'{year:04d}-{month:02d}-{last_day:02d}'
        return f'{year:04d}-{month:02d}-01', f'{year:04d}-{month:02d}-{last_day:02d}'

    def _get_facturas(self):
        self.ensure_one()
        date_from, date_to = self._get_fechas()
        return self.env['account.move'].search([
            ('move_type', 'in', ('in_invoice', 'in_refund')),
            ('state', '=', 'posted'),
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('company_id', '=', self.company_id.id),
        ], order='invoice_date, name')

    def _get_retencion_prov(self, move_id):
        """Retorna el primer comprobante ve.wh.iva.prov activo vinculado a la factura."""
        return self.env['ve.wh.iva.prov'].search([
            ('invoice_id', '=', move_id),
            ('state', '!=', 'anulado'),
        ], limit=1)

    def _compute_line_data(self, move, op_nro=0):
        """Devuelve dict con todos los campos del libro para una factura."""
        sign = -1 if move.move_type == 'in_refund' else 1

        base_16 = iva_16 = 0.0
        base_8  = iva_8  = 0.0
        base_exenta = 0.0

        for line in move.invoice_line_ids:
            if line.display_type in ('line_section', 'line_note'):
                continue
            taxes = line.tax_ids
            if not taxes:
                base_exenta += line.price_subtotal
                continue
            for tax in taxes:
                if not hasattr(tax, 'amount'):
                    continue
                if tax.amount == 0:
                    base_exenta += line.price_subtotal
                elif abs(tax.amount - 16.0) < 0.01:
                    base_16 += line.price_subtotal
                    iva_16  += line.price_subtotal * 0.16
                elif abs(tax.amount - 8.0) < 0.01:
                    base_8 += line.price_subtotal
                    iva_8  += line.price_subtotal * 0.08

        # NC/ND
        nro_nc = nro_nd = nro_fact_afectada = ''
        if move.move_type == 'in_refund':
            nro_nc = move.ref or move.name
            if move.reversed_entry_id:
                nro_fact_afectada = (
                    move.reversed_entry_id.ref or move.reversed_entry_id.name
                )

        # Retención emitida por nosotros (ve.wh.iva.prov)
        wh_prov = self._get_retencion_prov(move.id)
        monto_ret     = wh_prov.monto_retenido if wh_prov else 0.0
        nro_comp_ret  = wh_prov.ref            if wh_prov else ''
        pct_ret       = wh_prov.porcentaje_retencion if wh_prov else 0.0

        total_sin_iva = sign * (base_16 + base_8 + base_exenta)
        total_iva     = sign * (iva_16 + iva_8)
        total_con_iva = total_sin_iva + total_iva

        return {
            'op_nro':           op_nro,
            'fecha':            move.invoice_date,
            'rif':              move.partner_id.vat or '',
            'proveedor':        move.partner_id.name or '',
            'nro_comprobante':  move.ref or '',          # N° de la factura del proveedor
            'nro_control':      getattr(move, 'nro_control', '') or '',
            'nro_nd':           nro_nd,
            'nro_nc':           nro_nc,
            'tipo_doc':         _TIPO_DOC.get(move.move_type, '01-reg'),
            'nro_fact_afectada': nro_fact_afectada,
            'total_con_iva':    total_con_iva,
            'base_exenta':      sign * base_exenta,
            'base_16':          sign * base_16,
            'iva_16':           sign * iva_16,
            'base_8':           sign * base_8,
            'iva_8':            sign * iva_8,
            'total_iva':        total_iva,
            'monto_ret':        sign * monto_ret,
            'pct_ret':          pct_ret,
            'nro_comp_ret':     nro_comp_ret,   # N° comprobante que nosotros emitimos
        }

    def _get_summary(self):
        """Totales por alícuota para la sección de resumen del PDF."""
        base_exenta = base_16 = iva_16 = ret_16 = base_8 = iva_8 = ret_8 = 0.0
        for move in self._get_facturas():
            d = self._compute_line_data(move)
            base_exenta += d['base_exenta']
            base_16     += d['base_16']
            iva_16      += d['iva_16']
            base_8      += d['base_8']
            iva_8       += d['iva_8']
            total_iva = d['iva_16'] + d['iva_8']
            if total_iva:
                ret_16 += d['monto_ret'] * (d['iva_16'] / total_iva) if d['iva_16'] else 0.0
                ret_8  += d['monto_ret'] * (d['iva_8']  / total_iva) if d['iva_8']  else 0.0
            else:
                ret_16 += d['monto_ret']
        return {
            'base_exenta': base_exenta,
            'base_16':     base_16, 'iva_16': iva_16, 'ret_16': ret_16,
            'base_8':      base_8,  'iva_8':  iva_8,  'ret_8':  ret_8,
            'total_base':  base_exenta + base_16 + base_8,
            'total_iva':   iva_16 + iva_8,
            'total_ret':   ret_16 + ret_8,
        }

    def action_imprimir_pdf(self):
        self.ensure_one()
        facturas = self._get_facturas()
        if not facturas:
            date_from, date_to = self._get_fechas()
            # Diagnóstico: ¿existen facturas en borrador o en otro estado?
            en_periodo = self.env['account.move'].search([
                ('move_type', 'in', ('in_invoice', 'in_refund')),
                ('invoice_date', '>=', date_from),
                ('invoice_date', '<=', date_to),
                ('company_id', '=', self.company_id.id),
            ])
            total_sistema = self.env['account.move'].search_count([
                ('move_type', 'in', ('in_invoice', 'in_refund')),
                ('state', '=', 'posted'),
                ('company_id', '=', self.company_id.id),
            ])
            if en_periodo:
                estados = dict(en_periodo.mapped(lambda m: (m.state, True)))
                detalle = ', '.join(estados.keys())
                raise UserError(
                    f'No hay facturas de compra CONFIRMADAS para {self.periodo} '
                    f'({date_from} → {date_to}).\n\n'
                    f'Sí existen {len(en_periodo)} factura(s) en ese período '
                    f'en estado(s): {detalle}.\n\n'
                    f'Confirme (publique) las facturas antes de generar el libro.'
                )
            raise UserError(
                f'No hay facturas de compra en {self.periodo} '
                f'({date_from} → {date_to}) para {self.company_id.name}.\n\n'
                f'El sistema tiene {total_sistema} factura(s) de compra confirmada(s) '
                f'en total (cualquier período).\n\n'
                f'Si acaba de ejecutar "Reiniciar Demo", vuelva a correrlo — '
                f'las facturas de proveedores fueron agregadas al demo recientemente.'
            )
        return self.env.ref(
            've_retencion_iva.action_report_libro_compras'
        ).report_action(self)

    def action_exportar_excel(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise UserError('La librería openpyxl no está instalada.')

        import base64
        facturas = self._get_facturas()
        date_from, date_to = self._get_fechas()

        wb = openpyxl.Workbook()
        ws = wb.active
        label_q = {'completo': 'Mes completo', '1Q': '1ra Quincena', '2Q': '2da Quincena'}
        ws.title = f'LibroCompras {self.periodo} {label_q.get(self.quincena,"")}'[:31]

        navy   = PatternFill('solid', fgColor='1F4E79')
        blue2  = PatternFill('solid', fgColor='2E75B6')
        light  = PatternFill('solid', fgColor='D6E4F0')
        gray   = PatternFill('solid', fgColor='F2F2F2')
        green  = PatternFill('solid', fgColor='E2EFDA')
        thin   = Side(style='thin')
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

        company = self.env.company
        titulo_periodo = (
            f'LIBRO DE COMPRAS — FORMATO SENIAT | '
            f'Período: {self.periodo} | {label_q.get(self.quincena,"")} | '
            f'{date_from} al {date_to}'
        )

        # ── Encabezado empresa ────────────────────────────────────────────────
        NCOLS = 17
        last = openpyxl.utils.get_column_letter(NCOLS)
        for row_txt, row_num in [
            (company.name or '',                                1),
            (f'R.I.F.: {company.vat or ""}',                   2),
            (titulo_periodo,                                    3),
        ]:
            ws.merge_cells(f'A{row_num}:{last}{row_num}')
            c = ws[f'A{row_num}']
            c.value = row_txt
            c.font = Font(bold=(row_num == 3), size=10 if row_num < 3 else 9)
            c.alignment = Alignment(horizontal='center')

        # ── Encabezados columnas ──────────────────────────────────────────────
        HDR_ROW = 5
        headers = [
            ('Op.\nN°',         5),
            ('Fecha\nFactura',  11),
            ('RIF',             14),
            ('Nombre /\nRazón Social', 32),
            ('N° de\nFactura',  18),
            ('N° de\nControl',  14),
            ('N° Nota\nDébito',  10),
            ('N° Nota\nCrédito', 10),
            ('Tipo\nTranscc.',    8),
            ('N° Fact.\nAfectada', 14),
            ('Total Compras\nIncl. IVA',   14),
            ('Compras sin\nCrédito IVA',   13),
            ('Base\nImponible',  14),
            ('%\nAlic.',          6),
            ('IVA',              12),
            ('IVA Retenido\nal Vendedor',  14),
            ('N° Comp.\nRetención',        16),
        ]
        for ci, (h, w) in enumerate(headers, 1):
            c = ws.cell(HDR_ROW, ci, h)
            c.fill = navy
            c.font = Font(color='FFFFFF', bold=True, size=8)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = bdr
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[HDR_ROW].height = 28

        # ── Filas de datos ────────────────────────────────────────────────────
        num_fmt  = '#,##0.00'
        date_fmt = 'DD/MM/YYYY'
        pct_fmt  = '#,##0.00'

        COL_FECHA = 2; COL_TOT = 11; COL_EXENTA = 12
        COL_BASE  = 13; COL_PCT = 14; COL_IVA = 15; COL_RET = 16

        totals = {
            'base_exenta': 0.0, 'base_16': 0.0, 'iva_16': 0.0,
            'base_8': 0.0,       'iva_8': 0.0,  'monto_ret': 0.0,
            'total_con_iva': 0.0,
        }

        for ri, move in enumerate(facturas, HDR_ROW + 1):
            d = self._compute_line_data(move, ri - HDR_ROW)
            for k in totals:
                totals[k] += d.get(k, 0.0)

            row_fill = gray if (ri - HDR_ROW) % 2 == 0 else None
            row_data = [
                d['op_nro'], d['fecha'], d['rif'], d['proveedor'],
                d['nro_comprobante'], d['nro_control'],
                d['nro_nd'], d['nro_nc'],
                d['tipo_doc'], d['nro_fact_afectada'],
                d['total_con_iva'], d['base_exenta'],
                d['base_16'] + d['base_8'],   # Base imponible total
                16.0 if d['base_16'] else (8.0 if d['base_8'] else 0.0),
                d['iva_16'] + d['iva_8'],
                d['monto_ret'],
                d['nro_comp_ret'],
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = bdr
                cell.font   = Font(size=8)
                if row_fill:
                    cell.fill = row_fill
                if ci == COL_FECHA and val:
                    cell.number_format = date_fmt
                elif ci in (COL_TOT, COL_EXENTA, COL_BASE, COL_IVA, COL_RET):
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='right')
                elif ci == COL_PCT:
                    cell.number_format = pct_fmt
                    cell.alignment = Alignment(horizontal='center')

        # ── Fila totales ──────────────────────────────────────────────────────
        tr = HDR_ROW + len(facturas) + 1
        lbl = ws.cell(tr, 1, 'TOTALES')
        lbl.font = Font(bold=True, size=8)
        ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=10)

        tot_data = {
            COL_TOT:    totals['total_con_iva'],
            COL_EXENTA: totals['base_exenta'],
            COL_BASE:   totals['base_16'] + totals['base_8'],
            COL_IVA:    totals['iva_16']  + totals['iva_8'],
            COL_RET:    totals['monto_ret'],
        }
        for ci, val in tot_data.items():
            c = ws.cell(tr, ci, val)
            c.fill   = light
            c.font   = Font(bold=True, size=8)
            c.border = bdr
            c.number_format = num_fmt
            c.alignment = Alignment(horizontal='right')

        # ── Resumen SENIAT ────────────────────────────────────────────────────
        sr = tr + 2
        ws.merge_cells(f'A{sr}:{last}{sr}')
        t = ws[f'A{sr}']
        t.value = 'RESUMEN'
        t.font  = Font(bold=True, size=9, color='FFFFFF')
        t.fill  = blue2
        t.alignment = Alignment(horizontal='center')

        sum_headers = ['Categoría', 'Base Imponible', 'Crédito Fiscal', 'IVA Retenido por el Comprador']
        sum_cols    = [1, 13, 15, 16]
        sr += 1
        for col, hdr in zip(sum_cols, sum_headers):
            c = ws.cell(sr, col, hdr)
            c.fill   = blue2
            c.font   = Font(bold=True, size=8, color='FFFFFF')
            c.border = bdr
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        def sum_row(row, label, base, cred, ret, fill=None):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row, 1, label)
            c.font = Font(size=8, bold=True)
            c.alignment = Alignment(horizontal='left')
            if fill:
                c.fill = fill
            for col, val in [(13, base), (15, cred), (16, ret)]:
                cell = ws.cell(row, col, val)
                cell.number_format = num_fmt
                cell.font   = Font(size=8)
                cell.border = bdr
                cell.alignment = Alignment(horizontal='right')
                if fill:
                    cell.fill = fill

        sr += 1
        sum_row(sr, 'Total: Compras Exentas y/o sin derecho a Crédito Fiscal',
                totals['base_exenta'], 0.0, 0.0)
        sr += 1
        sum_row(sr, 'Total Compras Importación Afectadas sólo Alícuota 16,00%', 0.0, 0.0, 0.0)
        sr += 1
        sum_row(sr, 'Total Compras Importación Afectadas sólo Alícuota  8,00%', 0.0, 0.0, 0.0)
        sr += 1
        sum_row(sr, 'Total Compras Internas Afectadas sólo Alícuota     16,00%',
                totals['base_16'], totals['iva_16'], totals['monto_ret'], fill=green)
        sr += 1
        sum_row(sr, 'Total Compras Internas Afectadas sólo Alícuota      8,00%',
                totals['base_8'],  totals['iva_8'],  0.0)
        sr += 1
        # Gran total
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=12)
        c = ws.cell(sr, 1, 'GRAN TOTAL')
        c.font  = Font(bold=True, size=8)
        c.fill  = light
        for col, val in [
            (13, totals['base_16'] + totals['base_8'] + totals['base_exenta']),
            (15, totals['iva_16']  + totals['iva_8']),
            (16, totals['monto_ret']),
        ]:
            cell = ws.cell(sr, col, val)
            cell.number_format = num_fmt
            cell.font   = Font(bold=True, size=8)
            cell.border = bdr
            cell.fill   = light
            cell.alignment = Alignment(horizontal='right')

        # ── Ajustes finales ───────────────────────────────────────────────────
        ws.auto_filter.ref = f'A{HDR_ROW}:{last}{HDR_ROW + len(facturas)}'
        ws.freeze_panes = f'A{HDR_ROW + 1}'

        output = io.BytesIO()
        wb.save(output)
        label_q2 = label_q.get(self.quincena, '').replace(' ', '_')
        fname = f'libro_compras_{self.periodo}_{label_q2}.xlsx'
        att = self.env['ir.attachment'].create({
            'name':      fname,
            'datas':     __import__('base64').b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id':    self.id,
            'mimetype':  ('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url':  f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
