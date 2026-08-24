import base64
import io
import re
from collections import defaultdict

from odoo import api, fields, models
from odoo.exceptions import UserError

MESES_NOMBRE = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

# estado_recepcion (NO state -- bug real confirmado 2026-08-22: state de
# ve.wh.iva nunca vale 'recibido'/'recibido_dif'/'confirmado_dif', esos son
# valores de estado_recepcion; comparar contra state solo hacía match con
# 'confirmado', dejando fuera de "Con Comprobante" los state='borrador'
# -- literalmente "Recibido" según su propia etiqueta -- y descuadraba
# contra el Dashboard, que sí usa estado_recepcion (ver _RECIBIDO_ESTADOS
# en ve_wh_iva.py, mismo criterio acá ahora) para "Recibido"/IOC) que
# implica que el comprobante físico SI llegó -- incluye los "_dif" porque
# el comprobante igual llegó, solo que con monto distinto.
ESTADOS_CON_COMPROBANTE = ('recibido', 'recibido_dif', 'confirmado', 'confirmado_dif')

COL_FILL = {
    'c/DIF': 'FDEBD0',
    'Solo-Excel': 'D6EAF8',
    'Solo-SmartIVA': 'FADBD8',
    'OK': 'E9F7EF',
}

CATEGORIA_ODOO_LABEL = {
    'duplicada': 'Duplicada — misma factura repetida (ver "Discrepancias" de la carga en Odoo para el detalle exacto)',
    'dato_faltante': 'Dato faltante en el archivo (sin N.Factura ni N.Control)',
    'fecha_invalida': 'Fecha no reconocida en el archivo',
    'registro_anulacion': 'Registro + Anulación (par que neta a Bs.0) — no es una discrepancia real',
    'documento_vacio': 'Documento vacío (Bs.0 en el archivo)',
    'error_posteo': 'Error al postear la factura en Odoo',
}

SIGNIFICADO = {
    'c/DIF': ('La retención existe en ambos lados, pero el monto no coincide — '
              'ver tabla "Por qué las diferencias" abajo para el motivo exacto de cada caso.'),
    'Solo-Excel': ('El archivo trae un monto de IVA Retenido para esa factura, pero en '
                   'SmartIVA no hay ninguna retención con ese RIF+N.Control/N.Factura — porque '
                   'el cliente no estaba marcado Agente de Retención al momento de facturar '
                   '(la marca puede llegar después, ej. al cargar SENIAT, sin efecto retroactivo), '
                   'o la factura nunca se creó en Odoo (fila bloqueada o con datos faltantes).'),
    'Solo-SmartIVA': ('SmartIVA sí generó la retención (el cliente está marcado Agente de '
                      'Retención), pero el archivo del cliente no trae monto de IVA Retenido '
                      'para esa fila — el cliente no la registró como retenida en su propio '
                      'Libro de Ventas, o el N.Control/N.Factura no coincide exactamente entre '
                      'ambas fuentes y no se pudieron emparejar.'),
    'OK': 'El monto coincide (o ambos lados coinciden en que no aplica retención).',
    'TOTAL': 'Suma de las 4 filas anteriores.',
}

TOL = 0.02


def _norm_rif(rif):
    return re.sub(r'[^A-Z0-9]', '', str(rif or '').upper())


def _norm_doc(doc):
    if doc is None:
        return ''
    s = str(doc).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s.upper()


def _quincena_of(d):
    return 1 if d.day <= 15 else 2


def _motivo_de(control, factura, contrib, iva_excel, iva_smartiva, estado, cat_odoo,
               base_excel=0, base_smartiva=0, wh_duplicado=False):
    """Clasifica cualquier fila no-OK (c/DIF, Solo-Excel o Solo-SmartIVA).

    Prioridad 1: si Odoo ya sabe por qué esa fila no generó una retención
    normal (categoria_discrepancia, calculado en la carga -- incluye
    'duplicada'), se usa esa razón real en vez de adivinar por los montos."""
    if cat_odoo:
        return CATEGORIA_ODOO_LABEL.get(cat_odoo, f'Categoría Odoo: {cat_odoo}')

    if wh_duplicado:
        return ('N.Control/N.Factura repetido en el archivo del cliente -- ya emparejado '
                'con otra línea de esta misma retención SmartIVA (ver esa fila para el '
                'monto real comparado); esta línea se cuenta como Solo-Excel para que '
                'esta tabla cuadre exacto contra la fila TOTAL de "Validación de carga"')

    if estado == 'Solo-SmartIVA':
        return ('SmartIVA generó la retención pero no se encontró en el archivo del '
                'cliente ninguna fila con ese RIF+N.Control/N.Factura — formato distinto '
                'entre archivo y Odoo, o la fila no está en el archivo')

    if estado == 'Solo-Excel':
        if contrib != 'S':
            return ('Factura sin match en SmartIVA y Contribuyente no es Agente de '
                    'Retención (E/N) en el archivo — consistente con que no exista retención')
        return ('Factura sin match en SmartIVA pese a Contribuyente=S — revisar caso a caso '
                '(posible formato de N.Control/N.Factura distinto entre archivo y Odoo)')

    # estado == 'c/DIF' -- ambos lados tienen retención, pero el monto difiere
    if not control and not factura:
        return ('Sin N.Control ni N.Factura — regla "100% sin comprobante" '
                '(SmartIVA aplica 100% de retención; el archivo puede traer otro %)')
    if contrib != 'S':
        return ('Contribuyente no es Agente de Retención (E/N) — '
                'SmartIVA calcula distinto a lo que trae el archivo')
    if (iva_excel or 0) == 0 and (iva_smartiva or 0) > 0:
        return ('Contribuyente S pero el archivo no registra retención — '
                'SmartIVA sí retiene')

    be, bs, ie, ismv = (base_excel or 0), (base_smartiva or 0), (iva_excel or 0), (iva_smartiva or 0)
    misma_base = abs(be - bs) < 1
    if be * bs < 0 and abs(abs(be) - abs(bs)) < 1:
        return ('Archivo con signo opuesto a SmartIVA, misma Base Imponible — probable '
                'Registro + Anulación que Odoo no emparejó automáticamente, revisar par')
    if misma_base and ie > 0 and ismv > 0 and ie < ismv * 0.5:
        return ('Archivo retiene mucho menos que SmartIVA con la misma Base Imponible — '
                'revisar si ese cliente/zona tiene un acuerdo de retención distinto al '
                '75% normal (ver columnas IVA Retenido/Esperado para el monto exacto)')
    if misma_base and ie > 0 and ismv > 0 and ie > be * 0.5:
        return ('Archivo retiene un monto cercano a toda la Base Imponible (no un % de '
                'retención real) — posible error de captura en la columna IVA Retenido '
                'del archivo del cliente')
    if not misma_base:
        return ('Base Imponible distinta entre archivo y SmartIVA pese a mismo '
                'N.Control/N.Factura — posible match incorrecto (dos facturas '
                'distintas con el mismo N° de identificación)')
    return 'Otro — revisar caso a caso'


class WizardConciliacionLibroVentas(models.TransientModel):
    _name = 've.conciliacion.libro.ventas.wizard'
    _description = 'Generar Conciliación Libro de Ventas vs SmartIVA (Excel)'

    company_id = fields.Many2one(
        'res.company', string='Compañía', required=True,
        default=lambda self: self.env.company)
    fecha_desde = fields.Date(string='Desde', required=True)
    fecha_hasta = fields.Date(string='Hasta', required=True)

    @api.model
    def default_get(self, fields_list):
        vals = super().default_get(fields_list)
        company_id = vals.get('company_id') or self.env.company.id
        Linea = self.env['ve.conecta.carga.ventas.linea']
        primera = Linea.search([('carga_id.company_id', '=', company_id), ('fecha', '!=', False)],
                                order='fecha asc', limit=1)
        ultima = Linea.search([('carga_id.company_id', '=', company_id), ('fecha', '!=', False)],
                               order='fecha desc', limit=1)
        if primera and ultima:
            vals.setdefault('fecha_desde', primera.fecha)
            vals.setdefault('fecha_hasta', ultima.fecha)
        return vals

    def _elegir_candidato(self, cands, mes_num):
        """De los wh_iva que comparten la misma clave RIF+Control/Factura,
        prefiere el que caiga en el MISMO MES que se está procesando -- ver
        scripts/demo_cementos/actualizar_conciliacion_completo.py (bug real
        2026-08-21: el cliente reutiliza N° de Factura entre facturas reales
        distintas de meses distintos)."""
        if len(cands) == 1:
            return cands[0]
        mismos_mes = [w for w in cands if w['_fecha'] and w['_fecha'].month == mes_num]
        return mismos_mes[0] if mismos_mes else cands[0]

    def action_generar(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Side
        except ImportError:
            raise UserError('La librería openpyxl no está instalada.')

        if self.fecha_desde > self.fecha_hasta:
            raise UserError('"Desde" no puede ser posterior a "Hasta".')

        company_id = self.company_id.id
        desde, hasta = self.fecha_desde, self.fecha_hasta

        Wh = self.env['ve.wh.iva']
        Move = self.env['account.move']
        Linea = self.env['ve.conecta.carga.ventas.linea']

        # ── 1) ve.wh.iva activos con factura en el rango ────────────────
        wh_records = Wh.search_read([
            ('company_id', '=', company_id), ('state', '!=', 'anulado'),
            ('invoice_id.invoice_date', '>=', desde), ('invoice_id.invoice_date', '<=', hasta),
        ], ['rif', 'nro_control', 'nro_documento', 'nro_factura', 'monto_retenido',
            'monto_recibido', 'monto_base', 'monto_base_red', 'monto_exento', 'invoice_id',
            'partner_id', 'state', 'estado_recepcion', 'name'])

        inv_ids = list({w['invoice_id'][0] for w in wh_records if w['invoice_id']})
        moves = {m['id']: m['invoice_date'] for m in Move.browse(inv_ids).read(['invoice_date'])}
        for w in wh_records:
            w['_fecha'] = moves.get(w['invoice_id'][0]) if w['invoice_id'] else None

        if not wh_records:
            raise UserError('No hay retenciones SmartIVA (ve.wh.iva) con factura en ese rango de fechas.')

        by_control = defaultdict(list)
        by_doc = defaultdict(list)
        for w in wh_records:
            if w['nro_control']:
                by_control[(_norm_rif(w['rif']), _norm_doc(w['nro_control']))].append(w)
            if w['nro_factura']:
                by_doc[(_norm_rif(w['rif']), _norm_doc(w['nro_factura']))].append(w)
            if w['nro_documento']:
                by_doc[(_norm_rif(w['rif']), _norm_doc(w['nro_documento']))].append(w)

        # ── 1bis) categoria_discrepancia de la carga -- fuente real de
        # "duplicada" y demás rechazos, ver _motivo_de() arriba.
        lineas_cat = Linea.search_read([
            '&', '|', ('categoria_discrepancia', '!=', False), ('es_anulacion_par', '=', True),
            ('eliminada_duplicado', '=', False), ('carga_id.company_id', '=', company_id),
        ], ['rif', 'nro_control', 'nro_documento', 'categoria_discrepancia', 'es_anulacion_par'])
        cat_by_control = {}
        cat_by_doc = {}
        for l in lineas_cat:
            cat = 'registro_anulacion' if l['es_anulacion_par'] else l['categoria_discrepancia']
            if l['nro_control']:
                cat_by_control[(_norm_rif(l['rif']), _norm_doc(l['nro_control']))] = cat
            if l['nro_documento']:
                cat_by_doc[(_norm_rif(l['rif']), _norm_doc(l['nro_documento']))] = cat

        def categoria_odoo(rif, control, factura):
            key_c = (_norm_rif(rif), _norm_doc(control)) if control else None
            key_d = (_norm_rif(rif), _norm_doc(factura)) if factura else None
            if key_c and key_c in cat_by_control:
                return cat_by_control[key_c]
            if key_d and key_d in cat_by_doc:
                return cat_by_doc[key_d]
            return None

        # ── 2) Líneas del Libro de Ventas (archivo del cliente, ya en Odoo
        # desde la carga original -- ver Cargar Libro de Ventas (Conecta)).
        lineas = Linea.search_read([
            ('carga_id.company_id', '=', company_id),
            ('fecha', '>=', desde), ('fecha', '<=', hasta),
        ], ['rif', 'nombre_cliente', 'zona', 'fecha', 'nro_control', 'nro_documento',
            'base_16', 'base_8', 'base_exento', 'monto_retenido', 'es_spe', 'nro_comp_retencion'])
        if not lineas:
            raise UserError('No hay líneas de Libro de Ventas cargadas en ese rango de fechas.')

        # Descubre dinámicamente qué quincenas existen en el rango (en vez
        # de una lista fija) -- agrupa por fecha calendario, igual que el
        # archivo del cliente (no por periodo_retencion de Odoo).
        lineas_por_quincena = defaultdict(list)
        for l in lineas:
            d = fields.Date.from_string(l['fecha'])
            quincena = f'{d.year}-{d.month:02d} {_quincena_of(d)}Q'
            lineas_por_quincena[quincena].append(l)
        quincenas_orden = sorted(lineas_por_quincena.keys())
        meses_orden = []
        mes_de_quincena = {}
        for q in quincenas_orden:
            anio, mes_num_str = q[:4], q[5:7]
            mes_num = int(mes_num_str)
            label = f'{MESES_NOMBRE[mes_num - 1]} {anio}'
            mes_de_quincena[q] = label
            if label not in meses_orden:
                meses_orden.append(label)

        # ── 3) Procesar cada quincena ────────────────────────────────────
        resultados = {}
        detalles = {}
        wh_contado_contrib_global = set()
        # Bug real confirmado 2026-08-22: el archivo del cliente puede traer
        # el mismo RIF+N.Control/N.Factura en 2+ líneas (ej. reutiliza el
        # N° entre facturas, o filas repetidas) mientras que del lado Odoo
        # ese wh_iva es único -- sin este set, cada línea del archivo que
        # reclama el mismo wh_iva vuelve a sumar su monto_retenido en
        # motivo_cnt, inflando "Por qué las diferencias" muy por encima de
        # iva_smartiva_total (que cuenta cada wh_iva una sola vez). Solo la
        # PRIMERA línea que reclama un wh_iva se cuenta como c/DIF real; las
        # siguientes se tratan como Solo-Excel (ver wh_duplicado abajo).
        wh_diff_ya_contado_global = set()

        for quincena in quincenas_orden:
            anio_str, mes_num_str = quincena[:4], quincena[5:7]
            mes_num = int(mes_num_str)
            anio = int(anio_str)
            q_num = 1 if '1Q' in quincena else 2

            matched_wh_ids = set()
            n_ok = n_dif = n_solo_excel = 0
            base_excel_total = 0.0
            iva_excel_total = 0.0
            detalle_filas = []
            motivo_cnt = defaultdict(lambda: [0, 0.0])
            contrib_cnt = defaultdict(lambda: [0, 0.0, 0.0, 0.0, 0.0])
            comp_excel = {'con': [0, 0.0], 'sin': [0, 0.0]}

            for l in lineas_por_quincena[quincena]:
                rif = l['rif']
                nombre = l['nombre_cliente']
                zona = l['zona']
                fecha = l['fecha']
                control = l['nro_control']
                factura = l['nro_documento']
                base_imponible = (l['base_16'] or 0) + (l['base_8'] or 0) + (l['base_exento'] or 0)
                iva_excel = l['monto_retenido'] or 0
                contrib = l['es_spe']
                comprobante = l['nro_comp_retencion']

                base_excel_total += base_imponible
                iva_excel_total += iva_excel

                if abs(iva_excel) > TOL:
                    key = 'con' if comprobante else 'sin'
                    comp_excel[key][0] += 1
                    comp_excel[key][1] += iva_excel

                wh_match = None
                if control:
                    cands = by_control.get((_norm_rif(rif), _norm_doc(control)))
                    if cands:
                        wh_match = self._elegir_candidato(cands, mes_num)
                if wh_match is None and factura:
                    cands = by_doc.get((_norm_rif(rif), _norm_doc(factura)))
                    if cands:
                        wh_match = self._elegir_candidato(cands, mes_num)

                wh_duplicado = wh_match is not None and wh_match['id'] in wh_diff_ya_contado_global
                if wh_match is not None and not wh_duplicado:
                    matched_wh_ids.add(wh_match['id'])
                    wh_diff_ya_contado_global.add(wh_match['id'])

                es_ok = False
                iva_smartiva = base_smartiva = 0.0
                if wh_match is None or wh_duplicado:
                    if abs(iva_excel) <= TOL:
                        n_ok += 1
                        es_ok = True
                    else:
                        n_solo_excel += 1
                        estado = 'Solo-Excel'
                else:
                    iva_smartiva = wh_match['monto_retenido'] or 0
                    base_smartiva = ((wh_match['monto_base'] or 0) + (wh_match['monto_base_red'] or 0)
                                      + (wh_match['monto_exento'] or 0))
                    if abs(iva_smartiva - iva_excel) <= TOL:
                        n_ok += 1
                        es_ok = True
                    else:
                        n_dif += 1
                        estado = 'c/DIF'

                cc = contrib_cnt[contrib or '(sin dato)']
                cc[0] += 1
                cc[1] += base_imponible
                cc[2] += iva_excel
                if wh_match is None or wh_match['id'] not in wh_contado_contrib_global:
                    cc[3] += base_smartiva
                    cc[4] += iva_smartiva
                    if wh_match is not None:
                        wh_contado_contrib_global.add(wh_match['id'])

                if es_ok:
                    # Pedido explícito 2026-08-22 (antes se saltaba OK acá
                    # con un `continue` -- el detalle solo mostraba
                    # problemas). Sin motivo/categoría: OK no es un
                    # problema a clasificar, motivo_cnt sigue siendo solo
                    # de las 3 categorías con diferencia real.
                    detalle_filas.append(dict(
                        estado='OK', rif=rif, nombre=nombre, zona=zona, fecha=fecha,
                        control=control, factura=factura, comprobante=(wh_match['name'] or '') if wh_match else '',
                        base_excel=base_imponible, base_smartiva=base_smartiva,
                        iva_excel=iva_excel, iva_smartiva=iva_smartiva,
                        diferencia=iva_smartiva - iva_excel, contrib=contrib, motivo='',
                    ))
                    continue

                cat_odoo = categoria_odoo(rif, control, factura)
                motivo = _motivo_de(control, factura, contrib, iva_excel, iva_smartiva, estado, cat_odoo,
                                     base_excel=base_imponible, base_smartiva=base_smartiva,
                                     wh_duplicado=wh_duplicado)
                motivo_cnt[motivo][0] += 1
                motivo_cnt[motivo][1] += (iva_smartiva - iva_excel)

                detalle_filas.append(dict(
                    estado=estado, rif=rif, nombre=nombre, zona=zona, fecha=fecha,
                    control=control, factura=factura, comprobante=(wh_match['name'] or '') if wh_match else '',
                    base_excel=base_imponible, base_smartiva=base_smartiva,
                    iva_excel=iva_excel, iva_smartiva=iva_smartiva,
                    diferencia=iva_smartiva - iva_excel, contrib=contrib, motivo=motivo,
                ))

            wh_periodo = [
                w for w in wh_records
                if w['_fecha'] and fields.Date.from_string(w['_fecha']).year == anio
                and fields.Date.from_string(w['_fecha']).month == mes_num
                and _quincena_of(fields.Date.from_string(w['_fecha'])) == q_num
            ]
            solo_smartiva_wh = [w for w in wh_periodo if w['id'] not in matched_wh_ids]
            n_solo_smartiva = len(solo_smartiva_wh)
            n_retenciones_smartiva = len(wh_periodo)
            base_smartiva_total = sum(
                (w['monto_base'] or 0) + (w['monto_base_red'] or 0) + (w['monto_exento'] or 0)
                for w in wh_periodo)
            iva_smartiva_total = sum(w['monto_retenido'] or 0 for w in wh_periodo)

            comp_smart = {'con': [0, 0.0], 'sin': [0, 0.0]}
            for w in wh_periodo:
                con = w.get('estado_recepcion') in ESTADOS_CON_COMPROBANTE
                key = 'con' if con else 'sin'
                comp_smart[key][0] += 1
                if con:
                    # monto_recibido (el REAL del comprobante físico) --
                    # pedido explícito 2026-08-22, para que cuadre exacto
                    # contra "Recibido" del Dashboard (que también usa
                    # monto_recibido, ver _serie_valor_recibido).
                    comp_smart[key][1] += w['monto_recibido'] or 0
            # Sin Comprobante = residuo (Total - Con), NO suma independiente
            # de monto_retenido. El Dashboard calcula "Faltan" del IOC como
            # Total(monto_retenido) - Recibido(monto_recibido) -- una resta
            # de campos distintos, no una partición limpia. La suma directa
            # de monto_retenido sobre los NO recibidos difiere del residuo
            # en el monto_retenido-menos-monto_recibido acumulado de los
            # comprobantes "_dif" ya recibidos. Replicar el residuo aquí
            # (mismo total/con ya verificados idénticos al Dashboard) hace
            # que Sin Comprobante cuadre exacto con "Faltan" del IOC por
            # construcción, no por coincidencia.
            comp_smart['sin'][1] = iva_smartiva_total - comp_smart['con'][1]

            for w in solo_smartiva_wh:
                factura_w = w['nro_factura'] or w['nro_documento']
                cat_odoo_sm = categoria_odoo(w['rif'], w['nro_control'], factura_w)
                motivo_sm = _motivo_de(w['nro_control'], factura_w, '', 0,
                                        w['monto_retenido'] or 0, 'Solo-SmartIVA', cat_odoo_sm)
                motivo_cnt[motivo_sm][0] += 1
                motivo_cnt[motivo_sm][1] += (w['monto_retenido'] or 0)
                detalle_filas.append(dict(
                    estado='Solo-SmartIVA', rif=w['rif'],
                    nombre=(w['partner_id'][1] if w['partner_id'] else ''),
                    zona='', fecha=w['_fecha'], control=w['nro_control'], factura=factura_w,
                    comprobante=w['name'] or '',
                    base_excel=0,
                    base_smartiva=(w['monto_base'] or 0) + (w['monto_base_red'] or 0) + (w['monto_exento'] or 0),
                    iva_excel=0, iva_smartiva=w['monto_retenido'] or 0,
                    diferencia=(w['monto_retenido'] or 0), contrib='', motivo=motivo_sm,
                ))
                cc = contrib_cnt['(sin dato)']
                cc[0] += 1
                cc[3] += (w['monto_base'] or 0) + (w['monto_base_red'] or 0) + (w['monto_exento'] or 0)
                cc[4] += w['monto_retenido'] or 0

            total_filas = n_ok + n_dif + n_solo_excel
            resultados[quincena] = dict(
                ok=n_ok, dif=n_dif, solo_excel=n_solo_excel, solo_smartiva=n_solo_smartiva,
                total_filas=total_filas, n_retenciones_smartiva=n_retenciones_smartiva,
                base_excel_total=base_excel_total, base_smartiva_total=base_smartiva_total,
                iva_excel_total=iva_excel_total, iva_smartiva_total=iva_smartiva_total,
                diferencia_total=iva_excel_total - iva_smartiva_total,
                motivo_cnt=motivo_cnt, contrib_cnt=contrib_cnt,
                comp_excel=comp_excel, comp_smart=comp_smart,
            )
            detalles[quincena] = detalle_filas

        for quincena in quincenas_orden:
            dif_rows = [d for d in detalles[quincena] if d['estado'] == 'c/DIF']
            r = resultados[quincena]
            r['dif_base_excel'] = sum(d['base_excel'] for d in dif_rows)
            r['dif_base_smartiva'] = sum(d['base_smartiva'] for d in dif_rows)
            r['dif_iva_excel'] = sum(d['iva_excel'] for d in dif_rows)
            r['dif_iva_smartiva'] = sum(d['iva_smartiva'] for d in dif_rows)

        # ── 4) Agregados del rango completo ─────────────────────────────
        agg_solo_excel_base = sum(d['base_excel'] for q in detalles for d in detalles[q] if d['estado'] == 'Solo-Excel')
        agg_solo_excel_iva = sum(d['iva_excel'] for q in detalles for d in detalles[q] if d['estado'] == 'Solo-Excel')
        agg_solo_smartiva_base = sum(d['base_smartiva'] for q in detalles for d in detalles[q] if d['estado'] == 'Solo-SmartIVA')
        agg_solo_smartiva_iva = sum(d['iva_smartiva'] for q in detalles for d in detalles[q] if d['estado'] == 'Solo-SmartIVA')

        agg = {
            'ok': sum(r['ok'] for r in resultados.values()),
            'dif': sum(r['dif'] for r in resultados.values()),
            'solo_excel': sum(r['solo_excel'] for r in resultados.values()),
            'solo_smartiva': sum(r['solo_smartiva'] for r in resultados.values()),
            'base_excel_total': sum(r['base_excel_total'] for r in resultados.values()),
            'base_smartiva_total': sum(r['base_smartiva_total'] for r in resultados.values()),
            'iva_excel_total': sum(r['iva_excel_total'] for r in resultados.values()),
            'iva_smartiva_total': sum(r['iva_smartiva_total'] for r in resultados.values()),
            'dif_base_excel': sum(r['dif_base_excel'] for r in resultados.values()),
            'dif_base_smartiva': sum(r['dif_base_smartiva'] for r in resultados.values()),
            'dif_iva_excel': sum(r['dif_iva_excel'] for r in resultados.values()),
            'dif_iva_smartiva': sum(r['dif_iva_smartiva'] for r in resultados.values()),
        }
        agg_motivo_cnt = defaultdict(lambda: [0, 0.0])
        for r in resultados.values():
            for motivo, (cnt, monto) in r['motivo_cnt'].items():
                agg_motivo_cnt[motivo][0] += cnt
                agg_motivo_cnt[motivo][1] += monto

        agg_contrib_cnt = defaultdict(lambda: [0, 0.0, 0.0, 0.0, 0.0])
        for r in resultados.values():
            for contrib, (cnt, be, ie, bs, ismv) in r['contrib_cnt'].items():
                acc = agg_contrib_cnt[contrib]
                acc[0] += cnt
                acc[1] += be
                acc[2] += ie
                acc[3] += bs
                acc[4] += ismv

        agg_comp_excel = {'con': [0, 0.0], 'sin': [0, 0.0]}
        agg_comp_smart = {'con': [0, 0.0], 'sin': [0, 0.0]}
        for r in resultados.values():
            for k in ('con', 'sin'):
                agg_comp_excel[k][0] += r['comp_excel'][k][0]
                agg_comp_excel[k][1] += r['comp_excel'][k][1]
                agg_comp_smart[k][0] += r['comp_smart'][k][0]
                agg_comp_smart[k][1] += r['comp_smart'][k][1]

        agg_mes = {}
        for mes in meses_orden:
            quincenas_del_mes = [q for q in quincenas_orden if mes_de_quincena[q] == mes]
            rs = [resultados[q] for q in quincenas_del_mes]
            agg_mes[mes] = dict(
                ok=sum(r['ok'] for r in rs), dif=sum(r['dif'] for r in rs),
                solo_excel=sum(r['solo_excel'] for r in rs), solo_smartiva=sum(r['solo_smartiva'] for r in rs),
                total_filas=sum(r['total_filas'] for r in rs),
                n_retenciones_smartiva=sum(r['n_retenciones_smartiva'] for r in rs),
                base_excel_total=sum(r['base_excel_total'] for r in rs),
                base_smartiva_total=sum(r['base_smartiva_total'] for r in rs),
                iva_excel_total=sum(r['iva_excel_total'] for r in rs),
                iva_smartiva_total=sum(r['iva_smartiva_total'] for r in rs),
                diferencia_total=sum(r['diferencia_total'] for r in rs),
            )

        # ── 5) Escribir workbook ─────────────────────────────────────────
        TITLE_FONT = Font(bold=True, size=13)
        SECTION_FONT = Font(bold=True)
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        HEADER_FILL = PatternFill('solid', fgColor='383A4E')
        HEADER_ALIGN = Alignment(horizontal='center', wrap_text=True)
        GROUP_FONT = Font(bold=True, color='FFFFFF', italic=True)
        GROUP_FILL_EXCEL = PatternFill('solid', fgColor='5B6169')
        GROUP_FILL_SMARTIVA = PatternFill('solid', fgColor='669999')
        GROUP_ALIGN = Alignment(horizontal='center')
        TOTAL_FONT = Font(bold=True)
        TOTAL_FILL = PatternFill('solid', fgColor='F2E7CC')
        THIN = Side(style='thin')
        BORDER_BOTTOM = openpyxl.styles.Border(bottom=THIN)
        MONEY_FMT = '#,##0.00'
        DATE_FMT = 'DD/MM/YYYY'

        def title_group_row(ws_, row, excel_cols, smartiva_cols):
            for label, (c0, c1), fill in (
                ('Libro de Ventas (Excel)', excel_cols, GROUP_FILL_EXCEL),
                ('SmartIVA', smartiva_cols, GROUP_FILL_SMARTIVA),
            ):
                cell = ws_.cell(row=row, column=c0, value=label)
                cell.font = GROUP_FONT
                cell.alignment = GROUP_ALIGN
                if c1 > c0:
                    ws_.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
                for c in range(c0, c1 + 1):
                    ws_.cell(row=row, column=c).fill = fill

        def write_header_row(ws_, row, headers):
            for c, h in enumerate(headers, start=1):
                cell = ws_.cell(row=row, column=c, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # -- Resumen --
        ws = wb.create_sheet('Resumen')
        ws['A1'] = f'Conciliación Libro de Ventas vs SmartIVA — {desde} a {hasta} ({self.company_id.name})'
        ws['A1'].font = TITLE_FONT
        ws['A2'] = ('RIF+N.Control (N1); si no hay Control, RIF+N.Factura (N2). Compara '
                    'IVA Retenido (columna del archivo del cliente) vs IVA Esperado '
                    '(calculado por SmartIVA). Tolerancia Bs 0,02. En toda tabla, columnas '
                    'del Libro de Ventas (Excel) a la izquierda, columnas de SmartIVA a la derecha.')

        row_i = 4
        ws.cell(row=row_i, column=1, value='Desglose por Tipo de Contribuyente').font = SECTION_FONT
        row_i += 1
        title_group_row(ws, row_i, (3, 4), (5, 6))
        row_i += 1
        write_header_row(ws, row_i, ['Contribuyente', 'Cantidad', 'Base Imponible', 'IVA Retenido',
                                      'Base Imponible', 'IVA Esperado', 'Diferencia'])
        row_i += 1
        first_contrib_row = row_i
        for contrib, (cnt, be, ie, bs, ismv) in sorted(agg_contrib_cnt.items(), key=lambda kv: -kv[1][0]):
            vals = [contrib, cnt, round(be, 2), round(ie, 2), round(bs, 2), round(ismv, 2), round(ismv - ie, 2)]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c, value=v)
                if c in (3, 4, 5, 6, 7):
                    cell.number_format = MONEY_FMT
            row_i += 1
        last_contrib_row = row_i - 1
        ws.cell(row=row_i, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 8):
            ws.cell(row=row_i, column=c).fill = TOTAL_FILL
        ws.cell(row=row_i, column=2, value=f'=SUM(B{first_contrib_row}:B{last_contrib_row})').font = TOTAL_FONT
        for c in (3, 4, 5, 6, 7):
            col = openpyxl.utils.get_column_letter(c)
            cell = ws.cell(row=row_i, column=c, value=f'=SUM({col}{first_contrib_row}:{col}{last_contrib_row})')
            cell.number_format = MONEY_FMT
            cell.font = TOTAL_FONT
        row_i += 2

        ws.cell(row=row_i, column=1, value='Validación de carga (todas las quincenas)').font = SECTION_FONT
        row_i += 1
        title_group_row(ws, row_i, (3, 4), (5, 6))
        row_i += 1
        write_header_row(ws, row_i, ['Estado', 'Cantidad', 'Base Imponible', 'IVA Retenido',
                                      'Base Imponible', 'IVA Esperado', 'Diferencia', 'Significado'])
        row_i += 1
        first_estado_row = row_i
        filas_estado = [
            ('c/DIF', agg['dif'], agg['dif_base_excel'], agg['dif_iva_excel'], agg['dif_base_smartiva'], agg['dif_iva_smartiva']),
            ('Solo-Excel', agg['solo_excel'], agg_solo_excel_base, agg_solo_excel_iva, 0, 0),
            ('Solo-SmartIVA', agg['solo_smartiva'], 0, 0, agg_solo_smartiva_base, agg_solo_smartiva_iva),
            ('OK', agg['ok'],
             agg['base_excel_total'] - agg['dif_base_excel'] - agg_solo_excel_base,
             agg['iva_excel_total'] - agg['dif_iva_excel'] - agg_solo_excel_iva,
             agg['base_smartiva_total'] - agg['dif_base_smartiva'] - agg_solo_smartiva_base,
             agg['iva_smartiva_total'] - agg['dif_iva_smartiva'] - agg_solo_smartiva_iva),
        ]
        for estado, cant, be, ie, bs, ismv in filas_estado:
            vals = [estado, cant, round(be, 2), round(ie, 2), round(bs, 2), round(ismv, 2), round(ismv - ie, 2),
                    SIGNIFICADO[estado]]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c, value=v)
                cell.fill = PatternFill('solid', fgColor=COL_FILL[estado])
                if c in (3, 4, 5, 6, 7):
                    cell.number_format = MONEY_FMT
            row_i += 1
        last_estado_row = row_i - 1
        total_row = row_i
        ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 9):
            ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=2, value=f'=SUM(B{first_estado_row}:B{last_estado_row})').font = TOTAL_FONT
        ws.cell(row=total_row, column=3, value=round(agg['base_excel_total'], 2)).number_format = MONEY_FMT
        ws.cell(row=total_row, column=4, value=round(agg['iva_excel_total'], 2)).number_format = MONEY_FMT
        ws.cell(row=total_row, column=5, value=round(agg['base_smartiva_total'], 2)).number_format = MONEY_FMT
        ws.cell(row=total_row, column=6, value=round(agg['iva_smartiva_total'], 2)).number_format = MONEY_FMT
        ws.cell(row=total_row, column=7, value=round(agg['iva_smartiva_total'] - agg['iva_excel_total'], 2)).number_format = MONEY_FMT
        ws.cell(row=total_row, column=8, value=SIGNIFICADO['TOTAL'])
        for c in range(3, 8):
            ws.cell(row=total_row, column=c).font = TOTAL_FONT
        row_i = total_row + 1
        ws.cell(row=row_i, column=1, value='% variación').font = TOTAL_FONT
        c_pct = ws.cell(row=row_i, column=5, value=f'=IFERROR((E{total_row}-C{total_row})/C{total_row},"")')
        c_pct.number_format = '0.0%'
        c_pct.font = TOTAL_FONT
        c_pct2 = ws.cell(row=row_i, column=7, value=f'=IFERROR(G{total_row}/D{total_row},"")')
        c_pct2.number_format = '0.0%'
        c_pct2.font = TOTAL_FONT
        row_i += 2

        ws.cell(row=row_i, column=1,
                value='Por qué las diferencias (c/DIF, Solo-Excel, Solo-SmartIVA) — cantidad y monto').font = SECTION_FONT
        row_i += 1
        write_header_row(ws, row_i, ['Cantidad', 'Monto Diferencia (Retención)', 'Motivo'])
        row_i += 1
        agg_total_cnt_motivo = sum(cnt for cnt, _m in agg_motivo_cnt.values())
        agg_total_monto_motivo = sum(monto for _c, monto in agg_motivo_cnt.values())
        for motivo, (cnt, monto) in sorted(agg_motivo_cnt.items(), key=lambda kv: -kv[1][0]):
            ws.cell(row=row_i, column=1, value=cnt).fill = PatternFill('solid', fgColor='FDEBD0')
            c2 = ws.cell(row=row_i, column=2, value=round(monto, 2))
            c2.number_format = MONEY_FMT
            c2.fill = PatternFill('solid', fgColor='FDEBD0')
            ws.cell(row=row_i, column=3, value=motivo).fill = PatternFill('solid', fgColor='FDEBD0')
            row_i += 1
        ws.cell(row=row_i, column=1, value=agg_total_cnt_motivo).font = TOTAL_FONT
        ws.cell(row=row_i, column=1).fill = TOTAL_FILL
        c2 = ws.cell(row=row_i, column=2, value=round(agg_total_monto_motivo, 2))
        c2.number_format = MONEY_FMT
        c2.font = TOTAL_FONT
        c2.fill = TOTAL_FILL
        ws.cell(row=row_i, column=3, value='TOTAL').font = TOTAL_FONT
        ws.cell(row=row_i, column=3).fill = TOTAL_FILL
        row_i += 2

        ws.cell(row=row_i, column=1, value='Retenido con y sin Comprobante').font = SECTION_FONT
        row_i += 1
        title_group_row(ws, row_i, (2, 3), (4, 5))
        row_i += 1
        write_header_row(ws, row_i, ['Estado', 'Cantidad', 'IVA Retenido', 'Cantidad', 'IVA Real/Esperado'])
        row_i += 1
        first_comp_row = row_i
        # "Con Comprobante" muestra el monto REAL (monto_recibido, igual
        # que "Recibido" del Dashboard); "Sin Comprobante" es el residuo
        # Total-Con (igual que "Faltan" del IOC del Dashboard), no una
        # suma directa de monto_retenido -- ver comentario en comp_smart
        # más arriba. Etiqueta de cada fila lo aclara a propósito
        # (pedido explícito 2026-08-22).
        for label, key in (('Con Comprobante (real)', 'con'), ('Sin Comprobante (esperado)', 'sin')):
            ce = agg_comp_excel[key]
            cs = agg_comp_smart[key]
            vals = [label, ce[0], round(ce[1], 2), cs[0], round(cs[1], 2)]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c, value=v)
                if c in (3, 5):
                    cell.number_format = MONEY_FMT
            row_i += 1
        last_comp_row = row_i - 1
        ws.cell(row=row_i, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 6):
            ws.cell(row=row_i, column=c).fill = TOTAL_FILL
        for c in (2, 3, 4, 5):
            col = openpyxl.utils.get_column_letter(c)
            cell = ws.cell(row=row_i, column=c, value=f'=SUM({col}{first_comp_row}:{col}{last_comp_row})')
            cell.font = TOTAL_FONT
            if c in (3, 5):
                cell.number_format = MONEY_FMT
        row_i += 2

        ws.cell(row=row_i, column=1, value='Por mes').font = SECTION_FONT
        row_i += 1
        title_group_row(ws, row_i, (2, 4), (5, 7))
        row_i += 1
        write_header_row(ws, row_i, ['Mes', 'Total Filas', 'Base Imponible', 'IVA Retenido',
                                      'Total Filas (N. Retenciones)', 'Base Imponible', 'IVA Esperado',
                                      'Diferencia', '% Diferencia', 'OK', 'c/DIF', 'Solo-Excel', 'Solo-SmartIVA'])
        row_i += 1
        first_mes_row = row_i
        for mes in meses_orden:
            m = agg_mes[mes]
            vals = [mes, m['total_filas'], round(m['base_excel_total'], 2), round(m['iva_excel_total'], 2),
                    m['n_retenciones_smartiva'], round(m['base_smartiva_total'], 2), round(m['iva_smartiva_total'], 2),
                    round(m['diferencia_total'], 2), None, m['ok'], m['dif'], m['solo_excel'], m['solo_smartiva']]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c, value=v)
                if c in (3, 4, 6, 7, 8):
                    cell.number_format = MONEY_FMT
            pct_cell = ws.cell(row=row_i, column=9, value=f'=IFERROR(H{row_i}/D{row_i},"")')
            pct_cell.number_format = '0.0%'
            row_i += 1
        last_mes_row = row_i - 1
        mes_total_row = row_i
        ws.cell(row=mes_total_row, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 14):
            ws.cell(row=mes_total_row, column=c).fill = TOTAL_FILL
        for c in (2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13):
            col = openpyxl.utils.get_column_letter(c)
            cell = ws.cell(row=mes_total_row, column=c, value=f'=SUM({col}{first_mes_row}:{col}{last_mes_row})')
            cell.font = TOTAL_FONT
            if c in (3, 4, 6, 7, 8):
                cell.number_format = MONEY_FMT
        pct_total = ws.cell(row=mes_total_row, column=9, value=f'=IFERROR(H{mes_total_row}/D{mes_total_row},"")')
        pct_total.number_format = '0.0%'
        pct_total.font = TOTAL_FONT

        ws.column_dimensions['A'].width = 16
        for col in 'BCDEFGHIJKLM':
            ws.column_dimensions[col].width = 15

        # -- Pestañas por quincena --
        HEADERS_VALID = ['Estado', 'Cantidad', 'Base Imponible', 'IVA Retenido',
                          'Base Imponible', 'IVA Esperado', 'Diferencia']
        HEADERS_MOTIVO = ['Cantidad', 'Monto Diferencia (Retención)', 'Motivo']
        HEADERS_DET = ['Estado', 'RIF', 'Nombre', 'Zona', 'Fecha', 'N.Control', 'N.Factura', 'N.Comprobante',
                        'Base Imponible', 'IVA Retenido', 'Base Imponible', 'IVA Esperado',
                        'Diferencia', 'Contribuyente (Excel)', 'Motivo (si c/DIF)']
        COL_WIDTHS_DET = [11, 16, 55, 16, 12, 16, 16, 16, 16, 16, 16, 16, 14, 14, 55]

        for quincena in quincenas_orden:
            ws_q = wb.create_sheet(quincena[:31])
            r = resultados[quincena]

            ws_q['A1'] = f'Conciliación Excel vs SmartIVA — {quincena}'
            ws_q['A1'].font = TITLE_FONT

            row_i = 3
            ws_q.cell(row=row_i, column=1, value='Validación de carga — filas por estado').font = SECTION_FONT
            row_i += 1
            title_group_row(ws_q, row_i, (3, 4), (5, 6))
            row_i += 1
            write_header_row(ws_q, row_i, HEADERS_VALID)
            row_i += 1
            first_estado_row_q = row_i
            se_excel = sum(d['base_excel'] for d in detalles[quincena] if d['estado'] == 'Solo-Excel')
            se_iva = sum(d['iva_excel'] for d in detalles[quincena] if d['estado'] == 'Solo-Excel')
            ss_base = sum(d['base_smartiva'] for d in detalles[quincena] if d['estado'] == 'Solo-SmartIVA')
            ss_iva = sum(d['iva_smartiva'] for d in detalles[quincena] if d['estado'] == 'Solo-SmartIVA')
            filas_estado_q = [
                ('c/DIF', r['dif'], r['dif_base_excel'], r['dif_iva_excel'], r['dif_base_smartiva'], r['dif_iva_smartiva']),
                ('Solo-Excel', r['solo_excel'], se_excel, se_iva, 0, 0),
                ('Solo-SmartIVA', r['solo_smartiva'], 0, 0, ss_base, ss_iva),
                ('OK', r['ok'],
                 r['base_excel_total'] - r['dif_base_excel'] - se_excel,
                 r['iva_excel_total'] - r['dif_iva_excel'] - se_iva,
                 r['base_smartiva_total'] - r['dif_base_smartiva'] - ss_base,
                 r['iva_smartiva_total'] - r['dif_iva_smartiva'] - ss_iva),
            ]
            for estado, cant, be, ie, bs, ismv in filas_estado_q:
                vals = [estado, cant, round(be, 2), round(ie, 2), round(bs, 2), round(ismv, 2), round(ismv - ie, 2)]
                for c, v in enumerate(vals, start=1):
                    cell = ws_q.cell(row=row_i, column=c, value=v)
                    cell.fill = PatternFill('solid', fgColor=COL_FILL[estado])
                    if c >= 3:
                        cell.number_format = MONEY_FMT
                row_i += 1
            last_estado_row_q = row_i - 1
            total_row_q = row_i
            ws_q.cell(row=total_row_q, column=1, value='TOTAL').font = TOTAL_FONT
            for c in range(1, 8):
                ws_q.cell(row=total_row_q, column=c).fill = TOTAL_FILL
            ws_q.cell(row=total_row_q, column=2, value=f'=SUM(B{first_estado_row_q}:B{last_estado_row_q})').font = TOTAL_FONT
            ws_q.cell(row=total_row_q, column=3, value=round(r['base_excel_total'], 2)).number_format = MONEY_FMT
            ws_q.cell(row=total_row_q, column=4, value=round(r['iva_excel_total'], 2)).number_format = MONEY_FMT
            ws_q.cell(row=total_row_q, column=5, value=round(r['base_smartiva_total'], 2)).number_format = MONEY_FMT
            ws_q.cell(row=total_row_q, column=6, value=round(r['iva_smartiva_total'], 2)).number_format = MONEY_FMT
            ws_q.cell(row=total_row_q, column=7, value=round(r['iva_smartiva_total'] - r['iva_excel_total'], 2)).number_format = MONEY_FMT
            for c in range(3, 8):
                ws_q.cell(row=total_row_q, column=c).font = TOTAL_FONT
            row_i = total_row_q + 2

            ws_q.cell(row=row_i, column=1,
                      value='Por qué las diferencias (c/DIF, Solo-Excel, Solo-SmartIVA) — cantidad y monto').font = SECTION_FONT
            row_i += 1
            write_header_row(ws_q, row_i, HEADERS_MOTIVO)
            row_i += 1
            motivo_cnt_q = r['motivo_cnt']
            total_cnt_motivo_q = sum(cnt for cnt, _m in motivo_cnt_q.values())
            total_monto_motivo_q = sum(monto for _c, monto in motivo_cnt_q.values())
            for motivo, (cnt, monto) in sorted(motivo_cnt_q.items(), key=lambda kv: -kv[1][0]):
                ws_q.cell(row=row_i, column=1, value=cnt).fill = PatternFill('solid', fgColor='FDEBD0')
                c2 = ws_q.cell(row=row_i, column=2, value=round(monto, 2))
                c2.number_format = MONEY_FMT
                c2.fill = PatternFill('solid', fgColor='FDEBD0')
                ws_q.cell(row=row_i, column=3, value=motivo).fill = PatternFill('solid', fgColor='FDEBD0')
                row_i += 1
            ws_q.cell(row=row_i, column=1, value=total_cnt_motivo_q).font = TOTAL_FONT
            ws_q.cell(row=row_i, column=1).fill = TOTAL_FILL
            c2 = ws_q.cell(row=row_i, column=2, value=round(total_monto_motivo_q, 2))
            c2.number_format = MONEY_FMT
            c2.font = TOTAL_FONT
            c2.fill = TOTAL_FILL
            ws_q.cell(row=row_i, column=3, value='TOTAL').font = TOTAL_FONT
            ws_q.cell(row=row_i, column=3).fill = TOTAL_FILL
            row_i += 2

            ws_q.cell(row=row_i, column=1, value='Retenido con y sin Comprobante').font = SECTION_FONT
            row_i += 1
            title_group_row(ws_q, row_i, (2, 3), (4, 5))
            row_i += 1
            write_header_row(ws_q, row_i, ['Estado', 'Cantidad', 'IVA Retenido', 'Cantidad', 'IVA Real/Esperado'])
            row_i += 1
            first_comp_row_q = row_i
            for label, key in (('Con Comprobante (real)', 'con'), ('Sin Comprobante (esperado)', 'sin')):
                ce = r['comp_excel'][key]
                cs = r['comp_smart'][key]
                vals = [label, ce[0], round(ce[1], 2), cs[0], round(cs[1], 2)]
                for c, v in enumerate(vals, start=1):
                    cell = ws_q.cell(row=row_i, column=c, value=v)
                    if c in (3, 5):
                        cell.number_format = MONEY_FMT
                row_i += 1
            last_comp_row_q2 = row_i - 1
            ws_q.cell(row=row_i, column=1, value='TOTAL').font = TOTAL_FONT
            for c in range(1, 6):
                ws_q.cell(row=row_i, column=c).fill = TOTAL_FILL
            for c in (2, 3, 4, 5):
                col = openpyxl.utils.get_column_letter(c)
                cell = ws_q.cell(row=row_i, column=c, value=f'=SUM({col}{first_comp_row_q}:{col}{last_comp_row_q2})')
                cell.font = TOTAL_FONT
                if c in (3, 5):
                    cell.number_format = MONEY_FMT
            row_i += 2

            ws_q.cell(row=row_i, column=1, value='Detalle fila por fila (incluye OK)').font = SECTION_FONT
            row_i += 1
            title_group_row(ws_q, row_i, (9, 10), (11, 12))
            row_i += 1
            write_header_row(ws_q, row_i, HEADERS_DET)
            row_i += 1
            for d in detalles[quincena]:
                vals = [d['estado'], d['rif'], d['nombre'], d['zona'], d['fecha'], d['control'], d['factura'],
                        d['comprobante'],
                        round(d['base_excel'], 2), round(d['iva_excel'], 2),
                        round(d['base_smartiva'], 2), round(d['iva_smartiva'], 2), round(d['diferencia'], 2),
                        d['contrib'], d['motivo']]
                fill = PatternFill('solid', fgColor=COL_FILL[d['estado']])
                for c, v in enumerate(vals, start=1):
                    cell = ws_q.cell(row=row_i, column=c, value=v)
                    cell.fill = fill
                    cell.border = BORDER_BOTTOM
                    if c in (9, 10, 11, 12, 13):
                        cell.number_format = MONEY_FMT
                    if c == 5 and v:
                        cell.number_format = DATE_FMT
                row_i += 1

            for c, w in enumerate(COL_WIDTHS_DET, start=1):
                ws_q.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        # -- Consolidado --
        ws_c = wb.create_sheet('Consolidado')
        ws_c['A1'] = f'Consolidado — todas las quincenas (incluye OK), {desde} a {hasta}'
        ws_c['A1'].font = TITLE_FONT
        ws_c['A2'] = ('Misma clasificación que cada pestaña de quincena, con columna Quincena agregada. '
                      'Usar el autofiltro para ubicar una categoría (ej. Motivo = Duplicada, o Estado = OK) '
                      'en todo el rango.')

        HEADERS_CONS = ['Quincena'] + HEADERS_DET
        row_i = 4
        title_group_row(ws_c, row_i, (10, 11), (12, 13))
        row_i += 1
        write_header_row(ws_c, row_i, HEADERS_CONS)
        header_row_c = row_i
        row_i += 1
        for quincena in quincenas_orden:
            for d in detalles[quincena]:
                vals = [quincena, d['estado'], d['rif'], d['nombre'], d['zona'], d['fecha'], d['control'], d['factura'],
                        d['comprobante'],
                        round(d['base_excel'], 2), round(d['iva_excel'], 2),
                        round(d['base_smartiva'], 2), round(d['iva_smartiva'], 2), round(d['diferencia'], 2),
                        d['contrib'], d['motivo']]
                fill = PatternFill('solid', fgColor=COL_FILL[d['estado']])
                for c, v in enumerate(vals, start=1):
                    cell = ws_c.cell(row=row_i, column=c, value=v)
                    cell.fill = fill
                    cell.border = BORDER_BOTTOM
                    if c in (10, 11, 12, 13, 14):
                        cell.number_format = MONEY_FMT
                    if c == 6 and v:
                        cell.number_format = DATE_FMT
                row_i += 1
        last_row_c = row_i - 1
        ws_c.auto_filter.ref = f'A{header_row_c}:{openpyxl.utils.get_column_letter(len(HEADERS_CONS))}{last_row_c}'
        ws_c.freeze_panes = f'A{header_row_c + 1}'
        ws_c.column_dimensions['A'].width = 11
        for c, w in enumerate(COL_WIDTHS_DET, start=2):
            ws_c.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        # Orden: Resumen, Consolidado, quincenas en orden cronológico.
        orden = ['Resumen', 'Consolidado'] + [q[:31] for q in quincenas_orden]
        wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 999)

        output = io.BytesIO()
        wb.save(output)
        fname = f'Conciliacion_Libro_Ventas_SmartIVA_{desde}_a_{hasta}.xlsx'.replace(' ', '_')
        att = self.env['ir.attachment'].create({
            'name': fname,
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
