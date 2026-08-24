import base64
import io
import re
from collections import defaultdict

from odoo import fields, models
from odoo.exceptions import UserError

MESES_NOMBRE = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
MESES_EN = ['January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December']


def _mes_en(mes_es):
    """'Enero 2026' -> 'January 2026' -- solo para la pestaña "Result" (en
    inglés, pedido explícito); el resto del reporte se queda en español."""
    nombre, anio = mes_es.rsplit(' ', 1)
    return f'{MESES_EN[MESES_NOMBRE.index(nombre)]} {anio}'


TOL = 0.02

ESTADOS_CON_COMPROBANTE = ('recibido', 'recibido_dif', 'confirmado', 'confirmado_dif')


class WizardReporteEjecutivo(models.TransientModel):
    _name = 've.reporte.ejecutivo.wizard'
    _description = 'Generar Reporte Ejecutivo SmartIVA (Excel)'

    company_id = fields.Many2one(
        'res.company', string='Compañía', required=True,
        default=lambda self: self.env.company)

    def action_generar(self):
        """Genera "Reporte Ejecutivo SmartIVA.xlsx" -- reemplaza el pipeline
        local (scripts/demo_cementos/rebuild_consolidado_resultado_vencement.py)
        con una versión nativa dentro de Odoo, para no depender de correr un
        script aparte cada vez (pedido explícito 2026-08-22). 4 pestañas:

        Pestañas/columnas del Excel en inglés (pedido explícito 2026-08-22),
        nombre de archivo fijo "SmartIVA Executive Summary - Vencement 2026",
        y ya no pregunta rango de fechas -- cubre siempre TODOS los períodos
        cargados de la compañía:

        - "Result": espejo del Dashboard -- tabla Cantidad/Monto por Mes de
          las 5 series Esperadas/Recibido/Declarado/SENIAT/Conciliado,
          gráfico de barras SOLO acumulado, brecha SENIAT sin match e
          IOC/TAC/BDS, calculados con LOS MISMOS métodos que usa
          ve.dashboard.iva (nunca una fórmula paralela).
        - "Monthly Totals": Excel (archivo) vs SmartIVA por mes (mismo
          criterio que el script local), con una sección SENIAT agregada
          a la derecha (mismas columnas de la vieja pestaña SENIAT,
          agregadas por mes en vez de por quincena).
        - "Declared Withholdings": ve.declarado.mensual, agrupado por año.
        - "Consolidated": universo completo de ve.conecta.carga.ventas.linea
          (todas las líneas cargadas, incluye duplicados/vacíos/bloqueantes
          -- es el archivo del cliente tal cual, no un resultado filtrado).
        """
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Side
            from openpyxl.chart import BarChart, Series, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.chart.marker import DataPoint
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError('La librería openpyxl no está instalada.')

        # Pedido explícito 2026-08-22: este reporte no pregunta rango de
        # fechas -- siempre cubre TODOS los períodos cargados de la
        # compañía (hoy Ene-Jun 2026 en Vencement; se extiende solo con
        # cargar más períodos, sin tocar código).
        company = self.company_id
        company_id = company.id
        Periodo = self.env['ve.conciliacion.periodo']
        periodos = Periodo.search([
            ('company_id', '=', company_id),
        ], order='periodo_retencion asc')
        if not periodos:
            raise UserError('No hay períodos de Conciliación SENIAT cargados para esta compañía.')
        desde = periodos[0].periodo_retencion
        hasta = periodos[-1].periodo_retencion

        quincenas_orden = periodos.mapped('periodo_retencion')
        mes_de_quincena = {}
        meses_orden = []
        for p in periodos:
            m = re.match(r'^(\d{4})-(\d{2}) \dQ$', p.periodo_retencion)
            if not m:
                continue
            anio, mes_num = m.group(1), int(m.group(2))
            label = f'{MESES_NOMBRE[mes_num - 1]} {anio}'
            mes_de_quincena[p.periodo_retencion] = label
            if label not in meses_orden:
                meses_orden.append(label)
        periodos_por_mes = defaultdict(lambda: Periodo)
        for p in periodos:
            mes = mes_de_quincena.get(p.periodo_retencion)
            if mes:
                periodos_por_mes[mes] |= p

        # ── Estilos ──────────────────────────────────────────────────────
        TITLE_FONT = Font(bold=True, size=13)
        SECTION_FONT = Font(bold=True)
        SUB_FONT = Font(italic=True, size=9, color='5B6169')
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        HEADER_FILL = PatternFill('solid', fgColor='383A4E')
        HEADER_ALIGN = Alignment(horizontal='center', wrap_text=True)
        GROUP_FONT = Font(bold=True, color='FFFFFF', italic=True)
        GROUP_FILL_EXCEL = PatternFill('solid', fgColor='5B6169')
        GROUP_FILL_SMARTIVA = PatternFill('solid', fgColor='669999')
        GROUP_FILL_SENIAT = PatternFill('solid', fgColor='7B4B94')
        GROUP_ALIGN = Alignment(horizontal='center')
        TOTAL_FONT = Font(bold=True)
        TOTAL_FILL = PatternFill('solid', fgColor='F2E7CC')
        MONEY_FMT = '#,##0.00'
        PCT_FMT = '0.0%'

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ══════════════════════════════════════════════════════════════
        # Sheet "Result" -- espejo del Dashboard
        # ══════════════════════════════════════════════════════════════
        Dash = self.env['ve.dashboard.iva']
        Declarado = self.env['ve.declarado.mensual']

        filas_result = []  # (mes, est_cnt, est_m, rec_cnt, rec_m, dec_cnt, dec_m, sen_cnt, sen_m, con_cnt, con_m)
        for mes in meses_orden:
            qs = periodos_por_mes[mes]
            est_m = round(sum(Dash._serie_valor_estimado(p) for p in qs), 2)
            est_cnt = sum(Dash._serie_cantidad_estimado(p) for p in qs)
            rec_m = round(sum(Dash._serie_valor_recibido(p) for p in qs), 2)
            rec_cnt = sum(Dash._serie_cantidad_recibido(p) for p in qs)
            con_m = round(sum(Dash._serie_valor_conciliado(p) for p in qs), 2)
            con_cnt = sum(Dash._serie_cantidad_conciliado(p) for p in qs)
            sen_m = round(sum(qs.mapped('total_seniat')), 2)
            sen_cnt = sum(qs.mapped('n_seniat'))
            if company.ve_declarado_manual:
                m_match = re.match(r'^(\d{4})-(\d{2}) \dQ$', qs[0].periodo_retencion) if qs else None
                anio_i = int(mes.split()[-1])
                mes_i = MESES_NOMBRE.index(mes.split()[0]) + 1
                rec_decl = Declarado.search([
                    ('company_id', '=', company_id), ('anio', '=', anio_i), ('mes', '=', mes_i),
                ], limit=1)
                dec_m = round(rec_decl.monto_declarado, 2) if rec_decl else 0.0
            else:
                dec_m = round(sum(Dash._serie_valor_declarado_auto(p) for p in qs), 2)
            filas_result.append((mes, est_cnt, est_m, rec_cnt, rec_m, None, dec_m, sen_cnt, sen_m, con_cnt, con_m))

        base_ytd = sum(f[2] for f in filas_result)
        recibido_ytd = sum(f[4] for f in filas_result)
        declarado_ytd = sum(f[6] for f in filas_result)
        seniat_ytd = sum(f[8] for f in filas_result)
        base_cnt_ytd = sum(f[1] for f in filas_result)
        recibido_cnt_ytd = sum(f[3] for f in filas_result)
        seniat_cnt_ytd = sum(f[7] for f in filas_result)

        sin_match_seniat, total_seniat_confirmado = Dash._solo_seniat_sin_match_bs()
        con_match_seniat = total_seniat_confirmado - sin_match_seniat
        pct_sin_match_seniat = (sin_match_seniat / total_seniat_confirmado) if total_seniat_confirmado else 0.0

        ws_r = wb.create_sheet('Result')
        ws_r['A1'] = f'SmartIVA Dashboard — Result — {desde} to {hasta} ({company.name})'
        ws_r['A1'].font = TITLE_FONT
        ws_r['A2'] = ('Mirrors the SmartIVA Dashboard: same computation methods '
                      '(ve.dashboard.iva), so these figures always match what is shown '
                      'on screen there.')
        row = 4
        ws_r.cell(row=row, column=1, value='Monthly Summary (Dashboard)').font = SECTION_FONT
        row += 1
        headers_r = ['Month',
                     'Qty Expected', 'Amount Expected',
                     'Qty Received', 'Amount Received',
                     'Qty Declared', 'Amount Declared',
                     'Qty SENIAT', 'Amount SENIAT',
                     'Qty Reconciled', 'Amount Reconciled']
        for c, h in enumerate(headers_r, start=1):
            cell = ws_r.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        header_row_r = row
        row += 1
        first_row_r = row
        for mes, est_cnt, est_m, rec_cnt, rec_m, dec_cnt, dec_m, sen_cnt, sen_m, con_cnt, con_m in filas_result:
            vals = [_mes_en(mes), est_cnt, est_m, rec_cnt, rec_m,
                    (dec_cnt if dec_cnt is not None else 'N/A'), dec_m,
                    sen_cnt, sen_m, con_cnt, con_m]
            for c, v in enumerate(vals, start=1):
                cell = ws_r.cell(row=row, column=c, value=v)
                if c in (3, 5, 7, 9, 11):
                    cell.number_format = MONEY_FMT
            row += 1
        last_row_r = row - 1
        ws_r.cell(row=row, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 12):
            ws_r.cell(row=row, column=c).fill = TOTAL_FILL
        for c in (2, 3, 4, 5, 7, 8, 9, 10, 11):
            col_letter = get_column_letter(c)
            cell = ws_r.cell(row=row, column=c, value=f'=SUM({col_letter}{first_row_r}:{col_letter}{last_row_r})')
            cell.font = TOTAL_FONT
            if c in (3, 5, 7, 9, 11):
                cell.number_format = MONEY_FMT
        ws_r.cell(row=row, column=6, value='N/A').font = TOTAL_FONT
        total_row_r = row
        row += 2

        # Bloque fuente del gráfico -- pedido explícito 2026-08-22: el
        # gráfico debe ser SOLO el acumulado (5 barras, una por métrica,
        # igual que RESUMEN Ene-Jun 2026 del Dashboard -- _resumen_ytd_bars_
        # html), no un gráfico agrupado por mes. Cada valor es una fórmula
        # que apunta a la fila TOTAL de arriba (=C{total_row_r} etc.), así
        # nunca puede desalinearse de "la tabla que le precede".
        chart_header_row = row
        chart_value_row = row + 1
        chart_cols = [('Expected', 3), ('Received', 5), ('Declared', 7), ('SENIAT', 9), ('Reconciled', 11)]
        colores_chart = ['b8860b', '1baf7a', '4a3aa7', 'eb6834', 'c23b7e']
        for i, (nombre, col_total) in enumerate(chart_cols):
            col_letter_total = get_column_letter(col_total)
            ws_r.cell(row=chart_header_row, column=1 + i, value=nombre).font = SECTION_FONT
            c = ws_r.cell(row=chart_value_row, column=1 + i,
                           value=f'={col_letter_total}{total_row_r}')
            c.number_format = '#,##0.00,,,"B"'
        ws_r.cell(row=chart_header_row, column=7, value=(
            '(chart source -- linked to the TOTAL row above, amounts in billions Bs.)')).font = SUB_FONT
        row = chart_value_row + 2

        chart = BarChart()
        chart.type = 'col'
        chart.grouping = 'clustered'
        chart.title = f'Accumulated Summary — {desde} to {hasta}'
        chart.y_axis.title = 'Bs. (billions)'
        chart.height = 9
        chart.width = 20
        cats = Reference(ws_r, min_col=1, max_col=5, min_row=chart_header_row, max_row=chart_header_row)
        data = Reference(ws_r, min_col=1, max_col=5, min_row=chart_value_row, max_row=chart_value_row)
        chart.add_data(data, titles_from_data=False, from_rows=True)
        chart.set_categories(cats)
        chart.series[0].tx = None
        chart.legend = None
        for i, color in enumerate(colores_chart):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            chart.series[0].data_points.append(pt)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.numFmt = '#,##0.00,,,"B"'
        ws_r.add_chart(chart, f'A{row}')
        row += 20

        # SENIAT Reconciliation Gap -- pedido explícito 2026-08-22: reemplaza
        # "Tax Credit"; es el mismo índice "Crédito Fiscal SENIAT sin match"
        # del Dashboard (_solo_seniat_sin_match_bs), mismos colores (verde
        # Con Match #2f7d4f / naranja Sin Match #eb6834).
        ws_r.cell(row=row, column=1,
                  value='SENIAT Reconciliation Gap (Crédito Fiscal SENIAT sin match)').font = SECTION_FONT
        row += 1
        ws_r.cell(row=row, column=1, value=(
            'SENIAT-confirmed withholdings never linked to any SmartIVA record -- '
            'real tax credit the client may not know they have.'))
        row += 1
        for label, monto, color in [
            ('Matched (linked to SmartIVA)', con_match_seniat, '2F7D4F'),
            ('Unmatched (real gap)', sin_match_seniat, 'EB6834'),
        ]:
            ws_r.cell(row=row, column=1, value=label).font = Font(bold=True, color=color)
            c = ws_r.cell(row=row, column=2, value=round(monto, 2))
            c.number_format = MONEY_FMT
            c.font = Font(bold=True, color=color)
            row += 1
        ws_r.cell(row=row, column=1, value='Total confirmed by SENIAT')
        c = ws_r.cell(row=row, column=2, value=round(total_seniat_confirmado, 2))
        c.number_format = MONEY_FMT
        row += 1
        ws_r.cell(row=row, column=1, value='% Unmatched')
        c = ws_r.cell(row=row, column=2, value=pct_sin_match_seniat)
        c.number_format = PCT_FMT
        row += 2

        # IOC / TAC / BDS -- mismos 3 indicadores/colores que las tarjetas
        # del Dashboard (_ioc_tac_bds_html): IOC azul #17A2B8, TAC verde
        # #28A745, BDS rojo #B5474D.
        ws_r.cell(row=row, column=1, value='IOC / TAC / BDS (Dashboard indicators)').font = SECTION_FONT
        row += 1
        ws_r.cell(row=row, column=1, value=(
            'Same methodology as the Dashboard IOC/TAC/BDS cards -- '
            'Base = YTD Expected Withholdings (Retenciones Esperadas).'))
        row += 1
        headers_idx = ['Indicator', 'Base (Amount)', 'Achieved (Amount)', '% Achieved',
                       'Missing (Amount)', '% Missing', 'Missing (Qty)']
        for c, h in enumerate(headers_idx, start=1):
            cell = ws_r.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        row += 1

        def _cnt_faltan(logrado_cnt):
            return base_cnt_ytd - logrado_cnt if logrado_cnt is not None else ''

        filas_idx = [
            ('IOC — Receipt Collection', recibido_ytd, recibido_cnt_ytd,
             'of expected withholdings still pending receipt.', '17A2B8'),
            ('TAC — Tax Credit Utilization', declarado_ytd, None,
             'of potential tax credit not yet declared/utilized.', '28A745'),
            ('BDS — Gap vs. SENIAT Portal', seniat_ytd, seniat_cnt_ytd,
             'that clients have not yet reported to SENIAT.', 'B5474D'),
        ]
        for label, logrado, logrado_cnt, descripcion, color in filas_idx:
            pct = (logrado / base_ytd) if base_ytd else 0.0
            faltante = base_ytd - logrado
            pct_falta = 1 - pct
            fill = PatternFill('solid', fgColor=color)
            c1 = ws_r.cell(row=row, column=1, value=label)
            c1.font = Font(bold=True, color='FFFFFF')
            c1.fill = fill
            c2 = ws_r.cell(row=row, column=2, value=round(base_ytd, 2))
            c2.number_format = MONEY_FMT
            c3 = ws_r.cell(row=row, column=3, value=round(logrado, 2))
            c3.number_format = MONEY_FMT
            c4 = ws_r.cell(row=row, column=4, value=pct)
            c4.number_format = PCT_FMT
            c4.font = Font(bold=True, color=color)
            c5 = ws_r.cell(row=row, column=5, value=round(faltante, 2))
            c5.number_format = MONEY_FMT
            c6 = ws_r.cell(row=row, column=6, value=pct_falta)
            c6.number_format = PCT_FMT
            ws_r.cell(row=row, column=7, value=_cnt_faltan(logrado_cnt))
            row += 1
            ws_r.cell(row=row, column=1, value=descripcion).font = SUB_FONT
            row += 1

        ws_r.column_dimensions['A'].width = 34
        for c in range(2, 12):
            ws_r.column_dimensions[get_column_letter(c)].width = 15

        # ══════════════════════════════════════════════════════════════
        # Sheet "Totales por MES" -- Excel vs SmartIVA + sección SENIAT
        # ══════════════════════════════════════════════════════════════
        Linea = self.env['ve.conecta.carga.ventas.linea']
        Wh = self.env['ve.wh.iva']
        Seniat = self.env['ve.seniat.retencion']

        lineas = Linea.search_read([
            ('carga_id.company_id', '=', company_id),
            ('fecha', '>=', periodos[0].fecha_inicio), ('fecha', '<=', periodos[-1].fecha_fin),
        ], ['fecha', 'base_16', 'base_8', 'base_exento', 'monto_retenido', 'nro_comp_retencion',
            'eliminada_duplicado'])
        wh_all = Wh.search_read([
            ('company_id', '=', company_id), ('periodo_retencion', 'in', quincenas_orden),
        ], ['monto_retenido', 'monto_base', 'monto_base_red', 'monto_exento', 'state',
            'periodo_retencion'])
        seniat_all = Seniat.search_read([
            ('company_id', '=', company_id), ('periodo_retencion', 'in', quincenas_orden),
        ], ['monto_retenido', 'estado', 'periodo_retencion'])

        def _mes_de_fecha(fecha_str):
            anio, mes_num = int(fecha_str[:4]), int(fecha_str[5:7])
            return f'{MESES_NOMBRE[mes_num - 1]} {anio}'

        por_mes_excel = defaultdict(lambda: [0, 0.0, 0.0, 0.0])  # cnt, base, con_comp, sin_comp
        for l in lineas:
            if l['eliminada_duplicado'] or not l['fecha']:
                continue
            mes = _mes_de_fecha(str(l['fecha']))
            ret = l['monto_retenido'] or 0
            base = (l['base_16'] or 0) + (l['base_8'] or 0) + (l['base_exento'] or 0)
            por_mes_excel[mes][0] += 1
            por_mes_excel[mes][1] += base
            if (l['nro_comp_retencion'] or '').strip():
                por_mes_excel[mes][2] += ret
            else:
                por_mes_excel[mes][3] += ret

        por_mes_smart = defaultdict(lambda: [0, 0.0, 0.0, 0.0, 0.0])  # cnt, base, confirmado, vencido/esperado, borrador
        for w in wh_all:
            if w['state'] == 'anulado':
                continue
            mes = mes_de_quincena.get(w['periodo_retencion'])
            if not mes:
                continue
            ret = w['monto_retenido'] or 0
            base = (w['monto_base'] or 0) + (w['monto_base_red'] or 0) + (w['monto_exento'] or 0)
            por_mes_smart[mes][0] += 1
            por_mes_smart[mes][1] += base
            if w['state'] == 'confirmado':
                por_mes_smart[mes][2] += ret
            elif w['state'] in ('vencido', 'esperado'):
                por_mes_smart[mes][3] += ret
            elif w['state'] == 'borrador':
                por_mes_smart[mes][4] += ret

        por_mes_seniat = defaultdict(lambda: [0, 0.0, 0, 0.0, 0, 0.0, 0, 0.0])
        # cnt, monto, conc_cnt, conc_monto, dif_cnt, dif_monto, noconc_cnt, noconc_monto
        for s in seniat_all:
            mes = mes_de_quincena.get(s['periodo_retencion'])
            if not mes:
                continue
            monto = s['monto_retenido'] or 0
            acc = por_mes_seniat[mes]
            acc[0] += 1
            acc[1] += monto
            if s['estado'] == 'conciliado':
                acc[2] += 1
                acc[3] += monto
            elif s['estado'] == 'diferencia':
                acc[4] += 1
                acc[5] += monto
            else:  # cargado, sin_match
                acc[6] += 1
                acc[7] += monto

        ws2 = wb.create_sheet('Monthly Totals')
        ws2['A1'] = f'Monthly Totals — {desde} to {hasta} ({company.name})'
        ws2['A1'].font = TITLE_FONT
        r = 3
        ws2.cell(row=r, column=2, value='CLIENT FILE (Sales Ledger)').font = GROUP_FONT
        ws2.cell(row=r, column=2).fill = GROUP_FILL_EXCEL
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        for c in range(2, 7):
            ws2.cell(row=r, column=c).fill = GROUP_FILL_EXCEL
        ws2.cell(row=r, column=2).alignment = GROUP_ALIGN
        ws2.cell(row=r, column=7, value='SMARTIVA (actual status in Odoo)').font = GROUP_FONT
        ws2.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
        for c in range(7, 13):
            ws2.cell(row=r, column=c).fill = GROUP_FILL_SMARTIVA
        ws2.cell(row=r, column=7).alignment = GROUP_ALIGN
        ws2.cell(row=r, column=13, value='SENIAT (portal)').font = GROUP_FONT
        ws2.merge_cells(start_row=r, start_column=13, end_row=r, end_column=20)
        for c in range(13, 21):
            ws2.cell(row=r, column=c).fill = GROUP_FILL_SENIAT
        ws2.cell(row=r, column=13).alignment = GROUP_ALIGN
        r += 1
        headers2 = ['Month', 'Taxable Base', 'Qty', 'Withheld w/ Receipt', 'Withheld w/o Receipt', 'Total Withheld',
                    'Taxable Base', 'Qty', 'Withheld Confirmed', 'Withheld Overdue/Not Received',
                    'Withheld Received (pending)', 'Total Withheld',
                    'Qty', 'Amount Withheld', 'Reconciled (Qty)', 'Reconciled (Amount)',
                    'Reconciled w/ Diff (Qty)', 'Reconciled w/ Diff (Amount)',
                    'Not Reconciled (Qty)', 'Not Reconciled (Amount)']
        for c, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=r, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        r += 1
        first_row2 = r
        tot2 = [0.0] * 17
        for mes in meses_orden:
            ecn, ebe, econ, esin = por_mes_excel.get(mes, [0, 0.0, 0.0, 0.0])
            scn, sba, sconf, svenc, sborr = por_mes_smart.get(mes, [0, 0.0, 0.0, 0.0, 0.0])
            xcn, xmt, xcc, xcm, xdc, xdm, xnc, xnm = por_mes_seniat.get(mes, [0, 0.0, 0, 0.0, 0, 0.0, 0, 0.0])
            etot = econ + esin
            stot = sconf + svenc + sborr
            vals = [mes, ebe, ecn, econ, esin, etot, sba, scn, sconf, svenc, sborr, stot,
                    xcn, xmt, xcc, xcm, xdc, xdm, xnc, xnm]
            money_cols = {2, 4, 5, 6, 7, 9, 10, 11, 12, 14, 16, 18, 20}
            for c, v in enumerate(vals, start=1):
                cell = ws2.cell(row=r, column=c, value=v)
                if c in money_cols:
                    cell.number_format = MONEY_FMT
            for i, v in enumerate([ebe, ecn, econ, esin, sba, scn, sconf, svenc, sborr,
                                    xcn, xmt, xcc, xcm, xdc, xdm, xnc, xnm]):
                tot2[i] += v
            r += 1
        last_row2 = r - 1
        ws2.cell(row=r, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, 21):
            ws2.cell(row=r, column=c).fill = TOTAL_FILL
        etot_tot = tot2[2] + tot2[3]
        stot_tot = tot2[6] + tot2[7] + tot2[8]
        fila_tot = tot2[:4] + [etot_tot] + tot2[4:9] + [stot_tot] + tot2[9:]
        money_cols_total = {2, 4, 5, 6, 7, 9, 10, 11, 12, 14, 16, 18, 20}
        for i, v in enumerate(fila_tot):
            c = 2 + i
            cell = ws2.cell(row=r, column=c, value=round(v, 2) if isinstance(v, float) else v)
            cell.font = TOTAL_FONT
            if c in money_cols_total:
                cell.number_format = MONEY_FMT
        r += 2
        nota2 = ws2.cell(row=r, column=1, value=(
            'Client File = complete universe of the uploaded Sales Ledger (Libro de Ventas) in '
            'SmartIVA, includes duplicates/blank documents. SmartIVA = actual status of each '
            'withholding in Odoo. SENIAT = ve.seniat.retencion aggregated by month (Reconciled = '
            'Reconciled status; Reconciled w/ Diff = With-Difference status; Not Reconciled = '
            'Pending Reconciliation + No Match).'))
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=20)
        ws2.row_dimensions[r].height = 45
        nota2.alignment = Alignment(wrap_text=True, vertical='top')
        widths2 = [10, 18, 12, 16, 16, 16, 18, 12, 18, 20, 18, 16, 12, 16, 14, 15, 16, 16, 14, 15]
        for i, w in enumerate(widths2):
            ws2.column_dimensions[get_column_letter(i + 1)].width = w

        # ══════════════════════════════════════════════════════════════
        # Sheet "Retenciones Declaradas"
        # ══════════════════════════════════════════════════════════════
        decl_recs = Declarado.search([('company_id', '=', company_id)], order='anio asc, mes asc')
        ws3 = wb.create_sheet('Declared Withholdings')
        ws3.cell(row=1, column=1, value='Year / Month').font = HEADER_FONT
        ws3.cell(row=1, column=1).fill = HEADER_FILL
        ws3.cell(row=1, column=2, value='Amount').font = HEADER_FONT
        ws3.cell(row=1, column=2).fill = HEADER_FILL
        r = 2
        anio_actual = None
        subtotal_anio = 0.0
        for d in decl_recs:
            if anio_actual is not None and d.anio != anio_actual:
                ws3.cell(row=r, column=1, value=f'Subtotal {anio_actual}').font = TOTAL_FONT
                c = ws3.cell(row=r, column=2, value=round(subtotal_anio, 2))
                c.number_format = MONEY_FMT
                c.font = TOTAL_FONT
                r += 1
                subtotal_anio = 0.0
            if anio_actual != d.anio:
                ws3.cell(row=r, column=1, value=str(d.anio)).font = SECTION_FONT
                r += 1
                anio_actual = d.anio
            ws3.cell(row=r, column=1, value=MESES_EN[d.mes - 1])
            c = ws3.cell(row=r, column=2, value=round(d.monto_declarado, 2))
            c.number_format = MONEY_FMT
            subtotal_anio += d.monto_declarado
            r += 1
        if anio_actual is not None:
            ws3.cell(row=r, column=1, value=f'Subtotal {anio_actual}').font = TOTAL_FONT
            c = ws3.cell(row=r, column=2, value=round(subtotal_anio, 2))
            c.number_format = MONEY_FMT
            c.font = TOTAL_FONT
        ws3.column_dimensions['A'].width = 16
        ws3.column_dimensions['B'].width = 18

        # ══════════════════════════════════════════════════════════════
        # Sheet "Consolidado" -- universo completo del Libro de Ventas
        # ══════════════════════════════════════════════════════════════
        lineas_full = Linea.search_read([
            ('carga_id.company_id', '=', company_id),
            ('fecha', '>=', periodos[0].fecha_inicio), ('fecha', '<=', periodos[-1].fecha_fin),
        ], ['rif', 'nombre_cliente', 'nro_documento', 'nro_control', 'total_documento',
            'base_16', 'base_8', 'base_exento', 'monto_iva', 'monto_retenido',
            'nro_comp_retencion', 'zona', 'estado_pago', 'es_spe', 'eliminada_duplicado',
            'invoice_id', 'fecha'], order='fecha asc, id asc')

        wh_full = Wh.search_read([
            ('company_id', '=', company_id), ('invoice_id', '!=', False),
        ], ['invoice_id', 'monto_retenido'])
        wh_by_invoice = {}
        for w in wh_full:
            wh_by_invoice.setdefault(w['invoice_id'][0], w['monto_retenido'] or 0)

        ws4 = wb.create_sheet('Consolidated')
        headers4 = ['No.', 'Date', 'Tax ID (RIF)', 'Name / Business Name', 'Invoice Number',
                    'Control Number', 'Total Sales Incl. VAT', 'Taxable Base',
                    '% VAT', 'VAT Amount', 'Withheld VAT', 'Withholding Receipt', 'Zone',
                    'Payment Status', 'Withholding Agent', 'Duplicate Row', 'Withheld w/o Receipt',
                    'Expected Withholding', 'Expected vs Withheld Difference', 'Expected Differs']
        for c, h in enumerate(headers4, start=1):
            cell = ws4.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for i, l in enumerate(lineas_full, start=1):
            base = (l['base_16'] or 0) + (l['base_8'] or 0) + (l['base_exento'] or 0)
            pct_iva = 16 if l['base_16'] else (8 if l['base_8'] else 0)
            monto_ret = l['monto_retenido'] or 0
            sin_comp = bool(monto_ret) and not (l['nro_comp_retencion'] or '').strip()
            iva_esperado = wh_by_invoice.get(l['invoice_id'][0], 0) if l['invoice_id'] else 0
            diferencia = round(iva_esperado - monto_ret, 2)
            row_i = i + 1
            # Campos Char de Odoo llegan como False (no '') cuando están
            # vacíos vía search_read -- sin el `or ''`, openpyxl escribe el
            # booleano crudo y Excel lo muestra como "FALSE" (bug real
            # confirmado 2026-08-22 en el mismo patrón del wizard Libro de
            # Ventas, columna N.Comprobante).
            vals = [i, l['fecha'], l['rif'] or '', l['nombre_cliente'] or '',
                    l['nro_documento'] or '', l['nro_control'] or '',
                    l['total_documento'], base, pct_iva, l['monto_iva'], monto_ret,
                    l['nro_comp_retencion'] or '', l['zona'] or '', l['estado_pago'] or '',
                    l['es_spe'] or '',
                    l['eliminada_duplicado'], sin_comp, iva_esperado, diferencia,
                    abs(diferencia) > TOL]
            for c, v in enumerate(vals, start=1):
                ws4.cell(row=row_i, column=c, value=v)
        widths4 = [6, 12, 14, 32, 14, 14, 16, 14, 8, 14, 14, 16, 12, 14, 12, 12, 14, 14, 16, 14]
        for i, w in enumerate(widths4):
            ws4.column_dimensions[get_column_letter(i + 1)].width = w
        if lineas_full:
            ws4.auto_filter.ref = f'A1:{get_column_letter(len(headers4))}{len(lineas_full) + 1}'
            ws4.freeze_panes = 'B2'

        output = io.BytesIO()
        wb.save(output)
        fname = 'SmartIVA Executive Summary - Vencement 2026.xlsx'
        att = self.env['ir.attachment'].create({
            'name': fname,
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
