import base64
import io
import re
from collections import defaultdict

from odoo import api, fields, models
from odoo.exceptions import UserError

MESES_NOMBRE = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

ESTADO_WH_LABEL = {
    'listo_declarar': 'Conciliado OK',
    # 'declarado' es el estado de la Declaración (Forma 30) del período, un
    # ciclo de vida DISTINTO del de conciliación contra SENIAT -- una vez
    # que Odoo concilió OK (listo_declarar), declarar el período no cambia
    # si matcheó o no contra SENIAT. Se pliega en 'Conciliado OK' para no
    # mezclar los dos conceptos en este reporte (corregido 2026-08-22,
    # feedback explícito de la usuaria: "el estado de la declaración no
    # tiene nada que ver con el estado de la conciliación").
    'declarado': 'Conciliado OK',
    'conciliada_norec': 'No Recibido SENIAT OK',
    'diferencia': 'Conciliado c/Dif',
    'solo_odoo': 'Sin SENIAT',
    'pendiente': 'Por Conciliar',
    # legacy -- "no se asignan en código nuevo" (ver comentario en
    # ve_wh_iva.py::estado_conciliacion), sinónimos viejos de
    # 'listo_declarar'. Mismo pliegue que 'declarado'.
    'conciliada': 'Conciliado OK',
    'aprobado_declarar': 'Conciliado OK',
}

# Orden de prioridad para las columnas del Resumen -- action_generar solo
# dibuja una columna para un estado si al menos una fila del rango elegido
# cayó ahí (ver estados_orden_run más abajo); esta lista solo fija el
# ORDEN cuando el estado sí aparece, nunca fuerza una columna vacía.
# Cualquier estado detectado que no esté acá se agrega al final -- nunca
# se pierde un monto por falta de columna.
ESTADOS_ORDEN = ['Sin SENIAT', 'Por Conciliar', 'Conciliado OK', 'Conciliado c/Dif',
                  'No Recibido SENIAT OK', 'Solo en SENIAT',
                  'Cliente no Contribuyente pero con retención en SENIAT']

# Únicos 2 estados sin ningún wh_iva real detrás -- la fila viene de un
# ve.seniat.retencion sin wh_iva_id (ver _estado_solo_seniat). Se usan
# solo para decidir el color de grupo de esas 2 columnas -- el total "IVA
# Esperado SmartIVA" (smart_total_directo_por_mes) no depende de esta
# lista, sale de recorrer wh_all directo (ver comentario ahí).
ESTADOS_SOLO_SENIAT = {
    'Solo en SENIAT', 'Cliente no Contribuyente pero con retención en SENIAT',
}

# Rediseño 2026-08-22 (pedido explícito de la usuaria: "quiero que cuadre
# con lo que se ve en el Dashboard, no con lo que tú crees que debe ser"):
# la columna "Monto" de cada Estado ahora muestra el lado SmartIVA
# (wh.monto_retenido) para TODOS los estados que tienen un wh_iva real
# detrás -- exactamente el mismo criterio que usa el Dashboard
# (_calc_concil_buckets sólo conoce monto_retenido, nunca el monto SENIAT
# del comprobante). Antes "Conciliado c/Dif"/"Conciliado OK"/"No Recibido
# SENIAT OK" mostraban el lado SENIAT (diseño 2026-08-21, para que
# sumaran exacto contra "SENIAT total") -- eso hacía que "Conciliado
# c/Dif" NUNCA pudiera coincidir con la dona "Diferencia de monto" del
# Dashboard (que sí es SmartIVA), aunque ambos midieran las mismas 669
# filas. Los únicos 2 estados que se quedan en SENIAT son
# ESTADOS_SOLO_SENIAT (no tienen NINGÚN wh_iva detrás -- no hay lado
# SmartIVA que mostrar, es lo mismo que hace la dona en su porción "Solo
# SENIAT").
def _lado_monto(estado):
    return 'SmartIVA' if estado not in ESTADOS_SOLO_SENIAT else 'SENIAT'


def _norm_rif(rif):
    return re.sub(r'[^A-Z0-9]', '', str(rif or '').upper())


class WizardConciliacionSmartivaSeniat(models.TransientModel):
    _name = 've.conciliacion.smartiva.seniat.wizard'
    _description = 'Generar Conciliación SmartIVA-SENIAT (Excel)'

    company_id = fields.Many2one(
        'res.company', string='Compañía', required=True,
        default=lambda self: self.env.company)
    periodo_desde_id = fields.Many2one(
        've.conciliacion.periodo', string='Desde',
        domain="[('company_id', '=', company_id)]", required=True)
    periodo_hasta_id = fields.Many2one(
        've.conciliacion.periodo', string='Hasta',
        domain="[('company_id', '=', company_id)]", required=True)

    @api.model
    def default_get(self, fields_list):
        vals = super().default_get(fields_list)
        company_id = vals.get('company_id') or self.env.company.id
        periodos = self.env['ve.conciliacion.periodo'].search(
            [('company_id', '=', company_id)], order='periodo_retencion asc')
        if periodos:
            vals.setdefault('periodo_desde_id', periodos[0].id)
            vals.setdefault('periodo_hasta_id', periodos[-1].id)
        return vals

    def _estado_fila(self, w):
        return ESTADO_WH_LABEL.get(w['estado_conciliacion'], w['estado_conciliacion'] or 'Sin Estado')

    def _estado_solo_seniat(self, rif_agente, agente_by_rif):
        agente = agente_by_rif.get(_norm_rif(rif_agente))
        if agente is False:
            return 'Cliente no Contribuyente pero con retención en SENIAT'
        return 'Solo en SENIAT'

    def action_generar(self):
        """Genera el Excel "Conciliación SmartIVA-SENIAT" para el rango de
        períodos elegido -- misma metodología validada en
        scripts/demo_cementos/generar_conciliacion_smartiva_seniat.py
        (2026-08-21): ve.seniat.retencion es la ÚNICA fuente recorrida para
        armar filas relacionadas a SENIAT (via su propio wh_iva_id, sin
        matching por texto); wh_iva solo se agrega aparte para "Sin
        SENIAT" (activos que jamás aparecen en ve.seniat.retencion). La
        columna Monto de cada Estado en Resumen usa el lado SmartIVA
        (wh.monto_retenido) -- mismo criterio que el Dashboard, para que
        cuadre exacto con lo que se ve ahí (pedido explícito 2026-08-22:
        "quiero que cuadre con el Dashboard, no con lo que tú crees que
        debe ser" -- diseño anterior, 2026-08-21, usaba el lado SENIAT y
        nunca podía coincidir con la dona de Salud de Conciliación en los
        estados con diferencia de monto). Solo ESTADOS_SOLO_SENIAT (sin
        ningún wh_iva real detrás) muestran el lado SENIAT, porque no hay
        otro lado que mostrar -- ver _lado_monto. Las columnas se generan
        dinámicamente según qué estados aparecen en el rango elegido (ver
        estados_orden_run) -- nunca se muestra una columna en 0.
        """
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Side
        except ImportError:
            raise UserError('La librería openpyxl no está instalada.')

        desde = self.periodo_desde_id.periodo_retencion
        hasta = self.periodo_hasta_id.periodo_retencion
        if desde > hasta:
            raise UserError('"Desde" no puede ser posterior a "Hasta".')

        company_id = self.company_id.id
        periodos = self.env['ve.conciliacion.periodo'].search([
            ('company_id', '=', company_id),
            ('periodo_retencion', '>=', desde),
            ('periodo_retencion', '<=', hasta),
        ], order='periodo_retencion asc')
        if not periodos:
            raise UserError('No hay períodos de Conciliación SENIAT en ese rango.')

        quincenas_orden = periodos.mapped('periodo_retencion')
        mes_de_quincena = {}
        meses_orden = []
        for p in periodos:
            m = re.match(r'^(\d{4})-(\d{2}) \dQ$', p.periodo_retencion)
            if not m:
                continue
            anio, mes_num = m.group(1), int(m.group(2))
            label = f'{MESES_NOMBRE[mes_num - 1]} {anio}'
            mes_de_quincena[p.periodo_retencion] = label
            if label not in meses_orden:
                meses_orden.append(label)

        Wh = self.env['ve.wh.iva']
        Seniat = self.env['ve.seniat.retencion']
        Partner = self.env['res.partner']

        wh_all = Wh.search_read([
            ('company_id', '=', company_id),
            ('periodo_retencion', 'in', quincenas_orden),
        ], ['id', 'rif', 'partner_id', 'nro_control', 'nro_factura', 'nro_factura_match', 'name',
            'monto_retenido', 'state', 'estado_conciliacion', 'periodo_retencion', 'nivel_match',
            'monto_seniat', 'seniat_rif', 'seniat_nro_control', 'seniat_nro_documento'])

        seniat_all = Seniat.search_read([
            ('company_id', '=', company_id),
            ('periodo_retencion', 'in', quincenas_orden),
        ], ['id', 'rif_agente', 'nombre_agente', 'nro_control', 'nro_documento', 'monto_retenido',
            'estado', 'wh_iva_id', 'periodo_retencion', 'nivel_match'])

        if not wh_all and not seniat_all:
            raise UserError('No hay datos de SmartIVA ni de SENIAT en ese rango de períodos.')

        partners = Partner.search_read(
            [('company_id', '=', company_id)], ['id', 'vat', 'es_agente_retencion'])
        agente_by_rif = {_norm_rif(p['vat']): p['es_agente_retencion'] for p in partners if p['vat']}

        # Suma DIRECTA de monto_retenido por mes -- mismo criterio que el
        # campo nativo ve.conciliacion.periodo.total_seniat.
        seniat_total_directo_por_mes = defaultdict(float)
        seniat_cnt_directo_por_mes = defaultdict(int)
        for s in seniat_all:
            mes = mes_de_quincena.get(s['periodo_retencion'])
            if mes:
                seniat_total_directo_por_mes[mes] += s['monto_retenido'] or 0
                seniat_cnt_directo_por_mes[mes] += 1

        # Suma DIRECTA de monto_retenido por mes del lado SmartIVA -- mismo
        # criterio que la suma de ve.conciliacion.periodo.total_odoo /
        # monto_conciliado (Dashboard). Bug real 2026-08-22: calcular "IVA
        # Esperado SmartIVA" acumulando fila['monto_esperado'] cada vez que
        # se recorre un ve.seniat.retencion (más abajo) duplica el monto de
        # cualquier wh_iva vinculado a MÁS DE UN registro SENIAT -- se
        # confirmó en Vencement que 11 wh_iva tienen 2 vínculos SENIAT cada
        # uno (mismo comprobante SENIAT cargado 2 veces, mismo monto/
        # período), lo que inflaba el total SmartIVA ~Bs. 2.36M contra
        # Dashboard/monto_conciliado (que cuenta cada wh_iva UNA vez). Este
        # total, en cambio, recorre wh_all una sola vez -- inmune a
        # duplicados del lado SENIAT.
        smart_total_directo_por_mes = defaultdict(float)
        smart_cnt_directo_por_mes = defaultdict(int)
        for w in wh_all:
            if w['state'] == 'anulado':
                continue
            mes = mes_de_quincena.get(w['periodo_retencion'])
            if mes:
                smart_total_directo_por_mes[mes] += w['monto_retenido'] or 0
                smart_cnt_directo_por_mes[mes] += 1

        # ── Armar filas ──────────────────────────────────────────────────
        filas_por_quincena = defaultdict(list)
        resumen_mes_estado = defaultdict(lambda: defaultdict(lambda: [0, 0.0, 0.0]))

        def _acumular(fila, periodo):
            mes = mes_de_quincena.get(periodo)
            if periodo:
                filas_por_quincena[periodo].append(fila)
            if mes:
                me = resumen_mes_estado[mes][fila['estado']]
                me[0] += 1
                me[1] += fila['monto_esperado']
                me[2] += fila['seniat_monto']

        wh_by_id = {w['id']: w for w in wh_all}
        seniat_linked_wh_ids = {s['wh_iva_id'][0] for s in seniat_all if s['wh_iva_id']}

        for s in seniat_all:
            periodo_seniat = s['periodo_retencion']
            monto_s = s['monto_retenido'] or 0
            w = wh_by_id.get(s['wh_iva_id'][0]) if s['wh_iva_id'] else None
            if w is not None:
                estado = self._estado_fila(w)
                monto_esperado = w['monto_retenido'] or 0
                fila = dict(
                    rif=w['rif'], contribuyente=bool(w['partner_id'] and agente_by_rif.get(_norm_rif(w['rif']))),
                    cliente=(w['partner_id'][1] if w['partner_id'] else ''),
                    nro_control=w['nro_control'], nro_factura=w['nro_factura'],
                    nro_factura_match=w['nro_factura_match'], nro_comprobante=w['name'] or '',
                    monto_esperado=monto_esperado,
                    seniat_rif=s['rif_agente'], seniat_control=s['nro_control'], seniat_doc=s['nro_documento'],
                    seniat_monto=monto_s,
                    normalizacion=s['nivel_match'] or w['nivel_match'] or '',
                    diferencia=round(monto_esperado - monto_s, 2),
                    estado=estado,
                )
                # Bug real confirmado 2026-08-22: la fila se acumulaba bajo
                # el período DEL COMPROBANTE SENIAT (s.periodo_retencion),
                # no el del wh_iva -- el matching es a nivel de compañía,
                # no de período (ver _do_conciliar), así que SENIAT puede
                # reportar un comprobante en un período distinto al que
                # Odoo generó la retención. Como "IVA Esperado SmartIVA
                # (total)" (smart_total_directo_por_mes, más arriba) SÍ
                # agrupa por el período del wh_iva, mezclar los dos
                # criterios en la misma fila de "Mes" descuadraba la suma
                # de Cantidad/Monto por Estado contra la columna "por Mes"
                # -- Bs. 0 de diferencia neta en el TOTAL general (se
                # compensaba entre meses) pero cada fila de mes quedaba
                # mal, hasta Bs. 63,9M de corrimiento en un mes puntual.
                # Verificado en vivo: agrupar por w.periodo_retencion deja
                # las 6 filas de mes exactas, Bs.0,00 de diferencia.
                periodo_fila = w['periodo_retencion']
            else:
                estado = self._estado_solo_seniat(s['rif_agente'], agente_by_rif)
                fila = dict(
                    rif='', contribuyente=False, cliente='', nro_control='', nro_factura='',
                    nro_factura_match='', nro_comprobante='',
                    monto_esperado=0,
                    seniat_rif=s['rif_agente'], seniat_control=s['nro_control'], seniat_doc=s['nro_documento'],
                    seniat_monto=monto_s,
                    normalizacion=s['nivel_match'] or '',
                    diferencia=None,
                    estado=estado,
                )
                periodo_fila = periodo_seniat
            _acumular(fila, periodo_fila)

        # "Sin SENIAT" -- wh_iva activos que jamás aparecen en ve.seniat.retencion.
        for w in wh_all:
            if w['id'] in seniat_linked_wh_ids or w['state'] == 'anulado':
                continue
            periodo = w['periodo_retencion']
            if not periodo:
                continue
            estado = self._estado_fila(w)
            monto_esperado = w['monto_retenido'] or 0
            fila = dict(
                rif=w['rif'], contribuyente=bool(w['partner_id'] and agente_by_rif.get(_norm_rif(w['rif']))),
                cliente=(w['partner_id'][1] if w['partner_id'] else ''),
                nro_control=w['nro_control'], nro_factura=w['nro_factura'],
                nro_factura_match=w['nro_factura_match'], nro_comprobante=w['name'] or '',
                monto_esperado=monto_esperado,
                seniat_rif='', seniat_control='', seniat_doc='', seniat_monto=0,
                normalizacion=w['nivel_match'] or '',
                diferencia=None,
                estado=estado,
            )
            _acumular(fila, periodo)

        # Columnas del Resumen = solo los estados que realmente aparecen en
        # el rango elegido (nunca una columna vacía en 0, ver feedback
        # 2026-08-22) -- ordenados por ESTADOS_ORDEN cuando se conoce el
        # estado, y cualquier estado detectado que no esté ahí (legacy sin
        # mapear, o uno nuevo el día de mañana) se agrega al final -- nunca
        # se pierde un monto por falta de columna.
        estados_detectados = {
            e for por_estado in resumen_mes_estado.values() for e in por_estado
        }
        estados_orden_run = (
            [e for e in ESTADOS_ORDEN if e in estados_detectados]
            + sorted(estados_detectados - set(ESTADOS_ORDEN))
        )

        # ── Escribir workbook ────────────────────────────────────────────
        TITLE_FONT = Font(bold=True, size=13)
        SECTION_FONT = Font(bold=True)
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        HEADER_FILL = PatternFill('solid', fgColor='383A4E')
        HEADER_ALIGN = Alignment(horizontal='center', wrap_text=True)
        GROUP_FONT = Font(bold=True, color='FFFFFF', italic=True)
        GROUP_FILL_SMART = PatternFill('solid', fgColor='5B6169')
        GROUP_FILL_SENIAT = PatternFill('solid', fgColor='669999')
        GROUP_FILL_TOTAL = PatternFill('solid', fgColor='7B4B94')
        GROUP_ALIGN = Alignment(horizontal='center')
        TOTAL_FONT = Font(bold=True)
        TOTAL_FILL = PatternFill('solid', fgColor='F2E7CC')
        THIN = Side(style='thin')
        BORDER_BOTTOM = openpyxl.styles.Border(bottom=THIN)
        MONEY_FMT = '#,##0.00'
        ESTADO_FILL = {
            'Conciliado OK': 'E9F7EF',
            'No Recibido SENIAT OK': 'E9F7EF',
            'Conciliado c/Dif': 'FDEBD0',
            'Sin SENIAT': 'D6EAF8',
            'Por Conciliar': 'EAECEE',
            'Cliente no Contribuyente pero con retención en SENIAT': 'F8D7DA',
            'Solo en SENIAT': 'FADBD8',
        }

        def title_group_row(ws_, row, smart_cols, seniat_cols):
            for label, (c0, c1), fill in (
                ('SmartIVA', smart_cols, GROUP_FILL_SMART),
                ('SENIAT', seniat_cols, GROUP_FILL_SENIAT),
            ):
                cell = ws_.cell(row=row, column=c0, value=label)
                cell.font = GROUP_FONT
                cell.alignment = GROUP_ALIGN
                if c1 > c0:
                    ws_.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
                for c in range(c0, c1 + 1):
                    ws_.cell(row=row, column=c).fill = fill

        def write_header_row(ws_, row, headers):
            for c, h in enumerate(headers, start=1):
                cell = ws_.cell(row=row, column=c, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet('Resumen')
        ws['A1'] = f'Conciliación SmartIVA – SENIAT — {desde} a {hasta} ({self.company_id.name})'
        ws['A1'].font = TITLE_FONT
        ws['A2'] = ("Compara ve.wh.iva (SmartIVA) contra ve.seniat.retencion (comprobantes reales "
                    "descargados del portal SENIAT). El match ya lo calcula Odoo (Conciliar SENIAT); "
                    "este reporte solo lo resume. Incluye el caso 'Cliente no Contribuyente pero con "
                    "retención en SENIAT' -- retenciones que SENIAT sí reporta, aunque el cliente no "
                    "esté categorizado como Contribuyente Especial.")

        row_i = 4
        ws.cell(row=row_i, column=1,
                value='IVA Esperado por Mes y Estado de Conciliación').font = SECTION_FONT
        row_i += 1

        col = 2
        smart_total_cols = (col, col + 1)
        col += 2
        estado_cols = {}
        for estado in estados_orden_run:
            estado_cols[estado] = (col, col + 1)
            col += 2
        seniat_total_cols = (col, col + 1)
        col += 2
        dif_cols = (col, col + 1)
        col += 2

        group_row = row_i

        def _grupo(c0, c1, label, fill):
            cell = ws.cell(row=group_row, column=c0, value=label)
            cell.font = GROUP_FONT
            cell.alignment = GROUP_ALIGN
            if c1 > c0:
                ws.merge_cells(start_row=group_row, start_column=c0, end_row=group_row, end_column=c1)
            for c in range(c0, c1 + 1):
                ws.cell(row=group_row, column=c).fill = fill

        _grupo(*smart_total_cols, 'IVA Esperado SmartIVA (total)', GROUP_FILL_TOTAL)
        for estado, (c0, c1) in estado_cols.items():
            fill = GROUP_FILL_SENIAT if estado in ESTADOS_SOLO_SENIAT else GROUP_FILL_SMART
            _grupo(c0, c1, f'{estado} (Bs. {_lado_monto(estado)})', fill)
        _grupo(*seniat_total_cols, 'SENIAT (total)', GROUP_FILL_TOTAL)
        _grupo(*dif_cols, 'Diferencia (SmartIVA - SENIAT)', GROUP_FILL_TOTAL)
        row_i += 1

        headers = ['Mes', 'Cantidad', 'Monto SmartIVA']
        for estado in estados_orden_run:
            headers += ['Cantidad', f'Monto {_lado_monto(estado)}']
        headers += ['Cantidad', 'Monto SENIAT', 'Cantidad', 'Monto']
        write_header_row(ws, row_i, headers)
        row_i += 1
        first_row = row_i
        for mes in meses_orden:
            por_estado = resumen_mes_estado.get(mes, {})
            # "IVA Esperado SmartIVA" sale del conteo/suma DIRECTOS sobre
            # wh_all (smart_cnt/total_directo_por_mes, arriba) -- no de
            # resumen_mes_estado, que puede contar un mismo wh_iva más de
            # una vez si tiene más de un ve.seniat.retencion vinculado (ver
            # comentario en smart_total_directo_por_mes).
            smart_cnt = smart_cnt_directo_por_mes.get(mes, 0)
            smart_monto = smart_total_directo_por_mes.get(mes, 0.0)
            seniat_monto_total = seniat_total_directo_por_mes.get(mes, 0.0)
            seniat_cnt_total = seniat_cnt_directo_por_mes.get(mes, 0)
            ws.cell(row=row_i, column=1, value=mes)
            ws.cell(row=row_i, column=2, value=smart_cnt)
            c = ws.cell(row=row_i, column=3, value=round(smart_monto, 2))
            c.number_format = MONEY_FMT
            for estado in estados_orden_run:
                c0, c1 = estado_cols[estado]
                cnt, m_smart, m_seniat = por_estado.get(estado, [0, 0.0, 0.0])
                # Lado a mostrar: SmartIVA para todo lo que tenga un
                # wh_iva real detrás -- igual que el Dashboard, que solo
                # conoce monto_retenido (nunca el monto SENIAT del
                # comprobante). Solo ESTADOS_SOLO_SENIAT (sin wh_iva
                # detrás) muestran SENIAT, porque no hay otro lado que
                # mostrar. Antes esto dependía del dato (m_seniat != 0) --
                # diseño 2026-08-21, revertido 2026-08-22 (pedido
                # explícito: "quiero que cuadre con el Dashboard", ver
                # _lado_monto arriba).
                monto_estado = m_seniat if estado in ESTADOS_SOLO_SENIAT else m_smart
                ws.cell(row=row_i, column=c0, value=cnt)
                cm = ws.cell(row=row_i, column=c1, value=round(monto_estado, 2))
                cm.number_format = MONEY_FMT
            ws.cell(row=row_i, column=seniat_total_cols[0], value=seniat_cnt_total)
            cs = ws.cell(row=row_i, column=seniat_total_cols[1], value=round(seniat_monto_total, 2))
            cs.number_format = MONEY_FMT
            smart_col_letter = openpyxl.utils.get_column_letter(2)
            smart_monto_letter = openpyxl.utils.get_column_letter(3)
            seniat_cnt_letter = openpyxl.utils.get_column_letter(seniat_total_cols[0])
            seniat_monto_letter = openpyxl.utils.get_column_letter(seniat_total_cols[1])
            ws.cell(row=row_i, column=dif_cols[0],
                    value=f'={smart_col_letter}{row_i}-{seniat_cnt_letter}{row_i}')
            cd = ws.cell(row=row_i, column=dif_cols[1],
                          value=f'={smart_monto_letter}{row_i}-{seniat_monto_letter}{row_i}')
            cd.number_format = MONEY_FMT
            row_i += 1
        last_row = row_i - 1

        ws.cell(row=row_i, column=1, value='TOTAL').font = TOTAL_FONT
        for c in range(1, dif_cols[1] + 1):
            ws.cell(row=row_i, column=c).fill = TOTAL_FILL
        suma_cols = [2, 3] + [c for cols in estado_cols.values() for c in cols] + list(seniat_total_cols)
        for c in suma_cols:
            col_letter = openpyxl.utils.get_column_letter(c)
            cell = ws.cell(row=row_i, column=c, value=f'=SUM({col_letter}{first_row}:{col_letter}{last_row})')
            cell.font = TOTAL_FONT
            if c not in ([2, seniat_total_cols[0]] + [estado_cols[e][0] for e in estados_orden_run]):
                cell.number_format = MONEY_FMT
        smart_col_letter = openpyxl.utils.get_column_letter(2)
        smart_monto_letter = openpyxl.utils.get_column_letter(3)
        seniat_cnt_letter = openpyxl.utils.get_column_letter(seniat_total_cols[0])
        seniat_monto_letter = openpyxl.utils.get_column_letter(seniat_total_cols[1])
        cdt_cnt = ws.cell(row=row_i, column=dif_cols[0],
                           value=f'={smart_col_letter}{row_i}-{seniat_cnt_letter}{row_i}')
        cdt_cnt.font = TOTAL_FONT
        cd_t = ws.cell(row=row_i, column=dif_cols[1],
                        value=f'={smart_monto_letter}{row_i}-{seniat_monto_letter}{row_i}')
        cd_t.number_format = MONEY_FMT
        cd_t.font = TOTAL_FONT

        ws.freeze_panes = f'A{first_row}'
        ws.column_dimensions['A'].width = 14
        for c in range(2, dif_cols[1] + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15

        # Headers/anchos compartidos por "Consolidado" y las pestañas por
        # quincena -- agregan "N.Factura (usado en Match)" (nro_factura_match,
        # ve_wh_iva.py) junto a "N.Factura": son campos DISTINTOS a propósito
        # (pedido explícito 2026-08-22, tras confundir el N2 de un match con
        # el N° de Factura real). "N.Factura" es el dato tal cual vino del
        # Libro de Ventas, sin fallback. "N.Factura (usado en Match)" es lo
        # que realmente usa el Nivel 2 de _do_conciliar (invoice_id.name si
        # hay factura vinculada, si no nro_documento) -- puede haber caído al
        # N° Control como respaldo si faltaba el dato real al cargar, así que
        # puede diferir de "N.Factura" (o estar poblado cuando ese está
        # vacío) sin que sea un error.
        #
        # "Monto SENIAT" (antes solo "Monto", pedido explícito 2026-08-22):
        # esta tabla de detalle SIEMPRE mostró los 2 lados por fila (Monto
        # Esperado = SmartIVA, Monto = SENIAT) -- eso está bien fila por
        # fila, pero la columna no decía de cuál lado era, y sumar/filtrar
        # esta columna para "Conciliado c/Dif" da Bs. 142.093.795,11 (lado
        # SENIAT) en vez de los Bs. 140.906.185,73 (lado SmartIVA) que
        # muestra la dona del Dashboard -- mismo problema que ya se corrigió
        # en el Resumen, sin tocar acá.
        HEADERS = ['RIF', 'Contribuyente', 'Cliente', 'N.Control', 'N.Factura',
                   'N.Factura (usado en Match)', 'N.Comprobante', 'Monto Esperado',
                   'RIF', 'N.Control', 'N.Doc SENIAT', 'Monto SENIAT',
                   'Normalización', 'Diferencia', 'Estado Conciliación']
        COL_WIDTHS = [15, 13, 40, 14, 14, 16, 16, 16, 15, 14, 14, 16, 13, 14, 50]

        def _fila_vals(f):
            return [f['rif'], ('✓' if f['contribuyente'] else ''), f['cliente'], f['nro_control'],
                    f['nro_factura'], f['nro_factura_match'], f['nro_comprobante'],
                    round(f['monto_esperado'], 2) if f['monto_esperado'] else 0,
                    f['seniat_rif'], f['seniat_control'], f['seniat_doc'],
                    round(f['seniat_monto'], 2) if f['seniat_monto'] else 0,
                    f['normalizacion'], (round(f['diferencia'], 2) if f['diferencia'] is not None else None),
                    f['estado']]

        # -- Consolidado: todas las retenciones SENIAT de todo el rango,
        # en una sola pestaña (pedido explícito 2026-08-22) -- mismas
        # columnas que las pestañas por quincena, más "Período" al frente
        # para poder distinguir/filtrar entre quincenas mezcladas.
        HEADERS_CONSOL = ['Período'] + HEADERS
        COL_WIDTHS_CONSOL = [12] + COL_WIDTHS
        todas_las_filas = [
            (periodo, f) for periodo in quincenas_orden for f in filas_por_quincena.get(periodo, [])
        ]
        if todas_las_filas:
            ws_c = wb.create_sheet('Consolidado')
            ws_c['A1'] = f'Conciliación SmartIVA – SENIAT — Consolidado {desde} a {hasta} ({self.company_id.name})'
            ws_c['A1'].font = TITLE_FONT
            row_i = 3
            title_group_row(ws_c, row_i, (2, 9), (10, 13))
            row_i += 1
            write_header_row(ws_c, row_i, HEADERS_CONSOL)
            header_row = row_i
            row_i += 1
            first_data_row = row_i
            todas_las_filas.sort(key=lambda pf: (pf[0], pf[1]['estado'], pf[1]['rif'] or pf[1]['seniat_rif'] or ''))
            for periodo, f in todas_las_filas:
                vals = [periodo] + _fila_vals(f)
                vals = [('' if v is False else v) for v in vals]
                fill = PatternFill('solid', fgColor=ESTADO_FILL.get(f['estado'], 'FFFFFF'))
                for c, v in enumerate(vals, start=1):
                    cell = ws_c.cell(row=row_i, column=c, value=v)
                    cell.fill = fill
                    cell.border = BORDER_BOTTOM
                    if c in (9, 13, 15):
                        cell.number_format = MONEY_FMT
                row_i += 1
            last_data_row = row_i - 1

            total_row = row_i
            ws_c.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
            for c in range(1, len(HEADERS_CONSOL) + 1):
                ws_c.cell(row=total_row, column=c).fill = TOTAL_FILL
            c_me = ws_c.cell(row=total_row, column=9, value=f'=SUM(I{first_data_row}:I{last_data_row})')
            c_me.number_format = MONEY_FMT
            c_me.font = TOTAL_FONT
            c_ms = ws_c.cell(row=total_row, column=13, value=f'=SUM(M{first_data_row}:M{last_data_row})')
            c_ms.number_format = MONEY_FMT
            c_ms.font = TOTAL_FONT
            c_df = ws_c.cell(row=total_row, column=15, value=f'=SUM(O{first_data_row}:O{last_data_row})')
            c_df.number_format = MONEY_FMT
            c_df.font = TOTAL_FONT

            ws_c.auto_filter.ref = f'A{header_row}:{openpyxl.utils.get_column_letter(len(HEADERS_CONSOL))}{last_data_row}'
            ws_c.freeze_panes = f'B{header_row + 1}'
            for c, w in enumerate(COL_WIDTHS_CONSOL, start=1):
                ws_c.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        # -- Pestañas por quincena --
        for quincena in quincenas_orden:
            filas = filas_por_quincena.get(quincena)
            if not filas:
                continue
            ws_q = wb.create_sheet(quincena[:31])
            ws_q['A1'] = f'Conciliación SmartIVA – SENIAT — {quincena}'
            ws_q['A1'].font = TITLE_FONT
            row_i = 3
            title_group_row(ws_q, row_i, (1, 8), (9, 12))
            row_i += 1
            write_header_row(ws_q, row_i, HEADERS)
            header_row = row_i
            row_i += 1
            first_data_row = row_i
            for f in sorted(filas, key=lambda x: (x['estado'], x['rif'] or x['seniat_rif'] or '')):
                vals = _fila_vals(f)
                vals = [('' if v is False else v) for v in vals]
                fill = PatternFill('solid', fgColor=ESTADO_FILL.get(f['estado'], 'FFFFFF'))
                for c, v in enumerate(vals, start=1):
                    cell = ws_q.cell(row=row_i, column=c, value=v)
                    cell.fill = fill
                    cell.border = BORDER_BOTTOM
                    if c in (8, 12, 14):
                        cell.number_format = MONEY_FMT
                row_i += 1
            last_data_row = row_i - 1

            total_row = row_i
            ws_q.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
            for c in range(1, len(HEADERS) + 1):
                ws_q.cell(row=total_row, column=c).fill = TOTAL_FILL
            c8 = ws_q.cell(row=total_row, column=8, value=f'=SUM(H{first_data_row}:H{last_data_row})')
            c8.number_format = MONEY_FMT
            c8.font = TOTAL_FONT
            c12 = ws_q.cell(row=total_row, column=12, value=f'=SUM(L{first_data_row}:L{last_data_row})')
            c12.number_format = MONEY_FMT
            c12.font = TOTAL_FONT
            c14 = ws_q.cell(row=total_row, column=14, value=f'=SUM(N{first_data_row}:N{last_data_row})')
            c14.number_format = MONEY_FMT
            c14.font = TOTAL_FONT

            ws_q.auto_filter.ref = f'A{header_row}:{openpyxl.utils.get_column_letter(len(HEADERS))}{last_data_row}'
            ws_q.freeze_panes = f'A{header_row + 1}'
            for c, w in enumerate(COL_WIDTHS, start=1):
                ws_q.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        output = io.BytesIO()
        wb.save(output)
        fname = f'Conciliacion_SmartIVA_SENIAT_{desde}_a_{hasta}.xlsx'.replace(' ', '_')
        att = self.env['ir.attachment'].create({
            'name': fname,
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
