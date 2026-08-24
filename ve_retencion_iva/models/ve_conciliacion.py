import calendar
import json
import logging
import re
import urllib.parse
import urllib.request
import urllib.error
from datetime import date, timedelta

from markupsafe import Markup
from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Patrones de texto plano que el RPA (AET) postea directo al chatter al
# terminar una extracción SENIAT (ver message_post() más abajo). Lista
# abierta a propósito -- si el RPA cambia de wording o agrega un mensaje
# nuevo no listado acá, simplemente no se reformatea ni actualiza el
# estado (se queda como nota normal, no rompe nada), no hace falta que la
# lista sea exhaustiva del primer intento.
_RPA_PATRONES_RESULTADO_EXTRACCION = [
    (re.compile(r'registradas?\s*:?\s*\d+', re.IGNORECASE), 'completada'),
    (re.compile(r'no\s+se\s+encontraron\s+retenciones', re.IGNORECASE), 'completada'),
    (re.compile(r'no\s+se\s+logr[oó]|no\s+se\s+pudo|fall[oó]', re.IGNORECASE), 'fallo'),
]


class VeConciliacionPeriodo(models.Model):
    _name = 've.conciliacion.periodo'
    # mail.thread usa _description para armar el mensaje automático "X
    # created" del chatter al crear el registro (dispara porque
    # estado_extraccion tiene tracking=True) -- "Declaración IVA Clientes"
    # confundía al crear un período desde "Retenciones SENIAT por Período",
    # pantalla que no tiene nada que ver con declarar (pedido explícito de
    # la usuaria 2026-08-05). Este mismo modelo cubre ambas pantallas
    # (Conciliación SENIAT completa y Retenciones SENIAT por Período), así
    # que el texto tiene que servir para las dos.
    _description = 'Período de Conciliación IVA'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'periodo_retencion desc'
    _rec_name = 'periodo_retencion'

    ref = fields.Char(
        string='Referencia',
        copy=False,
        readonly=True,
        help='Número interno auto-generado (DECL-IVA/YYYY/NN).',
    )
    name = fields.Char(string='Nombre')
    company_id = fields.Many2one(
        'res.company',
        string='Empresa',
        required=True,
        default=lambda self: self.env.company,
        index=True,
    )
    periodo = fields.Char(string='Período Fiscal', required=True)
    periodo_retencion = fields.Char(
        string='Período Declaración',
        required=True,
        help='Formato: yyyy-mm 1Q ó yyyy-mm 2Q  (ej: 2026-04 1Q)',
    )
    fecha_inicio = fields.Date(string='Fecha Inicio', required=True)
    fecha_fin = fields.Date(string='Fecha Fin', required=True)

    estado = fields.Selection([
        ('borrador',   'Borrador'),
        ('en_proceso', 'En Proceso'),
        ('revision',   'En Revisión'),
        ('aprobado',   'Conciliación Aprobada'),
        ('declarado',  'Declarado SENIAT'),
    ], string='Estado', default='borrador')

    # Estado de la EXTRACCIÓN del RPA (independiente de `estado`, que es el
    # workflow de declaración). "Pendiente/Declarado" se muestra derivando
    # `estado` directamente en la vista, no hace falta un campo aparte.
    estado_extraccion = fields.Selection([
        ('pendiente',  'Pendiente'),
        ('iniciada',   'Iniciada'),
        ('completada', 'Completada'),
        ('fallo',      'Falló'),
    ], string='Estado Extracción SENIAT', default='pendiente', copy=False, tracking=True)
    fecha_estado_extraccion = fields.Datetime(
        string='Última Actualización Extracción', copy=False)
    mensaje_estado_extraccion = fields.Char(string='Detalle Extracción', copy=False)
    extraccion_repetidas = fields.Integer(
        string='Repetidas en la Extracción', default=0, copy=False,
        help='Cuántas retenciones de esta corrida ya existían (mismo N° '
             'Control + RIF Agente) -- el propio SENIAT a veces exporta la '
             'misma retención más de una vez en el mismo mes. Se resetea a '
             '0 cada vez que se pulsa Extraer SENIAT.',
    )
    extraccion_repetidas_detalle = fields.Text(
        string='Detalle de Repetidas', copy=False,
        help='N° Control + RIF Agente de cada retención que llegó repetida '
             'en esta corrida (ver extraccion_repetidas) -- antes solo se '
             'guardaba el conteo, sin forma de saber cuáles eran. Se '
             'resetea junto con extraccion_repetidas al pulsar Extraer '
             'SENIAT (pedido explícito de la usuaria 2026-08-05).',
    )
    extraccion_estancada = fields.Boolean(
        string='Extracción Estancada', compute='_compute_extraccion_estancada',
        help='Iniciada hace más de 1 hora sin resultado — el RPA todavía no '
             'reporta fallas a mitad de camino, esto es una señal provisional '
             'de que puede haberse colgado.',
    )
    notas_extraccion = fields.Char(
        string='Notas', compute='_compute_notas_extraccion',
        help='Mensaje de la última extracción SENIAT (RPA), o aviso de que '
             'lleva más de 1 hora sin respuesta.',
    )
    rifs_seniat_no_spe = fields.Text(
        string='RIFs SENIAT sin Contribuyente Especial',
        copy=False,
        help='RIF — Nombre, uno por línea. Se llena al cargar Retenciones '
             'SENIAT (XLSX o RPA) cuando aparece un RIF con retención '
             'SENIAT real pero el cliente NO está marcado como Contribuyente '
             'Especial en Odoo -- contradicción entre el Libro de Ventas del '
             'cliente y SENIAT que amerita revisión manual antes de marcarlo '
             '(pedido explícito 2026-08-18, ver res_partner.py::'
             '_detectar_agentes_retencion_por_rif). Se pisa completo en cada '
             'carga nueva (no se acumula), igual que mensaje_estado_extraccion. '
             'Texto plano para el chatter/Notas -- ver también '
             'partners_seniat_no_spe_ids (mismo dato, como relación real para '
             'el botón "Ver Ret. no SPE").',
    )
    partners_seniat_no_spe_ids = fields.Many2many(
        'res.partner', string='Clientes SENIAT sin Contribuyente Especial',
        copy=False,
        help='Mismos clientes que rifs_seniat_no_spe, como relación real (no '
             'texto) para que el botón "Ver Ret. no SPE" pueda abrir una '
             'lista de verdad -- pedido explícito 2026-08-18, el chatter no '
             'era práctico para revisar esto.',
    )

    wh_iva_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id', string='Retenciones Odoo')
    wh_iva_no_recibidas_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Retenciones No Recibidas',
        domain=[('state', 'in', ['esperado', 'vencido'])],
    )
    wh_iva_recibidas_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Retenciones Recibidas',
        domain=[('state', 'in', ['borrador', 'confirmado']),
                ('estado_declaracion', '=', 'no_declarado')],
    )
    wh_iva_conciliadas_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Retenciones Listas para Declarar',
        domain=[('state', '=', 'confirmado'),
                ('estado_conciliacion', 'in',
                 ['listo_declarar', 'conciliada_norec', 'conciliada', 'declarado', 'aprobado_declarar']),
                ('estado_declaracion', '=', 'no_declarado')],
    )
    wh_iva_declaradas_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Comprobantes Declarados',
        domain=[('estado_declaracion', '=', 'declarado')],
    )
    wh_iva_activas_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Comprobantes Activos',
        domain=[('state', '!=', 'anulado')],
    )
    wh_iva_periodo_ids = fields.One2many(
        've.wh.iva', 'conciliacion_id',
        string='Facturación Período',
        compute='_compute_wh_iva_periodo_ids',
    )
    seniat_ids = fields.One2many(
        've.seniat.retencion', 'conciliacion_id', string='Retenciones SENIAT')
    declaracion_iva_id = fields.Many2one(
        've.declaracion.iva',
        string='Declaración IVA',
        copy=False,
        readonly=True,
        help='Declaración IVA (Forma 30) asociada a este período. Se crea automáticamente.',
    )
    fecha_declaracion = fields.Datetime(
        string='Fecha Declaración',
        related='declaracion_iva_id.fecha_declaracion',
        store=True,
    )

    total_odoo = fields.Float(
        string='Total Comp. Esperados/Recibidos',
        compute='_compute_totales', store=True, digits=(16, 2))
    total_recibidos = fields.Float(
        string='Total Comp. Recibidos',
        compute='_compute_totales', store=True, digits=(16, 2))
    total_seniat = fields.Float(
        string='Total Comp. en SENIAT',
        compute='_compute_totales', store=True, digits=(16, 2))
    n_seniat = fields.Integer(
        string='N.Retenciones', compute='_compute_totales', store=True,
        help='Cantidad de retenciones SENIAT vinculadas a este período -- '
             'no confundir con "Retenciones SENIAT" (total_seniat), que es '
             'el monto en Bs, no la cantidad de filas.')
    diferencia = fields.Float(
        string='Dif. Recibido-SENIAT',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Total Recibidos − Total en SENIAT')
    diferencia_esperado_seniat = fields.Float(
        string='Dif. Esperado/Recibido-SENIAT',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Total Esperados+Recibidos − Total en SENIAT')
    total_no_recibido_prev = fields.Float(
        string='No Recibidos Per. Anterior',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Comprobantes de períodos anteriores arrastrados a este período')
    conciliadas = fields.Integer(
        string='Conciliadas', compute='_compute_totales', store=True)
    monto_conciliado = fields.Float(
        string='Monto Conciliado',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Suma de monto_retenido de los comprobantes en estado '
             'Conciliada/Conciliada sin Recibir/Listo Declarar/Declarado/'
             'Aprobado Declarar -- mismo criterio que "Conciliadas" pero en '
             'monto en vez de cantidad.')
    monto_conciliado_seniat = fields.Float(
        string='Conciliado del Período',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Parte de "SENIAT" (total_seniat) de este período cuyo match '
             'es una retención Odoo que TAMBIÉN pertenece a este período. '
             'Junto con Conciliado Fuera de Período + Conciliado C/Dif + '
             'No Conciliadas, suma exacto el total SENIAT de la fila (el '
             'match es por universo de compañía, no por período -- ver '
             '_do_conciliar).')
    monto_conciliado_fuera_periodo = fields.Float(
        string='Conciliado Fuera de Período',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Parte de "SENIAT" de este período cuyo match es una '
             'retención Odoo de OTRO período (match por universo de '
             'compañía). Ese monto se cuenta como "Conciliado del '
             'Período" en la fila del período al que sí pertenece la '
             'retención Odoo.')
    monto_conciliado_diferencia = fields.Float(
        string='Conciliado C/Dif',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Parte de "SENIAT" de este período que sí encontró match en '
             'Odoo, pero el monto no coincide (estado Diferencia) -- no '
             'se cuenta como conciliado limpio en ningún período.')
    monto_sin_match_seniat = fields.Float(
        string='No Conciliadas',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Parte de "SENIAT" de este período sin ninguna retención '
             'Odoo vinculada todavía.')
    count_normalizados = fields.Integer(
        string='Conciliadas por Normalización', compute='_compute_totales', store=True,
        help='Cuántas retenciones de este período solo conciliaron con el '
             'SENIAT después de normalizar N° de Control/Factura (quitar '
             'ceros a la izquierda, guiones, prefijos) -- el texto crudo no '
             'coincidía. Pedido explícito 2026-08-11.')
    monto_normalizado = fields.Float(
        string='Monto Conciliado por Normalización',
        compute='_compute_totales', store=True, digits=(16, 2),
        help='Monto retenido de las retenciones de count_normalizados.')
    solo_seniat = fields.Integer(
        string='Solo en SENIAT', compute='_compute_totales', store=True)
    solo_odoo = fields.Integer(
        string='Solo en Odoo', compute='_compute_totales', store=True)
    con_diferencia = fields.Integer(
        string='Con Diferencia', compute='_compute_totales', store=True)
    total_comp_recibidos = fields.Integer(
        string='Comp. Recibidos', compute='_compute_totales', store=True)
    total_comp_por_recibir = fields.Integer(
        string='Comp. Por Recibir', compute='_compute_totales', store=True)
    total_comp_activos = fields.Integer(
        string='Total Comprobantes', compute='_compute_totales', store=True,
        help='Total de comprobantes activos (No Recibidos + Recibidos, excluye Anulados).')

    aprobado_por = fields.Many2one('res.users', string='Aprobado por')
    fecha_aprobacion = fields.Datetime(string='Fecha Aprobación')
    invoice_ids = fields.Many2many(
        'account.move',
        compute='_compute_invoice_ids',
        string='Facturas',
    )

    @api.depends('wh_iva_ids.invoice_id')
    def _compute_invoice_ids(self):
        for rec in self:
            rec.invoice_ids = rec.wh_iva_ids.mapped('invoice_id').filtered(lambda i: i)

    @api.depends('wh_iva_ids', 'wh_iva_ids.periodo', 'periodo')
    def _compute_wh_iva_periodo_ids(self):
        for rec in self:
            rec.wh_iva_periodo_ids = rec.wh_iva_ids.filtered(
                lambda r: r.periodo == rec.periodo
            )

    # ── Alertas contextuales para Declaración IVA ───────────────────────────
    alerta_vencidos   = fields.Char(compute='_compute_alertas', store=False)
    alerta_esperados  = fields.Char(compute='_compute_alertas', store=False)
    alerta_borradores = fields.Char(compute='_compute_alertas', store=False)

    @api.depends('wh_iva_ids.state', 'wh_iva_ids.monto_retenido')
    def _compute_alertas(self):
        for rec in self:
            vencidos   = rec.wh_iva_ids.filtered(lambda r: r.state == 'vencido')
            esperados  = rec.wh_iva_ids.filtered(lambda r: r.state == 'esperado')
            borradores = rec.wh_iva_ids.filtered(lambda r: r.state == 'borrador')
            sv = sum(vencidos.mapped('monto_retenido'))
            se = sum(esperados.mapped('monto_retenido'))
            sb = sum(borradores.mapped('monto_retenido'))
            rec.alerta_vencidos = (
                f'{len(vencidos)} comprobante(s) vencido(s) sin recibir — '
                f'Bs. {sv:,.2f}'
            ) if vencidos else ''
            rec.alerta_esperados = (
                f'{len(esperados)} comprobante(s) pendiente(s) de recibir — '
                f'Bs. {se:,.2f}'
            ) if esperados else ''
            rec.alerta_borradores = (
                f'{len(borradores)} comprobante(s) recibido(s) sin confirmar — '
                f'Bs. {sb:,.2f}'
            ) if borradores else ''

    @staticmethod
    def _fechas_desde_periodo_retencion(periodo_retencion):
        """Calcula fecha_inicio y fecha_fin desde el período.
        Formatos soportados: 'yyyy-mm 1Q', 'yyyy-mm 2Q', 'yyyy-mm'."""
        if not periodo_retencion:
            return None, None
        m = re.match(r'^(\d{4})-(\d{2})(?:\s+(1Q|2Q))?$', periodo_retencion.strip())
        if not m:
            return None, None
        year, month, quinquena = int(m.group(1)), int(m.group(2)), m.group(3)
        last_day = calendar.monthrange(year, month)[1]
        if quinquena == '1Q':
            return (
                fields.Date.from_string(f'{year:04d}-{month:02d}-01'),
                fields.Date.from_string(f'{year:04d}-{month:02d}-15'),
            )
        elif quinquena == '2Q':
            return (
                fields.Date.from_string(f'{year:04d}-{month:02d}-16'),
                fields.Date.from_string(f'{year:04d}-{month:02d}-{last_day:02d}'),
            )
        else:
            return (
                fields.Date.from_string(f'{year:04d}-{month:02d}-01'),
                fields.Date.from_string(f'{year:04d}-{month:02d}-{last_day:02d}'),
            )

    def _asegurar_periodo(self, company, fecha):
        """Busca o crea el período quincenal que corresponde a `fecha` —
        NO necesariamente "hoy". Usado por CONECTA-14 (Ventas/Compras)
        para vincular cada retención al período que le toca por su
        propia fecha de factura, en vez de barrer todo al período activo
        de HOY sin mirar la fecha real de cada fila.

        Bug real encontrado 2026-07-29: una carga con facturas de 2
        quincenas distintas (ej. 1Q y 2Q del mismo mes) dejaba TODAS las
        retenciones vinculadas a la quincena de hoy — la Declaración de
        esa quincena terminaba contando C.66 (retenciones) de facturas
        cuyo débito fiscal (C.42/43/49) no aparecía ahí, porque ese
        cálculo vuelve a filtrar por `invoice_date` contra el rango de
        fechas real de la quincena (ver `ve_declaracion_iva.py::
        _compute_reporte_seniat`) — una Declaración internamente
        inconsistente consigo misma."""
        Periodo = self.env['ve.conciliacion.periodo'].sudo()
        if isinstance(fecha, str):
            fecha = fields.Date.from_string(fecha)
        if not fecha:
            fecha = fields.Date.today()
        quincena = '1Q' if fecha.day <= 15 else '2Q'
        periodo_retencion = f'{fecha.year:04d}-{fecha.month:02d} {quincena}'
        periodo = Periodo.search([
            ('periodo_retencion', '=', periodo_retencion),
            ('company_id', '=', company.id),
        ], limit=1)
        if periodo:
            return periodo
        ini, fin = self._fechas_desde_periodo_retencion(periodo_retencion)
        # Contexto que desactiva el barrido de "todas las huérfanas de la
        # compañía" del create() de este modelo — acá se asigna cada
        # huérfana explícitamente por su propia fecha (ver el llamador),
        # no en bloque (ver comentario en create()).
        return Periodo.with_context(ve_periodo_asignacion_manual=True).create({
            'periodo': f'{fecha.year:04d}-{fecha.month:02d}',
            'periodo_retencion': periodo_retencion,
            'fecha_inicio': ini,
            'fecha_fin': fin,
            'company_id': company.id,
        })

    @api.onchange('periodo_retencion')
    def _onchange_periodo_retencion(self):
        ini, fin = self._fechas_desde_periodo_retencion(self.periodo_retencion)
        if ini:
            self.fecha_inicio = ini
            self.fecha_fin = fin
        # Deriva "Período Fiscal" (yyyy-mm) del prefijo de periodo_retencion
        # (yyyy-mm 1Q/2Q) -- creación manual vía el botón "Nuevo" nativo
        # solo necesita escribir un campo, no dos consistentes entre sí.
        m = re.match(r'^(\d{4}-\d{2})', (self.periodo_retencion or '').strip())
        if m:
            self.periodo = m.group(1)

    @api.model_create_multi
    def create(self, vals_list):
        seq = self.env['ir.sequence'].next_by_code
        for vals in vals_list:
            if not vals.get('ref'):
                vals['ref'] = seq('ve.conciliacion.periodo') or ''
            if not vals.get('name'):
                vals['name'] = vals.get('periodo_retencion') or vals.get('periodo') or ''
            # Auto-inicializar fechas si no se proporcionaron
            if vals.get('periodo_retencion') and not vals.get('fecha_inicio'):
                ini, fin = self._fechas_desde_periodo_retencion(vals['periodo_retencion'])
                if ini:
                    vals['fecha_inicio'] = ini
                    vals['fecha_fin'] = fin
        records = super().create(vals_list)
        # `_asegurar_periodo()` (usado por CONECTA-14 para vincular cada
        # retención al período que le corresponde por su PROPIA fecha)
        # crea períodos uno por uno y asigna cada huérfana explícitamente
        # después — el barrido de abajo, si corriera acá también, agarraría
        # TODAS las huérfanas de la compañía (incluidas las de OTRA
        # quincena, aún sin asignar en ese momento) apenas se crea el
        # primer período nuevo, arruinando el reparto por fecha. Se
        # desactiva solo ese barrido con este contexto — bug real
        # 2026-07-29 — el resto (declaración companion, vincular SENIAT)
        # sigue corriendo igual.
        barrer_huerfanas = not self.env.context.get('ve_periodo_asignacion_manual')
        for rec in records:
            if barrer_huerfanas:
                # Toma TODAS las retenciones pendientes sin conciliar de la
                # MISMA compañía (incluye períodos anteriores) — sin el
                # filtro de company_id, un período nuevo de otra compañía
                # (ej. sandbox de piloto) se apropiaría de retenciones
                # huérfanas ajenas (MULTI-02).
                retenciones = self.env['ve.wh.iva'].search([
                    ('conciliacion_id', '=', False),
                    ('estado_conciliacion', '=', 'pendiente'),
                    ('company_id', '=', rec.company_id.id),
                ])
                if retenciones:
                    retenciones.write({'conciliacion_id': rec.id})
            rec._vincular_seniat_sin_link()
            # Crear declaración IVA companion si no existe
            self.env['ve.declaracion.iva'].sudo()._get_or_create_for_periodo(rec.id)
        return records

    @api.depends('estado_extraccion', 'fecha_estado_extraccion')
    def _compute_extraccion_estancada(self):
        limite = fields.Datetime.now() - timedelta(hours=1)
        for rec in self:
            rec.extraccion_estancada = bool(
                rec.estado_extraccion == 'iniciada'
                and rec.fecha_estado_extraccion
                and rec.fecha_estado_extraccion < limite
            )

    @api.depends('mensaje_estado_extraccion', 'extraccion_estancada',
                 'extraccion_repetidas_detalle', 'rifs_seniat_no_spe')
    def _compute_notas_extraccion(self):
        for rec in self:
            partes = []
            if rec.rifs_seniat_no_spe:
                partes.append('Revisar retenciones registradas en SENIAT no SPE')
            if rec.extraccion_estancada:
                partes.append('Sin respuesta del RPA hace más de 1 hora — revisar manualmente.')
            if rec.mensaje_estado_extraccion:
                partes.append(rec.mensaje_estado_extraccion)
            # El detalle completo (N° Control + N° Documento + RIF de cada
            # repetida) vive en el chatter, no acá -- Notas solo remite ahí
            # (pedido explícito de la usuaria 2026-08-05, para no duplicar
            # el mismo texto largo en 2 lugares distintos de la ficha).
            if rec.extraccion_repetidas_detalle:
                partes.append('ver Chatter para detalle de las repetidas')
            rec.notas_extraccion = ' · '.join(partes) if partes else False

    @api.constrains('company_id', 'periodo_retencion')
    def _check_periodo_unico(self):
        # Chequeo a nivel ORM (no _sql_constraints) a propósito: una
        # restricción SQL dura podría romper el deploy si ya existieran
        # duplicados cargados antes de este cambio (ver memoria del
        # proyecto sobre riesgos de migración). Esto protege toda escritura
        # nueva sin tocar filas existentes.
        for rec in self:
            dup = self.search([
                ('id', '!=', rec.id),
                ('company_id', '=', rec.company_id.id),
                ('periodo_retencion', '=', rec.periodo_retencion),
            ], limit=1)
            if dup:
                raise UserError(
                    f'Ya existe un período "{rec.periodo_retencion}" para '
                    f'{rec.company_id.name} (ID {dup.id}).')

    @api.depends(
        'wh_iva_ids.monto_retenido', 'wh_iva_ids.monto_recibido', 'wh_iva_ids.monto_seniat',
        'wh_iva_ids.estado_conciliacion', 'wh_iva_ids.periodo',
        'wh_iva_ids.incluir_declaracion', 'wh_iva_ids.state', 'wh_iva_ids.estado_recepcion',
        'wh_iva_ids.matched_por_normalizacion',
        'seniat_ids.estado', 'seniat_ids.monto_retenido',
        'seniat_ids.wh_iva_id', 'seniat_ids.wh_iva_id.conciliacion_id')
    def _compute_totales(self):
        for rec in self:
            activos = rec.wh_iva_ids.filtered(lambda r: r.state != 'anulado')
            # Separar retenciones del período actual de las de períodos
            # anteriores -- pedido explícito 2026-08-11: comparar por
            # QUINCENA real (día <=15/>15 de la fecha de factura), no solo
            # por mes. `r.periodo` es 'yyyy-mm' sin quincena, así que una
            # retención Vencida que se corrió de 1Q a 2Q del MISMO mes
            # (ver ve_conciliacion.py::_do_conciliar, barrida de
            # "pendientes_ant") seguía marcando `r.periodo == rec.periodo`
            # como verdadero en el período 2Q -- se contaba como "del
            # período actual" en vez de "arrastrada de un período
            # anterior", dejando total_no_recibido_prev siempre en 0 para
            # este caso, el más común. `r.periodo_retencion` NO sirve para
            # esto -- es un compute que refleja conciliacion_id.
            # periodo_retencion (SIEMPRE el período actual), no la
            # quincena original de la factura.
            def _quincena_real(r, _rec=rec):
                fecha_ref = r.invoice_id.invoice_date if r.invoice_id else False
                if not fecha_ref:
                    # Sin factura vinculada no hay forma de saber el día
                    # real -- se asume que pertenece al período actual (no
                    # se puede hacer mejor sin esa fecha).
                    return _rec.periodo_retencion
                quincena = '1Q' if fecha_ref.day <= 15 else '2Q'
                return f'{fecha_ref.year:04d}-{fecha_ref.month:02d} {quincena}'
            activos_periodo = activos.filtered(lambda r: _quincena_real(r) == rec.periodo_retencion)
            activos_prev    = activos - activos_periodo
            recibidos = activos.filtered(lambda r: r.estado_recepcion in r._RECIBIDO_ESTADOS)
            rec.total_odoo             = sum(activos_periodo.mapped('monto_retenido'))
            rec.total_no_recibido_prev = sum(activos_prev.mapped('monto_retenido'))
            rec.total_recibidos        = sum(recibidos.mapped('monto_recibido'))
            # total_seniat = suma directa de todos los registros SENIAT del período
            # (visible incluso antes de conciliar; no depende del match)
            rec.total_seniat    = sum(rec.seniat_ids.mapped('monto_retenido'))
            rec.n_seniat        = len(rec.seniat_ids)
            rec.diferencia                 = rec.total_recibidos - rec.total_seniat
            rec.diferencia_esperado_seniat = rec.total_odoo - rec.total_seniat
            conciliadas_recs = activos.filtered(
                lambda r: r.estado_conciliacion in (
                    'conciliada', 'conciliada_norec', 'listo_declarar', 'declarado',
                    'aprobado_declarar'))
            rec.conciliadas = len(conciliadas_recs)
            rec.monto_conciliado = sum(conciliadas_recs.mapped('monto_retenido'))
            # ── Partición exhaustiva de total_seniat (pedido explícito
            # 2026-08-18): a diferencia de monto_conciliado (arriba, mira
            # wh_iva_ids -- las retenciones de ESTE período), estas 4
            # miran seniat_ids -- los SENIAT homed a ESTE período -- y por
            # construcción SIEMPRE suman exacto total_seniat, sin importar
            # si el match cae en este período, en otro, o no hay match.
            seniat_conciliado = rec.seniat_ids.filtered(lambda s: s.estado == 'conciliado')
            seniat_diferencia = rec.seniat_ids.filtered(lambda s: s.estado == 'diferencia')
            seniat_del_periodo = seniat_conciliado.filtered(
                lambda s: s.wh_iva_id and s.wh_iva_id.conciliacion_id.id == rec.id)
            seniat_fuera_periodo = seniat_conciliado - seniat_del_periodo
            seniat_no_conciliadas = rec.seniat_ids - seniat_conciliado - seniat_diferencia
            rec.monto_conciliado_seniat = sum(seniat_del_periodo.mapped('monto_retenido'))
            rec.monto_conciliado_fuera_periodo = sum(seniat_fuera_periodo.mapped('monto_retenido'))
            rec.monto_conciliado_diferencia = sum(seniat_diferencia.mapped('monto_retenido'))
            rec.monto_sin_match_seniat = sum(seniat_no_conciliadas.mapped('monto_retenido'))
            rec.solo_odoo = len(activos.filtered(
                lambda r: r.estado_conciliacion == 'solo_odoo'))
            rec.solo_seniat = len(rec.seniat_ids.filtered(
                lambda s: s.estado == 'sin_match'))
            rec.con_diferencia = len(activos.filtered(
                lambda r: r.estado_conciliacion == 'diferencia'))
            rec.total_comp_recibidos  = len(recibidos)
            rec.total_comp_por_recibir = len(activos.filtered(
                lambda r: r.state in ('esperado', 'vencido')))
            rec.total_comp_activos = len(activos)
            normalizadas = activos.filtered('matched_por_normalizacion')
            rec.count_normalizados = len(normalizadas)
            rec.monto_normalizado = sum(normalizadas.mapped('monto_retenido'))

    def action_reiniciar_demo(self):
        self.ensure_one()
        return {
            'name': 'Reinicializar Demo',
            'type': 'ir.actions.act_window',
            'res_model': 've.reset.demo.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_conciliacion_id': self.id},
        }

    def action_cargar_seniat_csv(self):
        self.ensure_one()
        return {
            'name': 'Cargar Retenciones SENIAT desde CSV',
            'type': 'ir.actions.act_window',
            'res_model': 've.seniat.wizard.carga',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_conciliacion_id': self.id},
        }

    def message_post(self, **kwargs):
        """Intercepta las notas de resultado que el RPA (AET) postea
        DIRECTO al chatter con su propia sesión — no pasa por ningún
        endpoint de este módulo (confirmado 2026-08-04: el mensaje de
        cierre real de Cementos es "Número de retenciones IVA registradas:
        N", no el JSON estructurado de cargar_retenciones que se asumía
        originalmente). El texto viene envuelto en HTML simple del propio
        AET (ej. `<span style="color:...">...</span>`, no texto plano
        puro) -- se le quitan las etiquetas antes de matchear. Si el
        resultado matchea un patrón conocido: (1) actualiza
        estado_extraccion/fecha/mensaje igual que si hubiera llegado por
        el endpoint, (2) reemplaza el mensaje por la misma tarjeta con la
        que ya se postea "Solicitud RPA — Extracción SENIAT" al iniciar,
        para que se vea consistente. Mensajes que no matchean ningún
        patrón (notas manuales, otros mensajes del sistema, o las propias
        tarjetas "Solicitud RPA"/"Resultado RPA" que este módulo arma, que
        nunca contienen estas palabras clave) pasan sin tocar."""
        body = kwargs.get('body')
        texto = body if isinstance(body, str) else None
        if texto:
            texto = re.sub(r'<[^>]+>', ' ', texto)
            texto = re.sub(r'\s+', ' ', texto).strip()
            resultado = None
            for patron, valor in _RPA_PATRONES_RESULTADO_EXTRACCION:
                if patron.search(texto):
                    resultado = valor
                    break
            if resultado:
                now_fmt = fields.Datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                # El RPA extrae el MES completo (ambas quincenas) pero este
                # mensaje llega a UN período puntual -- "N registradas" del
                # texto de AET es el total del mes, no lo que realmente
                # quedó vinculado a ESTE período tras repartirse por fecha
                # real entre 1Q/2Q. Se muestran ambos números para que no
                # se lea como una discrepancia (pedido explícito de la
                # usuaria 2026-08-04, tras el bug real de la quincena).
                extraidas_match = re.search(r'(\d+)', texto)
                extraidas = extraidas_match.group(1) if extraidas_match else None

                def _armar_mensaje(periodo_rec):
                    # "repetidas" explica por qué extraídas != suma de
                    # vinculadas entre 1Q/2Q -- confirmado 2026-08-04
                    # (Cementos, enero) contra el archivo real de SENIAT:
                    # el propio SENIAT exporta la misma retención (mismo
                    # N° Control + RIF Agente) más de una vez en el mismo
                    # mes en algunos casos; Odoo la reconoce y no la
                    # duplica -- no es un error de carga.
                    vinculadas = self.env['ve.seniat.retencion'].search_count(
                        [('conciliacion_id', '=', periodo_rec.id)])
                    # "en el archivo" era impreciso acá -- a diferencia del
                    # wizard XLSX (que sí sube un archivo real), este camino
                    # es el RPA: no hay archivo, es SENIAT devolviendo la
                    # misma retención más de una vez dentro de la misma
                    # extracción (pedido explícito de la usuaria 2026-08-05).
                    rep_txt = (f', {periodo_rec.extraccion_repetidas} repetidas en la extracción SENIAT'
                               if periodo_rec.extraccion_repetidas else '')
                    if extraidas:
                        return f'{extraidas} extraídas del mes, {vinculadas} vinculadas a este período{rep_txt}'
                    return f'{texto} — {vinculadas} vinculadas a este período{rep_txt}'

                def _card_resultado(periodo_rec, msg):
                    # Detalle de repetidas (N° Control + N° Documento + RIF
                    # de cada una) va en el chatter, dentro de la misma
                    # tarjeta -- antes vivía como campo siempre visible en
                    # la ficha; la usuaria pidió moverlo acá para no
                    # duplicar texto largo en 2 lugares (2026-08-05).
                    partes = [
                        '<b>Resultado RPA — Extracción SENIAT</b><br/>',
                        f'<b>Fecha / Hora:</b> {now_fmt} UTC<br/>',
                        f'<b>{Markup.escape(msg)}</b>',
                    ]
                    if periodo_rec.extraccion_repetidas_detalle:
                        lineas_html = '<br/>'.join(
                            Markup.escape(l) for l in
                            periodo_rec.extraccion_repetidas_detalle.split('\n'))
                        partes.append(
                            f'<br/><br/><b>Detalle de repetidas:</b><br/>{lineas_html}')
                    return Markup(''.join(partes))

                def _no_spe_datos(periodo_rec):
                    # Mismo criterio que wizard_carga_seniat.py -- acá se
                    # evalúa en el punto de cierre real (cuando llega el
                    # mensaje final de AET), no por cada retención
                    # individual que llama api__rpa.py (llamado 1 por 1,
                    # ver comentario grande más arriba en este mismo
                    # archivo/api__rpa.py), sobre TODAS las retenciones ya
                    # vinculadas a este período. Devuelve (texto, recordset)
                    # -- texto para el chatter, recordset para el botón
                    # "Ver Ret. no SPE" (pedido explícito 2026-08-18).
                    rifs = set(periodo_rec.seniat_ids.mapped('rif_agente'))
                    no_spe = self.env['res.partner']._detectar_agentes_retencion_por_rif(
                        periodo_rec.company_id, rifs)
                    texto = '\n'.join(f'{p.vat} — {p.name}' for p in no_spe) if no_spe else False
                    return texto, no_spe

                mensaje = _armar_mensaje(self)
                texto_no_spe, partners_no_spe = _no_spe_datos(self)
                self.write({
                    'estado_extraccion': resultado,
                    'fecha_estado_extraccion': fields.Datetime.now(),
                    'mensaje_estado_extraccion': mensaje,
                    'rifs_seniat_no_spe': texto_no_spe,
                    'partners_seniat_no_spe_ids': [(6, 0, partners_no_spe.ids)],
                })
                # El RPA solo conoce el ID del período desde el que se
                # pulsó "Extraer SENIAT" (Odoo_Conciliacion_ID) -- si a
                # mitad de la extracción se creó un período HERMANO (ej.
                # 2Q auto-creado por _asegurar_periodo al llegar la
                # primera retención con fecha ≥16), AET nunca se entera de
                # su ID y su mensaje final JAMÁS le llega directamente. Sin
                # esto, ese hermano queda para siempre en 'iniciada' pese a
                # haber recibido datos reales (bug real 2026-08-04,
                # Cementos, extracción de enero: período 2Q con 1057
                # retenciones vinculadas, nunca marcado 'completada'). Se
                # infiere que pertenece a la MISMA corrida por compañía +
                # mismo mes fiscal (`periodo`) y se cierra igual.
                # Bug real encontrado 2026-08-05 (Cementos, re-extracción de
                # 2025-12): el filtro estado_extraccion=='iniciada' solo
                # cubre el caso "hermano recién creado". Si se REPITE una
                # extracción de un mes ya cargado antes (ej. reprocesar
                # 2025-12 pulsando Extraer SENIAT desde 1Q), la 2Q ya estaba
                # en 'completada' desde la corrida anterior -- nunca pasa
                # por 'iniciada' en esta corrida nueva, así que quedaba
                # fuera del cierre y se congelaba con el dato viejo aunque
                # AET sí reprocesó el mes completo. AET siempre procesa el
                # MES entero sin importar el estado previo de cada quincena,
                # así que el cierre debe cubrir cualquier hermano no
                # declarado (declarado = inmutable, no se toca).
                hermanos = self.search([
                    ('id', '!=', self.id),
                    ('company_id', '=', self.company_id.id),
                    ('periodo', '=', self.periodo),
                    ('estado', '!=', 'declarado'),
                ])
                for hermano in hermanos:
                    mensaje_hermano = (
                        f'{_armar_mensaje(hermano)} '
                        f'(inferido del resultado de {self.periodo_retencion})'
                    )
                    texto_no_spe_h, partners_no_spe_h = _no_spe_datos(hermano)
                    hermano.write({
                        'estado_extraccion': resultado,
                        'fecha_estado_extraccion': fields.Datetime.now(),
                        'mensaje_estado_extraccion': mensaje_hermano,
                        'rifs_seniat_no_spe': texto_no_spe_h,
                        'partners_seniat_no_spe_ids': [(6, 0, partners_no_spe_h.ids)],
                    })
                    # El hermano no recibía NINGÚN mensaje propio en su
                    # chatter antes (solo el período desde el que se pulsó
                    # "Extraer SENIAT" tenía la tarjeta) -- si el hermano
                    # tuvo sus propias repetidas, quedaban sin rastro visible.
                    hermano.message_post(
                        body=_card_resultado(hermano, mensaje_hermano),
                        message_type='comment', subtype_xmlid='mail.mt_note',
                    )
                kwargs = dict(kwargs, body=_card_resultado(self, mensaje))
        return super().message_post(**kwargs)

    def action_extraer_seniat(self):
        self.ensure_one()
        if self.estado == 'declarado':
            raise UserError(
                f'El período {self.periodo_retencion} ya está declarado — '
                'no tiene sentido volver a extraer del SENIAT.'
            )
        cfg      = self.env['ir.config_parameter'].sudo()
        base_url = cfg.get_param('ve_retencion_iva.rpa_base_url', '').rstrip('/')
        username = cfg.get_param('ve_retencion_iva.rpa_username', '')
        password = cfg.get_param('ve_retencion_iva.rpa_password', '')

        missing = [k for k, v in {
            've_retencion_iva.rpa_base_url': base_url,
            've_retencion_iva.rpa_username':  username,
            've_retencion_iva.rpa_password':  password,
        }.items() if not v]
        if missing:
            raise UserError(
                'El RPA no está completamente configurado.\n'
                'Vaya a Ajustes → Parámetros del sistema y complete:\n'
                + '\n'.join(f'  {k}' for k in missing)
            )
        if not self.company_id.vat:
            raise UserError(
                f'La compañía "{self.company_id.name}" no tiene RIF configurado '
                '(Ajustes → Compañías) — el RPA lo necesita para identificar '
                'con qué credenciales entrar al SENIAT de este cliente.'
            )
        rif_cliente_aet = self.company_id.ve_rif_cliente_aet()

        auth_url    = f'{base_url}/authenticate'
        execute_url = f'{base_url}/execute'
        logout_url  = f'{base_url}/Logout'

        # ── Paso 1: autenticar (POST con credenciales en query string) ───────────
        qs = urllib.parse.urlencode({'username': username, 'password': password})
        auth_req = urllib.request.Request(
            f'{auth_url}?{qs}', data=b'',
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        try:
            with urllib.request.urlopen(auth_req, timeout=30) as resp:
                token_data = json.loads(resp.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            body = e.read()[:300].decode('utf-8', errors='replace')
            raise UserError(f'Error autenticando con el RPA (HTTP {e.code}): {body}')
        except Exception as e:
            raise UserError(f'No se pudo autenticar con el RPA: {e}')

        session_token = token_data.get('sessionToken', '')
        if not session_token:
            raise UserError(
                f'El RPA no retornó sessionToken. Respuesta: {str(token_data)[:300]}'
            )
        tenant   = token_data.get('tenant') or {}
        org_code = tenant.get('orgCode', '')
        user_id  = username   # login name: "controlador"
        _logger.info('RPA auth OK — orgCode=%r  userId=%r  token=%r',
                     org_code, user_id, session_token[:10] + '...')

        # ── Paso 2: ejecutar proceso de extracción ─────────────────────────────
        periodo   = self.periodo or ''          # formato yyyy-mm
        mes       = periodo[5:7] if len(periodo) >= 7 else ''
        ano       = periodo[:4]  if len(periodo) >= 4 else ''
        now       = fields.Datetime.now()
        source_id = f'Odoo-{mes}{ano}-{now.strftime("%H%M")}'

        odoo_base_url = cfg.get_param('web.base.url', '').rstrip('/')
        odoo_api_key  = cfg.get_param('ve_retencion_iva.rpa_api_key', '')

        payload = json.dumps({
            'orgCode':      org_code,
            'workflowName': 'Retenciones IVA x Clientes',
            'userId':       user_id,
            'sourceId':     source_id,
            'source':       'Odoo',
            'params': [
                {'name': 'Mes_Declaracion',      'value': mes,                           'type': 'String'},
                {'name': 'Ano_Declaracion',      'value': ano,                           'type': 'String'},
                {'name': 'Odoo_Conciliacion_ID', 'value': str(self.id),                  'type': 'String'},
                {'name': 'Odoo_Base_URL',        'value': odoo_base_url,                 'type': 'String'},
                {'name': 'Odoo_API_Key',         'value': odoo_api_key,                  'type': 'String'},
                # Dos niveles — AET tiene una BD Postgres por CLIENTE (ver
                # res.company.ve_rif_cliente, típicamente el RIF del
                # despacho contable) y dentro de esa BD, credenciales SENIAT
                # por EMPRESA (cada RIF real). RIF_Cliente le dice a AET en
                # qué BD buscar; RIF_Empresa, qué credenciales usar dentro
                # de esa BD. Revisado 2026-07-27: ya NO usa
                # res.company.parent_id (Sucursales nativas de Odoo) —
                # semánticamente mal (un cliente de despacho no es una
                # sucursal) y además Odoo bloquea cambiar parent_id una vez
                # que la compañía tiene contabilidad posteada.
                {'name': 'P_RIF_Cliente',        'value': rif_cliente_aet,              'type': 'String'},
                {'name': 'P_RIF_Empresa',        'value': self.company_id.vat,          'type': 'String'},
            ],
        }).encode('utf-8')

        exec_req = urllib.request.Request(
            execute_url, data=payload,
            headers={
                'Content-Type':    'application/json',
                'X-session-token': session_token,
            },
        )
        _logger.info('RPA execute payload: %s', payload.decode('utf-8'))
        try:
            with urllib.request.urlopen(exec_req, timeout=30) as resp:
                exec_data = json.loads(resp.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            body = e.read()[:300].decode('utf-8', errors='replace')
            self._rpa_logout(logout_url, session_token)
            raise UserError(f'Error ejecutando el RPA (HTTP {e.code}): {body}')
        except Exception as e:
            self._rpa_logout(logout_url, session_token)
            raise UserError(f'No se pudo ejecutar el RPA: {e}')

        _logger.info('RPA execute response: %s', json.dumps(exec_data))

        # ── Paso 3: liberar token ──────────────────────────────────────────────
        self._rpa_logout(logout_url, session_token)

        req_id        = exec_data.get('automationRequestId', '?')
        success       = exec_data.get('success', False)
        periodo_label = self.periodo_retencion or self.periodo
        now_fmt       = fields.Datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        if success:
            detalle = (
                'Las retenciones estarán disponibles en la pestaña '
                '<b>Retenciones SENIAT</b> una vez que el RPA complete la extracción, '
                'lo cual puede tomar unos minutos.'
            )
            estado_extraccion_nuevo = 'iniciada'
            mensaje_extraccion_nuevo = f'Solicitud aceptada — Request ID {req_id}'
        else:
            code   = exec_data.get('responseCode', '')
            detail = exec_data.get('errorDetails') or exec_data.get('responseDetails') or ''
            detalle = (
                '<span style="color:#7B0000;">'
                'El RPA rechazó la solicitud — verifique los parámetros de configuración.<br/>'
                f'Código: {code}'
                + (f'<br/>Detalle: {detail}' if detail else '')
                + '</span>'
            )
            estado_extraccion_nuevo = 'fallo'
            mensaje_extraccion_nuevo = f'Rechazado — código {code}' + (f': {detail}' if detail else '')

        self.write({
            'estado_extraccion': estado_extraccion_nuevo,
            'fecha_estado_extraccion': fields.Datetime.now(),
            'mensaje_estado_extraccion': mensaje_extraccion_nuevo,
            'extraccion_repetidas': 0,
            'extraccion_repetidas_detalle': False,
        })

        self.message_post(
            body=Markup(
                f'<b>Solicitud RPA — Extracción SENIAT</b><br/>'
                f'<b>Fecha / Hora:</b> {now_fmt} UTC<br/>'
                f'<b>Período:</b> {periodo_label}<br/>'
                f'<b>Request ID:</b> {req_id}<br/><br/>'
                f'{detalle}'
            ),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )
        # No devolver un ir.actions.act_window apuntando al mismo registro: el
        # botón se pulsa con el formulario YA abierto, y hacerlo apila un
        # nivel nuevo de breadcrumb cada vez (el título termina duplicando
        # "/ 2026-07 2Q / 2026-07 2Q..."). True basta para que Odoo refresque
        # el formulario en el sitio.
        return True

    def action_extraer_seniat_multi(self):
        """Extraer SENIAT para varios períodos seleccionados a la vez (acción
        de lista, ver binding_model_id en ve_conciliacion_views.xml). Los
        declarados se omiten en silencio (no tiene sentido re-extraerlos);
        los que fallan se acumulan y se reportan juntos al final para que un
        error no corte el resto del lote."""
        errores = []
        for rec in self:
            if rec.estado == 'declarado':
                continue
            try:
                rec.action_extraer_seniat()
            except UserError as e:
                errores.append(f'{rec.periodo_retencion}: {e}')
        if errores:
            raise UserError('Algunos períodos fallaron:\n' + '\n'.join(errores))

    @staticmethod
    def _rpa_logout(logout_url, session_token):
        """Libera el sessionToken. No lanza excepción si falla."""
        try:
            req = urllib.request.Request(
                logout_url, data=b'',
                headers={
                    'Content-Type':    'application/json',
                    'X-session-token': session_token,
                },
            )
            with urllib.request.urlopen(req, timeout=10):
                pass
        except Exception:
            pass

    def _vincular_seniat_sin_link(self):
        """Vincula retenciones SENIAT sin conciliacion_id que coincidan con
        este período -- solo asigna el "período hogar" para reportes/UI
        (ej. "Retenciones SENIAT por Período"), no afecta si puede
        conciliar (ver _do_conciliar, matching por universo de compañía).

        Bug real corregido 2026-08-16: el OR original (`periodo_retencion
        == X OR periodo == mes-de-X`) hacía que conciliar la quincena 1Q
        de un mes se robara TAMBIÉN las SENIAT de la 2Q del mismo mes que
        ya traían su propio periodo_retencion -- confirmado en el
        incidente 2026-08-05/06 de Cementos (corregido a mano por RPC en
        ese momento, nunca en el código). Ahora: si el registro SENIAT
        trae periodo_retencion (quincena exacta), matchea SOLO por eso; el
        fallback a periodo (mes, sin quincena) es exclusivamente para
        registros legado que no tienen periodo_retencion."""
        self.ensure_one()
        base = [('conciliacion_id', '=', False)]
        if self.periodo_retencion:
            domain = [('periodo_retencion', '=', self.periodo_retencion)] + base
        elif self.periodo:
            domain = [('periodo_retencion', '=', False),
                      ('periodo', '=', self.periodo)] + base
        else:
            return
        sin_link = self.env['ve.seniat.retencion'].search(domain)
        if sin_link:
            sin_link.write({'conciliacion_id': self.id})

    @staticmethod
    def _norm_rif(rif):
        return (rif or '').upper().replace('-', '').replace(' ', '').replace('.', '').strip()

    @staticmethod
    def _norm_ctrl(ctrl):
        """Clave canónica de N° de Control: solo dígitos, sin ceros a la
        izquierda. Pedido explícito 2026-08-11, medido con datos reales de
        Cementos (Enero 2026): el match literal anterior (solo strip+upper)
        solo encontraba 1.7% de las retenciones "Solo en SENIAT"; esta
        normalización sube eso a 72% -- el N° de Control que el cliente
        reporta al SENIAT casi nunca trae los mismos ceros de relleno ni
        guiones que el Libro de Ventas, pero el número real es el mismo."""
        digits = re.sub(r'\D', '', ctrl or '')
        return digits.lstrip('0') or ('0' if digits else '')

    @staticmethod
    def _norm_factura(fact):
        """Clave canónica de N° de Factura/Documento: quita prefijo
        alfabético y separadores, el ".0" final (artefacto típico de Excel
        al leer un número como float), y ceros a la izquierda. Mismo
        pedido/medición que _norm_ctrl -- usada como Nivel 2 (RIF+Factura)
        cuando la retención no tiene N° de Control."""
        fact = (fact or '').strip().upper()
        fact = re.sub(r'\.0$', '', fact)
        digits = re.sub(r'\D', '', fact)
        return digits.lstrip('0') or ('0' if digits else '')

    def action_conciliar(self):
        self.ensure_one()
        n_borrador = len(self.wh_iva_ids.filtered(lambda r: r.state == 'borrador'))
        if n_borrador:
            return {
                'type': 'ir.actions.act_window',
                'name': 'Confirmar Conciliación',
                'res_model': 've.conciliar.confirm.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_conciliacion_id': self.id,
                    'default_n_borrador': n_borrador,
                },
            }
        return self._do_conciliar()

    def _do_conciliar(self):
        self.ensure_one()
        self.estado = 'en_proceso'

        # Incorporar retenciones sin período asignado — scoped a la propia
        # compañía del período: sin este filtro, conciliar CUALQUIER
        # período de CUALQUIER compañía absorbía huérfanas de TODAS las
        # compañías de la base (bug real, mismo patrón que MULTI-01/02).
        #
        # Bug real encontrado 2026-07-30: este barrido metía TODAS las
        # huérfanas al período que se está conciliando en ESTE momento,
        # sin mirar la fecha de cada una — a diferencia de CONECTA-14
        # (ve_conecta_carga_ventas.py, ya corregido antes en la misma
        # sesión), este método no se había tocado. Una retención de 2026
        # que por algún motivo quedara huérfana, si el usuario conciliaba
        # el período 2025-07 1Q en ese momento, se la llevaba por delante
        # — quedaba con `periodo` (Período Fiscal) correcto en 2026-07
        # pero `conciliacion_id` apuntando a 2025-07 1Q. Ahora cada
        # huérfana se vincula al período que le corresponde por su propia
        # fecha (invoice_id.invoice_date, o el campo `periodo` como
        # respaldo si no hay factura vinculada) — ver ve.conciliacion.
        # periodo::_asegurar_periodo.
        sin_periodo = self.env['ve.wh.iva'].search([
            ('conciliacion_id', '=', False),
            ('estado_conciliacion', '=', 'pendiente'),
            ('company_id', '=', self.company_id.id),
        ])
        Periodo = self.env['ve.conciliacion.periodo']
        for wh in sin_periodo:
            fecha_ref = wh.invoice_id.invoice_date if wh.invoice_id else False
            if not fecha_ref and wh.periodo:
                try:
                    y, m = int(wh.periodo[:4]), int(wh.periodo[5:7])
                    fecha_ref = date(y, m, 15)
                except (ValueError, IndexError):
                    fecha_ref = False
            periodo_dest = Periodo._asegurar_periodo(self.company_id, fecha_ref) if fecha_ref else self
            wh.conciliacion_id = periodo_dest.id

        self._vincular_seniat_sin_link()

        # ── Conciliación por UNIVERSO de la compañía, no por período
        # (pedido explícito 2026-08-16): el criterio de match es
        # RIF+Control / RIF+Factura, nunca el período — así que el período
        # (conciliacion_id) de una retención Odoo NO se mueve para
        # "competir" (se queda fijo en el mes real de su factura, que es
        # lo que determina legalmente a qué Declaración pertenece, Art. 13
        # COT). Antes había un parche que arrastraba las pendientes
        # esperado/vencido de períodos cerrados anteriores hacia self.id
        # para que pudieran competir en el matching de self -- ya no hace
        # falta moverlas: el matching busca directo contra todo el universo
        # no conciliado de la compañía. Esto además cierra 2 huecos reales
        # que tenía el diseño anterior: (1) una confirmada/borrador que no
        # encontró match SENIAT en su propio período (`solo_odoo`) nunca se
        # reintentaba en corridas futuras de OTRO período; (2) una SENIAT
        # `sin_match` de un período ya cerrado nunca competía contra
        # confirmadas nuevas de otros períodos. Efecto secundario esperado:
        # correr "Conciliar SENIAT" en CUALQUIER período ahora puede
        # resetear/rematchear registros de OTROS períodos también -- es
        # intencional, no un bug.
        #
        # No resetear comprobantes ya declarados — bloqueados
        # permanentemente. estado_declaracion (no `state`) es la fuente de
        # verdad desde la Etapa 3 del rediseño de 3 ejes: declarar ya no
        # mueve `state` a 'declarado', así que un registro declarado normal
        # sigue en state='confirmado' — sin este filtro por
        # estado_declaracion, volvería a entrar aquí y "Conciliar SENIAT"
        # le pisaría el match ya cerrado (riesgo real detectado al diseñar
        # ese rediseño).
        WhIva = self.env['ve.wh.iva']
        SeniatRet = self.env['ve.seniat.retencion']
        # Participan en el match: confirmado, borrador, esperado, vencido —
        # y NUNCA un registro ya declarado (mismo motivo que el filtro de
        # arriba).
        # - confirmado/borrador: SENIAT OK → listo_declarar; C.66 usa monto_recibido
        # - esperado/vencido: SENIAT OK → conciliada_norec (C.66 = 0 por defecto;
        #   si el usuario activa incluir_declaracion usa monto_retenido/esperado)
        #
        # Excluir también 'listo_declarar'/'conciliada_norec' -- bug real
        # confirmado 2026-08-18 (Vencement, 9.731 de 9.732 matches rotos):
        # el lado SENIAT se congela una vez `estado == 'conciliado'` (línea
        # de abajo, `!= 'conciliado'`), pero el lado Odoo se reseteaba
        # SIEMPRE, sin excepción. Resultado: cada corrida de "Conciliar
        # SENIAT" reseteaba una retención ya bien matcheada a 'pendiente',
        # intentaba rematchear, pero su SENIAT correcto ya estaba excluido
        # del universo (congelado) -- no lo volvía a encontrar, quedaba
        # 'solo_odoo', y el SENIAT viejo se quedaba con un `wh_iva_id`
        # fantasma apuntando a un match que ya no existe del lado Odoo.
        # Ahora ambos lados se congelan con el mismo criterio: una vez
        # matcheado limpio, ninguno de los dos se vuelve a tocar hasta que
        # cambie algo real (ej. declarar, o un ajuste manual que mueva
        # estado_conciliacion fuera de estos dos estados).
        para_conciliar = WhIva.search([
            ('company_id', '=', self.company_id.id),
            ('state', 'in', ('confirmado', 'borrador', 'esperado', 'vencido')),
            ('estado_declaracion', '!=', 'declarado'),
            ('estado_conciliacion', 'not in', ('listo_declarar', 'conciliada_norec')),
        ])
        seniat_universo = SeniatRet.search([
            ('company_id', '=', self.company_id.id),
            ('estado', '!=', 'conciliado'),
        ])
        para_conciliar.write({
            'estado_conciliacion': 'pendiente',
            'monto_seniat': 0,
            'fecha_conciliacion': False,
            'nivel_match': False,
            'matched_por_normalizacion': False,
        })
        seniat_universo.write({'estado': 'cargado', 'wh_iva_id': False, 'nivel_match': False})

        norm_rif = self._norm_rif
        norm_ctrl = self._norm_ctrl
        norm_factura = self._norm_factura
        # ── Índices por diccionario en vez de .filtered() dentro del loop
        # (pedido explícito 2026-08-18, medido en Cementos: 29.901 wh_iva ×
        # 12.902 SENIAT = ~386M comparaciones por corrida, varios minutos).
        # Un solo recorrido O(m) de seniat_universo arma 2 diccionarios
        # (clave RIF+Control y RIF+Factura, ambos normalizados); el loop
        # principal de abajo queda en O(n) con lookups O(1) -- mismo
        # resultado, sin recorrer seniat_universo completo por cada wh.
        SeniatEmpty = self.env['ve.seniat.retencion']
        by_ctrl = {}
        by_factura = {}
        seniat_norm_factura = {}
        # Bug real confirmado 2026-08-22 (Vencement, RIF J-00000453-6/
        # FAPECA): by_ctrl/by_factura se arman UNA sola vez antes del loop
        # y nunca se actualizaban a medida que cada wh_iva reclamaba un
        # candidato -- si SENIAT reutiliza un N°Control entre 2
        # comprobantes reales distintos del mismo agente, 2 wh_iva
        # DISTINTOS podían encontrar el mismo ve.seniat.retencion como
        # candidato (uno por N1/Control, otro por N2/Factura) y el que se
        # procesaba último se lo robaba al otro en silencio -- sin pasar
        # por la salvaguarda de "más de 1 candidato" (esa solo mira
        # ambigüedad DENTRO de la resolución de un mismo wh, no conflictos
        # ENTRE wh distintos). seniat_ya_asignado trackea qué
        # ve.seniat.retencion ya se le asignó a un wh en ESTA corrida, para
        # excluirlo de los candidatos de cualquier wh que se procese
        # después.
        seniat_ya_asignado = set()
        for s in seniat_universo:
            s_rif = norm_rif(s.rif_agente)
            s_ctrl = norm_ctrl(s.nro_control)
            s_factura = norm_factura(s.nro_documento)
            seniat_norm_factura[s.id] = s_factura
            if s_ctrl:
                by_ctrl.setdefault((s_rif, s_ctrl), []).append(s)
            if s_factura:
                by_factura.setdefault((s_rif, s_factura), []).append(s)

        # ── Cascada de 2 niveles + normalización (pedido explícito
        # 2026-08-11, medido con datos reales de Cementos Enero 2026 antes
        # de implementar -- ver _norm_ctrl/_norm_factura arriba):
        #   N1 = RIF + N° Control (normalizados) -- match fuerte.
        #   N2 = RIF + N° Factura (normalizados) -- SOLO cuando la
        #     retención no tiene N° de Control, o N1 no encontró nada.
        # Si CUALQUIER nivel encuentra más de 1 candidato tras intentar
        # ambos, se deja SIN match para revisión manual en vez de tomar
        # el primero sin criterio -- cambio de comportamiento a propósito
        # respecto a la versión anterior (que sí tomaba seniat_match[0]
        # como último recurso): un match ambiguo mal resuelto es peor que
        # uno pendiente de revisar, porque terminaría declarado.
        for wh in para_conciliar:
            wh_rif = norm_rif(wh.rif)
            wh_ctrl_raw = (wh.nro_control or '').strip().upper()
            wh_ctrl_norm = norm_ctrl(wh.nro_control)
            wh_factura_raw = (wh.invoice_id.name if wh.invoice_id else (wh.nro_documento or '')).strip().upper()
            wh_factura_norm = norm_factura(wh_factura_raw)

            seniat_match = SeniatEmpty
            nivel = False
            normalizado = False

            if wh_ctrl_norm:
                candidatos_list = [
                    s for s in by_ctrl.get((wh_rif, wh_ctrl_norm), [])
                    if s.id not in seniat_ya_asignado
                ]
                # Desambiguar por N° Factura cuando RIF+Control normalizados
                # encuentran MÁS DE UN candidato -- bug real confirmado
                # 2026-08-05 (Cementos): SENIAT a veces usa un N°Control
                # genérico ("00") para varias retenciones reales distintas
                # del MISMO agente.
                if len(candidatos_list) > 1 and wh_factura_norm:
                    por_doc = [
                        s for s in candidatos_list
                        if seniat_norm_factura[s.id] == wh_factura_norm
                    ]
                    if por_doc:
                        candidatos_list = por_doc
                if candidatos_list:
                    seniat_match = SeniatEmpty.browse([s.id for s in candidatos_list])
                    nivel = 'n1'
                    normalizado = not any(
                        (s.nro_control or '').strip().upper() == wh_ctrl_raw for s in candidatos_list)

            if not seniat_match and wh_factura_norm:
                candidatos_list = [
                    s for s in by_factura.get((wh_rif, wh_factura_norm), [])
                    if s.id not in seniat_ya_asignado
                ]
                if candidatos_list:
                    seniat_match = SeniatEmpty.browse([s.id for s in candidatos_list])
                    nivel = 'n2'
                    normalizado = not any(
                        (s.nro_documento or '').strip().upper() == wh_factura_raw for s in candidatos_list)

            if len(seniat_match) > 1:
                # Sigue ambiguo tras N1/N2 -- no adivinar, dejar sin match.
                seniat_match = self.env['ve.seniat.retencion']
                nivel = False
                normalizado = False

            if not seniat_match:
                wh.estado_conciliacion = 'solo_odoo'
                wh.nivel_match = False
                wh.matched_por_normalizacion = False
                wh.seniat_rif = False
                wh.seniat_nro_control = False
                wh.seniat_nro_documento = False
                # NO se toca incluir_declaracion aquí. El hook que crea la
                # retención (account_move.py) ya la deja en C.66=False por
                # defecto — no hace falta que Conciliar SENIAT lo vuelva a
                # imponer en cada corrida, y hacerlo pisaba cualquier marca
                # manual de la usuaria (bug real reportado en vivo: marcó
                # C.66 en 2 retenciones No Recibidas, Conciliar SENIAT se lo
                # apagó a la que no tenía match SENIAT). Mismo espíritu que
                # el fix anterior para confirmado/borrador — "Conciliar
                # SENIAT" no debe decidir C.66 por nadie, ni a favor ni en
                # contra.
            else:
                seniat = seniat_match[0]
                seniat_ya_asignado.add(seniat.id)
                seniat.wh_iva_id = wh.id
                seniat.nivel_match = nivel
                wh.nivel_match = nivel
                wh.matched_por_normalizacion = normalizado
                wh.monto_seniat = seniat.monto_retenido
                wh.seniat_rif = seniat.rif_agente
                wh.seniat_nro_control = seniat.nro_control
                wh.seniat_nro_documento = seniat.nro_documento
                wh.fecha_conciliacion = fields.Datetime.now()
                if abs(wh.monto_retenido - seniat.monto_retenido) < 0.01:
                    if wh.state in ('confirmado', 'borrador'):
                        wh.estado_conciliacion = 'listo_declarar'
                    else:
                        # esperado/vencido: SENIAT OK pero sin comprobante físico
                        wh.estado_conciliacion = 'conciliada_norec'
                    seniat.estado = 'conciliado'
                else:
                    wh.estado_conciliacion = 'diferencia'
                    seniat.estado = 'diferencia'

        # Limpiar `name` ("N° Comprobante") en lo que queda Solo en SENIAT --
        # bug confirmado 2026-08-11 en el script del RPA (fuera de este
        # repo): manda el N° de Control como si fuera el N° de Comprobante
        # ("nro_comprobante": nroControl). Una retención "Solo en SENIAT"
        # (sin contraparte en SmartIVA) no tiene forma de conocer un
        # comprobante real -- se limpia acá en cada corrida para no
        # depender de una limpieza manual por RPC cada vez que aparecen
        # residuales nuevos (ver project_rpa_bug_nro_comprobante_seniat).
        seniat_universo.filtered(
            lambda s: s.estado == 'cargado').write({'estado': 'sin_match', 'name': False})
        self.estado = 'revision'
        # Ver comentario en action_extraer_seniat sobre por qué no se
        # devuelve un ir.actions.act_window aquí.
        return True

    def action_confirmar_todos(self):
        """Confirma todas las retenciones en estado borrador del período."""
        self.ensure_one()
        borradores = self.wh_iva_ids.filtered(lambda r: r.state == 'borrador')
        if not borradores:
            raise UserError('No hay retenciones en estado "Recibido" (sin confirmar) en este período.')
        confirmados = 0
        errores = []
        for wh in borradores:
            try:
                with self.env.cr.savepoint():
                    wh.action_confirmar()
                confirmados += 1
            except Exception as e:
                errores.append(
                    f'• {wh.partner_id.name or "—"} ({wh.nro_control or "—"}): {e}'
                )
        msg = (
            f'<b>Confirmar Todos</b> — {confirmados} retención(es) confirmadas '
            f'por <b>{self.env.user.name}</b>.'
        )
        if errores:
            msg += (
                f'<br/>⚠ {len(errores)} no se pudieron confirmar:<br/>'
                + '<br/>'.join(errores)
            )
        self.message_post(
            body=Markup(msg),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )
        if errores and not confirmados:
            raise UserError('Ninguna retención pudo confirmarse:\n' + '\n'.join(errores))
        # Ver comentario en action_extraer_seniat.
        return True

    def action_aprobar(self):
        self.ensure_one()
        if self.estado != 'revision':
            raise UserError('Solo se puede aprobar un período en estado "En Revisión".')
        n_borrador = len(self.wh_iva_ids.filtered(lambda r: r.state == 'borrador'))
        if n_borrador:
            raise UserError(
                f'No se puede aprobar: hay {n_borrador} retención(es) en estado '
                f'"Recibido" sin confirmar.\n'
                f'Confirme todos los comprobantes antes de aprobar.'
            )
        self.estado = 'aprobado'
        self.aprobado_por = self.env.user.id
        self.fecha_aprobacion = fields.Datetime.now()
        # Aprobar es 100% un checkpoint de período — desde la Etapa 3 del
        # rediseño de 3 ejes ya no mueve `state` de las retenciones.
        # "Conciliado" es una condición derivada (state='confirmado' +
        # estado_conciliacion con match SENIAT), no un `state` propio que
        # haya que asignar aquí.
        _SENIAT_OK = frozenset({
            'listo_declarar', 'conciliada_norec', 'declarado',
            'conciliada', 'aprobado_declarar',
        })
        confirmadas = self.wh_iva_ids.filtered(lambda r: r.state == 'confirmado')
        con_match = confirmadas.filtered(lambda r: r.estado_conciliacion in _SENIAT_OK)
        sin_match = confirmadas - con_match
        msg = (
            '<b>Conciliación aprobada</b> por {usuario}.<br/>'
            '{n} retención(es) quedaron <b>Conciliadas</b> (match con SENIAT).'
        ).format(usuario=self.env.user.name, n=len(con_match))
        if sin_match:
            msg += (
                '<br/>⚠ {n} retención(es) permanecen <b>sin coincidencia con SENIAT</b>.'
            ).format(n=len(sin_match))
        self.message_post(
            body=Markup(msg),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )
        # Ver comentario en action_extraer_seniat.
        return True

    def action_reabrir(self):
        self.ensure_one()
        if self.estado != 'aprobado':
            raise UserError('Solo se puede reabrir un período en estado "Conciliación Aprobada".')
        if self.declaracion_iva_id and self.declaracion_iva_id.estado == 'presentada':
            raise UserError(
                'No se puede reabrir: este período ya fue declarado al SENIAT en Declaración IVA.'
            )
        # Desde la Etapa 3 del rediseño de 3 ejes, Aprobar ya no mueve
        # `state` de las retenciones a 'conciliado' — no hay nada que
        # revertir en `state` aquí (antes esto causaba un bug real:
        # las retenciones quedaban huérfanas porque `state='conciliado'`
        # nunca volvía a entrar en el domain de _do_conciliar).
        self.wh_iva_ids.write({
            'estado_conciliacion': 'pendiente',
            'monto_seniat': 0,
            'fecha_conciliacion': False,
        })
        self.seniat_ids.write({'estado': 'cargado', 'wh_iva_id': False})
        self.estado = 'borrador'
        self.message_post(
            body=Markup('<b>Período reabierto</b> por {u}.').format(u=self.env.user.name),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )

    def action_abrir_declaracion_iva(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 've.declaracion.iva',
            'view_mode': 'form',
            'res_id': self.declaracion_iva_id.id,
            'target': 'current',
        }

    def action_ver_ret_no_spe(self):
        """Abre los clientes de partners_seniat_no_spe_ids en una lista real
        -- pedido explícito 2026-08-18, el chatter (donde vivía este mismo
        detalle en texto) no era práctico para revisar caso por caso."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Ret. no SPE — {self.periodo_retencion}',
            'res_model': 'res.partner',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.partners_seniat_no_spe_ids.ids)],
        }

    def action_ver_normalizadas(self):
        """Abre las retenciones de este período cuyo cruce con el SENIAT
        solo funcionó después de normalizar N° de Control/Factura (quitar
        ceros a la izquierda, guiones, prefijos) -- para que el contador
        audite que el cruce es correcto antes de declarar. Pedido explícito
        2026-08-11, ver count_normalizados/_do_conciliar."""
        self.ensure_one()
        list_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_list_normalizadas')
        form_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_form')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Retenciones Conciliadas por Normalización — {self.periodo_retencion or self.periodo}',
            'res_model': 've.wh.iva',
            'views': [(list_view.id, 'list'), (form_view.id, 'form')],
            'domain': [('conciliacion_id', '=', self.id), ('matched_por_normalizacion', '=', True)],
        }

    def action_volver_revision(self):
        """Aprobado → En Revisión (sin deshacer la conciliación SENIAT)."""
        self.ensure_one()
        if self.estado != 'aprobado':
            raise UserError('Solo se puede volver a revisión desde estado "Aprobado".')
        # Desde la Etapa 3 del rediseño de 3 ejes, Aprobar ya no mueve
        # `state` — nada que revertir aquí.
        self.aprobado_por = False
        self.fecha_aprobacion = False
        self.estado = 'revision'
        self.message_post(
            body=Markup('<b>Período devuelto a En Revisión</b> por {u}.').format(
                u=self.env.user.name),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )

    def action_reporte_conciliacion(self):
        self.ensure_one()
        return self.env.ref(
            've_retencion_iva.action_report_conciliacion'
        ).report_action(self)

    def action_reporte_declaracion(self):
        self.ensure_one()
        return self.env.ref(
            've_retencion_iva.action_report_declaracion'
        ).report_action(self)

    def action_exportar_visual_excel(self):
        """Exporta la pestaña Conciliación Visual a XLSX."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise UserError('La librería openpyxl no está instalada.')

        import io, base64

        activos = self.wh_iva_ids.filtered(
            lambda r: r.state != 'anulado'
        ).sorted(key=lambda r: r.monto_retenido, reverse=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Conciliacion {self.periodo_retencion or self.periodo or ""}'[:31]

        navy  = PatternFill('solid', fgColor='1F4E79')
        light = PatternFill('solid', fgColor='D6E4F0')
        green = PatternFill('solid', fgColor='C6EFCE')
        yell  = PatternFill('solid', fgColor='FFEB9C')
        red   = PatternFill('solid', fgColor='FFC7CE')
        thin  = Side(style='thin')
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 13 columnas: añade Monto Odoo + Monto Recibidos, cambia Diferencia → Recibidos vs SENIAT
        headers = [
            'Factura', 'Fecha Rec.', 'Cliente', 'RIF', 'N° Control',
            'N° Comprobante', 'Estado Comp.',
            'Monto Esperado', 'Monto Recibido', 'Monto SENIAT', 'Recibidos vs SENIAT',
            'Conc. SENIAT', 'C.66',
        ]
        col_w = [18, 12, 35, 16, 12, 20, 14, 14, 14, 14, 16, 20, 6]
        NCOLS = len(headers)
        last_col_letter = openpyxl.utils.get_column_letter(NCOLS)

        # ── Título ──────────────────────────────────────────────────────────────
        ws.merge_cells(f'A1:{last_col_letter}1')
        t = ws['A1']
        t.value = (f'CONCILIACION VISUAL — {self.env.company.name} — '
                   f'{self.periodo_retencion or self.periodo or ""}')
        t.font = Font(bold=True, size=11, color='FFFFFF')
        t.fill = PatternFill('solid', fgColor='1F4E79')
        t.alignment = Alignment(horizontal='center')

        # ── Encabezados ──────────────────────────────────────────────────────────
        estado_labels = dict(self.env['ve.wh.iva']._fields['state'].selection)
        conc_labels   = dict(self.env['ve.wh.iva']._fields['estado_conciliacion'].selection)

        for ci, (h, w) in enumerate(zip(headers, col_w), 1):
            c = ws.cell(2, ci, h)
            c.fill = navy
            c.font = Font(color='FFFFFF', bold=True, size=9)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = bdr
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # ── Filas de datos ───────────────────────────────────────────────────────
        num_fmt  = '#,##0.00'
        date_fmt = 'DD/MM/YYYY'
        conc_fill = {
            'conciliada':    green, 'listo_declarar': green,
            'diferencia':    yell,  'conciliada_norec': yell,
            'solo_odoo':     red,
        }

        # Índice de columnas (1-based)
        COL_MONTO_ODOO     = 8
        COL_MONTO_RECIBIDO = 9
        COL_MONTO_SENIAT   = 10
        COL_DIFERENCIA     = 11
        COL_C66            = 13

        for ri, r in enumerate(activos, 3):
            monto_recibido = r.monto_recibido
            diferencia_rec_seniat = monto_recibido - r.monto_seniat
            row = [
                r.invoice_id.name if r.invoice_id else '',    # 1 Factura
                r.fecha,                                       # 2 Fecha Rec.
                r.partner_id.name if r.partner_id else '',     # 3 Cliente
                r.rif or '',                                   # 4 RIF
                r.nro_control or '',                           # 5 N° Control
                r.name or '',                                  # 6 N° Comprobante
                estado_labels.get(r.state, r.state),           # 7 Estado Comp.
                r.monto_retenido,                              # 8 Monto Odoo
                monto_recibido,                                # 9 Monto Recibidos
                r.monto_seniat,                                # 10 Monto SENIAT
                diferencia_rec_seniat,                         # 11 Recibidos vs SENIAT
                conc_labels.get(r.estado_conciliacion, r.estado_conciliacion),  # 12 Conc. SENIAT
                True if r.incluir_declaracion else False,      # 13 C.66 (bool para SUMIF)
            ]
            fill = conc_fill.get(r.estado_conciliacion)
            if not r.incluir_declaracion:
                fill = PatternFill('solid', fgColor='E2E3E5')
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = bdr
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill
                if ci == 2 and val:
                    cell.number_format = date_fmt
                elif ci in (COL_MONTO_ODOO, COL_MONTO_RECIBIDO, COL_MONTO_SENIAT, COL_DIFERENCIA):
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='right')
                elif ci == COL_C66:
                    # Mostrar ✓ / — pero dejar valor bool para fórmulas
                    cell.value = '✓' if val else '—'

        # ── Fila de Totales ──────────────────────────────────────────────────────
        last_data_row = len(activos) + 2   # última fila de datos
        tr = last_data_row + 1

        lbl = ws.cell(tr, 1, 'TOTALES')
        lbl.font = Font(bold=True, size=9)
        ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=7)

        # Columna C.66 usa la letra correspondiente para la fórmula SUMIF
        c66_col_letter  = openpyxl.utils.get_column_letter(COL_C66)
        rec_col_letter  = openpyxl.utils.get_column_letter(COL_MONTO_RECIBIDO)
        data_start = 3
        data_end   = last_data_row

        totals = {
            COL_MONTO_ODOO:     self.total_odoo,
            COL_MONTO_RECIBIDO: self.total_recibidos,
            COL_MONTO_SENIAT:   self.total_seniat,
            COL_DIFERENCIA:     self.diferencia,
            # C.66: suma de Monto Recibidos donde C.66 = "✓"
            COL_C66: f'=SUMPRODUCT(({c66_col_letter}{data_start}:{c66_col_letter}{data_end}="✓")*{rec_col_letter}{data_start}:{rec_col_letter}{data_end})',
        }
        for ci, val in totals.items():
            c = ws.cell(tr, ci, val)
            c.fill = light
            c.font = Font(bold=True, size=9)
            c.border = bdr
            if isinstance(val, (int, float)):
                c.number_format = num_fmt
                c.alignment = Alignment(horizontal='right')
            elif isinstance(val, str) and val.startswith('='):
                c.number_format = num_fmt
                c.alignment = Alignment(horizontal='right')

        # ── AutoFiltro por columnas ──────────────────────────────────────────────
        ws.auto_filter.ref = f'A2:{last_col_letter}{last_data_row}'

        # ── Fijar encabezado ─────────────────────────────────────────────────────
        ws.freeze_panes = 'A3'

        output = io.BytesIO()
        wb.save(output)
        fname = f'conciliacion_visual_{self.periodo_retencion or self.periodo or "periodo"}.xlsx'
        att = self.env['ir.attachment'].create({
            'name': fname,
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': ('application/vnd.openxmlformats-officedocument'
                         '.spreadsheetml.sheet'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{att.id}?download=true',
            'target': 'self',
        }

    def action_enviar_recordatorios_pendientes(self):
        """Crea actividades diferenciadas según las banderas necesita_* de cada
        retención (mismo criterio que los métodos de selección masiva en
        ve_wh_iva.py — antes este botón de período tenía su propia
        clasificación local ligeramente distinta, ej. "vencido sin SENIAT"
        caía en Reportar SENIAT aquí pero en Pedir Comprobante en el otro
        dispatcher; unificado para que los puntos de entrada de
        recordatorios se comporten igual). A diferencia de la selección
        manual (que desde 2026-07-30 separa 'normal' de 'SENIAT' a
        propósito), este botón de período sigue mandando los 3 tipos
        mezclados — es "atender todo lo pendiente del período", no una
        selección deliberada de un tipo."""
        self.ensure_one()
        pendientes = self.wh_iva_ids.filtered(
            lambda r: r.necesita_envio_comp or r.necesita_aclarar_dif_seniat
            or r.necesita_reportar_seniat
        )
        if not pendientes:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Sin pendientes',
                    'message': 'No hay comprobantes que requieran recordatorio en este período.',
                    'type': 'info',
                    'sticky': False,
                },
            }
        # Bug real 2026-07-28: llamaba a action_recordatorio_dif_seniat()/
        # rep_seniat()/envio_comp() — esos métodos ABREN un wizard de
        # confirmación (devuelven un ir.actions.act_window), que en un
        # bucle sobre varias retenciones se descarta en cada vuelta sin
        # enviar nada. El método correcto para envío directo (sin wizard)
        # es _enviar_recordatorio_tipo(), el mismo que ya usan los
        # métodos de selección masiva en ve_wh_iva.py.
        for wh in pendientes:
            if wh.necesita_aclarar_dif_seniat:
                wh._enviar_recordatorio_tipo('dif_seniat')
            elif wh.necesita_reportar_seniat:
                wh._enviar_recordatorio_tipo('rep_seniat')
            elif wh.necesita_envio_comp:
                wh._enviar_recordatorio_tipo('envio_comp')

        # Registrar en chatter del período
        clientes = ', '.join(
            f'{w.partner_id.name} ({w.invoice_id.name if w.invoice_id else "—"})'
            for w in pendientes[:10]
        )
        if len(pendientes) > 10:
            clientes += f' … y {len(pendientes) - 10} más'
        self.message_post(
            body=Markup(
                f'<b>Recordatorios enviados ({len(pendientes)} comprobante(s))</b><br/>'
                f'{clientes}'
            ),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )
        # Ver comentario en action_extraer_seniat.
        return True

    def action_registrar_declaracion(self, nro_declaracion, fecha_declaracion):
        """Callback RPA — delega a la Declaración IVA companion."""
        self.ensure_one()
        decl = self.declaracion_iva_id
        if not decl:
            raise UserError(
                f'El período {self.periodo_retencion} no tiene Declaración IVA asociada.'
            )
        decl.action_registrar_declaracion_rpa(nro_declaracion, fecha_declaracion)
