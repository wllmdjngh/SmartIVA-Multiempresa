import base64
import csv
import io
import logging
import re
import unicodedata

from markupsafe import Markup
from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# ── Helpers de parseo (mismo criterio que wizard_carga_seniat.py) ─────────────

def _norm_rif(rif):
    """Mismo criterio que ve_conciliacion.py::_norm_rif — sin esto, un RIF
    con guión en el archivo cargado ('J-12345678-9') nunca matchea contra
    un contacto ya guardado sin guión ('J123456789', o viceversa), y el
    cliente existente se trata como nuevo (duplicado). Bug real 2026-07-28."""
    return (rif or '').upper().replace('-', '').replace(' ', '').strip()


def _codigo_diario_zona(zona):
    """Código corto y estable para el diario de venta dedicado a una Zona
    (2026-08-14, fix Causa C definitivo: Odoo exige N° Factura único por
    diario ENTRE FACTURAS POSTEADAS -- restricción account_move_unique_name
    -- sin importar Zona; el criterio de Zona en _compute_partner_id solo
    evitaba el falso positivo en la VISTA PREVIA, pero Move.create() +
    action_post() seguían chocando de verdad cuando dos zonas reusaban el
    mismo N° de Factura. Un diario por Zona resuelve el choque real, mismo
    patrón que journal_nc/NCVTA para Registro+Anulación). Código de diario
    limitado a pocos caracteres en Odoo -- se recorta, no se trunca en
    medio de una palabra si se puede evitar, y dos Zonas cuyo código
    recortado coincida (raro, pero posible con nombres largos parecidos)
    terminan compartiendo diario -- aceptable, es el mismo comportamiento
    de fallback que ya existía (compartir diario) para ese caso límite."""
    letras = re.sub(r'[^A-Za-z0-9]', '', zona or '').upper()
    return ('Z' + letras)[:10]


def _formatear_rif(rif):
    """Acepta un RIF sin guión (ej. 'J411947830', formato común en
    exportes de terceros) y lo devuelve con el formato estándar
    LETRA-12345678-9. Si ya trae guión, o si su forma no coincide con el
    patrón esperado (letra + 9 dígitos), se devuelve tal cual — no se
    inventa formato sobre algo que no se reconoce con certeza. Pedido
    explícito 2026-08-05."""
    limpio = (rif or '').upper().strip()
    if not limpio or '-' in limpio:
        return limpio
    m = re.match(r'^([VEJPG])(\d{9})$', limpio)
    if not m:
        return limpio
    letra, digitos = m.groups()
    return f'{letra}-{digitos[:8]}-{digitos[8]}'


def _norm_header(text):
    text = text.lower().strip()
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    # Encabezados reales tipo "Base 16,00 %" (decimal venezolano) — mismo
    # bug real 2026-08-02 encontrado en ve_conecta_carga_compras.py (ver
    # ese archivo para el detalle): sin este paso, la coma se convertía en
    # espacio más abajo y dejaba "base 16 00" en vez de "base 16".
    text = re.sub(r'(\d+),(\d{2})\b', r'\1', text)
    # Colapsar acrónimos con puntos ("i.v.a." -> "iva") ANTES de convertir
    # el resto de la puntuación en espacio — mismo fix que
    # ve_conecta_carga_compras.py (ver ese archivo para el detalle del bug
    # real encontrado 2026-07-29).
    text = re.sub(r'\b(?:[a-z]\.){2,}', lambda m: m.group(0).replace('.', ''), text)
    # 'º' (ordinal masculino, U+00BA) es visualmente parecido a '°' (grado,
    # U+00B0) pero es OTRO carácter — "Nº Factura" no matcheaba igual que
    # "N° Factura" antes de este fix (2026-07-31). La coma se incluye acá
    # también, como red de seguridad, por si queda alguna suelta que el
    # regex decimal de arriba no cubrió.
    text = re.sub(r'[°º#.,%\-_/\\]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _parse_amount(val):
    """Bug real encontrado 2026-08-12 (Cementos, pares Registro+Anulación
    de zona Invecem): la regex vieja `[^\\d.,]` borraba CUALQUIER carácter
    que no fuera dígito/punto/coma -- incluido el signo "-". Un monto
    negativo del archivo ("-1.074.060,00") perdía el signo y quedaba
    positivo, indistinguible de su Registro original. Se detecta el signo
    ANTES de limpiar el resto del string."""
    if val is None:
        return 0.0
    s = str(val).strip()
    negativo = s.startswith('-') or (s.startswith('(') and s.endswith(')'))
    s = re.sub(r'[^\d.,]', '', s)
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        v = float(s)
        return -v if negativo else v
    except ValueError:
        return 0.0


def _parse_date(val, formato='dmy'):
    """formato: 'dmy' (usar tal cual, default) o 'mdy' (invertir día/mes).

    Ampliado 2026-07-31 (segunda vuelta con Cementos): el primer intento de
    este fix solo cubría fechas-TEXTO "N/N/AAAA" ambiguas, asumiendo que una
    fecha real de Excel (rama hasattr(strftime)) nunca podía estar mal. Caso
    real encontrado: `Libro De Ventas Nov 1Q 2025 (Formato1).xlsx` trae
    fechas REALES de Excel (datetime ya resuelto por openpyxl) pero con
    día/mes invertidos en el propio dato — ej. datetime(2025, 7, 11) para
    una factura que en realidad es del 7 de noviembre (día=7, mes=11), no
    el 11 de julio. Confirmado en vivo: todas las fechas de muestra tenían
    día=11 fijo con el mes variando (1,7,8,12) — exactamente al revés de lo
    esperable en un archivo "Nov 1Q" (mes=11 fijo, día 1-15 variando). Es un
    error de quien generó el archivo de origen, no un problema de texto
    ambiguo — por eso 'mdy' ahora también corrige fechas reales/seriales,
    no solo texto."""
    if val is None:
        return False
    if hasattr(val, 'strftime'):
        if formato == 'mdy':
            try:
                from datetime import date as _date
                return _date(val.year, val.day, val.month).strftime('%Y-%m-%d')
            except ValueError:
                return False
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)) and 1000 < val < 100000:
        try:
            from datetime import date, timedelta
            d = date(1899, 12, 30) + timedelta(days=int(val))
            if formato == 'mdy':
                d = date(d.year, d.day, d.month)
            return d.strftime('%Y-%m-%d')
        except Exception:
            pass
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        a, b, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
        dia, mes = (b, a) if formato == 'mdy' else (a, b)
        if not (1 <= mes <= 12 and 1 <= dia <= 31):
            # Bloqueante existente "Fecha no reconocida" (ver
            # _compute_partner_id) en vez de escribir una fecha inválida
            # en silencio — mismo principio de auditoría de inputs.
            return False
        return f"{anio}-{mes:02d}-{dia:02d}"
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return False


def _fecha_candidatos(val):
    """Devuelve lista de (año, mes, día) — normal e invertido — para el
    valor crudo de la celda, filtrando los que no son fecha calendario
    válida. Usado por el modo 'auto' (ver _resolver_fechas_auto): caso
    real 2026-07-31 (Cementos) donde el MISMO archivo mezclaba filas ya
    correctas con filas invertidas (no un solo formato consistente para
    todo el archivo, por eso un simple interruptor dmy/mdy no alcanzaba)."""
    if val is None:
        return []
    if hasattr(val, 'strftime'):
        anio, mes, dia = val.year, val.month, val.day
    elif isinstance(val, (int, float)) and 1000 < val < 100000:
        from datetime import date, timedelta
        d = date(1899, 12, 30) + timedelta(days=int(val))
        anio, mes, dia = d.year, d.month, d.day
    else:
        s = str(val).strip()
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
        if not m:
            return []
        dia, mes, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
    from datetime import date as _date
    candidatos = []
    try:
        _date(anio, mes, dia)
        candidatos.append((anio, mes, dia))
    except ValueError:
        pass
    if mes != dia:
        try:
            _date(anio, dia, mes)
            candidatos.append((anio, dia, mes))
        except ValueError:
            pass
    return candidatos


def _normalizar_tipo_transaccion(val):
    """Normaliza el valor crudo de la columna Tipo de Transacción a los 3
    códigos SENIAT (01 Factura Regular, 02 Nota de Débito, 03 Nota de
    Crédito) — acepta el código numérico directo o el nombre en texto
    (con o sin acentos/mayúsculas). Devuelve False si no reconoce nada,
    en vez de adivinar. Insumo para AJUSTE-FISCAL-01/02 (ver REQUISITOS.md),
    no bloquea ni reemplaza todavía la detección Registro+Anulación."""
    s = (val or '').strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    if s in ('01', '1'):
        return '01'
    if s in ('02', '2'):
        return '02'
    if s in ('03', '3'):
        return '03'
    if 'CREDITO' in s or s in ('NC',):
        return '03'
    if 'DEBITO' in s or s in ('ND',):
        return '02'
    if 'FACTURA' in s or 'REGULAR' in s:
        return '01'
    return False


# Mapeo de encabezados reconocidos — CONECTA-13: columnas del Libro de Ventas
# que el motor necesita (ver REQUISITOS.md sección 11, Bloque 1). Extender
# esta lista con los sinónimos que use cada cliente real, en vez de construir
# un modelo de mapping configurable aparte (mismo patrón ya probado en
# wizard_carga_seniat.py).
_HEADER_MAP = {
    'rif': 'rif', 'rif cliente': 'rif', 'nro rif': 'rif', 'n rif': 'rif',
    'nombre': 'nombre_cliente', 'cliente': 'nombre_cliente',
    'razon social': 'nombre_cliente', 'nombre cliente': 'nombre_cliente',
    'nombre o razon social': 'nombre_cliente',
    'n control': 'nro_control', 'nro control': 'nro_control',
    'control': 'nro_control', 'numero control': 'nro_control',
    'n de control': 'nro_control', 'numero de control': 'nro_control',
    'nro de control': 'nro_control',
    'n factura': 'nro_documento', 'nro factura': 'nro_documento',
    'factura': 'nro_documento', 'numero factura': 'nro_documento',
    'n documento': 'nro_documento', 'nro documento': 'nro_documento',
    'n de factura': 'nro_documento', 'numero de factura': 'nro_documento',
    'nro de factura': 'nro_documento',
    'fecha': 'fecha', 'fecha factura': 'fecha', 'fecha emision': 'fecha',
    'fecha documento': 'fecha', 'fecha de la factura': 'fecha',
    'base 16': 'base_16', 'base imponible 16': 'base_16',
    'base general': 'base_16', 'base gravada 16': 'base_16',
    'base 8': 'base_8', 'base imponible 8': 'base_8',
    'base reducida': 'base_8', 'base gravada 8': 'base_8',
    'monto retenido': 'monto_retenido', 'iva retenido': 'monto_retenido',
    'retencion': 'monto_retenido', 'retenido': 'monto_retenido',
    'iva ret x comp': 'monto_retenido', 'iva ret x comprobante': 'monto_retenido',
    # Venta exenta/no gravada (formato oficial del Libro de Ventas del
    # propio módulo, ver wizard_libro_ventas.py/report ve_libro_ventas.xml)
    'v int no grav': 'base_exento', 'v internas no gravadas': 'base_exento',
    'ventas internas no gravadas': 'base_exento', 'exento': 'base_exento',
    'no gravado': 'base_exento',
    # Estado de pago/cobranza — opcional, solo si el cliente activa
    # seguimiento proactivo (ver MEJORA-CONTACTO-01, CHECKLIST_ARRANQUE_
    # PILOTO.md sección 2). Informativo por ahora, no afecta la confirmación.
    'estadopago': 'estado_pago', 'estado pago': 'estado_pago',
    'estado de pago': 'estado_pago', 'cobranza': 'estado_pago',
    'estado cobranza': 'estado_pago', 'estado de cobranza': 'estado_pago',
    # Zona/Planta/Sucursal — opcional, solo para clientes que facturan desde
    # varios puntos bajo un único RIF (un solo Odoo, una sola Declaración
    # IVA) y quieren seguimiento de comprobantes separado por punto. Ver
    # account.move.zona / ve.wh.iva.zona.
    'zona': 'zona', 'planta': 'zona', 'sucursal': 'zona',
    'centro': 'zona', 'ubicacion': 'zona', 'punto de venta': 'zona',
    # Contribuyente Especial/SPE — opcional, columna explícita Sí/No.
    # Pedido explícito 2026-08-01: sin esta columna, "Agente de
    # Retención" se infiere solo de si ALGUNA fila del cliente trae
    # monto_retenido>0 — si la primera factura de ese cliente en el
    # archivo no trae retención pero una posterior sí, la primera queda
    # sin retención esperada (orden del archivo importa). Esta columna
    # es una señal ADICIONAL, no reemplaza la existente — ver
    # action_confirmar (primera pasada, antes de crear ninguna factura).
    'contribuyente especial': 'es_spe', 'es contribuyente especial': 'es_spe',
    'agente de retencion': 'es_spe', 'es agente de retencion': 'es_spe',
    'spe': 'es_spe', 'contribuyente': 'es_spe',
    # Validado SENIAT — opcional, si el cliente ya cruzó manualmente ese RIF
    # contra el portal SENIAT (independiente de la conciliación automática
    # de retenciones). Pedido explícito 2026-08-05: campo persistente en el
    # Cliente (ver res_partner.py), blanco/Sí/No, se sincroniza con lo que
    # traiga cada carga del Libro de Ventas.
    'validado seniat': 'validado_seniat',
    # Formato "largo" (una sola columna Base Imponible + % Alíc./IVA por
    # fila, en vez de Base 16%/Base 8% separadas) — mismo patrón que ya
    # usa ve_conecta_carga_compras.py. Se reparte a base_16/base_8 en
    # _leer_filas según la alícuota de esa misma fila. Encontrado real
    # 2026-07-31 (Libro Ventas Cemento.xlsx, prospecto nuevo).
    'base imponible': 'base_generica',
    'alic': 'alicuota_pct', 'alicuota': 'alicuota_pct', 'alic iva': 'alicuota_pct',
    # '% IVA' normaliza a 'iva' a secas (el % se convierte en espacio) —
    # solo mapea si el encabezado es EXACTAMENTE eso, no colisiona con
    # 'iva retenido'/'impuesto iva' (multi-palabra, no matchean acá).
    'iva': 'alicuota_pct',
    # N° de Comprobante de Retención YA emitido por el cliente del
    # cliente (el agente de retención que compra) — NO lo genera
    # SmartIVA. Si el feed lo trae, la retención no está "esperando"
    # nada: ya fue recibida y hay que marcarla recibida+confirmada al
    # confirmar la carga (ver action_confirmar), no dejarla en blanco
    # para que el sistema invente un número nuevo.
    'comprobante de retencion': 'nro_comp_retencion',
    'n comprobante retencion': 'nro_comp_retencion',
    'nro comprobante retencion': 'nro_comp_retencion',
    'numero comprobante retencion': 'nro_comp_retencion',
    'n comp retencion': 'nro_comp_retencion', 'nro comp retencion': 'nro_comp_retencion',
    # Total e IVA tal cual vienen en el archivo — informativos, NO
    # participan en la creación de la factura (esa se sigue construyendo
    # desde base_16/base_8/base_exento con los impuestos de Odoo). Se
    # capturan solo para la tabla "Consistencia por Zona" (pedido
    # explícito 2026-08-02): comparar el dato crudo del archivo contra lo
    # que Odoo terminó calculando, sin depender de reconstruir el total
    # desde los mismos campos que ya se usaron para crear la factura
    # (eso sería una comprobación circular).
    'total de ventas incluyendo iva': 'total_documento',
    'total documento': 'total_documento', 'total factura': 'total_documento',
    'monto documento': 'total_documento',
    'impuesto iva': 'monto_iva', 'iva debito fiscal': 'monto_iva',
    'monto iva': 'monto_iva', 'iva debito': 'monto_iva',
    # Tipo de Transacción SENIAT (01 Factura Regular, 02 Nota de Débito,
    # 03 Nota de Crédito) — insumo para AJUSTE-FISCAL-01/02. Ver
    # _normalizar_tipo_transaccion arriba para los valores aceptados.
    'tipo de transaccion': 'tipo_transaccion', 'tipo transaccion': 'tipo_transaccion',
    'tipo tr': 'tipo_transaccion', 'tipo trans': 'tipo_transaccion',
    'cod transaccion': 'tipo_transaccion', 'codigo transaccion': 'tipo_transaccion',
    'tipo doc seniat': 'tipo_transaccion', 'tipo documento seniat': 'tipo_transaccion',
}

# Segunda pasada de reconocimiento por frase ancla — mismo criterio que
# ve_conecta_carga_compras.py (ver ese archivo para el detalle completo):
# cada cliente redacta sus encabezados distinto, "razon social" es una
# señal confiable para el nombre del cliente sin depender de la frase
# exacta ("Nombre o Razón Social" vs "Cliente Razón Social", etc).
_HEADER_FALLBACK = [
    ('razon social', 'nombre_cliente'),
]

_AMOUNT_FIELDS = {'base_16', 'base_8', 'base_exento', 'monto_retenido',
                  'base_generica', 'alicuota_pct', 'total_documento', 'monto_iva'}

_ESTADOS_PAGADO = {'pagada', 'pagado', 'paid', 'cancelada', 'cancelado'}


def _es_pagado(estado_pago):
    """True si el texto de EstadoPago del feed indica que la factura ya
    se cobró — se usa para registrar un pago real en Odoo (ver
    action_confirmar), único punto de integración: payment_state nativo
    es lo que ya leen Cobranza vs. Comprobante, Crédito Fiscal en
    Tránsito y la Lista de Trabajo (ve_wh_iva.py/ve_dashboard_iva.py),
    así que no hace falta tocar esas 3 vistas por separado."""
    return bool(estado_pago) and _norm_header(estado_pago) in _ESTADOS_PAGADO


_SPE_VERDADERO = {'si', 's', 'spe', 'especial', 'contribuyente especial',
                  'true', '1', 'yes'}
_SPE_FALSO = {'no', 'n', 'e', 'exento', 'x', 'false', '0'}
# Regla confirmada por la usuaria/contador 2026-08-14: la retención IVA
# SOLO corresponde si Contribuyente=S (Sujeto Pasivo Especial confirmado
# por SENIAT) -- no importa qué diga cualquier otra columna del archivo,
# incluido monto_retenido. 'n' (No), 'e' (Exento, RIF que empieza por
# V/C/E/A) y 'x' (RIF no encontrado en SENIAT) son todos NO-agente de
# forma EXPLÍCITA, no "sin señal" -- ver el uso de _es_spe_verdadero() más
# abajo, donde un False explícito ahora bloquea el heurístico viejo de
# monto_retenido>0 en vez de dejarlo pasar sin más. Hasta 2026-08-13 'x'
# no estaba en ningún set a propósito (ver historial) para que devolviera
# None; ese hueco es justo lo que dejaba que filas Exento/No-encontrado
# con monto retenido en el Excel igual marcaran al cliente como agente de
# retención -- el bug real que esta regla corrige.


def _es_spe_verdadero(valor):
    """Interpreta la columna opcional Contribuyente Especial/SPE — None si
    la celda viene vacía o con un texto no reconocido (no es una señal,
    se aplica la lógica existente sin esta columna); True/False si el
    texto es claramente reconocible como Sí/No."""
    if not valor:
        return None
    norm = _norm_header(str(valor))
    if norm in _SPE_VERDADERO:
        return True
    if norm in _SPE_FALSO:
        return False
    return None


_VALIDADO_SI = {'si', 's', 'true', '1', 'yes'}
_VALIDADO_NO = {'no', 'n', 'false', '0'}
# S vale más que N, que vale más que en blanco — ni una fila en blanco
# borra un S/N ya visto (en el archivo o ya guardado en el Cliente), ni un
# N revierte un S ya confirmado. Mismo criterio "ratchet" que ya usa
# es_agente_retencion (ver acción action_confirmar): una señal positiva no
# se pierde por una fila posterior menos concluyente.
_VALIDADO_PRIORIDAD = {'si': 2, 'no': 1}


def _validado_seniat_norm(valor):
    """Interpreta la columna opcional Validado_SENIAT — None si la celda
    viene vacía o con texto no reconocido (sin señal); 'si'/'no' si el
    texto es claramente reconocible."""
    if not valor:
        return None
    norm = _norm_header(str(valor))
    if norm in _VALIDADO_SI:
        return 'si'
    if norm in _VALIDADO_NO:
        return 'no'
    return None


class VeConectaCargaVentas(models.Model):
    _name = 've.conecta.carga.ventas'
    _description = 'Carga de Libro de Ventas — SmartIVA Conecta (CONECTA-14)'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    name = fields.Char(string='Referencia', copy=False, readonly=True, default='Nueva Carga')
    company_id = fields.Many2one(
        'res.company', string='Empresa', required=True,
        default=lambda self: self.env.company, index=True,
    )
    archivo = fields.Binary(string='Archivo (CSV/XLSX)', attachment=True)
    archivo_nombre = fields.Char(string='Nombre del Archivo')
    mes_archivo = fields.Char(
        string='Mes', compute='_compute_mes_archivo', store=True,
        help='Extraído del nombre del archivo (ej. "DEMO MAR 1Q part2.xlsx" '
             '→ "Mar") -- pedido explícito 2026-08-12 para filtrar/agrupar '
             'la lista de cargas por mes, ya que cada quincena llega '
             'partida en varios archivos "partN" (ver split_carga_lotes.py) '
             'y la lista crece rápido.')
    formato_fecha = fields.Selection([
        ('dmy', 'Normal (usar la fecha tal cual viene)'),
        ('mdy', 'Día y Mes invertidos en TODO el archivo'),
        ('auto', 'Detectar automáticamente (archivo con mezcla de ambos)'),
    ], string='Formato de Fecha', default='dmy', required=True,
        help='"Día y Mes invertidos": todo el archivo tiene el mismo '
             'problema (ej. exportado en MM/DD/AAAA). '
             '"Detectar automáticamente": el archivo mezcla filas ya '
             'correctas con filas invertidas (caso real 2026-07-31, '
             'Cementos — un mismo Libro de Ventas traía ambas) — cada '
             'fila se corrige comparando contra el mes que domina el '
             'resto del archivo, no un solo criterio para todas. Aplica '
             'tanto a fechas de TEXTO como a fechas REALES de Excel ya '
             'resueltas por openpyxl.')

    estado = fields.Selection([
        ('borrador',    'Borrador — Vista Previa'),
        ('confirmado_discrepancias', 'Confirmado c/Discrepancias'),
        ('confirmado',  'Confirmado'),
    ], string='Estado', default='borrador', required=True, tracking=True,
        help='"Confirmado c/Discrepancias" (pedido explícito 2026-08-12): '
             'se usó "Confirmar y Omitir Bloqueadas" -- las filas limpias ya '
             'son facturas reales, las bloqueadas (ej. falso positivo de '
             'zona) siguen pendientes en la carga para resolver después, '
             'sin tener retenida toda la carga por unas pocas filas.')

    linea_ids = fields.One2many(
        've.conecta.carga.ventas.linea', 'carga_id', string='Filas',
        domain=[('eliminada_duplicado', '=', False)],
        help='Excluye a propósito las filas soft-eliminadas por '
             '"Eliminar Filas Duplicadas" (eliminada_duplicado=True) -- '
             'así todo el código que recorre self.linea_ids (confirmar, '
             'conteos, chequeo de bloqueantes) las ignora automáticamente, '
             'sin tener que filtrarlas en cada sitio. Para verlas, usar '
             'action_ver_duplicados_eliminados (busca sin este domain).')
    count_lineas = fields.Integer(compute='_compute_counts', store=True)
    count_brechas = fields.Integer(compute='_compute_counts', store=True)
    count_bloqueantes = fields.Integer(compute='_compute_counts', store=True)
    count_duplicados = fields.Integer(compute='_compute_counts', store=True)
    count_duplicados_eliminados = fields.Integer(compute='_compute_counts', store=True)
    count_discrepancias = fields.Integer(
        compute='_compute_counts', store=True,
        help='Filas con categoria_discrepancia distinta de False -- para el '
             'botón "Discrepancias" (2026-08-14), junta bloqueantes e '
             'informativas de todas las categorías en una sola vista '
             'filtrable/agrupable, en vez de tener que buscarlas a mano '
             'en la grilla completa de linea_ids.')
    count_anulacion_omitida = fields.Integer(
        compute='_compute_counts', store=True,
        help='Candidatas para "Ver Registro+Anulación Omitidos" -- pares con '
             'el mismo N° Factura cuyo monto neta a cero (ver es_anulacion_par '
             'en la línea). No están excluidas de linea_ids (a diferencia de '
             'eliminada_duplicado), solo se filtran aparte.')
    count_sin_retencion = fields.Integer(
        compute='_compute_counts_post_confirmacion',
        help='Candidatas para "Ver Facturas sin Retención (revisar)" -- '
             'no se puede calcular antes de confirmar (depende de que las '
             'facturas/ve.wh.iva ya existan), por eso va separado de '
             '_compute_counts. Gate real del botón, antes solo chequeaba '
             'estado == confirmado sin mirar si había alguna fila.')
    count_diferencias_archivo = fields.Integer(
        compute='_compute_counts_post_confirmacion',
        help='Candidatas para "Ver Facturas con Diferencia" -- mismo '
             'criterio que el domain del botón (diferencia_vs_archivo '
             '!= 0 real, no aceptada). Antes el botón también incluía '
             'filas sin N° de Control aunque el monto coincidiera '
             '(diferencia=0) -- quitado, pedido explícito 2026-08-12, '
             'ver action_ver_diferencias_archivo.')
    count_ret_no_spe = fields.Integer(
        compute='_compute_counts_post_confirmacion',
        string='N° Ret. no SPE',
        help='Candidatas para "Ret. no SPE" -- caso inverso a Sin Retención: '
             'retenciones (ve.wh.iva) que YA existen para facturas cuyo '
             'cliente actualmente NO está marcado como Agente de Retención. '
             'No aparecen en Dif. Retención porque el monto puede coincidir '
             'exacto con el archivo (diferencia_vs_archivo == 0) -- el '
             'problema no es el monto, es que la retención no debería '
             'existir. Caso real 2026-08-18 (Vencement): tras fusionar '
             'clientes duplicados por RIF y corregir es_agente_retencion a '
             'mano en un RIF que empieza por V (persona natural, no aplica '
             'SPE), sus 18 retenciones ya creadas quedaban invisibles con '
             'las herramientas existentes.')

    # ── Auditoría (principio SmartIVA: DJCS no responde por inputs, sí por
    #    cálculos/Dashboards — toda aprobación del cliente queda registrada) ──
    confirmado_por_id = fields.Many2one('res.users', string='Confirmado por', readonly=True)
    fecha_confirmacion = fields.Datetime(string='Fecha Confirmación', readonly=True)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') or vals.get('name') == 'Nueva Carga':
                vals['name'] = self.env['ir.sequence'].next_by_code(
                    've.conecta.carga.ventas') or 'CARGA-VTA/nueva'
        return super().create(vals_list)

    _MESES_ARCHIVO = {
        'ENE': 'Ene', 'FEB': 'Feb', 'MAR': 'Mar', 'ABR': 'Abr',
        'MAY': 'May', 'JUN': 'Jun', 'JUL': 'Jul', 'AGO': 'Ago',
        'SEP': 'Sep', 'OCT': 'Oct', 'NOV': 'Nov', 'DIC': 'Dic',
    }

    @api.depends('archivo_nombre')
    def _compute_mes_archivo(self):
        for rec in self:
            m = re.search(r'\b(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b',
                           (rec.archivo_nombre or '').upper())
            rec.mes_archivo = self._MESES_ARCHIVO[m.group(1)] if m else False

    @api.depends('linea_ids.brecha', 'linea_ids.bloqueante', 'linea_ids.es_duplicado_factura',
                 'linea_ids.eliminada_duplicado', 'linea_ids.es_anulacion_par',
                 'linea_ids.categoria_discrepancia')
    def _compute_counts(self):
        Linea = self.env['ve.conecta.carga.ventas.linea'].sudo()
        for rec in self:
            rec.count_lineas = len(rec.linea_ids)
            rec.count_brechas = len(rec.linea_ids.filtered('brecha'))
            rec.count_bloqueantes = len(rec.linea_ids.filtered('bloqueante'))
            rec.count_duplicados = len(rec.linea_ids.filtered('es_duplicado_factura'))
            rec.count_anulacion_omitida = len(rec.linea_ids.filtered('es_anulacion_par'))
            rec.count_discrepancias = len(rec.linea_ids.filtered('categoria_discrepancia'))
            # linea_ids ya excluye las soft-eliminadas (ver domain del campo) --
            # para contarlas hace falta un search aparte, sin ese domain.
            rec.count_duplicados_eliminados = Linea.search_count(
                [('carga_id', '=', rec.id), ('eliminada_duplicado', '=', True)])

    @api.depends('estado', 'linea_ids.invoice_id')
    def _compute_counts_post_confirmacion(self):
        WhIva = self.env['ve.wh.iva'].sudo()
        for rec in self:
            if rec.estado not in ('confirmado', 'confirmado_discrepancias'):
                rec.count_sin_retencion = 0
                rec.count_diferencias_archivo = 0
                rec.count_ret_no_spe = 0
                continue
            lineas_candidatas = rec.linea_ids.filtered(
                lambda l: l.invoice_id and l.partner_id and l.partner_id.es_agente_retencion)
            inv_ids = lineas_candidatas.mapped('invoice_id').ids
            con_retencion_ids = WhIva.search(
                [('invoice_id', 'in', inv_ids)]).mapped('invoice_id').ids
            rec.count_sin_retencion = len(
                [i for i in inv_ids if i not in con_retencion_ids])

            rec.count_diferencias_archivo = WhIva.search_count([
                ('invoice_id', 'in', rec.linea_ids.mapped('invoice_id').ids),
                ('diferencia_archivo_aceptada', '=', False),
                ('diferencia_vs_archivo', '!=', 0),
            ])

            lineas_no_spe = rec.linea_ids.filtered(
                lambda l: l.invoice_id and l.partner_id and not l.partner_id.es_agente_retencion)
            rec.count_ret_no_spe = WhIva.search_count([
                ('invoice_id', 'in', lineas_no_spe.mapped('invoice_id').ids),
            ])

    # ─────────────────────────────────────────────────────────────────────
    # Parseo
    # ─────────────────────────────────────────────────────────────────────

    def _leer_filas(self):
        """Devuelve lista de dicts crudos (una entrada por fila de datos)."""
        if not self.archivo:
            raise UserError('Suba un archivo (CSV o XLSX) primero.')
        raw = base64.b64decode(self.archivo)
        nombre = (self.archivo_nombre or '').lower()

        if nombre.endswith('.csv'):
            texto = raw.decode('utf-8-sig', errors='replace')
            sniffer_sample = texto[:2048]
            try:
                dialecto = csv.Sniffer().sniff(sniffer_sample, delimiters=',;')
            except csv.Error:
                dialecto = csv.excel
            filas = list(csv.reader(io.StringIO(texto), dialecto))
        else:
            try:
                import openpyxl
            except ImportError:
                raise UserError('La librería openpyxl no está disponible en este servidor.')
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(raw), read_only=True, data_only=True)
                hoja = wb[wb.sheetnames[0]]
                filas = list(hoja.iter_rows(values_only=True))
                wb.close()
            except Exception as e:
                raise UserError(f'No se pudo abrir el archivo: {e}')
            # Bug real encontrado 2026-08-05 (Cementos, Q1-01-2026.xlsx,
            # 1328 filas reales): read_only=True depende del metadato interno
            # <dimension> del XLSX para saber hasta dónde iterar -- si ese
            # archivo lo trae mal (algunos exportadores no lo escriben bien),
            # openpyxl corta la lectura casi de inmediato (acá, a 1 sola
            # fila con 1 celda) sin lanzar ningún error, dejando "Columnas
            # detectadas: (ninguna)" más abajo sin pista de la causa real.
            # Reintentar sin read_only (carga todo en memoria, más lento
            # pero confiable) solo cuando la primera lectura salió
            # sospechosamente corta.
            if len(filas) <= 1:
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(raw), read_only=False, data_only=True)
                    hoja = wb[wb.sheetnames[0]]
                    filas_reintento = list(hoja.iter_rows(values_only=True))
                    wb.close()
                    if len(filas_reintento) > len(filas):
                        filas = filas_reintento
                except Exception:
                    pass

        if not filas:
            raise UserError('El archivo está vacío.')

        encabezado = None
        data_start = 0
        for i, fila in enumerate(filas):
            if any(c is not None and str(c).strip() for c in fila):
                encabezado = fila
                data_start = i + 1
                break
        if encabezado is None:
            raise UserError('El archivo no tiene encabezados.')

        col_map = {}
        for i, h in enumerate(encabezado):
            if h is None:
                continue
            norm = _norm_header(str(h))
            if norm in _HEADER_MAP:
                col_map[i] = _HEADER_MAP[norm]

        mapeadas_ya = set(col_map.values())
        for i, h in enumerate(encabezado):
            if h is None or i in col_map:
                continue
            norm = _norm_header(str(h))
            for ancla, field in _HEADER_FALLBACK:
                if field in mapeadas_ya:
                    continue
                if ancla in norm:
                    col_map[i] = field
                    mapeadas_ya.add(field)
                    break

        mapeadas = set(col_map.values())
        faltantes = {'rif', 'fecha', 'nro_documento'} - mapeadas
        if faltantes:
            raise UserError(
                f'Columnas obligatorias faltantes: {", ".join(sorted(faltantes))}.\n'
                f'Columnas detectadas: {", ".join(sorted(mapeadas)) or "(ninguna)"}')

        raw_rows = []
        for row_num, row in enumerate(filas[data_start:], start=data_start + 1):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            # Char con ceros a la izquierda, no Integer — los campos de
            # texto se alinean a la izquierda por defecto (encabezado
            # incluido), a diferencia de los numéricos, que Odoo fuerza a
            # la derecha sin que un simple class="text-start" lo pueda
            # anular en la columna de encabezado (pedido 2026-07-24). El
            # padding a 4 dígitos mantiene el orden correcto como texto
            # (hasta 9999 filas).
            vals = {'fila': f'{row_num:04d}'}
            for col_idx, field in col_map.items():
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if cell is None or str(cell).strip() == '':
                    continue
                if field in _AMOUNT_FIELDS:
                    # SUMAR, no sobreescribir — mismo fix que
                    # ve_conecta_carga_compras.py (ver ese archivo para el
                    # detalle): un header duplicado (2 secciones con el
                    # mismo nombre de columna) no debe perder el monto de
                    # la sección que NO sea la última en aparecer.
                    vals[field] = vals.get(field, 0.0) + _parse_amount(cell)
                elif field == 'fecha':
                    if self.formato_fecha == 'auto':
                        # Resuelto recién al final (_resolver_fechas_auto),
                        # necesita ver TODAS las filas primero para saber
                        # cuál mes domina el archivo.
                        vals['_fecha_raw'] = cell
                    else:
                        vals[field] = _parse_date(cell, self.formato_fecha) or False
                elif field == 'rif':
                    vals[field] = _formatear_rif(str(cell).strip())
                elif field == 'tipo_transaccion':
                    vals[field] = _normalizar_tipo_transaccion(str(cell))
                else:
                    vals[field] = str(cell).strip()
            # Formato "largo" (Base Imponible + % Alíc./IVA genéricos por
            # fila, ej. Libro Ventas Cemento.xlsx) → repartir a base_16/
            # base_8 según la alícuota de ESTA fila, mismo criterio que
            # ve_conecta_carga_compras.py. No se mezcla con el formato
            # "ancho" (columnas ya separadas): solo actúa si base_generica
            # vino y base_16/base_8 no.
            base_gen = vals.pop('base_generica', None)
            alic = vals.pop('alicuota_pct', None)
            if alic is not None and 0 < alic <= 1:
                # Bug real 2026-07-31: el sistema de origen de este cliente
                # exporta la alícuota como fracción (0.16), no como entero
                # (16) — sin normalizar, ni 8 ni 16 matcheaban y todo caía
                # al default.
                alic = alic * 100
            if base_gen and not vals.get('base_16') and not vals.get('base_8'):
                if alic is not None and abs(alic - 8.0) < 0.5:
                    vals['base_8'] = base_gen
                else:
                    vals['base_16'] = base_gen  # 16% o alícuota no reconocida: default 16%
            raw_rows.append(vals)
        if self.formato_fecha == 'auto':
            self._resolver_fechas_auto(raw_rows)
        return raw_rows

    def _resolver_fechas_auto(self, raw_rows):
        """Resuelve vals['fecha'] para cada fila usando el mes que domina
        el archivo como referencia — mismo criterio ya validado 2026-07-31
        contra un archivo real de Cementos (1180 filas, 100% resuelto sin
        ambigüedad): las filas con una sola fecha calendario válida
        determinan el mes dominante; las filas ambiguas (día y mes ambos
        ≤12, ambas interpretaciones válidas) se resuelven a favor de la
        que coincide con ese mes. Si ninguna coincide, se deja la
        interpretación normal (queda visible para revisión, no se pierde
        en silencio)."""
        from collections import Counter
        conteo_meses = Counter()
        candidatos_por_fila = []
        for vals in raw_rows:
            cands = _fecha_candidatos(vals.get('_fecha_raw'))
            candidatos_por_fila.append(cands)
            if len(cands) == 1:
                conteo_meses[cands[0][1]] += 1
        mes_dominante = conteo_meses.most_common(1)[0][0] if conteo_meses else None
        for vals, cands in zip(raw_rows, candidatos_por_fila):
            vals.pop('_fecha_raw', None)
            elegido = None
            if cands:
                elegido = cands[0]
                if len(cands) > 1 and mes_dominante is not None:
                    for c in cands:
                        if c[1] == mes_dominante:
                            elegido = c
                            break
            vals['fecha'] = (
                f"{elegido[0]:04d}-{elegido[1]:02d}-{elegido[2]:02d}" if elegido else False
            )

    def action_ver_duplicados_pendientes(self):
        """Abre las filas de ESTA carga marcadas 'Factura duplicada'
        TODAVÍA sin eliminar, con el motivo del choque (N° Control/N°
        Factura, contra qué factura existente, misma Zona o distinta).
        Pedido explícito 2026-08-12: antes "Eliminar Filas Duplicadas"
        borraba de un clic sin poder revisar primero -- ahora este botón
        reemplaza a ese paso directo: el usuario revisa acá y elige
        eliminar (todas o solo las seleccionadas) desde la propia lista,
        con action_eliminar_seleccionadas en el header."""
        self.ensure_one()
        # Recompute en vivo antes de mostrar -- si esto es una carga
        # "Confirmado c/Discrepancias" y hubo un fix de código desde la
        # última vez, la vista debe reflejarlo ya (ver mismo criterio en
        # action_confirmar; no se hace en Borrador, ya viene fresco de
        # Previsualizar).
        if self.estado == 'confirmado_discrepancias':
            self.linea_ids.filtered(lambda l: not l.invoice_id)._compute_partner_id()
        list_view = self.env.ref('ve_retencion_iva.ve_conecta_carga_ventas_linea_view_list_pendientes')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Duplicadas a Revisar — {self.name}',
            'res_model': 've.conecta.carga.ventas.linea',
            'views': [(list_view.id, 'list')],
            'domain': [('carga_id', '=', self.id), ('es_duplicado_factura', '=', True),
                       ('eliminada_duplicado', '=', False)],
        }

    def action_ver_discrepancias(self):
        """Abre TODAS las discrepancias de facturación de ESTA carga (las 6
        categorías: Duplicada, Dato faltante, Fecha inválida, Registro+
        Anulación, Documento vacío, Error al postear) en una sola vista
        agrupada por categoría con totales -- pedido explícito 2026-08-14:
        antes había que revisar la grilla completa de linea_ids (hasta
        1000 filas por carga) para encontrar las problemáticas a mano.

        Recompute en vivo antes de mostrar, en 'confirmado' TAMBIÉN (no solo
        'confirmado_discrepancias' como action_ver_duplicados_pendientes) --
        una carga 'confirmado' puede tener líneas que se bloquearon DESPUÉS
        de confirmarse (ver 0dfb0d5), y el botón debe reflejarlo."""
        self.ensure_one()
        if self.estado in ('confirmado', 'confirmado_discrepancias'):
            self.linea_ids.filtered(lambda l: not l.invoice_id)._compute_partner_id()
        list_view = self.env.ref('ve_retencion_iva.ve_conecta_carga_ventas_linea_view_list_discrepancias')
        search_view = self.env.ref('ve_retencion_iva.ve_conecta_carga_ventas_linea_view_search_discrepancias')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Discrepancias — {self.name}',
            'res_model': 've.conecta.carga.ventas.linea',
            'views': [(list_view.id, 'list')],
            'search_view_id': search_view.id,
            'domain': [('carga_id', '=', self.id), ('categoria_discrepancia', '!=', False)],
            'context': {'search_default_group_categoria': 1},
        }

    def action_ver_duplicados_eliminados(self):
        """Abre las filas de ESTA carga marcadas eliminada_duplicado=True
        por ve.conecta.carga.ventas.linea::action_eliminar_seleccionadas
        (desde la lista "Duplicadas a Revisar") -- soft-delete, no
        unlink() real (pedido explícito 2026-08-12)."""
        self.ensure_one()
        list_view = self.env.ref('ve_retencion_iva.ve_conecta_carga_ventas_linea_view_list_eliminadas')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Duplicadas Eliminadas — {self.name}',
            'res_model': 've.conecta.carga.ventas.linea',
            'views': [(list_view.id, 'list')],
            'domain': [('carga_id', '=', self.id), ('eliminada_duplicado', '=', True)],
        }

    def action_ver_anulaciones_omitidas(self):
        """Abre las filas de ESTA carga marcadas es_anulacion_par=True --
        pares Registro+Anulación (mismo N° Factura, monto neto cero)
        omitidos sin bloquear al confirmar. Pedido explícito 2026-08-12,
        mismo patrón que "Ver Duplicadas Eliminadas": el chatter solo
        muestra el conteo agregado, el detalle fila por fila vive acá."""
        self.ensure_one()
        list_view = self.env.ref('ve_retencion_iva.ve_conecta_carga_ventas_linea_view_list_eliminadas')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Registro + Anulación Omitidos — {self.name}',
            'res_model': 've.conecta.carga.ventas.linea',
            'views': [(list_view.id, 'list')],
            'domain': [('carga_id', '=', self.id), ('es_anulacion_par', '=', True)],
        }

    def action_previsualizar(self):
        self.ensure_one()
        # Bug real encontrado 2026-08-13 (Cementos, 7 cargas): este guard
        # solo miraba 'confirmado', no 'confirmado_discrepancias' -- una
        # carga en ese estado intermedio (algunas filas ya con factura
        # real, otras bloqueadas pendientes) SÍ podía volver a
        # previsualizarse, y el unlink() de abajo borraba TODAS las líneas
        # sin distinguir, dejando huérfanas las facturas ya creadas (la
        # línea nueva las ve como "ya existe" y las bloquea como
        # duplicado). El propio comentario de abajo ya advertía esto, pero
        # el `if` nunca lo cubrió.
        if self.estado in ('confirmado', 'confirmado_discrepancias'):
            raise UserError('Esta carga ya fue confirmada — no se puede volver a previsualizar.')
        raw_rows = self._leer_filas()
        Linea = self.env['ve.conecta.carga.ventas.linea']
        # Reset completo, incluidas las ya soft-eliminadas (linea_ids las
        # excluye por su domain, ver ese campo) -- una re-previsualización
        # es un reinicio total desde el archivo, no tendría sentido dejar
        # residuo huérfano de una ronda de "Eliminar Duplicadas" anterior
        # acumulándose en cada vuelta.
        Linea.search([('carga_id', '=', self.id)]).unlink()
        for vals in raw_rows:
            vals['carga_id'] = self.id
            Linea.create(vals)
        return self._reload()

    def _journal_zona(self, company, zona, journal_default, crear=False):
        """Diario de venta de esta Zona -- get-or-create si crear=True (al
        facturar de verdad, dentro de action_confirmar), solo búsqueda si
        crear=False (dentro de _compute_partner_id, que no debe tener
        efectos secundarios). Sin Zona, usa journal_default (compatibilidad
        con clientes/cargas que no traen esa columna -- comparten diario
        como siempre). Ver _codigo_diario_zona para el criterio del código."""
        if not zona:
            return journal_default
        Journal = self.env['account.journal'].sudo()
        code = _codigo_diario_zona(zona)
        journal = Journal.search(
            [('type', '=', 'sale'), ('company_id', '=', company.id),
             ('code', '=', code)], limit=1)
        if not journal and crear:
            journal = Journal.create({
                'name': f'Ventas — Zona {zona}',
                'code': code,
                'type': 'sale',
                'company_id': company.id,
            })
        return journal or journal_default

    def action_confirmar(self, omitir_bloqueadas=False):
        self.ensure_one()
        # Bug real 2026-08-05 (Cementos, carga de 1327 filas — casi 6
        # minutos de procesamiento): dos peticiones action_confirmar
        # llegaron con ~0.3s de diferencia (probable reintento automático
        # de red/proxy ante una petición tan larga, sin doble clic de la
        # usuaria) y corrieron en paralelo -- ambas leían estado='borrador'
        # antes de que ninguna alcanzara a guardar 'confirmado', así que el
        # chequeo de abajo no las distinguía. Chocaron escribiendo la misma
        # fila de ve.conecta.carga.ventas.linea ("could not serialize
        # access due to concurrent update"), abortando toda la transacción
        # de la que perdió la carrera. El dato en sí quedó bien (la que
        # ganó completó normal) pero el error confundía sin explicar la
        # causa. FOR UPDATE NOWAIT hace que la segunda falle YA, con un
        # mensaje claro, en vez de competir por las mismas filas.
        try:
            self.env.cr.execute(
                'SELECT id FROM ve_conecta_carga_ventas WHERE id = %s FOR UPDATE NOWAIT',
                (self.id,))
        except Exception:
            raise UserError(
                'Esta carga ya se está confirmando en otra solicitud '
                '(puede tardar varios minutos en cargas grandes) -- '
                'espere a que termine antes de reintentar.')
        if self.estado == 'confirmado':
            if not omitir_bloqueadas:
                raise UserError('Esta carga ya fue confirmada.')
            # 2026-08-14: una carga 'confirmado' puede tener líneas que se
            # bloquearon DESPUÉS de confirmarse -- ej. una carga posterior
            # generó un choque de Zona con una factura de ESTA (ver
            # 0dfb0d5, 423 casos reales encontrados en Cementos). Antes
            # esto quedaba atascado sin ninguna forma de reintentar desde
            # la UI, porque el guard de arriba bloqueaba cualquier reintento
            # apenas veía estado='confirmado'. Recompute en vivo (mismo
            # criterio que la rama 'confirmado_discrepancias' más abajo)
            # antes de decidir si de verdad hay algo pendiente.
            self.linea_ids.filtered(lambda l: not l.invoice_id)._compute_partner_id()
            # Bug real encontrado en esta misma prueba (2026-08-14, Cementos
            # cargas 030-036): chequear solo 'bloqueante' está mal -- el
            # recompute de arriba puede DESBLOQUEAR una línea (ej. la carga
            # que la chocaba ya se resolvió con su propio diario de Zona)
            # sin que eso signifique que ya tiene factura. Con el chequeo
            # viejo, esas líneas quedaban con bloqueante=False y SIN
            # invoice_id para siempre -- este método cortaba acá pensando
            # que no había nada pendiente, sin llegar nunca a crearles la
            # factura real.
            if not self.linea_ids.filtered(lambda l: not l.invoice_id):
                raise UserError(
                    'Esta carga ya fue confirmada y no tiene filas pendientes.')
        if not self.linea_ids:
            raise UserError('No hay filas para confirmar — previsualice primero.')
        # Toda la data de input es responsabilidad del cliente — no se
        # inventa ni completa nada faltante (RIF, N° Factura) ni se permite
        # duplicar una factura/retención ya existente. Bloqueante, no
        # silencioso: la fila se queda marcada en la vista previa hasta que
        # el cliente la corrija.
        #
        # omitir_bloqueadas=True (pedido explícito 2026-08-12, botón
        # "Confirmar y Omitir Bloqueadas"): en vez de rechazar TODA la
        # carga por unas pocas filas bloqueadas (ej. falso positivo de
        # zona, ver [[project_bug_no_retencion_entre_spe]] hermano de
        # zona), confirma lo que SÍ está limpio y deja las bloqueadas
        # pendientes en la carga -- estado queda "Confirmado c/
        # Discrepancias" en vez de "Confirmado". Re-invocable: solo procesa
        # las que YA no están bloqueadas y aún no tienen factura -- no
        # reintenta las que ya tienen invoice_id. No se puede volver a
        # Previsualizar una carga parcialmente confirmada (reconstruiría
        # las líneas ya vinculadas a factura real, ver action_previsualizar)
        # -- por eso, SOLO en un reintento sobre una carga ya "Confirmado
        # c/Discrepancias", se fuerza el recompute acá mismo antes de
        # decidir qué sigue bloqueado, para que un fix de código nuevo (ej.
        # el de zona pendiente) sí se refleje en las filas pendientes sin
        # necesitar re-Previsualizar. No se hace en el primer intento
        # (estado='borrador') para no pagar ese costo de recompute extra
        # en cargas grandes que ya vienen frescas de Previsualizar (ver
        # bug de rendimiento 2026-08-05, carga de 1327 filas).
        if self.estado == 'confirmado_discrepancias':
            self.linea_ids.filtered(lambda l: not l.invoice_id)._compute_partner_id()
        bloqueadas = self.linea_ids.filtered('bloqueante')
        if bloqueadas and not omitir_bloqueadas:
            detalle = '\n'.join(
                f'  Fila {l.fila}: {l.brecha}' for l in bloqueadas[:15])
            raise UserError(
                f'{len(bloqueadas)} fila(s) bloqueada(s) — corrija o elimine antes de confirmar:\n'
                f'{detalle}')
        # Con omitir_bloqueadas, el bucle de creación de más abajo NUNCA
        # toca una fila bloqueante ni una que ya tenga factura (reintento
        # sobre una carga "Confirmado c/Discrepancias") -- sin bloqueadas
        # de por medio (caso normal), esto es simplemente self.linea_ids.
        lineas_a_procesar = self.linea_ids.filtered(
            lambda l: not l.bloqueante and not l.invoice_id)

        # Bug real encontrado 2026-07-30: sin este contexto, el create() de
        # ve.wh.iva (ver ese archivo) enganchaba cada retención nueva al
        # "período abierto más reciente" apenas se posteaba la factura —
        # antes de que el bucle de más abajo (que sí vincula por la fecha
        # real de cada factura) llegara a mirarla. Con dos lotes de años
        # distintos, el segundo lote quedaba pegado al período que acababa
        # de crear el primero, sin pasar nunca por _asegurar_periodo.
        self = self.with_context(ve_periodo_asignacion_manual=True)
        Partner = self.env['res.partner'].sudo()
        Move = self.env['account.move'].sudo()
        journal = self.env['account.journal'].sudo().search(
            [('type', '=', 'sale'), ('company_id', '=', self.company_id.id)], limit=1)
        # Diario dedicado para Notas de Crédito (Registro+Anulación) —
        # pedido explícito 2026-08-13: la NC debe postear con el mismo N°
        # de Documento que trae el Excel (igual al de la factura Registro
        # que anula), y Odoo exige nombre único por (name, journal_id)
        # ENTRE POSTEADAS sin importar move_type -- si NC y factura
        # comparten diario, el nombre repetido siempre choca. Diario
        # separado es además la práctica contable real (facturas y notas
        # de crédito en series distintas). Se crea solo/una vez por
        # compañía (get-or-create, igual patrón que el resto del módulo).
        journal_nc = self.env['account.journal'].sudo().search(
            [('type', '=', 'sale'), ('company_id', '=', self.company_id.id),
             ('code', '=', 'NCVTA')], limit=1)
        if not journal_nc:
            journal_nc = self.env['account.journal'].sudo().create({
                'name': 'Notas de Crédito de Ventas',
                'code': 'NCVTA',
                'type': 'sale',
                'company_id': self.company_id.id,
                'refund_sequence': True,
            })
        account = self.env['account.account'].sudo().search(
            [('account_type', 'in', ('income', 'income_other')),
             ('company_ids', 'in', [self.company_id.id]), ('currency_id', '=', False)], limit=1)
        if not (journal and account):
            raise UserError(
                'Falta configurar el diario o la cuenta de ventas de esta compañía '
                '(Contabilidad → Ajustes) antes de confirmar la carga.')
        tax_16 = self.env['account.tax'].sudo().search(
            [('type_tax_use', '=', 'sale'), ('amount', '=', 16.0),
             ('amount_type', '=', 'percent'), ('company_id', '=', self.company_id.id)], limit=1)
        tax_8 = self.env['account.tax'].sudo().search(
            [('type_tax_use', '=', 'sale'), ('amount', '=', 8.0),
             ('amount_type', '=', 'percent'), ('company_id', '=', self.company_id.id)], limit=1)
        # Cuenta por Cobrar de la compañía — mismo criterio que
        # wizard_reset_piloto.py::_asegurar_partners: sin fijarla en el
        # partner nuevo, queda mezclada con el default global en vez de la
        # propia de esta compañía (reportado en vivo 2026-07-24, "faltan
        # Cuenta por Cobrar/Pagar por defecto al crear el Cliente").
        receivable_acct = self.env['account.account'].sudo().search(
            [('account_type', '=', 'asset_receivable'),
             ('company_ids', 'in', [self.company_id.id])], limit=1)
        payable_acct = self.env['account.account'].sudo().search(
            [('account_type', '=', 'liability_payable'),
             ('company_ids', 'in', [self.company_id.id])], limit=1)

        creadas = nuevos_partners = nuevos_agentes = pagos_registrados = 0
        retenciones_confirmadas = n_anulacion_omitida = n_notas_credito = 0
        errores = []
        # (wh_id, monto_retenido del feed en esa fila) por cada retención
        # creada en esta carga — el bucket final (Confirmado vs No
        # Recibido/Vencido) se calcula DESPUÉS de todo el procesamiento
        # (incluido el sync de Vencido), usando el estado REAL de cada
        # retención en vez de asumir que "traía monto en el feed" siempre
        # implica que quedó confirmada (pedido explícito 2026-08-01).
        wh_tracking = []
        # Lineas creadas (con invoice_id) para las que el hook nativo NO
        # generó ninguna ve.wh.iva -- pedido explícito 2026-08-11, para
        # desglosar "Facturas sin Retención" en el Resumen por el motivo
        # real (Cliente no es Agente de Retención vs. anomalía a revisar),
        # en vez de un solo número sin explicación.
        sin_retencion_lineas = []
        WhIva = self.env['ve.wh.iva'].sudo()
        # RIF normalizado -> partner ya creado en ESTE mismo confirmar. Bug
        # real 2026-07-30: sin esta caché, un cliente nuevo (sin partner
        # previo) que aparece en varias filas del mismo archivo (lo normal —
        # varias facturas del mismo cliente) creaba un partner DISTINTO por
        # cada fila, porque `linea.partner_id` (calculado en la vista previa,
        # antes de que existiera ninguno) era False en todas por igual.
        # Primera pasada — determinar qué RIF son Agente de Retención ANTES
        # de crear ninguna factura. Bug real detectado 2026-08-01: si se
        # decide fila por fila (como antes), la PRIMERA factura de un
        # cliente en el archivo podía quedar sin retención esperada si esa
        # fila puntual no traía monto, aunque una fila POSTERIOR del mismo
        # cliente sí lo trajera — el hook nativo ya había posteado esa
        # primera factura antes de que el sistema supiera que ese cliente
        # es agente. Resuelto sin depender del orden: se mira TODO el
        # archivo primero.
        #
        # 2026-08-14: la columna opcional SPE/Contribuyente Especial dejó
        # de ser solo una señal ADICIONAL — un 'N'/'E'/'X' explícito ahora
        # BLOQUEA el heurístico viejo de monto_retenido>0 para ese RIF
        # (regla confirmada: solo Contribuyente=S calcula retención, sin
        # importar qué diga ninguna otra columna). 'S' sigue ganando
        # siempre (ratchet: una señal positiva en cualquier fila del RIF
        # no se pierde por una fila posterior menos concluyente, ni una
        # fila 'N/E/X' revierte un 'S' ya visto para ese mismo RIF).
        rif_es_agente = {}
        rif_spe_signal = {}
        # Validado SENIAT — mismo criterio "mirar TODO el archivo primero"
        # que rif_es_agente arriba: si el mismo RIF aparece en varias filas
        # con valores distintos, gana el de mayor prioridad (Sí > No >
        # blanco), sin depender del orden de las filas.
        rif_validado = {}
        for linea in self.linea_ids:
            if not linea.rif:
                continue
            key = _norm_rif(linea.rif)
            spe = _es_spe_verdadero(linea.es_spe)
            if spe is True:
                rif_spe_signal[key] = True
            elif spe is False and rif_spe_signal.get(key) is not True:
                rif_spe_signal[key] = False
            if spe is not False and (linea.monto_retenido > 0 or spe is True):
                rif_es_agente[key] = True
            val = _validado_seniat_norm(linea.validado_seniat)
            if val and _VALIDADO_PRIORIDAD[val] > _VALIDADO_PRIORIDAD.get(rif_validado.get(key), 0):
                rif_validado[key] = val
        # Aplicar el bloqueo: un RIF con señal 'N/E/X' explícita en alguna
        # fila y ningún 'S' en ninguna fila no es agente, aunque el
        # heurístico de monto_retenido>0 lo haya marcado True arriba antes
        # de ver esa señal (el orden de las filas no debe importar).
        for key, signal in rif_spe_signal.items():
            if signal is False:
                rif_es_agente.pop(key, None)

        partners_creados_lote = {}
        for linea in lineas_a_procesar:
            if linea.es_anulacion_par:
                base_linea_par = (linea.base_16 or 0) + (linea.base_8 or 0) + (linea.base_exento or 0)
                if base_linea_par >= 0:
                    # Es el "Registro" (positivo) del par -- pedido
                    # explícito 2026-08-13: ya NO se omite, se factura
                    # normal (cae al flujo de abajo, igual que cualquier
                    # otra fila) -- el cliente sí emitió este documento
                    # real, omitirlo dejaría el Libro de Ventas de Odoo
                    # incompleto frente a SENIAT aunque el neto del par
                    # sea cero.
                    pass
                elif linea.par_linea_id:
                    # Es la "Anulación" (negativa), con línea hermana
                    # conocida -- se crea como Nota de Crédito REAL
                    # (move_type='out_refund'), vinculada a la factura del
                    # Registro vía reversed_entry_id, con el N° de
                    # Documento tal cual viene del archivo (mismo criterio
                    # que toda factura -- no se autogenera). Si el
                    # Registro todavía no tiene factura (aparece más
                    # adelante en el archivo, aún no se procesó en esta
                    # misma pasada), se deja pendiente para el próximo
                    # reintento en vez de forzar el orden.
                    par = linea.par_linea_id
                    factura_registro = par.invoice_id
                    if not factura_registro:
                        errores.append(
                            f'Fila {linea.fila}: Nota de Crédito pendiente — su '
                            f'Registro (fila {par.fila}) aún no tiene factura, se '
                            f'reintentará')
                        continue
                    lineas_nc = []
                    if linea.base_16:
                        lv = {'name': f'Anulación — fila {linea.fila}', 'quantity': 1,
                              'price_unit': abs(linea.base_16), 'account_id': account.id}
                        if tax_16:
                            lv['tax_ids'] = [(6, 0, [tax_16.id])]
                        lineas_nc.append((0, 0, lv))
                    if linea.base_8:
                        lv = {'name': f'Anulación 8% — fila {linea.fila}', 'quantity': 1,
                              'price_unit': abs(linea.base_8), 'account_id': account.id}
                        if tax_8:
                            lv['tax_ids'] = [(6, 0, [tax_8.id])]
                        lineas_nc.append((0, 0, lv))
                    if linea.base_exento:
                        lineas_nc.append((0, 0, {
                            'name': f'Anulación (exento) — fila {linea.fila}',
                            'quantity': 1, 'price_unit': abs(linea.base_exento),
                            'account_id': account.id,
                        }))
                    nc = Move.create({
                        'name': linea.nro_documento or linea.nro_control,
                        'move_type': 'out_refund',
                        'reversed_entry_id': factura_registro.id,
                        'partner_id': factura_registro.partner_id.id,
                        'invoice_date': linea.fecha or fields.Date.today(),
                        'invoice_date_due': linea.fecha or fields.Date.today(),
                        'journal_id': journal_nc.id,
                        'company_id': self.company_id.id,
                        'currency_id': self.company_id.currency_id.id,
                        'nro_control': linea.nro_control or False,
                        'nro_factura': linea.nro_documento or False,
                        'zona': linea.zona or False,
                        'invoice_line_ids': lineas_nc,
                    })
                    try:
                        with self.env.cr.savepoint():
                            nc.action_post()
                        linea.invoice_id = nc.id
                        creadas += 1
                        n_notas_credito += 1
                        # Bug real encontrado 2026-08-20: esta rama nunca
                        # revisaba si el hook nativo generó (o no) una
                        # ve.wh.iva para la Nota de Crédito -- sin esto, una
                        # NC sin retención quedaba invisible tanto en
                        # wh_tracking como en sin_retencion_lineas, y el
                        # "TOTAL Sin Retención" del Resumen (creadas -
                        # retenciones_creadas) terminaba más alto que la suma
                        # real de las 2 filas del desglose por motivo.
                        wh_creada_nc = WhIva.search([('invoice_id', '=', nc.id)], limit=1)
                        if wh_creada_nc:
                            wh_tracking.append((wh_creada_nc.id, linea.monto_retenido))
                            wh_creada_nc.write({
                                'monto_retenido_archivo': linea.monto_retenido,
                                'monto_iva_archivo': linea.monto_iva,
                                'viene_de_libro_ventas': True,
                            })
                        else:
                            sin_retencion_lineas.append(linea)
                    except Exception as exc:
                        errores.append(f'Fila {linea.fila} (Nota de Crédito): {exc}')
                        # Si el posteo falla (ej. choque de nombre con OTRA
                        # NC ya posteada), no dejar el Borrador huérfano sin
                        # vínculo -- bug real encontrado 2026-08-13 en la
                        # primera prueba (43 Notas de Crédito huérfanas).
                        nc.sudo().unlink()
                        linea.write({
                            'categoria_discrepancia': 'error_posteo',
                            'brecha': f'Error al postear Nota de Crédito: {str(exc)[:200]}',
                            'bloqueante': False,
                        })
                    continue
                else:
                    # Anulación emparejada solo contra una factura PRE-
                    # EXISTENTE detectada por monto (no contra una línea
                    # hermana, ver match_existente en _compute_partner_id)
                    # -- comportamiento original sin cambios por ahora
                    # (se omite sin crear nada), pendiente de extender
                    # también a Nota de Crédito.
                    n_anulacion_omitida += 1
                    continue
            partner = linea.partner_id
            partner_creado = agente_marcado = False
            rif_key = _norm_rif(linea.rif) if linea.rif else False
            es_agente_este_rif = bool(rif_key and rif_es_agente.get(rif_key))
            validado_este_rif = rif_validado.get(rif_key) if rif_key else None
            if not partner and linea.rif:
                partner = partners_creados_lote.get(rif_key)
            if not partner and linea.rif:
                # Revalida en vivo contra la DB antes de crear -- linea.
                # partner_id es un compute store=True que puede haber
                # quedado obsoleto si se calculó ANTES de que OTRA carga
                # (distinta corrida de action_confirmar) creara este mismo
                # cliente por el mismo RIF. Bug real encontrado 2026-08-18
                # en Vencement: 230 RIFs con más de un contacto duplicado,
                # dejando huérfanas (sin retención, con saldo repartido) las
                # facturas del contacto "perdedor". RIF normalizado (sin
                # guiones/espacios) igual que en el resto del módulo (ver
                # _norm_rif). El RIF es el identificador único del
                # contribuyente -- si ya existe un cliente con ese RIF, ES
                # ese cliente, sin ambigüedad que revisar: se vincula acá
                # (nunca se crea un segundo registro para el mismo RIF) y
                # sigue el flujo normal más abajo (autocorrección de
                # formato, ratchet de validado_seniat/es_agente_retencion),
                # igual que si `partner` se hubiera encontrado desde el
                # principio.
                partner = Partner.search([
                    ('vat', '!=', False),
                    '|', ('company_id', '=', False), ('company_id', '=', self.company_id.id),
                ]).filtered(lambda p: _norm_rif(p.vat) == rif_key)[:1]
                if partner and linea.rif:
                    partners_creados_lote[rif_key] = partner
            if not partner:
                # company_id explícito (mismo criterio ya documentado en
                # wizard_reset_piloto.py::_asegurar_partners) — sin esto el
                # partner queda "global" (company_id=False) y el cliente
                # web activa otras compañías (ej. DJCS) en el selector al
                # navegar a sus facturas. Causa real del bug reportado
                # 2026-07-24 ("se habilita DJCS al ir a Cliente>Factura").
                vals_partner = {
                    'name': linea.nombre_cliente or linea.rif,
                    'vat': linea.rif,
                    'company_type': 'company',
                    'company_id': self.company_id.id,
                    'customer_rank': 1,
                    'es_agente_retencion': es_agente_este_rif,
                    'validado_seniat': validado_este_rif or False,
                }
                if receivable_acct:
                    vals_partner['property_account_receivable_id'] = receivable_acct.id
                if payable_acct:
                    vals_partner['property_account_payable_id'] = payable_acct.id
                partner = Partner.with_company(self.company_id).create(vals_partner)
                nuevos_partners += 1
                partner_creado = True
                if linea.rif:
                    partners_creados_lote[rif_key] = partner
                if es_agente_este_rif:
                    nuevos_agentes += 1
                    agente_marcado = True
            else:
                # Autocorrige el RIF guardado en un cliente ya existente si
                # quedó sin guión (dato legado, de antes de _formatear_rif) —
                # sin esto, _validar_para_confirmar rechaza el "recibir" con
                # "RIF no tiene el formato correcto" aunque el archivo SÍ
                # traiga el RIF bien. Bug real 2026-08-06: 10 facturas
                # creadas quedaron sin recibir por esto.
                rif_fmt = _formatear_rif(linea.rif) if linea.rif else False
                if rif_fmt and partner.vat != rif_fmt and _norm_rif(partner.vat) == rif_key:
                    partner.vat = rif_fmt
                if es_agente_este_rif and not partner.es_agente_retencion:
                    partner.es_agente_retencion = True
                    nuevos_agentes += 1
                    agente_marcado = True
                # Ratchet Sí > No > blanco -- nunca se sobreescribe con un
                # valor de menor prioridad (ver _VALIDADO_PRIORIDAD).
                if validado_este_rif and _VALIDADO_PRIORIDAD[validado_este_rif] > \
                        _VALIDADO_PRIORIDAD.get(partner.validado_seniat, 0):
                    partner.validado_seniat = validado_este_rif
            # Bug reportado 2026-07-24: la línea nunca quedaba enlazada al
            # partner recién creado/actualizado (solo la variable local
            # `partner` se usaba para la factura) — la brecha "Cliente
            # nuevo" se quedaba así para siempre en la vista previa aunque
            # el cliente ya existiera.
            linea.write({'partner_id': partner.id, 'partner_creado': partner_creado,
                         'agente_marcado': agente_marcado})

            lineas_factura = []
            if linea.base_16:
                lv = {'name': f'Carga Libro de Ventas — fila {linea.fila}',
                      'quantity': 1, 'price_unit': linea.base_16, 'account_id': account.id}
                if tax_16:
                    lv['tax_ids'] = [(6, 0, [tax_16.id])]
                lineas_factura.append((0, 0, lv))
            if linea.base_8:
                lv = {'name': f'Carga Libro de Ventas 8% — fila {linea.fila}',
                      'quantity': 1, 'price_unit': linea.base_8, 'account_id': account.id}
                if tax_8:
                    lv['tax_ids'] = [(6, 0, [tax_8.id])]
                lineas_factura.append((0, 0, lv))
            if linea.base_exento:
                # Venta exenta/no gravada — sin tax_ids a propósito, no
                # genera IVA (mismo criterio que wizard_reset_piloto.py).
                lineas_factura.append((0, 0, {
                    'name': f'Carga Libro de Ventas (exento) — fila {linea.fila}',
                    'quantity': 1, 'price_unit': linea.base_exento, 'account_id': account.id,
                }))
            if not lineas_factura:
                # No es un fallo de parseo — la única forma de llegar acá es
                # que base_16/base_8/base_exento sean 0 los 3 (si el feed
                # trae un monto real en cualquiera, alguna rama de arriba ya
                # lo agregó). Confirmado en vivo 2026-07-31 (Cementos, 10 de
                # 1180 filas): el propio "Total de Ventas Incluyendo IVA"
                # también estaba en 0 en esas filas — transacción realmente
                # vacía del cliente, no un gap del mapeo de columnas.
                errores.append(
                    f'Fila {linea.fila}: Base Imponible/Total en 0 en el archivo — '
                    'sin monto que facturar, fila omitida (no es un error de carga)')
                continue

            journal_linea = self._journal_zona(self.company_id, linea.zona, journal, crear=True)
            inv = Move.create({
                # N° Factura del cliente se respeta tal cual — SmartIVA NO
                # genera su propio número, la factura ya existe y ya tiene
                # su numeración legal real (pedido explícito 2026-07-24).
                # Si falta N° Factura pero SÍ trae N° Control, se usa este
                # último como respaldo (pedido explícito 2026-08-12 — N°
                # Factura/N° Control se tratan como intercambiables para
                # identificar la fila, ver el bloqueo arriba en
                # _compute_partner_id, que solo exige que exista AL MENOS
                # uno de los dos).
                'name': linea.nro_documento or linea.nro_control,
                'move_type': 'out_invoice',
                'partner_id': partner.id,
                'invoice_date': linea.fecha or fields.Date.today(),
                # Bug real encontrado 2026-07-30: sin fijar esto, Odoo lo
                # calculaba solo (según términos de pago del partner, o su
                # propio fallback si no hay ninguno) y terminaba dando HOY
                # en vez de la fecha real de la factura — con un Libro de
                # Ventas histórico (ej. 2025), la Fecha Límite Entrega del
                # comprobante (que ancla en invoice_date_due, ver
                # ve_wh_iva.py::_compute_fecha_vencimiento, diseño a
                # propósito por Art. 13) salía calculada desde HOY (2026)
                # pese a que Fecha Factura sí mostraba 2025 correctamente.
                # Por defecto se asume contado (vencimiento = misma fecha
                # de la factura) — si el feed llegara a traer un plazo de
                # pago real en el futuro, ajustar acá.
                'invoice_date_due': linea.fecha or fields.Date.today(),
                'journal_id': journal_linea.id,
                'company_id': self.company_id.id,
                'currency_id': self.company_id.currency_id.id,
                'nro_control': linea.nro_control or False,
                'nro_factura': linea.nro_documento or False,
                'zona': linea.zona or False,
                'invoice_line_ids': lineas_factura,
            })
            try:
                with self.env.cr.savepoint():
                    inv.action_post()
                linea.invoice_id = inv.id
                creadas += 1
            except Exception as exc:
                errores.append(f'Fila {linea.fila}: {exc}')
                # Persistir la categoría acá mismo (no alcanza con dejar que
                # el compute la re-derive luego): categoria_discrepancia
                # solo PRESERVA 'error_posteo' si ya estaba puesta antes --
                # la primera vez que pasa, el compute no tiene forma de
                # saberlo por su cuenta. Sin este write, quedaba visible
                # solo en el resumen del momento de confirmar (mismo bug
                # que las 329 de documento vacío, ver categoria_discrepancia).
                linea.write({
                    'categoria_discrepancia': 'error_posteo',
                    'brecha': f'Error al postear: {str(exc)[:200]}',
                    'bloqueante': False,
                })
                continue

            # Consistencia: registrar TODA retención esperada que el hook
            # nativo haya creado para esta factura, haya venido o no con
            # monto en el feed — pedido explícito 2026-08-01, tras
            # encontrar en vivo un caso real (530 retenciones en Odoo vs.
            # 426 filas con monto en el Excel): un cliente marcado Agente
            # de Retención genera retención esperada en TODAS sus
            # facturas (diseño correcto del hook, no un duplicado), no
            # solo en las filas que traen monto en esa transacción
            # puntual. Sin este desglose, la diferencia parece un bug.
            wh_creada = WhIva.search([('invoice_id', '=', inv.id)], limit=1)
            if wh_creada:
                wh_tracking.append((wh_creada.id, linea.monto_retenido))
                # Guardar el monto tal cual venía en el archivo -- permite
                # comparar después contra lo que Odoo calculó sin tener que
                # volver a abrir esta carga (ver diferencia_vs_archivo,
                # action_ver_diferencias_archivo). Pedido explícito
                # 2026-08-05.
                wh_creada.write({
                    'monto_retenido_archivo': linea.monto_retenido,
                    'monto_iva_archivo': linea.monto_iva,
                    'viene_de_libro_ventas': True,
                })
            else:
                sin_retencion_lineas.append(linea)

            # N° de Comprobante de Retención YA emitido por el agente de
            # retención (comprador) — el feed lo trae hecho, SmartIVA no lo
            # genera (aclaración explícita de la usuaria 2026-07-31). La
            # factura recién posteada disparó el hook nativo
            # (_ve_crear_retencion_esperada) que crea la retención en
            # 'esperado' — se busca esa MISMA retención por invoice_id
            # (única por diseño) y se avanza a Recibido (+Confirmado SOLO
            # si el monto del archivo cuadra con lo que calculó Odoo), en
            # vez de dejarla "No Recibida" esperando el Buzón/OCR de algo
            # que ya llegó hace tiempo.
            #
            # Bug real corregido 2026-08-06: "Montos según Comprobante"
            # (comp_base_16/comp_iva_16/comp_base_8/comp_iva_8/
            # comp_monto_retenido) se llenaban con el propio monto que
            # Odoo YA había calculado (wh.monto_base/monto_iva/...) -- una
            # comparación contra sí mismo, nunca podía dar diferente.
            # _compute_estado_recepcion (ve_wh_iva.py) ya compara
            # comp_monto_retenido vs. monto_retenido para marcar
            # recibido_dif/confirmado_dif (banner "Diferencia de monto con
            # el comprobante físico" en la ficha, badges en las listas) --
            # con el bug, esa detección nunca podía dispararse para
            # retenciones creadas por esta carga. Ahora comp_* se llena con
            # lo que el ARCHIVO trae de verdad: linea.base_16/base_8 (ya
            # vienen repartidos por alícuota desde el parseo, ver
            # _leer_filas) y linea.monto_retenido para comp_monto_retenido.
            # El IVA por alícuota no viene separado en el archivo (una sola
            # columna combinada) -- se deriva con la tasa legal fija
            # (16%/8%) sobre la base de cada alícuota, no se inventa un
            # reparto arbitrario del combinado.
            #
            # Segundo ajuste, mismo día, pedido explícito 2026-08-06: si
            # hay diferencia real, la retención NO se confirma sola -- se
            # queda en Recibido (con Dif), visible en "Facturas con
            # Diferencia", para que el usuario decida (Aceptar Monto del
            # Archivo, o confirmar tal cual el cálculo de Odoo). Antes
            # confirmaba siempre, lo que bloqueaba (is_locked) exactamente
            # los casos que más necesitaban revisión.
            if linea.nro_comp_retencion:
                wh = wh_creada
                if wh:
                    try:
                        with self.env.cr.savepoint():
                            wh.write({
                                'name': linea.nro_comp_retencion,
                                'canal_recepcion': 'libro_ventas',
                                'fecha': linea.fecha,
                                'comp_base_16': linea.base_16,
                                'comp_iva_16': round(linea.base_16 * 0.16, 2),
                                'comp_base_8': linea.base_8,
                                'comp_iva_8': round(linea.base_8 * 0.08, 2),
                                'comp_monto_retenido': linea.monto_retenido,
                            })
                            wh.action_recibir()
                            if abs(linea.monto_retenido - wh.monto_retenido) <= 0.01:
                                wh.action_confirmar()
                        retenciones_confirmadas += 1
                    except Exception as exc:
                        errores.append(
                            f'Fila {linea.fila}: factura creada pero la retención no se pudo '
                            f'recibir/confirmar con el N° de Comprobante '
                            f'"{linea.nro_comp_retencion}": {exc}')

            # EstadoPago "Pagada" → registrar el pago real (único punto de
            # integración: payment_state nativo alimenta Cobranza vs.
            # Comprobante, Crédito Fiscal en Tránsito y Lista de Trabajo —
            # no hace falta tocar esas 3 vistas, ver _es_pagado arriba).
            if _es_pagado(linea.estado_pago):
                pago_wizard = False
                try:
                    with self.env.cr.savepoint():
                        pago_wizard = self.env['account.payment.register'].sudo().with_context(
                            active_model='account.move', active_ids=inv.ids,
                        ).create({'payment_date': linea.fecha or fields.Date.today()})
                        pago_wizard._create_payments()
                    pagos_registrados += 1
                except Exception as exc:
                    errores.append(f'Fila {linea.fila}: factura creada pero el pago no se pudo '
                                    f'registrar (EstadoPago="{linea.estado_pago}"): {exc}')
                finally:
                    # Bug reportado 2026-07-24: el wizard transitorio
                    # (account.payment.register) quedaba vivo en la base
                    # después de usarlo — bloqueaba borrar el cliente
                    # después ("Otro modelo usa el registro: 'Pagar'").
                    if pago_wizard:
                        pago_wizard.unlink()

        # Vincular cada retención al período que le corresponde por su
        # PROPIA fecha de factura — el hook nativo (_ve_crear_retencion_
        # esperada) deja conciliacion_id vacío a propósito (en Odoo-nativo
        # ese vínculo lo completa "Conciliar SENIAT" más adelante).
        #
        # Bug real encontrado 2026-07-29: antes se barrían TODAS las
        # huérfanas al "período activo de hoy" sin mirar la fecha de cada
        # una — una carga con facturas de 2 quincenas distintas (ej. 1Q y
        # 2Q del mismo mes) dejaba las de 1Q vinculadas a la quincena de
        # HOY (2Q). La Declaración de esa quincena terminaba contando su
        # C.66 (retenciones) aunque su débito fiscal (C.42/43/49) no
        # aparecía ahí — ese cálculo vuelve a filtrar por `invoice_date`
        # contra el rango real de fechas de la quincena. Ahora cada
        # huérfana se vincula al período de SU propia `invoice_id.
        # invoice_date` (ver ve.conciliacion.periodo::_asegurar_periodo).
        n_vinculadas = 0
        periodos_usados = set()
        Periodo = self.env['ve.conciliacion.periodo'].sudo()
        huerfanas = self.env['ve.wh.iva'].sudo().search([
            ('company_id', '=', self.company_id.id), ('conciliacion_id', '=', False),
        ])
        for wh in huerfanas:
            fecha_factura = (wh.invoice_id.invoice_date if wh.invoice_id else False) or fields.Date.today()
            periodo = Periodo._asegurar_periodo(self.company_id, fecha_factura)
            wh.conciliacion_id = periodo.id
            periodos_usados.add(periodo.periodo_retencion)
            n_vinculadas += 1

        # Sincronizar Vencido de inmediato — sin esto, una carga histórica
        # (facturas cuya Fecha Límite ya pasó, ej. un Libro de Ventas de un
        # período cerrado) se queda en "No Recibido" hasta el próximo tick
        # del cron diario (`ir_cron_actualizar_estado_vencido`), que puede
        # tardar hasta 24h — o no correr nunca si el proyecto es un trial
        # de Odoo.sh sin worker de cron activo (bug real reportado en vivo
        # 2026-07-30). No tiene sentido esperar al cron para algo que ya se
        # sabe vencido en el momento de la carga.
        self.env['ve.wh.iva'].sudo()._cron_actualizar_estado_vencido()

        # Bucket final por ESTADO REAL de cada retención (Confirmado vs
        # No Recibido/Vencido) — se calcula recién ahora, después del
        # sync de Vencido de arriba, para que "vencido" ya esté aplicado.
        # No se asume que "traía monto en el feed" siempre implica
        # "quedó confirmada" (ni al revés) — se lee el estado real de
        # cada registro.
        retenciones_creadas = len(wh_tracking)
        retenciones_confirmadas_bucket = retenciones_pendientes_bucket = 0
        monto_feed_confirmadas = monto_odoo_confirmadas = 0.0
        monto_feed_pendientes = monto_odoo_pendientes = 0.0

        # Totales Archivo vs. Odoo (Base Imponible / Monto IVA / Monto
        # Retenido) para TODAS las facturas creadas, sin depender de que
        # traigan Zona ni de que hayan generado retención -- pedido
        # explícito 2026-08-06, para poder cuadrar el Resumen contra el
        # propio Libro de Ventas de un vistazo, no solo por Zona.
        base_archivo_tot = base_odoo_tot = 0.0
        iva_archivo_tot = iva_odoo_tot = 0.0
        retenido_archivo_tot = 0.0
        for linea in self.linea_ids:
            if not linea.invoice_id:
                continue
            # Bug real encontrado 2026-08-20: omitía base_exento (ventas
            # internas no gravadas) -- linea.invoice_id.amount_untaxed SI la
            # incluye (suma TODAS las líneas de la factura, gravadas o no),
            # así que sin esto "Diferencia" salía falsa cada vez que el
            # archivo traía esa columna con datos.
            base_archivo_tot += linea.base_16 + linea.base_8 + linea.base_exento
            iva_archivo_tot += linea.monto_iva
            retenido_archivo_tot += linea.monto_retenido
            # Segundo bug real 2026-08-20: una Nota de Crédito (Anulación de
            # un par Registro+Anulación) trae base_16/base_8/monto_iva NEGATIVOS
            # en el archivo (neteando contra su Registro), pero
            # amount_untaxed/amount_tax de un move out_refund en Odoo SIEMPRE
            # son positivos (misma convención que una factura normal, no un
            # signo invertido) -- sumarlos tal cual duplicaba el monto del
            # lado Odoo en vez de netearlo, inflando "Diferencia" en cada par.
            signo = -1 if linea.invoice_id.move_type == 'out_refund' else 1
            base_odoo_tot += signo * linea.invoice_id.amount_untaxed
            iva_odoo_tot += signo * linea.invoice_id.amount_tax

        # Consistencia por Zona — pedido explícito 2026-08-02, tras validar
        # a mano por RPC (Cementos, Nov 1Q 2025) que Total/IVA/Monto
        # Retenido cuadraban exacto por zona entre el archivo y Odoo. Se
        # arma solo si la carga trae al menos una zona (no aplica a
        # clientes sin ese dato, ej. DJCS-piloto). total_documento/
        # monto_iva son opcionales — si el feed no los trae, esas 2
        # columnas de archivo quedan en 0 y el Resultado lo aclara en vez
        # de marcar "Con diferencias" en falso.
        hay_zonas = any(l.zona for l in self.linea_ids)
        zonas_data = {}

        def _zbucket(zona):
            zona = zona or '(sin zona)'
            return zonas_data.setdefault(zona, {
                'leidas': 0, 'creadas': 0, 'confirmadas': 0,
                'total_archivo': 0.0, 'iva_archivo': 0.0, 'retenido_archivo': 0.0,
                'total_odoo': 0.0, 'iva_odoo': 0.0, 'retenido_odoo': 0.0,
            })

        if hay_zonas:
            for linea in self.linea_ids:
                b = _zbucket(linea.zona)
                b['leidas'] += 1
                b['total_archivo'] += linea.total_documento
                b['iva_archivo'] += linea.monto_iva
                b['retenido_archivo'] += linea.monto_retenido
                if linea.invoice_id:
                    b['creadas'] += 1
                    b['total_odoo'] += linea.invoice_id.amount_total
                    b['iva_odoo'] += linea.invoice_id.amount_tax

        # Desglose por estado_recepcion — pedido explícito 2026-08-06, para
        # que el Resumen deje ver de un vistazo cuántas retenciones quedaron
        # en cada estado (no solo Confirmado/pendiente en bloque) y que la
        # suma cuadre exacto contra "Facturas creadas" junto con "Facturas
        # sin Retención" (no CE, u otra razón).
        por_estado_recepcion = {}
        monto_por_estado_recepcion = {}
        # N° de Control: si falta, Odoo aplica 100% de retención por regla
        # legal (ver account_move.py::_ve_crear_retencion_esperada) -- eso
        # casi siempre difiere del monto que traía el archivo. Desglose
        # aparte pedido explícito 2026-08-11, para que esa diferencia no se
        # pierda mezclada dentro de "Confirmadas"/"Pendientes" en bloque.
        con_control_n = sin_control_n = 0
        con_control_feed = con_control_odoo = 0.0
        sin_control_feed = sin_control_odoo = 0.0

        if wh_tracking:
            whs_finales = {w.id: w for w in WhIva.browse([t[0] for t in wh_tracking])}
            for wh_id, monto_feed in wh_tracking:
                wh = whs_finales[wh_id]
                por_estado_recepcion[wh.estado_recepcion] = (
                    por_estado_recepcion.get(wh.estado_recepcion, 0) + 1)
                monto_por_estado_recepcion[wh.estado_recepcion] = (
                    monto_por_estado_recepcion.get(wh.estado_recepcion, 0.0) + wh.monto_retenido)
                if wh.nro_control:
                    con_control_n += 1
                    con_control_feed += monto_feed
                    con_control_odoo += wh.monto_retenido
                else:
                    sin_control_n += 1
                    sin_control_feed += monto_feed
                    sin_control_odoo += wh.monto_retenido
                # Bug real encontrado 2026-08-05 (Cementos, carga histórica
                # de enero — la mayoría de los comprobantes ya estaban
                # "Vencido" por fecha límite pasada al momento de cargar):
                # retenido_odoo por zona solo sumaba las Confirmadas,
                # dejando afuera Vencidas/No Recibidas -- el monto_retenido
                # de una retención se fija al CREARSE (viene del cálculo de
                # impuestos de la factura), no depende de si después se
                # recibió/confirmó el papel físico. total_odoo/iva_odoo
                # (arriba) ya suman TODAS las facturas creadas sin ese
                # filtro; retenido_odoo debe hacer lo mismo para que la
                # comparación con el archivo sea real, no una manzana
                # contra una naranja.
                if hay_zonas:
                    b = _zbucket(wh.zona)
                    b['retenido_odoo'] += wh.monto_retenido
                if wh.state == 'confirmado':
                    retenciones_confirmadas_bucket += 1
                    monto_feed_confirmadas += monto_feed
                    monto_odoo_confirmadas += wh.monto_retenido
                    if hay_zonas:
                        b['confirmadas'] += 1
                else:
                    retenciones_pendientes_bucket += 1
                    monto_feed_pendientes += monto_feed
                    monto_odoo_pendientes += wh.monto_retenido

        # bloqueadas se calculó ANTES del bucle de creación -- nada en ese
        # bucle toca el campo bloqueante de otra fila, así que sigue
        # reflejando exactamente qué quedó sin procesar.
        self.write({
            'estado': 'confirmado_discrepancias' if bloqueadas else 'confirmado',
            'confirmado_por_id': self.env.user.id,
            'fecha_confirmacion': fields.Datetime.now(),
        })

        def _n(v):
            return f'{v:,}'

        def _m(v):
            return f'{v:,.2f}' if v else '—'

        diferencia_confirmadas = monto_odoo_confirmadas - monto_feed_confirmadas
        th = ('style="border:1px solid #ccc; padding:3px 8px; background:#f0f0f0; '
              'text-align:left;"')
        td = 'style="border:1px solid #ccc; padding:3px 8px;"'
        tdr = 'style="border:1px solid #ccc; padding:3px 8px; text-align:right;"'

        def _fila(concepto, cant, m_archivo='—', m_odoo='—', dif='—', color=None):
            estilo_dif = f' style="color:{color};"' if color else ''
            return (
                f'<tr><td {td}>{concepto}</td>'
                f'<td {tdr}>{cant}</td>'
                f'<td {tdr}>{m_archivo}</td>'
                f'<td {tdr}>{m_odoo}</td>'
                f'<td {tdr}><span{estilo_dif}>{dif}</span></td></tr>'
            )

        diferencia_control = sin_control_odoo - sin_control_feed
        tabla_consistencia = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Concepto</th><th {th}>Cantidad</th>'
            f'<th {th}>Monto Archivo</th><th {th}>Monto Odoo</th>'
            f'<th {th}>Diferencia</th></tr>'
            + _fila('Filas leídas', _n(len(self.linea_ids)))
            + _fila('Facturas creadas', _n(creadas))
            + _fila('Facturas rechazadas (ver desglose por motivo más abajo)',
                    _n(len(self.linea_ids) - creadas))
            + _fila('Facturas sin Retención generada en SmartIVA (ver desglose más abajo)',
                    _n(creadas - retenciones_creadas))
            + _fila('Facturas con monto retenido y Retención creada como Confirmado',
                    _n(retenciones_confirmadas_bucket), _m(monto_feed_confirmadas),
                    _m(monto_odoo_confirmadas),
                    _m(diferencia_confirmadas) if abs(diferencia_confirmadas) > 0.01 else 'cuadra',
                    color='#dc3545' if abs(diferencia_confirmadas) > 0.01 else '#198754')
            + _fila('Facturas sin monto retenido y Retención creada como No Recibido/Vencido',
                    _n(retenciones_pendientes_bucket), '—', _m(monto_odoo_pendientes))
            + _fila('Retenciones SIN N° de Control (retención forzada al 100%, regla legal SPE)',
                    _n(sin_control_n), _m(sin_control_feed), _m(sin_control_odoo),
                    _m(diferencia_control) if abs(diferencia_control) > 0.01 else 'cuadra',
                    color='#dc3545' if abs(diferencia_control) > 0.01 else '#198754')
            + _fila('Retenciones creadas (total)', _n(retenciones_creadas), '—',
                    _m(monto_odoo_confirmadas + monto_odoo_pendientes))
            + '</table>'
        )

        # Estados de las Retenciones — pedido explícito 2026-08-06, con
        # columna Monto y separada de "Facturas sin Retención" agregado
        # explícito 2026-08-11: "Confirmado · c/Dif" etc. son estados de
        # una retención que SÍ existe (ve.wh.iva real) -- mezclarlos en la
        # misma tabla con las filas de "Sin Retención" (que no tienen
        # NINGÚN registro) confundía la lectura, ver conversación.
        _ESTADO_RECEPCION_LABEL = {
            'esperado': 'No Recibido', 'vencido': 'Vencido',
            'recibido': 'Recibido', 'recibido_dif': 'Recibido · c/Dif',
            'confirmado': 'Confirmado', 'confirmado_dif': 'Confirmado · c/Dif',
            'anulado': 'Anulado',
        }
        filas_estado = ''.join(
            f'<tr><td {td}>{_ESTADO_RECEPCION_LABEL.get(estado, estado)}</td>'
            f'<td {tdr}>{_n(n)}</td>'
            f'<td {tdr}>{_m(monto_por_estado_recepcion.get(estado, 0.0))}</td></tr>'
            for estado, n in sorted(por_estado_recepcion.items(), key=lambda kv: -kv[1])
        )
        tabla_estados = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Estado de la Retención</th><th {th}>Cantidad</th>'
            f'<th {th}>Monto Retenido (Odoo)</th></tr>'
            + filas_estado
            + f'<tr style="font-weight:700; border-top:2px solid #999;">'
              f'<td {td}>TOTAL Retenciones Generadas</td>'
              f'<td {tdr}>{_n(retenciones_creadas)}</td>'
              f'<td {tdr}>{_m(monto_odoo_confirmadas + monto_odoo_pendientes)}</td></tr>'
            + '</table>'
        )

        # Facturas SIN Retención generada — desglose por motivo (NO es un
        # "estado" de una retención, es la ausencia total de un registro
        # ve.wh.iva para esa factura). Pedido explícito 2026-08-11: el
        # número solo ("Facturas sin Retención") no dejaba ver si era
        # normal (cliente no calificaba) o una anomalía a revisar (cliente
        # SÍ es Agente de Retención pero el hook nativo no generó nada).
        sin_ret_agente_true = sin_ret_agente_false = 0
        base_agente_true = base_agente_false = 0.0
        for linea in sin_retencion_lineas:
            base = linea.base_16 + linea.base_8
            if linea.partner_id and linea.partner_id.es_agente_retencion:
                sin_ret_agente_true += 1
                base_agente_true += base
            else:
                sin_ret_agente_false += 1
                base_agente_false += base
        sin_retencion = len(sin_retencion_lineas)

        # Rechazadas — pedido explícito 2026-08-20: la tabla de abajo antes
        # solo mostraba "Sin Retención" (facturas SÍ creadas sin retención),
        # dejando las rechazadas (NUNCA llegaron a ser factura) fuera de
        # cualquier desglose por motivo -- solo vivían como texto libre en
        # "Filas con error"/contadores sueltos. Se categorizan acá con el
        # mismo campo `categoria_discrepancia` que ya usa la pestaña
        # Discrepancias (fuente única de verdad, ya se recalcula sola en
        # cada línea -- no hay que rastrear listas nuevas durante el bucle
        # de arriba).
        rechazadas_lineas = self.linea_ids.filtered(lambda l: not l.invoice_id)
        CATEGORIA_LABEL = dict(
            self.env['ve.conecta.carga.ventas.linea']._fields['categoria_discrepancia'].selection)
        rechazadas_por_categoria = {}
        for linea in rechazadas_lineas:
            cat = linea.categoria_discrepancia or False
            bucket = rechazadas_por_categoria.setdefault(cat, {'n': 0, 'base': 0.0})
            bucket['n'] += 1
            bucket['base'] += linea.base_16 + linea.base_8 + linea.base_exento
        rechazadas = len(rechazadas_lineas)

        # CHEQUEO ahora contra Filas Leídas (universo completo del archivo),
        # no solo contra Facturas creadas -- pedido explícito: Retenciones
        # Generadas + Sin Retención + Rechazadas debe sumar Filas Leídas.
        total_filas = len(self.linea_ids)
        no_generadas = sin_retencion + rechazadas
        suma_cuadra = (retenciones_creadas + no_generadas) == total_filas

        filas_rechazadas_html = ''
        # Orden fijo (no por cantidad) para que la tabla no "salte" de
        # posición entre cargas distintas -- más fácil de leer de un
        # vistazo cuando se repite la prueba varias veces.
        ORDEN_CATEGORIAS = ['duplicada', 'dato_faltante', 'fecha_invalida',
                            'registro_anulacion', 'documento_vacio', 'error_posteo', False]
        for cat in ORDEN_CATEGORIAS:
            if cat not in rechazadas_por_categoria:
                continue
            b = rechazadas_por_categoria[cat]
            etiqueta = CATEGORIA_LABEL.get(cat, 'Sin categoría (revisar)')
            filas_rechazadas_html += (
                f'<tr><td {td}><span style="color:#dc3545;">Rechazada — {etiqueta}</span></td>'
                f'<td {tdr}>{_n(b["n"])}</td><td {tdr}>{_m(b["base"])}</td></tr>'
            )

        tabla_sin_retencion = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Motivo</th><th {th}>Cantidad</th>'
            f'<th {th}>Base Imponible (informativa)</th></tr>'
            f'<tr><td {td}>Cliente NO es Agente de Retención (no le correspondía retener)</td>'
            f'<td {tdr}>{_n(sin_ret_agente_false)}</td><td {tdr}>{_m(base_agente_false)}</td></tr>'
            f'<tr><td {td}><span style="color:#dc3545;">Cliente SÍ es Agente de Retención pero no '
            f'se generó retención (revisar)</span></td>'
            f'<td {tdr}>{_n(sin_ret_agente_true)}</td><td {tdr}>{_m(base_agente_true)}</td></tr>'
            + filas_rechazadas_html
            + f'<tr style="font-weight:700; border-top:2px solid #999;">'
              f'<td {td}>TOTAL Sin Retención + Rechazadas</td>'
              f'<td {tdr}>{_n(no_generadas)}</td>'
              f'<td {tdr}>{_m(base_agente_true + base_agente_false + sum(b["base"] for b in rechazadas_por_categoria.values()))}</td></tr>'
            + '</table>'
            + f'<p style="font-size:0.75rem; color:#666;">CHEQUEO: Retenciones Generadas '
              f'({_n(retenciones_creadas)}) + Sin Retención + Rechazadas ({_n(no_generadas)}) = '
              f'<span style="color:{"#198754" if suma_cuadra else "#dc3545"};">'
              f'{_n(retenciones_creadas + no_generadas)}</span> '
              f'— debe cuadrar con Filas Leídas ({_n(total_filas)})</p>'
        )

        # Tabla "Estimadas vs. Retenido Real" ELIMINADA (pedido explícito
        # 2026-08-20) -- confundía más de lo que aclaraba en la demo. La
        # comparación real Excel/Odoo ya vive en "Montos: Archivo vs. Odoo"
        # (arriba) y "Retenciones SIN N° de Control" (tabla Consistencia).

        # Montos: Archivo vs. Odoo — pedido explícito 2026-08-06, 2 filas
        # (Archivo/Odoo) + 1 de Diferencia, mismas 3 columnas que ya usa la
        # vista "Facturas con Diferencia" (Base Imponible, Monto IVA, Monto
        # Retenido) -- para toda la carga, no solo las filas con diferencia.
        retenido_odoo_tot = monto_odoo_confirmadas + monto_odoo_pendientes

        def _fila_monto(etiqueta, base, iva, retenido, color=None):
            estilo = f' style="color:{color};"' if color else ''
            return (
                f'<tr><td {td}><span{estilo}>{etiqueta}</span></td>'
                f'<td {tdr}><span{estilo}>{_m(base)}</span></td>'
                f'<td {tdr}><span{estilo}>{_m(iva)}</span></td>'
                f'<td {tdr}><span{estilo}>{_m(retenido)}</span></td></tr>'
            )

        dif_base = base_odoo_tot - base_archivo_tot
        dif_iva = iva_odoo_tot - iva_archivo_tot
        dif_retenido = retenido_odoo_tot - retenido_archivo_tot
        tabla_montos = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}></th><th {th}>Base Imponible</th>'
            f'<th {th}>Monto IVA</th><th {th}>Monto Retenido</th></tr>'
            + _fila_monto('Leído del Archivo', base_archivo_tot, iva_archivo_tot, retenido_archivo_tot)
            + _fila_monto('Odoo', base_odoo_tot, iva_odoo_tot, retenido_odoo_tot)
            + _fila_monto('Diferencia', dif_base, dif_iva, dif_retenido,
                          color='#198754' if not (abs(dif_base) > 0.01 or abs(dif_iva) > 0.01
                                                   or abs(dif_retenido) > 0.01) else '#dc3545')
            + '</table>'
        )

        tabla_zona = ''
        if hay_zonas:
            MARGEN = 0.01
            hay_total_archivo = any(l.total_documento for l in self.linea_ids)
            hay_iva_archivo = any(l.monto_iva for l in self.linea_ids)

            def _fila_zona(zona, b):
                diffs = [b['retenido_odoo'] - b['retenido_archivo']]
                if hay_total_archivo:
                    diffs.append(b['total_odoo'] - b['total_archivo'])
                if hay_iva_archivo:
                    diffs.append(b['iva_odoo'] - b['iva_archivo'])
                cuadra = all(abs(d) <= MARGEN for d in diffs)
                resultado = 'Cuadra' if cuadra else 'Con diferencias'
                color = '#198754' if cuadra else '#dc3545'
                return (
                    f'<tr><td {td}>{zona}</td>'
                    f'<td {tdr}>{_n(b["leidas"])}</td>'
                    f'<td {tdr}>{_n(b["creadas"])}</td>'
                    f'<td {tdr}>{_n(b["confirmadas"])}</td>'
                    f'<td {tdr}>{_m(b["total_archivo"]) if hay_total_archivo else "—"}</td>'
                    f'<td {tdr}>{_m(b["iva_archivo"]) if hay_iva_archivo else "—"}</td>'
                    f'<td {tdr}>{_m(b["retenido_archivo"])}</td>'
                    f'<td {tdr}>{_m(b["total_odoo"]) if hay_total_archivo else "—"}</td>'
                    f'<td {tdr}>{_m(b["iva_odoo"]) if hay_iva_archivo else "—"}</td>'
                    f'<td {tdr}>{_m(b["retenido_odoo"])}</td>'
                    f'<td {tdr}><span style="color:{color};">{resultado}</span></td></tr>'
                )

            zonas_ordenadas = sorted(zonas_data.items(), key=lambda kv: -kv[1]['total_odoo'])
            total_b = {k: 0.0 for k in next(iter(zonas_data.values()))}
            for _zona, b in zonas_ordenadas:
                for k in total_b:
                    total_b[k] += b[k]

            tabla_zona = (
                f'<table style="border-collapse:collapse; font-size:0.85rem;">'
                f'<tr><th {th}>Zona</th><th {th}>Facturas Leídas</th>'
                f'<th {th}>Facturas Creadas</th><th {th}>Retenciones Confirmadas</th>'
                f'<th {th}>Total c/IVA (Archivo)</th><th {th}>Monto IVA (Archivo)</th>'
                f'<th {th}>Monto Retenido (Archivo)</th>'
                f'<th {th}>Total c/IVA (Odoo)</th><th {th}>Monto IVA (Odoo)</th>'
                f'<th {th}>Monto Retenido (Odoo)</th><th {th}>Resultado</th></tr>'
                + ''.join(_fila_zona(zona, b) for zona, b in zonas_ordenadas)
                + _fila_zona('TOTAL', total_b)
                + '</table>'
            )
            if not (hay_total_archivo and hay_iva_archivo):
                tabla_zona += (
                    '<p style="font-size:0.75rem; color:#666;">'
                    'El archivo no trae columna de Total y/o Impuesto Iva — esas '
                    'columnas de Archivo quedan en "—" y el Resultado se calcula '
                    'solo con Monto Retenido.</p>'
                )

        tabla_detalle = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Concepto</th><th {th}>Valor</th></tr>'
            f'<tr><td {td}>Pagos registrados (EstadoPago="Pagada")</td>'
            f'<td {tdr}>{_n(pagos_registrados)}</td></tr>'
            f'<tr><td {td}>Retenciones recibidas con N° de Comprobante real '
            f'(confirmadas si el monto coincidía con Odoo, si no quedan '
            f'Recibidas c/Dif para revisar)</td>'
            f'<td {tdr}>{_n(retenciones_confirmadas)}</td></tr>'
            f'<tr><td {td}>Retenciones vinculadas a su período (por fecha de factura)</td>'
            f'<td {tdr}>{_n(n_vinculadas)} — {", ".join(sorted(periodos_usados)) or "—"}</td></tr>'
            f'<tr><td {td}>Clientes nuevos (marcados Agente de Retención)</td>'
            f'<td {tdr}>{_n(nuevos_partners)} ({_n(nuevos_agentes)})</td></tr>'
            f'<tr><td {td}>Notas de Crédito creadas (Anulación real de un Registro '
            f'-- mismo archivo u otra carga con factura ya existente, N° Documento '
            f'tal cual el archivo)</td>'
            f'<td {tdr}>{_n(n_notas_credito)}</td></tr>'
            f'<tr><td {td}>Pares Registro + Anulación omitidos (emparejados solo por '
            f'monto contra una factura preexistente, sin línea hermana — no es '
            f'duplicado real, ver "Filas con error" para el detalle fila por fila)</td>'
            f'<td {tdr}>{_n(n_anulacion_omitida)}</td></tr>'
            f'</table>'
        )

        titulo = ('Carga de Libro de Ventas confirmada CON DISCREPANCIAS'
                  if bloqueadas else 'Carga de Libro de Ventas confirmada')
        cuerpo = (
            f'<b>{titulo}</b><br/>'
            f'<b>Archivo:</b> {self.archivo_nombre or "—"}<br/>'
            f'<b>Confirmado por:</b> {self.env.user.name}<br/>'
            + (f'<b>Filas bloqueadas pendientes (sin procesar, ver "Ver '
               f'Duplicadas (Revisar)"):</b> {len(bloqueadas)}<br/>' if bloqueadas else '')
            + '<br/>'
            f'<b>— Consistencia —</b><br/>{tabla_consistencia}<br/>'
            f'<b>— Retenciones Generadas — Estado Actual —</b><br/>{tabla_estados}<br/>'
            f'<b>— Facturas SIN Retención Generada y Rechazadas — Desglose por Motivo —</b><br/>{tabla_sin_retencion}<br/>'
            f'<b>— Montos: Archivo vs. Odoo —</b><br/>{tabla_montos}<br/>'
            + (f'<b>— Consistencia por Zona —</b><br/>{tabla_zona}<br/>' if hay_zonas else '')
            + f'<b>— Detalle —</b><br/>{tabla_detalle}'
        )
        if errores:
            cuerpo += '<br/><b>Filas con error:</b><br/>' + '<br/>'.join(errores)
        self.message_post(body=Markup(cuerpo), message_type='comment', subtype_xmlid='mail.mt_note')
        return self._reload()

    def action_confirmar_omitir_bloqueadas(self):
        """Wrapper de action_confirmar(omitir_bloqueadas=True) -- los
        botones de vista solo invocan el método sin argumentos extra, así
        que la variante se expone aparte. Pedido explícito 2026-08-12: con
        9 cargas de Cementos completamente atascadas (¡~6.800 filas
        legítimas retenidas por solo ~340 filas bloqueadas del choque de
        zona Ocumare/Pertigalete!) porque action_confirmar es todo-o-nada,
        esto confirma lo que SÍ está limpio y deja las bloqueadas
        pendientes para resolver después sin re-tocar lo ya confirmado."""
        return self.action_confirmar(omitir_bloqueadas=True)

    def action_ver_diferencias_archivo(self):
        """Abre Retenciones IVA Clientes filtrada a las de ESTA carga cuyo
        Monto Esperado (calculado por Odoo) difiere del Monto Retenido que
        traía el archivo -- para que el usuario compare fila por fila y
        decida si corrige el N° Control en el sistema origen o si el
        archivo del cliente subestimó la retención.

        Antes (pedido 2026-08-05) el domain también incluía las filas sin
        N° de Control aunque el monto coincidiera exacto (diferencia=0),
        para que el usuario las revisara igual -- pero eso hacía que
        "Ver Facturas con Diferencia" mostrara filas sin ninguna
        diferencia real, confuso. Pedido explícito 2026-08-12: el botón
        solo debe mostrar diferencia_vs_archivo != 0 real."""
        self.ensure_one()
        inv_ids = self.linea_ids.mapped('invoice_id').ids
        list_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_list_diferencias_archivo')
        form_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_form')
        search_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_search_diferencias_archivo')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Dif. Retención — {self.name}',
            'res_model': 've.wh.iva',
            'views': [(list_view.id, 'list'), (form_view.id, 'form')],
            'search_view_id': search_view.id,
            'domain': [
                ('invoice_id', 'in', inv_ids),
                ('diferencia_archivo_aceptada', '=', False),
                ('diferencia_vs_archivo', '!=', 0),
            ],
        }

    def action_ver_diferencias_aceptadas(self):
        """Igual que action_ver_diferencias_archivo pero al revés -- las de
        ESTA carga donde el usuario ya usó "Aceptar Monto del Archivo".
        Aparte porque el domain del botón de pendientes excluye a propósito
        las ya aceptadas (piden desaparecer de ahí), así que no hay forma
        de verlas desde esa misma pantalla. Pedido explícito 2026-08-06."""
        self.ensure_one()
        inv_ids = self.linea_ids.mapped('invoice_id').ids
        list_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_list_diferencias_archivo')
        form_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_form')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Diferencias Aceptadas — {self.name}',
            'res_model': 've.wh.iva',
            'views': [(list_view.id, 'list'), (form_view.id, 'form')],
            'domain': [
                ('invoice_id', 'in', inv_ids),
                ('diferencia_archivo_aceptada', '=', True),
            ],
        }

    def action_ver_sin_retencion_revisar(self):
        """Abre las facturas de ESTA carga cuyo cliente SÍ está marcado
        Agente de Retención pero para las que el hook nativo NO generó
        ninguna ve.wh.iva -- anomalía real a revisar (un cliente Agente de
        Retención debería generar retención esperada en TODAS sus
        facturas, ver account_move.py::_ve_crear_retencion_esperada).
        Pedido explícito 2026-08-11: antes esta cifra ("Facturas sin
        Retención") solo existía como número en el texto del chatter, sin
        forma de ver ni una sola fila concreta para investigar por qué."""
        self.ensure_one()
        lineas_candidatas = self.linea_ids.filtered(
            lambda l: l.invoice_id and l.partner_id and l.partner_id.es_agente_retencion)
        inv_ids = lineas_candidatas.mapped('invoice_id').ids
        con_retencion_ids = self.env['ve.wh.iva'].sudo().search(
            [('invoice_id', 'in', inv_ids)]).mapped('invoice_id').ids
        sin_retencion_ids = [i for i in inv_ids if i not in con_retencion_ids]
        list_view = self.env.ref('ve_retencion_iva.ve_account_move_view_list_sin_retencion')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Facturas sin Retención — Agente de Retención (revisar) — {self.name}',
            'res_model': 'account.move',
            'views': [(list_view.id, 'list'), (False, 'form')],
            'domain': [('id', 'in', sin_retencion_ids)],
        }

    def action_ver_ret_no_spe(self):
        """Abre las retenciones IVA (ve.wh.iva) de ESTA carga que YA existen
        para facturas cuyo cliente actualmente NO está marcado Agente de
        Retención -- caso inverso a "Sin Retención". No aparecen en "Dif.
        Retención" porque el monto puede coincidir exacto con el archivo
        (el problema no es el monto, es que la retención no debería
        existir). Pedido explícito 2026-08-18, tras encontrar 18 así en
        Vencement luego de fusionar un cliente duplicado y corregir su
        es_agente_retencion a mano (RIF persona natural, no aplica SPE)."""
        self.ensure_one()
        lineas_no_spe = self.linea_ids.filtered(
            lambda l: l.invoice_id and l.partner_id and not l.partner_id.es_agente_retencion)
        inv_ids = lineas_no_spe.mapped('invoice_id').ids
        list_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_list_no_spe')
        form_view = self.env.ref('ve_retencion_iva.ve_wh_iva_view_form')
        return {
            'type': 'ir.actions.act_window',
            'name': f'Ret. no SPE — {self.name}',
            'res_model': 've.wh.iva',
            'views': [(list_view.id, 'list'), (form_view.id, 'form')],
            'domain': [('invoice_id', 'in', inv_ids)],
        }

    def action_deshacer(self):
        """Revierte TODO lo que action_confirmar() creó para ESTA carga
        específica — pagos registrados, retenciones (con su asiento
        contable si estaban confirmadas) y las facturas mismas — y la
        deja de vuelta en Borrador para corregir y volver a confirmar.

        Construida 2026-07-31 (bug real: 1170 facturas de Cementos
        cargadas con fecha invertida por _parse_date, ver commit del
        formato de fecha) — no existía forma de deshacer una carga real
        sin usar 'Borrar Todo' de wizard_reset_piloto.py, que es
        específico de la compañía PILOTO de QA (busca por un RIF
        hardcodeado) y no sirve para un cliente real como este.

        A propósito NO borra los partners creados por la carga — no
        tienen ningún dato erróneo en sí mismos, solo el N° de factura/
        fecha/montos de la transacción. Mismo criterio de savepoint por
        registro que wizard_reset_piloto.py::_borrar_todo_piloto (un
        fallo puntual no debe tumbar el resto de la reversión)."""
        self.ensure_one()
        if self.estado not in ('confirmado', 'confirmado_discrepancias'):
            raise UserError('Solo se puede deshacer una carga ya confirmada.')

        Payment = self.env['account.payment'].sudo()
        WhIva = self.env['ve.wh.iva'].sudo()

        facturas = self.linea_ids.mapped('invoice_id').sudo()
        errores = []
        n_pagos = n_ret = n_fact = 0

        for inv in facturas:
            # 1. Pago(s) reconciliados con esta factura (EstadoPago=
            #    "Pagada") — hay que desligarlos/borrarlos ANTES de tocar
            #    la factura, si no el unreconcile bloquea el button_draft.
            try:
                pagos = inv._get_reconciled_payments()
            except AttributeError:
                pagos = Payment.browse()
                if inv.payment_state != 'not_paid':
                    errores.append(
                        f'Fila con factura {inv.name or inv.id}: no se pudo '
                        f'ubicar el pago reconciliado automáticamente (revisar '
                        f'a mano, payment_state={inv.payment_state})')
            for pago in pagos:
                try:
                    with self.env.cr.savepoint():
                        if pago.state == 'paid' and hasattr(pago, 'action_draft'):
                            pago.action_draft()
                        pago.unlink()
                    n_pagos += 1
                except Exception as exc:
                    errores.append(f'Pago de factura {inv.name or inv.id} no eliminado: {exc}')

            # 2. Retención(es) IVA de esta factura — si ya estaba
            #    Confirmada (comprobante real, ver nro_comp_retencion),
            #    tiene un asiento contable posteado que hay que cancelar
            #    primero.
            for wh in WhIva.search([('invoice_id', '=', inv.id)]):
                try:
                    with self.env.cr.savepoint():
                        if wh.asiento_id and wh.asiento_id.state == 'posted':
                            wh.asiento_id.button_cancel()
                        wh.unlink()
                    n_ret += 1
                except Exception as exc:
                    errores.append(f'Retención de factura {inv.name or inv.id} no eliminada: {exc}')

            # 3. La factura misma.
            try:
                with self.env.cr.savepoint():
                    if inv.state == 'posted':
                        inv.button_draft()
                    inv.unlink()
                n_fact += 1
            except Exception as exc:
                errores.append(f'Factura {inv.name or inv.id} no eliminada: {exc}')

        # 3b. Facturas HUÉRFANAS de esta carga — creadas por Move.create()
        #     pero que nunca llegaron a linea.invoice_id porque
        #     action_post() falló (ej. N° de Factura duplicado dentro del
        #     mismo archivo, ver es_dup_nro_en_archivo en _compute_partner_id
        #     más abajo). El paso 3 de arriba no las toca porque solo mira
        #     self.linea_ids.mapped('invoice_id'). Bug real 2026-08-06
        #     (Cementos, Q1-11-2025.xlsx): 143 facturas borrador quedaron
        #     huérfanas tras Deshacer Carga hasta que se limpiaron a mano
        #     por RPC. Se identifican por N° de Factura (coincide con algún
        #     nro_documento de ESTA carga) + borrador + mismo diario/
        #     compañía — nunca llegaron a postearse, así que borrarlas no
        #     afecta contabilidad real.
        # Diario por Zona (2026-08-14): las huérfanas de esta carga pueden
        # estar en CUALQUIER diario de venta (el compartido de siempre, o
        # el de la Zona correspondiente), no solo en uno fijo.
        journals_venta = self.env['account.journal'].sudo().search(
            [('type', '=', 'sale'), ('company_id', '=', self.company_id.id)])
        nros_carga = [n for n in self.linea_ids.mapped('nro_documento') if n]
        n_huerfanas = 0
        if journals_venta and nros_carga:
            huerfanas = self.env['account.move'].sudo().search([
                ('company_id', '=', self.company_id.id),
                ('journal_id', 'in', journals_venta.ids),
                ('move_type', '=', 'out_invoice'),
                ('state', '=', 'draft'),
                ('name', 'in', nros_carga),
                ('id', 'not in', facturas.ids),
            ])
            for inv in huerfanas:
                try:
                    with self.env.cr.savepoint():
                        inv.unlink()
                    n_huerfanas += 1
                except Exception as exc:
                    errores.append(f'Factura huérfana {inv.name or inv.id} no eliminada: {exc}')

        # 4. Períodos de conciliación que quedaron vacíos por esta
        #    reversión (creados solo para las fechas de esta carga) — no
        #    se tocan si otra carga/retención todavía los usa.
        #
        # Bug real encontrado 2026-08-05 (Cementos, incidente de producción):
        # esta consulta buscaba TODOS los períodos de la compañía, sin
        # limitarse a los que esta carga pudo haber creado, y solo
        # verificaba ve.wh.iva antes de borrar -- nunca revisaba si había
        # ve.seniat.retencion todavía enganchadas. En una compañía donde la
        # mayoría de los períodos vienen de la carga SENIAT (no del Libro de
        # Ventas), CUALQUIER "Deshacer Carga" borraba TODOS los períodos de
        # la compañía de un solo golpe -- por ondelete='set null' eso
        # desenganchó 18.650 retenciones SENIAT de sus períodos. Doble fix:
        # (a) limitar la búsqueda a los períodos que las líneas de ESTA
        # carga realmente tocaron, (b) exigir también 0 retenciones SENIAT
        # antes de borrar un período.
        Seniat = self.env['ve.seniat.retencion'].sudo()
        Periodo = self.env['ve.conciliacion.periodo'].sudo()
        periodos_retencion_carga = {
            self.env['ve.seniat.retencion']._periodo_from_fecha(f)[1]
            for f in self.linea_ids.mapped('fecha') if f
        }
        n_periodos = 0
        if periodos_retencion_carga:
            candidatos = Periodo.search([
                ('company_id', '=', self.company_id.id),
                ('periodo_retencion', 'in', list(periodos_retencion_carga)),
            ])
            for periodo in candidatos:
                if WhIva.search_count([('conciliacion_id', '=', periodo.id)]):
                    continue
                if Seniat.search_count([('conciliacion_id', '=', periodo.id)]):
                    continue
                try:
                    with self.env.cr.savepoint():
                        if periodo.declaracion_iva_id:
                            periodo.declaracion_iva_id.unlink()
                        periodo.unlink()
                    n_periodos += 1
                except Exception as exc:
                    errores.append(f'Período {periodo.periodo_retencion} no eliminado: {exc}')

        self.linea_ids.write({
            'invoice_id': False, 'partner_creado': False, 'agente_marcado': False,
        })
        self.write({
            'estado': 'borrador', 'confirmado_por_id': False, 'fecha_confirmacion': False,
        })

        cuerpo = (
            f'<b>Carga deshecha</b><br/>'
            f'<b>Facturas eliminadas:</b> {n_fact} de {len(facturas)}'
            + (f' (+{n_huerfanas} huérfanas de intentos fallidos anteriores)'
               if n_huerfanas else '') + '<br/>'
            f'<b>Pagos eliminados:</b> {n_pagos}<br/>'
            f'<b>Retenciones eliminadas:</b> {n_ret}<br/>'
            f'<b>Períodos vacíos eliminados:</b> {n_periodos}<br/>'
            f'<b>Clientes creados por esta carga:</b> se conservan (no tienen '
            f'dato erróneo propio) — revisar a mano si ya no hacen falta.<br/>'
        )
        if errores:
            cuerpo += '<br/><b>Filas con error:</b><br/>' + '<br/>'.join(errores)
        self.message_post(body=Markup(cuerpo), message_type='comment', subtype_xmlid='mail.mt_note')
        return self._reload()

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }


class VeConectaCargaVentasLinea(models.Model):
    _name = 've.conecta.carga.ventas.linea'
    _description = 'Fila de Carga de Libro de Ventas — SmartIVA Conecta'
    _order = 'fila'

    carga_id = fields.Many2one(
        've.conecta.carga.ventas', string='Carga', required=True, ondelete='cascade')
    estado_carga = fields.Selection(related='carga_id.estado', string='Estado Carga')
    fila = fields.Char(string='Fila')

    rif = fields.Char(string='RIF')
    nombre_cliente = fields.Char(string='Cliente')
    nro_control = fields.Char(string='N° Control')
    nro_documento = fields.Char(string='N° Factura')
    fecha = fields.Date(string='Fecha')
    base_16 = fields.Float(string='Base 16%', digits=(16, 2))
    base_8 = fields.Float(string='Base 8%', digits=(16, 2))
    base_exento = fields.Float(
        string='Exento/No Gravado', digits=(16, 2),
        help='Venta exenta o no gravada — sin IVA. Una factura 100% exenta '
             'no tiene Base 16%/8%, solo este monto.')
    monto_retenido = fields.Float(
        string='Monto Retenido', digits=(16, 2),
        help='Si el feed trae el monto realmente retenido, se usa para marcar '
             'al cliente como Agente de Retención automáticamente.')
    estado_pago = fields.Char(
        string='Estado Pago',
        help='Informativo, tal cual viene en el feed (ej. Pagada/Pendiente/Vencida) '
             '— no afecta la confirmación. Insumo para seguimiento proactivo '
             '(MEJORA-CONTACTO-01), todavía no conectado a ningún flujo.')
    nro_comp_retencion = fields.Char(
        string='N° Comprobante (Retención)',
        help='Si el feed lo trae, es el N° de Comprobante YA emitido por el '
             'agente de retención (comprador) — SmartIVA no lo genera. Al '
             'confirmar, la retención de esta fila se marca Recibida y '
             'Confirmada de una vez con este número, en vez de quedar '
             '"No Recibida" esperando el Buzón/OCR.')
    total_documento = fields.Float(
        string='Total c/IVA (Archivo)', digits=(16, 2),
        help='Opcional — "Total de Ventas Incluyendo IVA" tal cual viene en '
             'el archivo. No se usa para crear la factura (eso sale de '
             'base_16/base_8/base_exento) — solo sirve para la tabla '
             '"Consistencia por Zona" del resumen al confirmar, como '
             'comprobación independiente contra lo que Odoo calculó.')
    monto_iva = fields.Float(
        string='Monto IVA (Archivo)', digits=(16, 2),
        help='Opcional — "Impuesto Iva" tal cual viene en el archivo. '
             'Mismo criterio que total_documento: solo informativo, para '
             'la tabla "Consistencia por Zona".')
    zona = fields.Char(
        string='Zona/Planta',
        help='Opcional — solo si el cliente factura desde varios puntos '
             '(plantas/sucursales) bajo un único RIF y quiere seguimiento '
             'de comprobantes separado por punto.')
    es_spe = fields.Char(
        string='Contribuyente Especial (SPE)',
        help='Opcional — Sí/No explícito. Si no viene, se infiere igual que '
             'antes (si el cliente tiene monto retenido en cualquier fila '
             'del archivo). Confirma la condición de Agente de Retención '
             'sin depender del orden de las filas — ver action_confirmar.')
    validado_seniat = fields.Char(
        string='Validado SENIAT',
        help='Opcional — en blanco/Sí/No. Se sincroniza al campo homónimo '
             'del Cliente (pestaña Retenciones IVA VE) al confirmar la '
             'carga. Sí y No prevalecen sobre blanco, y Sí prevalece sobre '
             'No — ver action_confirmar.')
    tipo_transaccion = fields.Selection([
        ('01', '01 - Factura Regular'),
        ('02', '02 - Nota de Débito'),
        ('03', '03 - Nota de Crédito'),
    ], string='Tipo de Transacción (TR)',
        help='Código SENIAT tal cual lo trae el archivo (ver '
             '_normalizar_tipo_transaccion). Opcional e informativo por '
             'ahora — insumo para AJUSTE-FISCAL-01/02 (Nota de Crédito/'
             'Débito no ajustan la retención), todavía no conectado a esa '
             'lógica ni a la detección Registro+Anulación existente.')

    partner_id = fields.Many2one(
        'res.partner', string='Cliente (match)',
        compute='_compute_partner_id', store=True, readonly=False,
        help='Cliente encontrado por RIF. Editable — corrija aquí si el match '
             'automático no es el correcto antes de confirmar.')
    es_partner_nuevo = fields.Boolean(compute='_compute_partner_id', store=True)

    # Brechas BLOQUEANTES — toda la data de input es responsabilidad del
    # cliente (2026-07-24): no se completa/inventa nada faltante, ni se
    # permite duplicar una factura o retención ya existente. La fila queda
    # marcada, visible, y no se puede confirmar hasta corregirla.
    es_duplicado_factura = fields.Boolean(
        compute='_compute_partner_id', store=True,
        help='Ya existe una factura con el mismo N° Control en esta compañía '
             '(N° Control es único por RIF emisor).')
    es_duplicado_retencion = fields.Boolean(
        compute='_compute_partner_id', store=True,
        help='Ya existe una retención IVA (ve.wh.iva) con el mismo N° Control '
             'en esta compañía, sin pasar por una factura de este lote.')
    bloqueante = fields.Boolean(compute='_compute_partner_id', store=True)

    # Brecha informativa (no bloquea) o el motivo del bloqueo — un solo
    # texto para la columna de la vista previa.
    brecha = fields.Char(compute='_compute_partner_id', store=True, string='Brecha')
    categoria_discrepancia = fields.Selection([
        ('duplicada', 'Duplicada'),
        ('dato_faltante', 'Dato faltante'),
        ('fecha_invalida', 'Fecha inválida'),
        ('registro_anulacion', 'Registro + Anulación'),
        ('documento_vacio', 'Documento vacío'),
        ('error_posteo', 'Error al postear'),
    ], compute='_compute_partner_id', store=True, string='Categoría',
        help='Agrupa brecha/bloqueante en 6 categorías para la pestaña '
             'Discrepancias (2026-08-14) -- False si la fila no tiene ningún '
             'problema (incluye "Cliente nuevo"/"Se marcará Agente", que son '
             'informativas, no discrepancias). "error_posteo" es la única '
             'que no se puede derivar de los datos de la fila -- la escribe '
             'action_confirmar cuando Move.create()/action_post() falla con '
             'una excepción no prevista por ninguna otra categoría; este '
             'compute la PRESERVA (no la pisa) mientras la fila siga sin '
             'invoice_id, para que no desaparezca en el próximo recompute.')
    es_anulacion_par = fields.Boolean(
        compute='_compute_partner_id', store=True, string='Registro + Anulación',
        help='Mismo N° de Factura que otra fila (mismo RIF), con Base '
             'Imponible que neta a cero -- la misma operación registrada y '
             'luego anulada, no un duplicado real. Caso confirmado '
             '2026-08-12 en zona Invecem (247 pares, todos netean exacto). '
             'Se omite al confirmar, sin bloquear ni crear factura.')
    par_linea_id = fields.Many2one(
        've.conecta.carga.ventas.linea', string='Línea Emparejada (Anulación)',
        help='La otra línea con el mismo N° Documento+RIF y monto que neta '
             'a cero -- puede estar en esta misma carga (se omite AUTOMÁTICO,'
             ' ver es_anulacion_par, comportamiento original validado '
             '2026-08-12) o en OTRA carga (pedido explícito 2026-08-13: caso '
             'real donde el Registro quedó como línea "silenciosa" sin '
             'facturar en una carga ya "Confirmado", y la Anulación llegó '
             'después en una carga de recuperación aparte). Cuando es OTRA '
             'carga, NO se resuelve solo -- la línea de allá podría ya '
             'estar declarada, y no hay certeza automática de que sea la '
             'pareja correcta -- queda bloqueante con esta referencia para '
             'que el usuario decida.')

    # Soft-delete de duplicadas (2026-08-12, pedido explícito de la usuaria
    # tras preguntar "cómo se pueden ver o saber cuáles filas duplicadas
    # fueron eliminadas" -- antes action_eliminar_duplicados hacía unlink()
    # real, sin dejar ningún registro consultable de CUÁLES filas se
    # quitaron). El campo linea_ids del padre (ve.conecta.carga.ventas)
    # tiene domain=[('eliminada_duplicado','=',False)] -- cualquier código
    # que recorra self.linea_ids (confirmar, conteos, bloqueadas, etc.) ya
    # excluye estas filas automáticamente, sin tener que tocar cada sitio.
    eliminada_duplicado = fields.Boolean(
        default=False, copy=False, string='Eliminada (Duplicado)')
    fecha_eliminacion_duplicado = fields.Datetime(
        copy=False, string='Fecha Eliminación')
    motivo_eliminacion = fields.Char(
        compute='_compute_motivo_eliminacion', string='Motivo',
        help='Explica CONTRA QUÉ chocó esta fila (N° Control o N° Factura) y '
             'la Zona de la factura ya existente en Odoo -- calculado en vivo, '
             'no guardado, porque el `brecha` original se pierde una vez la '
             'carga queda Confirmada (ver _compute_partner_id). Pedido '
             'explícito 2026-08-12: la mayoría de las "duplicadas" resultaron '
             'ser choques entre zonas distintas con la misma numeración '
             '(ej. Ocumare/Pertigalete), no duplicados reales.')

    invoice_id = fields.Many2one('account.move', string='Factura Creada', readonly=True)

    # Escritos explícitamente en action_confirmar (no son compute) — sin
    # esto, después de confirmar no hay forma de distinguir "este cliente
    # ya existía" de "se acaba de crear ahora mismo", porque una búsqueda
    # posterior por RIF encuentra el partner de cualquier manera.
    partner_creado = fields.Boolean(default=False, readonly=True)
    agente_marcado = fields.Boolean(default=False, readonly=True)

    @api.depends('rif', 'nro_control', 'nro_documento', 'monto_retenido', 'fecha',
                 'carga_id.company_id', 'carga_id.estado', 'partner_creado', 'agente_marcado')
    def _compute_partner_id(self):
        Partner = self.env['res.partner']
        Move = self.env['account.move']
        WhIva = self.env['ve.wh.iva']
        for linea in self:
            categoria_previa = linea.categoria_discrepancia
            company = linea.carga_id.company_id
            partner = False
            if linea.rif:
                # Sin filtro de compañía, esta búsqueda podía enganchar un
                # contacto con el mismo RIF de OTRA compañía (ej. una PILOTO
                # archivada) — como el campo es store=True, quedaba pegado
                # ahí para siempre y usuarios sin acceso a esa compañía
                # veían "Error de acceso" al confirmar. Ver [[project_smartiva_conecta]].
                # Match NORMALIZADO (sin guión/espacios, mayúsculas) — un
                # match exacto contra `vat` no encontraba el cliente ya
                # existente si el formato del RIF en el archivo no coincidía
                # letra por letra con el guardado, y lo creaba duplicado.
                rif_norm = _norm_rif(linea.rif)
                candidatos = Partner.search([
                    ('vat', '!=', False),
                    '|', ('company_id', '=', False), ('company_id', '=', company.id),
                ])
                partner = candidatos.filtered(lambda p: _norm_rif(p.vat) == rif_norm)[:1]
            linea.partner_id = partner
            linea.es_partner_nuevo = bool(linea.rif) and not partner

            if linea.carga_id.estado in ('confirmado', 'confirmado_discrepancias') and not linea.invoice_id:
                # Fila que quedó pendiente sin facturar, en CUALQUIERA de
                # los dos estados terminales -- originalmente esto solo
                # cubría 'confirmado_discrepancias' (tras "Confirmar y
                # Omitir Bloqueadas", pedido 2026-08-12). Bug real
                # encontrado 2026-08-13: una carga 'confirmado' (100% "sin
                # discrepancias") TAMBIÉN puede tener líneas sin factura y
                # sin ningún aviso -- 1.708 líneas así en 49 cargas, sin
                # error ni bloqueo, porque la rama de abajo las congelaba
                # en bloqueante=False para siempre sin volver a evaluarlas
                # (típicamente el Registro de un par Registro+Anulación
                # cuya Anulación llegó después, en otra carga -- ver
                # par_linea_id). Se sigue evaluando en vivo, mismo camino
                # que una carga en Borrador, para que estos casos (y el
                # choque de Zona de la Causa C) puedan bloquear con motivo
                # real en vez de quedar invisibles.
                pass
            elif linea.carga_id.estado in ('confirmado', 'confirmado_discrepancias'):
                # Ya se resolvió — no repetir el chequeo de duplicados (la
                # factura de ESTA fila ahora existe y matchearía consigo
                # misma) ni la brecha "se creará" (ya se creó). Bug
                # reportado 2026-07-24: antes se quedaba con el texto viejo
                # para siempre.
                linea.es_duplicado_factura = False
                linea.es_duplicado_retencion = False
                linea.bloqueante = False
                if linea.es_anulacion_par:
                    # Bug real encontrado 2026-08-21: esta condición exigía
                    # "and not linea.invoice_id" -- leftover de una versión
                    # anterior donde la Anulación se omitía sin factura
                    # propia. Hoy (ver action_confirmar) AMBOS lados del par
                    # -- Registro Y Anulación -- se facturan de verdad (la
                    # Anulación como Nota de Crédito real, con su propio
                    # invoice_id), así que la condición nunca se cumplía y
                    # categoria_discrepancia caía siempre al 'False' de más
                    # abajo -- 335 de 438 pares en Vencement quedaron con la
                    # etiqueta perdida (par SÍ resuelto correctamente, solo
                    # la clasificación de la pestaña Discrepancias mentía).
                    linea.categoria_discrepancia = 'registro_anulacion'
                    continue
                linea.categoria_discrepancia = False
                partes = []
                if linea.partner_creado:
                    partes.append('Cliente creado')
                if linea.agente_marcado:
                    partes.append('Agente de Retención marcado')
                linea.brecha = ' · '.join(partes) if partes else False
                continue

            # Bug real encontrado 2026-08-06 (Cementos, Q1-11-2025.xlsx): 143
            # de 793 filas eran copias exactas de otra fila del mismo archivo
            # (mismo RIF/fecha/montos, N° Factura repetido) — este chequeo
            # solo miraba N° Control, así que ninguna se marcó "bloqueante"
            # en la vista previa y todas llegaron a action_confirmar, que
            # las rechazó una por una contra la restricción de Postgres
            # (nombre de factura único por diario) sin que se viera la razón
            # real hasta revisar "Filas con error" (y ahí se truncaba a 10).
            journal_venta = self.env['account.journal'].sudo().search(
                [('type', '=', 'sale'), ('company_id', '=', company.id)],
                limit=1) if company else False
            # Diario de Zona (2026-08-14, fix Causa C definitivo -- ver
            # _journal_zona): una factura nueva de esta fila iría a este
            # diario (el de su Zona, o el compartido si no trae Zona).
            journal_zona = (
                linea.carga_id._journal_zona(company, linea.zona, journal_venta)
                if company else False)
            # Criterio Zona (pedido explícito 2026-08-12, caso real Cementos
            # zonas Ocumare/Pertigalete: numeran su Libro de Ventas cada una
            # por su cuenta, sin N° Control, así que un mismo N° de Factura
            # se repite entre ambas por coincidencia -- antes se marcaba como
            # duplicado real). Solo se descarta el match cuando hay EVIDENCIA
            # POSITIVA de que son zonas distintas (ambas informadas y
            # diferentes) -- si a cualquiera de los dos lados le falta el
            # dato de Zona, se sigue bloqueando como antes (no se asume
            # coincidencia sin evidencia, mismo principio de auditoría de
            # inputs que el resto del módulo).
            zona_domain = (
                ['|', ('zona', '=', False), ('zona', '=', linea.zona)]
                if linea.zona else [])
            # Choque real de numeración (pedido explícito 2026-08-13, caso real
            # Cementos): Odoo exige N° Factura único por diario ENTRE FACTURAS
            # POSTEADAS (restricción SQL account_move_unique_name). Con diario
            # por Zona ya activo (2026-08-14), dos búsquedas distintas:
            # (a) en el diario PROPIO de esta Zona -- cualquier factura ahí
            # es un choque real sin importar su campo zona (ese diario ya es
            # exclusivo de esta Zona, no hace falta relajar por zona_domain);
            # (b) en el diario COMPARTIDO histórico -- ahí conviven todas las
            # Zonas de antes de este fix, así que SÍ hace falta el mismo
            # zona_domain que ya usa match_existente más abajo -- si no, un
            # choque contra una factura vieja de OTRA Zona en el compartido
            # bloquearía igual que antes del fix (bug real encontrado en la
            # prueba de CARGA-VTA/2026/047: Ocumare vs. Pertigalete, ambas
            # en el diario compartido, bloqueaba aunque ya había diario de
            # Zona disponible). Se bloquea ACÁ, antes de intentar crear nada.
            match_existente_posteada = Move.browse()
            if linea.nro_documento and journal_zona and journal_zona != journal_venta:
                match_existente_posteada = Move.search([
                    ('company_id', '=', company.id),
                    ('move_type', '=', 'out_invoice'),
                    ('journal_id', '=', journal_zona.id),
                    ('name', '=', linea.nro_documento),
                    ('state', '=', 'posted'),
                ], limit=1)
            if not match_existente_posteada and linea.nro_documento and journal_venta:
                match_existente_posteada = Move.search([
                    ('company_id', '=', company.id),
                    ('move_type', '=', 'out_invoice'),
                    ('journal_id', '=', journal_venta.id),
                    ('name', '=', linea.nro_documento),
                    ('state', '=', 'posted'),
                ] + zona_domain, limit=1)
            journal_ids_chequeo = list({
                j.id for j in (journal_zona, journal_venta) if j})
            match_existente = match_existente_posteada or (
                linea.nro_documento and journal_ids_chequeo and Move.search([
                    ('company_id', '=', company.id),
                    ('move_type', '=', 'out_invoice'),
                    ('journal_id', 'in', journal_ids_chequeo),
                    ('name', '=', linea.nro_documento),
                ] + zona_domain, limit=1)) or Move.browse()
            match_en_archivo = (
                linea.nro_documento and self.search([
                    ('carga_id', '=', linea.carga_id.id),
                    ('nro_documento', '=', linea.nro_documento),
                    ('fila', '<', linea.fila),
                ] + zona_domain, limit=1)) or self.browse()
            # N° Control repetido DENTRO del mismo archivo -- pedido
            # explícito 2026-08-12: existía el equivalente para N° Factura
            # (match_en_archivo arriba) pero no para N° Control, así que dos
            # filas del mismo archivo con igual N° Control y N° Factura
            # distinto (o en blanco) no se detectaban entre sí.
            match_ctrl_en_archivo = (
                linea.nro_control and self.search([
                    ('carga_id', '=', linea.carga_id.id),
                    ('nro_control', '=', linea.nro_control),
                    ('fila', '<', linea.fila),
                ] + zona_domain, limit=1)) or self.browse()
            es_dup_nro_existente = bool(match_existente)
            es_dup_nro_en_archivo = bool(match_en_archivo)
            es_dup_ctrl_en_archivo = bool(match_ctrl_en_archivo)
            linea.es_duplicado_factura = bool(
                (linea.nro_control and company and Move.search([
                    ('company_id', '=', company.id),
                    ('move_type', '=', 'out_invoice'),
                    ('nro_control', '=', linea.nro_control),
                ] + zona_domain, limit=1))
                or es_dup_nro_existente or es_dup_nro_en_archivo or es_dup_ctrl_en_archivo)
            # N° Factura contra retenciones existentes -- pedido explícito
            # 2026-08-12: ve.wh.iva también tiene nro_documento (no solo
            # nro_control), así que el chequeo de "Retención duplicada"
            # antes se quedaba corto si el N° Control venía distinto/en
            # blanco pero el N° Factura sí coincidía con una retención ya
            # creada por otra vía (ej. Buzón de Comprobantes/OCR).
            linea.es_duplicado_retencion = bool(
                company and not linea.es_duplicado_factura and (
                    (linea.nro_control and WhIva.search([
                        ('company_id', '=', company.id),
                        ('nro_control', '=', linea.nro_control),
                        ('state', '!=', 'anulado'),
                    ] + zona_domain, limit=1))
                    or (linea.nro_documento and WhIva.search([
                        ('company_id', '=', company.id),
                        ('nro_documento', '=', linea.nro_documento),
                        ('state', '!=', 'anulado'),
                    ] + zona_domain, limit=1))))

            # Par Registro + Anulación (mismo N° Factura, monto neto CERO) --
            # pedido explícito 2026-08-12, caso real confirmado en zona
            # Invecem: 247 pares, TODOS netean exacto a cero (mismo RIF,
            # Base/IVA con signo opuesto). No es un duplicado real -- es la
            # misma operación anulada, se omite la fila SIN bloquear (a
            # diferencia de "Factura duplicada", que sí bloquea) — ver el
            # skip informativo en action_confirmar. Búsqueda SIMÉTRICA
            # (cualquier otra fila, no solo 'fila < linea.fila' como
            # match_en_archivo/es_dup_nro_en_archivo de arriba) -- si no,
            # solo la SEGUNDA fila del par se detectaba y la primera
            # (el "Registro") se colaba como factura real.
            linea.es_anulacion_par = False
            linea.par_linea_id = False
            base_propia = (linea.base_16 or 0) + (linea.base_8 or 0) + (linea.base_exento or 0)
            if base_propia and linea.nro_documento and linea.rif:
                # Busca en TODA la empresa, no solo en esta carga (pedido
                # explícito 2026-08-13: caso real confirmado -- el Registro
                # quedó como línea "silenciosa" sin facturar en una carga ya
                # Confirmado, y la Anulación llegó meses después en una carga
                # de recuperación aparte. Antes de este cambio, limitar la
                # búsqueda a `carga_id` propio hacía que nunca se cruzaran, y
                # cada una intentaba postear una factura real por separado --
                # una de las dos (la negativa) siempre fallaba al postear
                # (Odoo no permite factura con total negativo).
                par_en_archivo = self.search([
                    ('carga_id.company_id', '=', company.id if company else 0),
                    ('nro_documento', '=', linea.nro_documento),
                    ('rif', '=', linea.rif),
                    ('id', '!=', linea.id if isinstance(linea.id, int) else 0),
                ] + zona_domain, limit=1)
                if par_en_archivo:
                    base_otra = ((par_en_archivo.base_16 or 0) + (par_en_archivo.base_8 or 0)
                                 + (par_en_archivo.base_exento or 0))
                    if abs(base_propia + base_otra) < 0.02:
                        linea.par_linea_id = par_en_archivo.id
                        # Se resuelve AUTOMÁTICO (Nota de Crédito real, ver
                        # action_confirmar) cuando la pareja está en el
                        # MISMO archivo (validado 2026-08-12, 247 pares
                        # reales) O cuando es la Anulación (base negativa)
                        # de una factura Registro que YA EXISTE posteada en
                        # OTRA carga -- pedido explícito 2026-08-13: la NC
                        # se crea con su propia fecha (la de ESTA fila, su
                        # propia quincena), vinculada por reversed_entry_id
                        # -- no toca ni reabre la carga vieja, así que no
                        # hay riesgo de pisar lo ya declarado a SENIAT en
                        # su período original. Si el Registro cruzado
                        # TODAVÍA no tiene factura (nadie lo procesó
                        # todavía), no hay nada seguro que vincular -- se
                        # deja bloqueada para revisión (rama de abajo).
                        if par_en_archivo.carga_id == linea.carga_id:
                            linea.es_anulacion_par = True
                        elif base_propia < 0 and par_en_archivo.invoice_id:
                            linea.es_anulacion_par = True
                elif match_existente:
                    # La factura ya existente en Odoo (de otra carga/parte
                    # del mismo archivo) siempre nace con base POSITIVA
                    # (Move.create() más abajo usa los montos tal cual del
                    # archivo -- si esta fila trae el signo opuesto, es la
                    # anulación de esa misma factura).
                    base_otra = match_existente.amount_untaxed
                    if abs(base_propia + base_otra) < 0.02:
                        linea.es_anulacion_par = True

            if linea.es_anulacion_par:
                # No cuenta como "Factura duplicada" (no infla count_duplicados
                # ni aparece en "Eliminar Filas Duplicadas") -- se resuelve
                # solo y sin bloquear en action_confirmar: el Registro se
                # factura normal y la Anulación se crea como Nota de
                # Crédito real (ver v19.0.2.14.74/.75), sea el par del
                # mismo archivo o de otra carga con factura ya existente.
                linea.es_duplicado_factura = False
                linea.bloqueante = False
                linea.categoria_discrepancia = 'registro_anulacion'
                linea.brecha = ('Registro + Anulación (neto cero) — se factura/crea '
                                 'Nota de Crédito real al confirmar, no bloquea')
            elif linea.par_linea_id:
                # Anulación cruzando cargas cuyo Registro TODAVÍA no tiene
                # factura (nadie procesó esa carga todavía) -- no hay nada
                # seguro que vincular con reversed_entry_id. Bloquea con la
                # referencia exacta (carga, fila, RIF, monto) para que el
                # usuario la revise y decida -- no se inventa la decisión.
                linea.bloqueante = True
                linea.categoria_discrepancia = 'registro_anulacion'
                otra = linea.par_linea_id
                linea.brecha = (
                    f'Posible Registro + Anulación con OTRA carga (mismo N° Documento '
                    f'+ RIF, neto cero) — {otra.carga_id.name}, fila {otra.fila}, '
                    f'Bs. {(otra.base_16 or 0) + (otra.base_8 or 0):,.2f} — requiere '
                    f'revisión y decisión manual, no se resuelve automáticamente')
            elif not linea.rif:
                linea.bloqueante = True
                linea.categoria_discrepancia = 'dato_faltante'
                linea.brecha = 'Sin RIF'
            elif not linea.nro_documento and not linea.nro_control:
                # Antes bloqueaba solo por falta de N° Factura, sin mirar
                # N° Control -- pedido explícito 2026-08-12: se aceptan
                # como intercambiables para identificar la fila (si falta
                # uno se usa el otro, ver Move.create() más abajo), solo
                # bloquea si AMBOS están en blanco.
                linea.bloqueante = True
                linea.categoria_discrepancia = 'dato_faltante'
                linea.brecha = 'Sin N° Factura ni N° Control'
            elif not linea.fecha:
                # Bug real encontrado 2026-07-30: si el valor de Fecha del
                # archivo no coincide con ninguno de los formatos que
                # reconoce _parse_date (datetime real, serial de Excel,
                # DD/MM/YYYY o YYYY-MM-DD), quedaba en False — y
                # action_confirmar() lo pisaba en silencio con "hoy"
                # (fields.Date.today()), sin ningún aviso. Con un Libro de
                # Ventas de 2025, esto hacía que TODAS las facturas
                # nacieran con fecha de HOY: la retención se vinculaba al
                # período de hoy (no al de 2025) y la Fecha Límite salía
                # calculada desde hoy también. Ahora bloquea en vez de
                # adivinar — mismo principio de auditoría de inputs del
                # módulo (nada se transforma sin que el usuario lo vea).
                linea.bloqueante = True
                linea.categoria_discrepancia = 'fecha_invalida'
                linea.brecha = 'Fecha no reconocida — revise el formato de la columna Fecha'
            elif linea.es_duplicado_factura:
                linea.bloqueante = True
                linea.categoria_discrepancia = 'duplicada'
                if match_existente_posteada:
                    linea.brecha = (
                        f'Factura duplicada — N° de Factura ya usado por la factura '
                        f'POSTEADA {match_existente_posteada.name} (diario '
                        f'{match_existente_posteada.journal_id.name}) — con diario por '
                        f'Zona ya activo, este choque es un duplicado real dentro de la '
                        f'misma Zona (o contra una factura histórica del diario '
                        f'compartido), requiere renumeración o revisión manual')
                elif es_dup_nro_en_archivo:
                    linea.brecha = ('Factura duplicada — mismo N° de Factura que otra '
                                     'fila de este mismo archivo')
                elif es_dup_ctrl_en_archivo:
                    linea.brecha = ('Factura duplicada — mismo N° de Control que otra '
                                     'fila de este mismo archivo')
                elif es_dup_nro_existente:
                    linea.brecha = ('Factura duplicada — ya existe una factura con ese '
                                     'N° de Factura en Odoo')
                else:
                    linea.brecha = 'Factura duplicada — ya existe (mismo N° Control)'
            elif linea.es_duplicado_retencion:
                linea.bloqueante = True
                linea.categoria_discrepancia = 'duplicada'
                linea.brecha = 'Retención duplicada — ya existe (mismo N° Control o N° Factura)'
            elif not partner:
                linea.bloqueante = False
                linea.categoria_discrepancia = False
                linea.brecha = 'Cliente nuevo — se creará'
            elif (linea.monto_retenido > 0 and not partner.es_agente_retencion
                  and _es_spe_verdadero(linea.es_spe) is not False):
                linea.bloqueante = False
                linea.categoria_discrepancia = False
                linea.brecha = 'Se marcará como Agente de Retención'
            elif (not linea.base_16 and not linea.base_8 and not linea.base_exento
                  and not linea.invoice_id):
                # Documento vacío -- las 3 bases en 0, no hay nada que
                # facturar (ver el mismo chequeo en action_confirmar, que
                # por eso omite la fila sin crear factura). Antes esto solo
                # se veía en el resumen del momento de confirmar (329 filas
                # así encontradas 2026-08-14, sin ningún rastro después);
                # ahora queda categorizado y visible en la pestaña
                # Discrepancias sin depender de haber estado presente
                # cuando se confirmó.
                linea.bloqueante = False
                linea.categoria_discrepancia = 'documento_vacio'
                linea.brecha = 'Base Imponible/Total en 0 — sin monto que facturar'
            elif categoria_previa == 'error_posteo' and not linea.invoice_id:
                # Preservar: esta fila falló al postear en un intento
                # anterior de action_confirmar (excepción no prevista por
                # ninguna categoría de arriba) -- ese motivo solo se puede
                # conocer intentando crear la factura, no se puede
                # re-derivar acá. Sin este caso, el próximo recompute (por
                # cualquier motivo, ej. otra carga cambiando un dato
                # relacionado) la dejaría caer al else de abajo y el
                # problema volvería a quedar invisible.
                linea.bloqueante = False
                linea.categoria_discrepancia = 'error_posteo'
            else:
                linea.bloqueante = False
                linea.categoria_discrepancia = False
                linea.brecha = False

    @api.model
    def action_relink_facturas_huerfanas(self, pares):
        """Recuperación puntual del bug de re-Previsualizar sobre una
        carga 'confirmado_discrepancias' (ver commit bba914f): esa carga
        podía volver a previsualizarse por RPC directo (sin botón visible
        en la UI), lo que borraba TODAS las líneas -- incluidas las que ya
        tenían invoice_id -- y las recreaba desde cero. La factura real ya
        creada quedaba huérfana (sin línea que le apunte) y la línea nueva
        la veía como "ya existe" y quedaba bloqueada como falso duplicado.

        `pares`: lista de [linea_id, invoice_id] ya verificados 1:1 (mismo
        RIF + N° Factura + Zona, factura sin ninguna línea que le apunte).
        Solo actúa sobre líneas que hoy están bloqueadas por
        es_duplicado_factura=True y sin invoice_id propio -- no pisa nada
        que ya esté resuelto ni algo distinto a este caso puntual. No crea
        ninguna factura nueva, solo vincula lo que ya existe."""
        lineas = self.browse([p[0] for p in pares])
        lineas_por_id = {l.id: l for l in lineas}
        tocadas = self.browse()
        for linea_id, invoice_id in pares:
            linea = lineas_por_id.get(linea_id)
            if not linea or not linea.es_duplicado_factura or linea.invoice_id:
                continue
            linea.invoice_id = invoice_id
            tocadas |= linea
        tocadas._compute_partner_id()
        return len(tocadas)

    @api.model
    def action_recompute_publico(self, linea_ids):
        """Fuerza el recompute de _compute_partner_id sobre líneas
        puntuales -- método público porque Odoo bloquea invocar métodos
        privados (con guion bajo) directo por RPC externo. Uso puntual de
        mantenimiento/diagnóstico (pedido explícito 2026-08-13, líneas
        "silenciosas" sin factura en cargas ya 'confirmado' que quedaron
        con el valor viejo de bloqueante/es_anulacion_par/par_linea_id
        guardado antes de este fix -- leer con read() no las recalcula,
        Odoo solo recomputa cuando cambia algo de lo que depende)."""
        lineas = self.browse(linea_ids)
        lineas._compute_partner_id()
        return len(lineas)

    def _compute_motivo_eliminacion(self):
        Move = self.env['account.move'].sudo()
        for linea in self:
            if linea.es_anulacion_par:
                linea.motivo_eliminacion = (
                    'Registro + Anulación — mismo N° de Factura, monto neta a cero '
                    '(no es duplicado real, fila omitida sin bloquear)')
                continue
            if not (linea.eliminada_duplicado or linea.es_duplicado_factura):
                linea.motivo_eliminacion = False
                continue
            company = linea.carga_id.company_id
            existente = False
            clave = False
            if linea.nro_control:
                existente = Move.search([
                    ('company_id', '=', company.id), ('nro_control', '=', linea.nro_control),
                ], limit=1)
                clave = 'N° Control'
            if not existente and linea.nro_documento:
                existente = Move.search([
                    ('company_id', '=', company.id), ('name', '=', linea.nro_documento),
                ], limit=1)
                clave = 'N° Factura'
            if not existente:
                linea.motivo_eliminacion = (
                    'N° de Factura repetido dentro del mismo archivo (sin match en Odoo)')
                continue
            if existente.zona and linea.zona and existente.zona != linea.zona:
                linea.motivo_eliminacion = (
                    f'{clave} duplicado con {existente.name} — Zona distinta '
                    f'({linea.zona or "—"} ↔ {existente.zona}), probable falso positivo')
            else:
                linea.motivo_eliminacion = f'{clave} duplicado con {existente.name} — misma Zona'

    def action_eliminar_seleccionadas(self):
        """Elimina (soft-delete) las filas seleccionadas en la lista
        "Duplicadas a Revisar" -- pedido explícito 2026-08-12: antes
        "Eliminar Filas Duplicadas" actuaba sobre TODAS de una carga sin
        poder revisar/elegir primero. Este botón vive en el header de esa
        lista (visible solo con filas seleccionadas, ver [[feedback_
        odoo_list_header_seleccion]]) y opera sobre self = la selección
        real del usuario, no sobre toda la carga."""
        if not self:
            return
        if any(l.carga_id.estado == 'confirmado' for l in self):
            raise UserError('No se puede eliminar duplicadas de una carga ya confirmada.')
        self.write({
            'eliminada_duplicado': True,
            'fecha_eliminacion_duplicado': fields.Datetime.now(),
        })
        for carga in self.mapped('carga_id'):
            n = len(self.filtered(lambda l: l.carga_id == carga))
            carga.message_post(
                body=Markup(
                    f'<b>Filas duplicadas eliminadas:</b> {n} (revisadas y '
                    f'seleccionadas manualmente) — ver detalle con el botón '
                    f'"Ver Duplicadas Eliminadas".'),
                message_type='comment', subtype_xmlid='mail.mt_note')

    def action_eliminar_discrepancia(self):
        """Elimina (soft-delete) filas seleccionadas desde la pestaña
        Discrepancias -- generaliza action_eliminar_seleccionadas (antes
        solo para categoria_discrepancia='duplicada') a cualquier
        categoría. Reusa eliminada_duplicado/fecha_eliminacion_duplicado en
        vez de un campo nuevo por categoría, para no duplicar el domain que
        ya excluye estas filas de linea_ids en toda la carga (ver comentario
        en eliminada_duplicado más arriba). El motivo que queda en el
        chatter es el propio `brecha` de cada fila -- ya es específico por
        categoría, no hace falta pedirlo aparte en un wizard."""
        if not self:
            return
        if any(l.invoice_id for l in self):
            raise UserError(
                'Una o más filas seleccionadas ya tienen factura creada -- '
                'no son discrepancias, no se pueden eliminar desde acá.')
        if any(not l.brecha for l in self):
            raise UserError(
                'Una o más filas seleccionadas no tienen ninguna discrepancia '
                'registrada que justifique eliminarlas.')
        self.write({
            'eliminada_duplicado': True,
            'fecha_eliminacion_duplicado': fields.Datetime.now(),
        })
        for carga in self.mapped('carga_id'):
            lineas_carga = self.filtered(lambda l: l.carga_id == carga)
            detalle = '<br/>'.join(
                f'Fila {l.fila} ({l.categoria_discrepancia}): {l.brecha}'
                for l in lineas_carga[:20])
            extra = ('' if len(lineas_carga) <= 20
                     else f'<br/>… y {len(lineas_carga) - 20} más.')
            carga.message_post(
                body=Markup(
                    f'<b>{len(lineas_carga)} discrepancia(s) eliminada(s) desde '
                    f'la pestaña Discrepancias:</b><br/>{detalle}{extra}'),
                message_type='comment', subtype_xmlid='mail.mt_note')
