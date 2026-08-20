import base64
import csv
import io
import logging
import re
import unicodedata

from markupsafe import Markup
from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# ── Helpers de parseo (idénticos a ve_conecta_carga_ventas.py — no se
#    comparten en un módulo aparte para no romper el patrón ya establecido
#    de duplicar estos helpers pequeños por archivo, ver wizard_carga_
#    seniat.py) ──────────────────────────────────────────────────────────

def _norm_rif(rif):
    """Mismo criterio que ve_conciliacion.py::_norm_rif — sin esto, un RIF
    con guión en el archivo cargado nunca matchea contra un contacto ya
    guardado sin guión (o viceversa), y el proveedor existente se trata
    como nuevo (duplicado). Bug real 2026-07-28."""
    return (rif or '').upper().replace('-', '').replace(' ', '').strip()


def _formatear_rif(rif):
    """Mismo criterio que ve_conecta_carga_ventas.py::_formatear_rif (ver
    ese archivo para el detalle) -- RIF sin guión (letra + 9 dígitos) se
    formatea a LETRA-12345678-9. Pedido explícito 2026-08-05."""
    limpio = (rif or '').upper().strip()
    if not limpio or '-' in limpio:
        return limpio
    m = re.match(r'^([VEJPG])(\d{9})$', limpio)
    if not m:
        return limpio
    letra, digitos = m.groups()
    return f'{letra}-{digitos[:8]}-{digitos[8]}'


def _norm_header(text):
    text = text.lower().strip()
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    # Encabezados reales tipo "Base 16,00 %" (formato SENIAT, decimal
    # venezolano) — sin esto, la coma se convertía en espacio más abajo y
    # dejaba "base 16 00" (con el "00" suelto como si fuera otra palabra)
    # en vez de "base 16", que es la clave real del mapa de sinónimos.
    # Bug real 2026-08-02 (Cementos, Libro de Compras): esas columnas
    # quedaban sin reconocer, ninguna factura tenía línea que crear con
    # monto y 0 facturas se creaban en toda la carga.
    text = re.sub(r'(\d+),(\d{2})\b', r'\1', text)
    # Colapsar acrónimos con puntos ("i.v.a." -> "iva", "r.i.f." -> "rif")
    # ANTES de convertir el resto de la puntuación en espacio — si no,
    # quedan letras sueltas ("i v a") que nunca matchean contra las claves
    # compactas del mapa. Bug real encontrado 2026-07-29: el formato real
    # de Profit Plus (ver ejemplos/Libro de compras.jpeg) usa "I.V.A." con
    # puntos justo en el header crítico "I.V.A. Retenido al Vendedor" —
    # quedaba sin reconocer, el monto retenido de la fila se perdía en
    # silencio (0), sin ningún error visible para el usuario.
    text = re.sub(r'\b(?:[a-z]\.){2,}', lambda m: m.group(0).replace('.', ''), text)
    # 'º' (ordinal masculino, U+00BA) es visualmente parecido a '°' (grado,
    # U+00B0) pero es OTRO carácter — mismo fix que ve_conecta_carga_ventas.py.
    # La coma se incluye acá también, como red de seguridad, por si queda
    # alguna suelta que el regex decimal de arriba no cubrió.
    text = re.sub(r'[°º#.,%\-_/\\]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _parse_amount(val):
    """Bug real encontrado 2026-08-12 (Cementos, ver mismo fix en
    ve_conecta_carga_ventas.py): la regex vieja `[^\\d.,]` borraba el
    signo "-" junto con el resto de caracteres no numéricos -- un monto
    negativo del archivo perdía el signo silenciosamente."""
    if val is None:
        return 0.0
    s = str(val).strip()
    negativo = s.startswith('-') or (s.startswith('(') and s.endswith(')'))
    s = re.sub(r'[^\d.,]', '', s)
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        v = float(s)
        return -v if negativo else v
    except ValueError:
        return 0.0


def _parse_date(val, formato='dmy'):
    """formato: 'dmy' (usar tal cual, default) o 'mdy' (invertir día/mes) —
    mismo fix que ve_conecta_carga_ventas.py (ver ese archivo para el
    detalle completo). Ampliado 2026-07-31: aplica tanto a fechas-texto
    ambiguas como a fechas REALES de Excel con día/mes invertidos en el
    propio dato de origen (no solo un problema de texto)."""
    if val is None:
        return False
    if hasattr(val, 'strftime'):
        if formato == 'mdy':
            try:
                from datetime import date as _date
                return _date(val.year, val.day, val.month).strftime('%Y-%m-%d')
            except ValueError:
                return False
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)) and 1000 < val < 100000:
        try:
            from datetime import date, timedelta
            d = date(1899, 12, 30) + timedelta(days=int(val))
            if formato == 'mdy':
                d = date(d.year, d.day, d.month)
            return d.strftime('%Y-%m-%d')
        except Exception:
            pass
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        a, b, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
        dia, mes = (b, a) if formato == 'mdy' else (a, b)
        if not (1 <= mes <= 12 and 1 <= dia <= 31):
            return False
        return f"{anio}-{mes:02d}-{dia:02d}"
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return False


def _fecha_candidatos(val):
    """Ver ve_conecta_carga_ventas.py::_fecha_candidatos — mismo helper,
    usado por el modo 'auto' de formato_fecha."""
    if val is None:
        return []
    if hasattr(val, 'strftime'):
        anio, mes, dia = val.year, val.month, val.day
    elif isinstance(val, (int, float)) and 1000 < val < 100000:
        from datetime import date, timedelta
        d = date(1899, 12, 30) + timedelta(days=int(val))
        anio, mes, dia = d.year, d.month, d.day
    else:
        s = str(val).strip()
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
        if not m:
            return []
        dia, mes, anio = int(m.group(1)), int(m.group(2)), int(m.group(3))
    from datetime import date as _date
    candidatos = []
    try:
        _date(anio, mes, dia)
        candidatos.append((anio, mes, dia))
    except ValueError:
        pass
    if mes != dia:
        try:
            _date(anio, dia, mes)
            candidatos.append((anio, dia, mes))
        except ValueError:
            pass
    return candidatos


# Mapeo de encabezados — formato estándar SENIAT del Libro de Compras, mismas
# columnas que ya genera wizard_reset_piloto.py (_HEADERS_COMPRAS) para su
# data sintética. A diferencia de Ventas, no hay un mapeo configurable por
# cliente porque el formato de Compras ya es estándar (CHECKLIST_ARRANQUE_
# PILOTO.md sección 2: "Compras no requiere esto, sigue el formato SENIAT").
_HEADER_MAP = {
    'rif': 'rif', 'rif proveedor': 'rif', 'nro rif': 'rif', 'n rif': 'rif',
    'nombre': 'nombre_proveedor', 'proveedor': 'nombre_proveedor',
    'razon social': 'nombre_proveedor', 'nombre proveedor': 'nombre_proveedor',
    'nombre razon social': 'nombre_proveedor',
    # "Nombre o Razón Social" (con "o") es el header real del formato SENIAT
    # (ver ejemplos/PXL_20260728_145144441.jpg) — sin este sinónimo exacto
    # nunca matcheaba contra "nombre razon social" (sin "o") y el proveedor
    # quedaba sin nombre (se creaba con el RIF como nombre, ver action_confirmar).
    'nombre o razon social': 'nombre_proveedor',
    'n control': 'nro_control', 'nro control': 'nro_control',
    'control': 'nro_control', 'numero control': 'nro_control',
    'n de control': 'nro_control', 'numero de control': 'nro_control',
    'nro de control': 'nro_control',
    # En Compras NO existe un "comprobante del proveedor" — el proveedor
    # solo emite la FACTURA (nro_documento). El único "comprobante" del
    # flujo es el de RETENCIÓN, y ese lo emite NUESTRO cliente (el
    # comprador), no el proveedor — ver nro_comp_retencion más abajo.
    # Corregido 2026-07-28: antes "Número de Comprobante"/"N Comprobante"
    # mapeaban acá por error (mismo campo que N° de Factura), pero en el
    # formato SENIAT real esa columna ES el N° de comprobante de retención
    # (ver ejemplos/PXL_20260728_145144441.jpg, export real Profit Plus).
    'n factura': 'nro_documento', 'nro factura': 'nro_documento',
    'factura': 'nro_documento', 'n de factura': 'nro_documento',
    'numero factura': 'nro_documento', 'numero de factura': 'nro_documento',
    'nro de factura': 'nro_documento',
    'fecha': 'fecha', 'fecha factura': 'fecha', 'fecha emision': 'fecha',
    'fecha documento': 'fecha', 'fecha de la factura': 'fecha',
    # Fecha en que NUESTRO cliente (el comprador/agente de retención)
    # realmente aplicó/pagó la retención — NO es la fecha de la factura
    # del proveedor. Determina la quincena SENIAT real de la retención
    # (bug real 2026-08-02, Cementos: sin reconocer esta columna, el
    # período se calculaba con "Fecha Documento" y quedaba disperso en 7
    # quincenas distintas en vez de caer todo en la quincena real de
    # aplicación).
    'fecha aplic retencion': 'fecha_aplic_retencion',
    'fecha aplicacion retencion': 'fecha_aplic_retencion',
    'fecha de aplicacion de la retencion': 'fecha_aplic_retencion',
    'base 16': 'base_16', 'base imponible 16': 'base_16',
    'base general': 'base_16', 'base gravada 16': 'base_16',
    'base 8': 'base_8', 'base imponible 8': 'base_8',
    'base reducida': 'base_8', 'base gravada 8': 'base_8',
    'iva retenido al vendedor': 'monto_retenido', 'monto retenido': 'monto_retenido',
    'iva retenido': 'monto_retenido', 'retencion': 'monto_retenido',
    'iva ret vendedor': 'monto_retenido',
    # Total tal cual viene en el archivo — informativo, NO participa en la
    # creación de la factura (esa sigue saliendo de base_16/base_8/
    # base_sin_credito). Solo para comparar contra amount_total de Odoo en
    # la tabla de Consistencia (pedido explícito 2026-08-02) — sin esto la
    # fila "Facturas creadas" no tenía ninguna cifra propia para verificar,
    # solo la cantidad. No se usa "Total Compras con IVA e IGTF" porque
    # este módulo no genera línea de IGTF en la factura — compararía
    # montos que nunca van a cuadrar por diseño.
    'total importe con iva': 'total_documento',
    'total documento': 'total_documento', 'total factura': 'total_documento',
    # N° de Comprobante (de Retención) — lo emite nuestro cliente, no el
    # proveedor. En el export real de Profit Plus esta columna se imprime
    # como "Número de Comprobante" a secas (no dice "Retención"), en la
    # fila de la retención — por eso esos sinónimos van acá y no a nro_documento.
    'n comp retencion': 'nro_comp_retencion', 'n comprobante retencion': 'nro_comp_retencion',
    'nro comp retencion': 'nro_comp_retencion', 'numero comprobante retencion': 'nro_comp_retencion',
    'n comp ret': 'nro_comp_retencion', 'numero de comprobante': 'nro_comp_retencion',
    'n de comprobante': 'nro_comp_retencion',
    'n comprobante': 'nro_comp_retencion', 'nro comprobante': 'nro_comp_retencion',
    'n comprobante proveedor': 'nro_comp_retencion', 'n comp proveedor': 'nro_comp_retencion',
    'compras sin credito iva': 'base_sin_credito', 'sin credito iva': 'base_sin_credito',
    'compras exentas y o sin derecho a credito fiscal': 'base_sin_credito',
    # Formato SENIAT "ancho" real (Cementos, 2026-08-02): 3 columnas
    # separadas para las distintas categorías sin IVA — legalmente
    # distintas (Exenta/Exonerada/No Sujeta) pero contablemente se tratan
    # igual acá (línea de factura sin tax_ids), mismo bucket que ya usa
    # "Compras Sin Derecho a Crédito IVA".
    'comp inter exentas o exoneradas': 'base_sin_credito',
    'compras exoneradas': 'base_sin_credito', 'compras no sujetas': 'base_sin_credito',
    # Formato "largo" (ej. export real de Profit Plus, ver ejemplos/Libro
    # de compras.jpeg): una sola columna "Base Imponible" + "% Alíc." por
    # fila, en vez de columnas separadas Base 16%/Base 8% del formato
    # SENIAT "ancho". Se reparte a base_16/base_8 en _leer_filas según el
    # valor de alicuota_pct de esa misma fila.
    'base imponible': 'base_generica',
    'alic': 'alicuota_pct', 'alicuota': 'alicuota_pct', 'alic iva': 'alicuota_pct',
    # N° de Factura Afectada — solo aparece en la fila del comprobante de
    # retención (formato de 2 filas por transacción, ver ejemplos/
    # PXL_20260728_145144441.jpg), apunta de vuelta al N° de Factura de la
    # fila que le dio origen. Campo real del modelo Linea — la vista previa
    # muestra las 2 filas TAL CUAL vienen en el archivo (pedido explícito
    # 2026-07-28: nada se funde en silencio antes de que el usuario lo vea).
    # La vinculación ocurre recién en action_confirmar().
    'n de factura afectada': 'nro_factura_afectada', 'factura afectada': 'nro_factura_afectada',
    'n factura afectada': 'nro_factura_afectada',
}

# Segunda pasada de reconocimiento — pedido explícito 2026-08-02: cada
# cliente redacta sus encabezados distinto (orden de palabras, "Proveedor"
# vs "Nombre" vs "Cliente" antes de "Razón Social"), y perseguir cada
# variante a mano en _HEADER_MAP siempre va un caso real por detrás (ver
# bug 2026-08-02: "Proveedor o Razón Social" de Cementos no matcheaba
# contra el sinónimo existente "Nombre o Razón Social"). En vez de eso,
# esta pasada busca una FRASE ANCLA dentro del encabezado ya normalizado
# — "razon social" casi nunca se usa para otra cosa que el nombre del
# tercero, así que es una señal confiable sin necesitar la frase exacta.
# Solo se usa sobre columnas que la primera pasada (match exacto) NO
# reconoció, y solo si ese campo todavía no tiene ninguna columna
# asignada — el match exacto siempre gana si ya existe.
_HEADER_FALLBACK = [
    ('razon social', 'nombre_proveedor'),
]

_AMOUNT_FIELDS = {'base_16', 'base_8', 'monto_retenido', 'base_sin_credito',
                  'base_generica', 'alicuota_pct', 'total_documento'}


class VeConectaCargaCompras(models.Model):
    _name = 've.conecta.carga.compras'
    _description = 'Carga de Libro de Compras — SmartIVA Conecta (CONECTA-14)'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    name = fields.Char(string='Referencia', copy=False, readonly=True, default='Nueva Carga')
    company_id = fields.Many2one(
        'res.company', string='Empresa', required=True,
        default=lambda self: self.env.company, index=True,
    )
    archivo = fields.Binary(string='Archivo (CSV/XLSX)', attachment=True)
    archivo_nombre = fields.Char(string='Nombre del Archivo')
    formato_fecha = fields.Selection([
        ('dmy', 'Normal (usar la fecha tal cual viene)'),
        ('mdy', 'Día y Mes invertidos en TODO el archivo'),
        ('auto', 'Detectar automáticamente (archivo con mezcla de ambos)'),
    ], string='Formato de Fecha', default='dmy', required=True,
        help='"Día y Mes invertidos": todo el archivo tiene el mismo '
             'problema. "Detectar automáticamente": el archivo mezcla '
             'filas ya correctas con filas invertidas — cada fila se '
             'corrige comparando contra el mes que domina el resto del '
             'archivo. Ver mismo campo en ve_conecta_carga_ventas.py — '
             'bug real 2026-07-31.')

    estado = fields.Selection([
        ('borrador',    'Borrador — Vista Previa'),
        ('confirmado',  'Confirmado'),
    ], string='Estado', default='borrador', required=True, tracking=True)

    linea_ids = fields.One2many(
        've.conecta.carga.compras.linea', 'carga_id', string='Filas')
    count_lineas = fields.Integer(compute='_compute_counts', store=True)
    count_brechas = fields.Integer(compute='_compute_counts', store=True)
    count_bloqueantes = fields.Integer(compute='_compute_counts', store=True)

    # ── Auditoría (mismo principio que Ventas — ver
    #    feedback_smartiva_auditoria_responsabilidad en memoria) ──────────
    confirmado_por_id = fields.Many2one('res.users', string='Confirmado por', readonly=True)
    fecha_confirmacion = fields.Datetime(string='Fecha Confirmación', readonly=True)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') or vals.get('name') == 'Nueva Carga':
                vals['name'] = self.env['ir.sequence'].next_by_code(
                    've.conecta.carga.compras') or 'CARGA-CMP/nueva'
        return super().create(vals_list)

    @api.depends('linea_ids.brecha', 'linea_ids.bloqueante')
    def _compute_counts(self):
        for rec in self:
            rec.count_lineas = len(rec.linea_ids)
            rec.count_brechas = len(rec.linea_ids.filtered('brecha'))
            rec.count_bloqueantes = len(rec.linea_ids.filtered('bloqueante'))

    # ─────────────────────────────────────────────────────────────────────
    # Parseo (idéntico al de Ventas, ver ve_conecta_carga_ventas.py)
    # ─────────────────────────────────────────────────────────────────────

    def _leer_filas(self):
        if not self.archivo:
            raise UserError('Suba un archivo (CSV o XLSX) primero.')
        raw = base64.b64decode(self.archivo)
        nombre = (self.archivo_nombre or '').lower()

        if nombre.endswith('.csv'):
            texto = raw.decode('utf-8-sig', errors='replace')
            sniffer_sample = texto[:2048]
            try:
                dialecto = csv.Sniffer().sniff(sniffer_sample, delimiters=',;')
            except csv.Error:
                dialecto = csv.excel
            filas = list(csv.reader(io.StringIO(texto), dialecto))
        else:
            try:
                import openpyxl
            except ImportError:
                raise UserError('La librería openpyxl no está disponible en este servidor.')
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(raw), read_only=True, data_only=True)
                hoja = wb[wb.sheetnames[0]]
                filas = list(hoja.iter_rows(values_only=True))
                wb.close()
            except Exception as e:
                raise UserError(f'No se pudo abrir el archivo: {e}')
            # Mismo bug real que ve_conecta_carga_ventas.py::_leer_filas (ver
            # ese archivo para el detalle) -- read_only=True puede cortar la
            # lectura casi de inmediato si el XLSX trae mal su metadato
            # <dimension>. Reintentar sin read_only solo si la primera
            # lectura salió sospechosamente corta.
            if len(filas) <= 1:
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(raw), read_only=False, data_only=True)
                    hoja = wb[wb.sheetnames[0]]
                    filas_reintento = list(hoja.iter_rows(values_only=True))
                    wb.close()
                    if len(filas_reintento) > len(filas):
                        filas = filas_reintento
                except Exception:
                    pass

        if not filas:
            raise UserError('El archivo está vacío.')

        encabezado = None
        data_start = 0
        for i, fila in enumerate(filas):
            if any(c is not None and str(c).strip() for c in fila):
                encabezado = fila
                data_start = i + 1
                break
        if encabezado is None:
            raise UserError('El archivo no tiene encabezados.')

        col_map = {}
        for i, h in enumerate(encabezado):
            if h is None:
                continue
            norm = _norm_header(str(h))
            if norm in _HEADER_MAP:
                col_map[i] = _HEADER_MAP[norm]

        mapeadas_ya = set(col_map.values())
        for i, h in enumerate(encabezado):
            if h is None or i in col_map:
                continue
            norm = _norm_header(str(h))
            for ancla, field in _HEADER_FALLBACK:
                if field in mapeadas_ya:
                    continue
                if ancla in norm:
                    col_map[i] = field
                    mapeadas_ya.add(field)
                    break

        mapeadas = set(col_map.values())
        faltantes = {'rif', 'fecha', 'nro_documento'} - mapeadas
        if faltantes:
            raise UserError(
                f'Columnas obligatorias faltantes: {", ".join(sorted(faltantes))}.\n'
                f'Columnas detectadas: {", ".join(sorted(mapeadas)) or "(ninguna)"}')

        raw_rows = []
        for row_num, row in enumerate(filas[data_start:], start=data_start + 1):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            vals = {'fila': f'{row_num:04d}'}
            for col_idx, field in col_map.items():
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if cell is None or str(cell).strip() == '':
                    continue
                if field in _AMOUNT_FIELDS:
                    # SUMAR, no sobreescribir — el formato SENIAT "ancho"
                    # real repite encabezados (ej. "Base 16,00 %" aparece
                    # 2 veces: sección de importación y sección de compras
                    # internas). Con overwrite, la última columna ganaba
                    # sin importar cuál traía el monto real — funcionaba
                    # de casualidad cuando la sección real era la última,
                    # pero perdía el dato en silencio si el cliente tiene
                    # montos reales en AMBAS secciones (bug real
                    # encontrado 2026-08-02 con "Comp Inter Exentas o
                    # Exoneradas"/"Compras Exoneradas"/"Compras No
                    # sujetas", las 3 mapeadas al mismo campo).
                    vals[field] = vals.get(field, 0.0) + _parse_amount(cell)
                elif field == 'fecha':
                    if self.formato_fecha == 'auto':
                        vals['_fecha_raw'] = cell
                    else:
                        vals[field] = _parse_date(cell, self.formato_fecha) or False
                elif field == 'fecha_aplic_retencion':
                    # Mismo formato que 'fecha' (dmy/mdy/auto). Bug real
                    # 2026-08-02 (Cementos): esta carga usaba formato_fecha
                    # 'auto', y la primera versión de este fix solo
                    # soportaba dmy/mdy acá — quedaba vacía en silencio y
                    # todo caía de vuelta a 'fecha' (Fecha Documento), el
                    # mismo bug que se estaba corrigiendo. 'auto' resuelve
                    # con SU PROPIO mes dominante (no el de 'fecha') en
                    # _resolver_fechas_auto — las 2 columnas pueden tener
                    # meses distintos (factura vs. aplicación real).
                    if self.formato_fecha == 'auto':
                        vals['_fecha_aplic_raw'] = cell
                    else:
                        vals[field] = _parse_date(cell, self.formato_fecha) or False
                elif field == 'rif':
                    vals[field] = _formatear_rif(str(cell).strip())
                else:
                    vals[field] = str(cell).strip()
            # Formato "largo" (Base Imponible + % Alíc. genéricos, ver
            # ejemplos/Libro de compras.jpeg — export real de Profit Plus)
            # → repartir a base_16/base_8 según la alícuota de esta fila.
            # No se mezcla con el formato "ancho" (columnas ya separadas):
            # solo actúa si base_generica vino y base_16/base_8 no.
            base_gen = vals.pop('base_generica', None)
            alic = vals.pop('alicuota_pct', None)
            if alic is not None and 0 < alic <= 1:
                # Normaliza fracción (0.16) a porcentaje (16) — algunos
                # sistemas de origen exportan la alícuota así en vez de
                # como entero (ver mismo fix en ve_conecta_carga_ventas.py).
                alic = alic * 100
            if base_gen and not vals.get('base_16') and not vals.get('base_8'):
                if alic is not None and abs(alic - 8.0) < 0.5:
                    vals['base_8'] = base_gen
                elif alic is not None and abs(alic - 16.0) < 0.5:
                    vals['base_16'] = base_gen
                elif alic in (None, 0.0):
                    vals['base_sin_credito'] = base_gen
                else:
                    vals['base_16'] = base_gen  # alícuota no reconocida: default 16%
            raw_rows.append(vals)
        if self.formato_fecha == 'auto':
            self._resolver_fechas_auto(raw_rows)
        return raw_rows

    def _resolver_fechas_auto(self, raw_rows):
        """Ver ve_conecta_carga_ventas.py::_resolver_fechas_auto — mismo
        criterio (mes dominante del archivo resuelve las filas ambiguas).

        Se aplica DOS VECES, una por cada columna de fecha ('fecha' y
        'fecha_aplic_retencion'), cada una con su PROPIO mes dominante —
        no se puede reusar el de 'fecha' para la otra columna, ya que
        representan momentos distintos (fecha de la factura del proveedor
        vs. fecha real de aplicación de la retención) y pueden caer en
        meses distintos (bug real 2026-08-02, Cementos: Fecha Documento
        dispersa en jul-nov, Fecha Aplic. Retención casi toda en nov)."""
        self._resolver_columna_fecha_auto(raw_rows, '_fecha_raw', 'fecha')
        self._resolver_columna_fecha_auto(raw_rows, '_fecha_aplic_raw', 'fecha_aplic_retencion')

    def _resolver_columna_fecha_auto(self, raw_rows, raw_key, out_key):
        from collections import Counter
        conteo_meses = Counter()
        candidatos_por_fila = []
        hay_columna = any(raw_key in vals for vals in raw_rows)
        for vals in raw_rows:
            cands = _fecha_candidatos(vals.get(raw_key))
            candidatos_por_fila.append(cands)
            if len(cands) == 1:
                conteo_meses[cands[0][1]] += 1
        mes_dominante = conteo_meses.most_common(1)[0][0] if conteo_meses else None
        for vals, cands in zip(raw_rows, candidatos_por_fila):
            vals.pop(raw_key, None)
            if not hay_columna:
                continue
            elegido = None
            if cands:
                elegido = cands[0]
                if len(cands) > 1 and mes_dominante is not None:
                    for c in cands:
                        if c[1] == mes_dominante:
                            elegido = c
                            break
            vals[out_key] = (
                f"{elegido[0]:04d}-{elegido[1]:02d}-{elegido[2]:02d}" if elegido else False
            )

    def action_previsualizar(self):
        self.ensure_one()
        if self.estado == 'confirmado':
            raise UserError('Esta carga ya fue confirmada — no se puede volver a previsualizar.')
        raw_rows = self._leer_filas()
        self.linea_ids.unlink()
        Linea = self.env['ve.conecta.carga.compras.linea']
        for vals in raw_rows:
            vals['carga_id'] = self.id
            Linea.create(vals)
        return self._reload()

    def action_confirmar(self):
        self.ensure_one()
        if self.estado == 'confirmado':
            raise UserError('Esta carga ya fue confirmada.')
        if not self.linea_ids:
            raise UserError('No hay filas para confirmar — previsualice primero.')
        bloqueadas = self.linea_ids.filtered('bloqueante')
        if bloqueadas:
            detalle = '\n'.join(
                f'  Fila {l.fila}: {l.brecha}' for l in bloqueadas[:15])
            raise UserError(
                f'{len(bloqueadas)} fila(s) bloqueada(s) — corrija o elimine antes de confirmar:\n'
                f'{detalle}')

        Partner = self.env['res.partner'].sudo()
        Move = self.env['account.move'].sudo()
        WhIvaProv = self.env['ve.wh.iva.prov'].sudo()
        journal = self.env['account.journal'].sudo().search(
            [('type', '=', 'purchase'), ('company_id', '=', self.company_id.id)], limit=1)
        account = self.env['account.account'].sudo().search(
            [('account_type', 'in', ('expense', 'expense_depreciation', 'expense_direct_cost')),
             ('company_ids', 'in', [self.company_id.id]), ('currency_id', '=', False)], limit=1)
        if not (journal and account):
            raise UserError(
                'Falta configurar el diario o la cuenta de compras de esta compañía '
                '(Contabilidad → Ajustes) antes de confirmar la carga.')
        tax_16 = self.env['account.tax'].sudo().search(
            [('type_tax_use', '=', 'purchase'), ('amount', '=', 16.0),
             ('amount_type', '=', 'percent'), ('company_id', '=', self.company_id.id)], limit=1)
        tax_8 = self.env['account.tax'].sudo().search(
            [('type_tax_use', '=', 'purchase'), ('amount', '=', 8.0),
             ('amount_type', '=', 'percent'), ('company_id', '=', self.company_id.id)], limit=1)
        payable_acct = self.env['account.account'].sudo().search(
            [('account_type', '=', 'liability_payable'),
             ('company_ids', 'in', [self.company_id.id])], limit=1)

        # Declaración IVA — ve.wh.iva.prov cuelga de ahí (declaracion_iva_id),
        # no directo del período. Se resuelve POR TRANSACCIÓN según la fecha
        # de la fila-factura (no una sola "declaración activa" para todo el
        # lote) — bug real 2026-07-29: un lote con facturas de 2 quincenas
        # distintas dejaba TODAS las retenciones en la declaración de hoy,
        # contando su C.66 en una quincena cuyo débito fiscal (C.42/43/49)
        # no las incluía. Ver ve_conecta_carga_ventas.py (mismo fix) y
        # ve.conciliacion.periodo::_asegurar_periodo.
        Periodo = self.env['ve.conciliacion.periodo'].sudo()
        DeclaracionIva = self.env['ve.declaracion.iva'].sudo()
        _declaraciones_cache = {}

        def _declaracion_para_fecha(fecha):
            periodo = Periodo._asegurar_periodo(self.company_id, fecha or fields.Date.today())
            if periodo.id not in _declaraciones_cache:
                _declaraciones_cache[periodo.id] = DeclaracionIva._get_or_create_for_periodo(periodo.id)
            return _declaraciones_cache[periodo.id]

        creadas = nuevos_partners = retenciones_creadas = 0
        monto_retenido_feed = monto_retenido_odoo = 0.0
        monto_facturas_feed = monto_facturas_odoo = 0.0
        errores = []
        # (rif, nro_documento) -> account.move ya creada en ESTE confirmar —
        # es lo que permite vincular la fila de retención (sin N° de Factura
        # propio) con la fila de la factura que la precede en el mismo
        # archivo (formato real de 2 filas, ver ejemplos/
        # PXL_20260728_145144441.jpg). El orden de `linea_ids` (_order =
        # 'fila') garantiza que la fila-factura se procese antes que su
        # fila-retención, igual que en el archivo de origen.
        facturas_creadas = {}
        lineas_factura_creadas = {}
        # RIF normalizado -> partner ya creado en ESTE mismo confirmar —
        # mismo bug/fix que Ventas (ver ese archivo): sin esta caché, un
        # proveedor nuevo que aparece en varias filas del mismo archivo
        # creaba un partner distinto por cada fila.
        partners_creados_lote = {}
        for linea in self.linea_ids:
            partner = linea.partner_id
            partner_creado = False
            if not partner and linea.rif:
                partner = partners_creados_lote.get(_norm_rif(linea.rif))
            if not partner:
                # company_id explícito — mismo motivo que en Ventas: sin
                # esto el partner queda "global" y el cliente web activa
                # otras compañías al navegar a sus documentos.
                vals_partner = {
                    'name': linea.nombre_proveedor or linea.rif,
                    'vat': linea.rif,
                    'company_type': 'company',
                    'company_id': self.company_id.id,
                    'supplier_rank': 1,
                }
                if payable_acct:
                    vals_partner['property_account_payable_id'] = payable_acct.id
                partner = Partner.with_company(self.company_id).create(vals_partner)
                nuevos_partners += 1
                partner_creado = True
                if linea.rif:
                    partners_creados_lote[_norm_rif(linea.rif)] = partner
            linea.write({'partner_id': partner.id, 'partner_creado': partner_creado})

            if not linea.nro_documento:
                # Fila de RETENCIÓN (sin N° de Factura propio) — no crea una
                # factura nueva, se vincula a la que ya se creó en esta
                # misma carga a partir de su N° de Factura Afectada.
                key = (linea.rif, linea.nro_factura_afectada)
                inv = facturas_creadas.get(key)
                linea_factura = lineas_factura_creadas.get(key)
                if not inv:
                    errores.append(
                        f'Fila {linea.fila}: no se encontró en esta carga la factura '
                        f'"{linea.nro_factura_afectada}" (N° de Factura Afectada) — '
                        'cárguela en la misma carga o corrija el dato.')
                    continue
                # Bug real encontrado 2026-07-29: en el formato de 2 filas,
                # Base 16%/8%/N° Control viven en la fila-FACTURA, no en
                # esta fila-retención (que solo trae N° de Comprobante/
                # Fecha/Monto Retenido) — pasar solo `linea` acá dejaba la
                # retención creada con Base Imponible=0 y % Retención
                # forzado a 100% (fallback erróneo por iva_total=0).
                # La declaración se resuelve por Fecha Aplic. Retención (fecha
                # real en que se aplicó/pagó la retención) si el feed la trae
                # — bug real 2026-08-02, Cementos: sin esa columna, se usaba
                # la fecha de la FACTURA y las retenciones quedaban dispersas
                # en 7 quincenas distintas en vez de caer en la quincena real
                # de aplicación. Si no viene, cae al criterio anterior (fecha
                # de la factura) — mismo comportamiento de siempre para
                # formatos que no traen esta columna (ej. Profit Plus).
                declaracion = _declaracion_para_fecha(
                    linea.fecha_aplic_retencion
                    or (linea_factura.fecha if linea_factura else linea.fecha))
                wh_prov = self._crear_retencion_prov(
                    linea_factura or linea, linea, inv, partner, declaracion, WhIvaProv)
                linea.write({'invoice_id': inv.id, 'wh_iva_prov_id': wh_prov.id})
                retenciones_creadas += 1
                monto_retenido_feed += linea.monto_retenido
                monto_retenido_odoo += wh_prov.monto_retenido
                continue

            # Fila de FACTURA — crea la factura de compra.
            lineas_factura = []
            if linea.base_16:
                lv = {'name': f'Carga Libro de Compras — fila {linea.fila}',
                      'quantity': 1, 'price_unit': linea.base_16, 'account_id': account.id}
                if tax_16:
                    lv['tax_ids'] = [(6, 0, [tax_16.id])]
                lineas_factura.append((0, 0, lv))
            if linea.base_8:
                lv = {'name': f'Carga Libro de Compras 8% — fila {linea.fila}',
                      'quantity': 1, 'price_unit': linea.base_8, 'account_id': account.id}
                if tax_8:
                    lv['tax_ids'] = [(6, 0, [tax_8.id])]
                lineas_factura.append((0, 0, lv))
            if linea.base_sin_credito:
                # Compra sin derecho a crédito fiscal — sin tax_ids.
                lineas_factura.append((0, 0, {
                    'name': f'Carga Libro de Compras (sin crédito) — fila {linea.fila}',
                    'quantity': 1, 'price_unit': linea.base_sin_credito, 'account_id': account.id,
                }))
            if not lineas_factura:
                errores.append(f'Fila {linea.fila}: sin base 16%/8%/sin-crédito — factura omitida')
                continue

            inv = Move.create({
                # A diferencia de Ventas: acá NO fijamos `name` con el
                # documento del proveedor — el proveedor es un tercero, su
                # numeración legal es asunto suyo, no algo que SmartIVA deba
                # preservar como "nuestro" número de documento. `name` es
                # solo la referencia interna de este movimiento contable
                # (se autoasigna del diario de compras); `ref` sí guarda el
                # N° de factura del proveedor tal cual, para trazabilidad
                # (decisión de diseño 2026-07-24, distinta de Ventas a
                # propósito).
                'move_type': 'in_invoice',
                'partner_id': partner.id,
                'invoice_date': linea.fecha or fields.Date.today(),
                # Mismo fix que ve_conecta_carga_ventas.py (bug real
                # 2026-07-30): sin esto, Odoo calculaba invoice_date_due
                # solo y podía terminar en HOY en vez de la fecha real de
                # la factura con un Libro de Compras histórico.
                'invoice_date_due': linea.fecha or fields.Date.today(),
                'ref': linea.nro_documento or False,
                'journal_id': journal.id,
                'company_id': self.company_id.id,
                'currency_id': self.company_id.currency_id.id,
                'nro_control': linea.nro_control or False,
                'invoice_line_ids': lineas_factura,
            })
            try:
                with self.env.cr.savepoint():
                    inv.action_post()
                linea.invoice_id = inv.id
                creadas += 1
                monto_facturas_feed += linea.total_documento
                monto_facturas_odoo += inv.amount_total
            except Exception as exc:
                errores.append(f'Fila {linea.fila}: {exc}')
                continue
            facturas_creadas[(linea.rif, linea.nro_documento)] = inv
            lineas_factura_creadas[(linea.rif, linea.nro_documento)] = linea

            # Retención IVA Proveedores — formato "1 fila" (compatibilidad
            # hacia atrás): si esta MISMA fila ya trae el monto retenido
            # (en vez de venir en una segunda fila aparte), se crea de una
            # vez. A diferencia de Ventas, no existe ningún hook nativo que
            # la cree sola al postear — el comprobante de retención lo
            # genera nuestro propio cliente al pagarle al proveedor, un
            # proceso externo a SmartIVA.
            if linea.monto_retenido > 0:
                # Fecha Aplic. Retención (si el feed la trae) determina el
                # período/quincena real, no la fecha de la factura del
                # proveedor — ver comentario detallado en la rama de 2 filas
                # más arriba (bug real 2026-08-02, Cementos).
                declaracion = _declaracion_para_fecha(linea.fecha_aplic_retencion or linea.fecha)
                wh_prov = self._crear_retencion_prov(linea, linea, inv, partner, declaracion, WhIvaProv)
                linea.wh_iva_prov_id = wh_prov.id
                retenciones_creadas += 1
                monto_retenido_feed += linea.monto_retenido
                monto_retenido_odoo += wh_prov.monto_retenido

        self.write({
            'estado': 'confirmado',
            'confirmado_por_id': self.env.user.id,
            'fecha_confirmacion': fields.Datetime.now(),
        })

        periodos_usados = sorted({
            d.conciliacion_id.periodo_retencion for d in _declaraciones_cache.values()
            if d.conciliacion_id
        })
        diferencia_monto = monto_retenido_odoo - monto_retenido_feed
        # Total archivo es opcional (columna "Total Importe con IVA" no
        # siempre viene en el feed) — si nunca se pobló, no se compara en
        # falso contra el total real de Odoo (mostraría "Con diferencias"
        # aunque el archivo simplemente no traiga ese dato).
        hay_total_archivo = any(l.total_documento for l in self.linea_ids)
        diferencia_facturas = monto_facturas_odoo - monto_facturas_feed

        # Tabla de consistencia — mismo estilo/columnas que
        # ve_conecta_carga_ventas.py::action_confirmar (pedido explícito
        # 2026-08-02, mismo criterio de auditoría en las 2 cargas).
        def _n(v):
            return f'{v:,}'

        def _m(v):
            return f'{v:,.2f}' if v else '—'

        th = ('style="border:1px solid #ccc; padding:3px 8px; background:#f0f0f0; '
              'text-align:left;"')
        td = 'style="border:1px solid #ccc; padding:3px 8px;"'
        tdr = 'style="border:1px solid #ccc; padding:3px 8px; text-align:right;"'

        def _fila(concepto, cant, m_archivo='—', m_odoo='—', dif='—', color=None):
            estilo_dif = f' style="color:{color};"' if color else ''
            return (
                f'<tr><td {td}>{concepto}</td>'
                f'<td {tdr}>{cant}</td>'
                f'<td {tdr}>{m_archivo}</td>'
                f'<td {tdr}>{m_odoo}</td>'
                f'<td {tdr}><span{estilo_dif}>{dif}</span></td></tr>'
            )

        tabla_consistencia = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Concepto</th><th {th}>Cantidad</th>'
            f'<th {th}>Monto Archivo</th><th {th}>Monto Odoo</th>'
            f'<th {th}>Diferencia</th></tr>'
            + _fila('Filas leídas', _n(len(self.linea_ids)))
            + (_fila('Facturas creadas', _n(creadas),
                      _m(monto_facturas_feed), _m(monto_facturas_odoo),
                      _m(diferencia_facturas) if abs(diferencia_facturas) > 0.01 else 'cuadra',
                      color='#dc3545' if abs(diferencia_facturas) > 0.01 else '#198754')
               if hay_total_archivo else _fila('Facturas creadas', _n(creadas)))
            + _fila('Facturas rechazadas', _n(len(self.linea_ids) - creadas))
            + _fila('Retenciones creadas', _n(retenciones_creadas),
                    _m(monto_retenido_feed), _m(monto_retenido_odoo),
                    _m(diferencia_monto) if abs(diferencia_monto) > 0.01 else 'cuadra',
                    color='#dc3545' if abs(diferencia_monto) > 0.01 else '#198754')
            + '</table>'
        )

        tabla_detalle = (
            f'<table style="border-collapse:collapse; font-size:0.85rem;">'
            f'<tr><th {th}>Concepto</th><th {th}>Valor</th></tr>'
            f'<tr><td {td}>Proveedores nuevos</td>'
            f'<td {tdr}>{_n(nuevos_partners)}</td></tr>'
            f'<tr><td {td}>Períodos usados</td>'
            f'<td {tdr}>{", ".join(periodos_usados) or "—"}</td></tr>'
            f'</table>'
        )

        cuerpo = (
            f'<b>Carga de Libro de Compras confirmada</b><br/>'
            f'<b>Archivo:</b> {self.archivo_nombre or "—"}<br/>'
            f'<b>Confirmado por:</b> {self.env.user.name}<br/><br/>'
            f'<b>— Consistencia —</b><br/>{tabla_consistencia}<br/>'
            f'<b>— Detalle —</b><br/>{tabla_detalle}'
        )
        if errores:
            cuerpo += '<br/><b>Filas con error:</b><br/>' + '<br/>'.join(errores[:10])
        self.message_post(body=Markup(cuerpo), message_type='comment', subtype_xmlid='mail.mt_note')
        return self._reload()

    def action_deshacer(self):
        """Revierte lo que action_confirmar() creó para ESTA carga
        específica — retenciones IVA Proveedores y las facturas mismas —
        y la deja de vuelta en Borrador. Mismo criterio que
        ve_conecta_carga_ventas.py::action_deshacer (ver ese archivo para
        el contexto completo del bug real que motivó construir esto,
        2026-07-31). No borra los partners creados por la carga."""
        self.ensure_one()
        if self.estado != 'confirmado':
            raise UserError('Solo se puede deshacer una carga ya confirmada.')

        WhIvaProv = self.env['ve.wh.iva.prov'].sudo()
        errores = []
        n_ret = n_fact = 0

        facturas = self.linea_ids.mapped('invoice_id').sudo()
        declaraciones_tocadas = self.env['ve.declaracion.iva']
        for inv in facturas:
            for wh in WhIvaProv.search([('invoice_id', '=', inv.id)]):
                declaraciones_tocadas |= wh.declaracion_iva_id
                try:
                    with self.env.cr.savepoint():
                        wh.unlink()
                    n_ret += 1
                except Exception as exc:
                    errores.append(f'Retención de factura {inv.name or inv.id} no eliminada: {exc}')
            try:
                with self.env.cr.savepoint():
                    if inv.state == 'posted':
                        inv.button_draft()
                    inv.unlink()
                n_fact += 1
            except Exception as exc:
                errores.append(f'Factura {inv.name or inv.id} no eliminada: {exc}')

        # Declaraciones/períodos que quedaron vacíos por esta reversión —
        # no se tocan si otra carga/retención todavía los usa. ve.conciliacion
        # .periodo se comparte con el lado Clientes/SENIAT (ver mismo fix en
        # ve_conecta_carga_ventas.py::action_deshacer, incidente real
        # 2026-08-05) -- un período de este lado Proveedores puede tener
        # igual retenciones SENIAT enganchadas, hay que revisar también.
        n_decl = 0
        for decl in declaraciones_tocadas:
            if not WhIvaProv.search_count([('declaracion_iva_id', '=', decl.id)]):
                periodo = decl.conciliacion_id
                try:
                    with self.env.cr.savepoint():
                        decl.unlink()
                        if periodo and not self.env['ve.wh.iva'].sudo().search_count(
                                [('conciliacion_id', '=', periodo.id)]) and not \
                                self.env['ve.seniat.retencion'].sudo().search_count(
                                    [('conciliacion_id', '=', periodo.id)]):
                            periodo.unlink()
                    n_decl += 1
                except Exception as exc:
                    errores.append(f'Declaración/período no eliminado: {exc}')

        self.linea_ids.write({'invoice_id': False, 'wh_iva_prov_id': False, 'partner_creado': False})
        self.write({
            'estado': 'borrador', 'confirmado_por_id': False, 'fecha_confirmacion': False,
        })

        cuerpo = (
            f'<b>Carga deshecha</b><br/>'
            f'<b>Facturas eliminadas:</b> {n_fact} de {len(facturas)}<br/>'
            f'<b>Retenciones eliminadas:</b> {n_ret}<br/>'
            f'<b>Declaraciones/períodos vacíos eliminados:</b> {n_decl}<br/>'
            f'<b>Proveedores creados por esta carga:</b> se conservan.<br/>'
        )
        if errores:
            cuerpo += '<br/><b>Filas con error:</b><br/>' + '<br/>'.join(errores[:15])
        self.message_post(body=Markup(cuerpo), message_type='comment', subtype_xmlid='mail.mt_note')
        return self._reload()

    def _crear_retencion_prov(self, linea_montos, linea_datos, inv, partner, declaracion, WhIvaProv):
        """Crea ve.wh.iva.prov a partir de los datos tal cual vienen en el
        archivo (sin recalcular nada). Recibe 2 filas porque el formato
        real de 2 filas por transacción (ver ejemplos/
        PXL_20260728_145144441.jpg) separa los datos en 2 registros
        distintos: `linea_montos` es la fila que trae Base 16%/8%/N°
        Control (la FACTURA), `linea_datos` es la que trae N° de
        Comprobante/Fecha/Monto Retenido (la RETENCIÓN). En el formato
        legado de 1 fila, ambos parámetros son la MISMA fila."""
        iva_total = round(linea_montos.base_16 * 0.16, 2) + round(linea_montos.base_8 * 0.08, 2)
        pct = round((linea_datos.monto_retenido / iva_total) * 100, 2) if iva_total else (
            75.0 if linea_montos.nro_control else 100.0)
        return WhIvaProv.create({
            # N° de Comprobante de Retención: lo emite NUESTRO cliente (el
            # comprador), no el proveedor — mismo principio que N° Factura
            # en Ventas, se respeta tal cual si el feed lo trae; si no, sí
            # se autogenera (no hay otra fuente posible para ese número).
            'name': linea_datos.nro_comp_retencion or False,
            'company_id': self.company_id.id,
            'partner_id': partner.id,
            'invoice_id': inv.id,
            'nro_control': linea_montos.nro_control or False,
            'fecha': linea_datos.fecha_aplic_retencion or linea_datos.fecha or fields.Date.today(),
            'declaracion_iva_id': declaracion.id,
            'monto_base_16': linea_montos.base_16,
            'monto_base_8': linea_montos.base_8,
            'porcentaje_retencion': pct,
        })

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }


class VeConectaCargaComprasLinea(models.Model):
    _name = 've.conecta.carga.compras.linea'
    _description = 'Fila de Carga de Libro de Compras — SmartIVA Conecta'
    _order = 'fila'

    carga_id = fields.Many2one(
        've.conecta.carga.compras', string='Carga', required=True, ondelete='cascade')
    estado_carga = fields.Selection(related='carga_id.estado', string='Estado Carga')
    fila = fields.Char(string='Fila')

    rif = fields.Char(string='RIF')
    nombre_proveedor = fields.Char(string='Proveedor')
    nro_control = fields.Char(string='N° Control')
    nro_documento = fields.Char(string='N° de Factura')
    nro_comp_retencion = fields.Char(
        string='N° de Comprobante (Retención)',
        help='Número del comprobante de retención que NUESTRO cliente le emite '
             'al proveedor — no lo genera SmartIVA, se toma tal cual del feed.')
    fecha = fields.Date(string='Fecha')
    fecha_aplic_retencion = fields.Date(
        string='Fecha Aplic. Retención',
        help='Opcional — fecha en que el comprador (nuestro cliente) realmente '
             'aplicó/pagó la retención. Determina la quincena SENIAT de la '
             'retención (ver action_confirmar); si no viene, se usa "Fecha" '
             '(fecha de la factura del proveedor) como respaldo.')
    base_16 = fields.Float(string='Base 16%', digits=(16, 2))
    base_8 = fields.Float(string='Base 8%', digits=(16, 2))
    base_sin_credito = fields.Float(
        string='Sin Crédito Fiscal', digits=(16, 2),
        help='Compras sin derecho a crédito fiscal — sin IVA deducible.')
    total_documento = fields.Float(
        string='Total Documento (Archivo)', digits=(16, 2),
        help='Opcional — "Total Importe con IVA" tal cual viene en el archivo. '
             'No se usa para crear la factura — solo para la tabla de '
             'Consistencia del resumen al confirmar, como comprobación '
             'independiente contra amount_total de Odoo.')
    monto_retenido = fields.Float(
        string='IVA Retenido al Vendedor', digits=(16, 2),
        help='Si el feed trae este monto, se crea la retención IVA Proveedores '
             '(ve.wh.iva.prov) directamente — no hay hook automático para Compras.')
    nro_factura_afectada = fields.Char(
        string='N° de Factura Afectada',
        help='Solo presente en la fila del COMPROBANTE DE RETENCIÓN (formato real '
             'de 2 filas por transacción, ver ejemplos/PXL_20260728_145144441.jpg) — '
             'apunta al N° de Factura de la fila que le dio origen. Esta fila no '
             'crea una factura propia; al confirmar se vincula a la factura ya '
             'creada en esta misma carga con ese N° de Factura.')

    partner_id = fields.Many2one(
        'res.partner', string='Proveedor (match)',
        compute='_compute_partner_id', store=True, readonly=False,
        help='Proveedor encontrado por RIF. Editable — corrija aquí si el match '
             'automático no es el correcto antes de confirmar.')
    es_partner_nuevo = fields.Boolean(compute='_compute_partner_id', store=True)

    es_duplicado_factura = fields.Boolean(
        compute='_compute_partner_id', store=True,
        help='Ya existe una factura de este proveedor con el mismo N° de Factura '
             '(cada proveedor tiene su propia numeración — la unicidad es por '
             'RIF del proveedor, no de la compañía).')
    es_duplicado_retencion = fields.Boolean(
        compute='_compute_partner_id', store=True,
        help='Ya existe una retención IVA Proveedores con el mismo N° Comprobante '
             'de Retención en esta compañía (ese número lo emite nuestro cliente, '
             'es único por compañía).')
    bloqueante = fields.Boolean(compute='_compute_partner_id', store=True)
    brecha = fields.Char(compute='_compute_partner_id', store=True, string='Brecha')

    invoice_id = fields.Many2one('account.move', string='Factura Creada', readonly=True)
    wh_iva_prov_id = fields.Many2one(
        've.wh.iva.prov', string='Retención Creada', readonly=True)

    partner_creado = fields.Boolean(default=False, readonly=True)

    @api.depends('rif', 'nro_control', 'nro_documento', 'nro_comp_retencion',
                 'nro_factura_afectada', 'fecha', 'carga_id.company_id', 'carga_id.estado',
                 'partner_creado')
    def _compute_partner_id(self):
        Partner = self.env['res.partner']
        Move = self.env['account.move']
        WhIvaProv = self.env['ve.wh.iva.prov']
        for linea in self:
            company = linea.carga_id.company_id
            partner = False
            if linea.rif:
                # Mismo fix que ve_conecta_carga_ventas.py — sin filtro de
                # compañía podía enganchar un contacto de OTRA compañía
                # (ej. PILOTO archivada) y quedar pegado ahí (store=True).
                # Match NORMALIZADO (sin guión/espacios) — un match exacto
                # contra `vat` no encontraba el proveedor ya existente si
                # el formato del RIF no coincidía letra por letra, y lo
                # creaba duplicado.
                rif_norm = _norm_rif(linea.rif)
                candidatos = Partner.search([
                    ('vat', '!=', False),
                    '|', ('company_id', '=', False), ('company_id', '=', company.id),
                ])
                partner = candidatos.filtered(lambda p: _norm_rif(p.vat) == rif_norm)[:1]
            linea.partner_id = partner
            linea.es_partner_nuevo = bool(linea.rif) and not partner

            if linea.carga_id.estado == 'confirmado':
                linea.es_duplicado_factura = False
                linea.es_duplicado_retencion = False
                linea.bloqueante = False
                linea.brecha = 'Proveedor creado' if linea.partner_creado else False
                continue

            # Vinculación y control de duplicados SIEMPRE por RIF + N° de
            # Factura (no por N° Control, que en la práctica a veces no
            # viene en el export del cliente — ver ejemplos/
            # PXL_20260728_145144441.jpg) — cada proveedor numera su propia
            # serie de facturas, por eso el chequeo incluye partner_id.
            linea.es_duplicado_factura = bool(
                linea.nro_documento and company and partner and Move.search([
                    ('company_id', '=', company.id),
                    ('move_type', '=', 'in_invoice'),
                    ('partner_id', '=', partner.id),
                    ('ref', '=', linea.nro_documento),
                ], limit=1))
            linea.es_duplicado_retencion = bool(
                linea.nro_comp_retencion and company and WhIvaProv.search([
                    ('company_id', '=', company.id),
                    ('name', '=', linea.nro_comp_retencion),
                ], limit=1))

            if not linea.rif:
                linea.bloqueante = True
                linea.brecha = 'Sin RIF'
            elif not linea.fecha:
                # Mismo bug real que ve_conecta_carga_ventas.py, encontrado
                # 2026-07-30: si Fecha no parsea (formato no reconocido por
                # _parse_date), action_confirmar() la pisaba en silencio
                # con "hoy" — tanto en la fila-factura (invoice_date) como
                # en la fila-retención (fecha del comprobante). Bloquea en
                # vez de adivinar.
                linea.bloqueante = True
                linea.brecha = 'Fecha no reconocida — revise el formato de la columna Fecha'
            elif not linea.nro_documento and not linea.nro_factura_afectada:
                linea.bloqueante = True
                linea.brecha = 'Sin N° de Factura'
            elif not linea.nro_documento and linea.nro_factura_afectada:
                # Fila de RETENCIÓN (formato real de 2 filas, ver
                # ejemplos/PXL_20260728_145144441.jpg) — no crea su propia
                # factura, se vincula a la fila-factura con ese mismo N° de
                # Factura al confirmar (action_confirmar). No bloquea por
                # "Sin N° de Factura" — es una fila válida de por sí.
                if linea.es_duplicado_retencion:
                    linea.bloqueante = True
                    linea.brecha = 'Retención duplicada — ya existe (mismo N° Comp. Retención)'
                else:
                    linea.bloqueante = False
                    linea.brecha = f'Retención de la factura {linea.nro_factura_afectada} — se vinculará al confirmar'
            elif linea.es_duplicado_factura:
                linea.bloqueante = True
                linea.brecha = 'Factura duplicada — ya existe (mismo RIF + N° de Factura de este proveedor)'
            elif linea.es_duplicado_retencion:
                linea.bloqueante = True
                linea.brecha = 'Retención duplicada — ya existe (mismo N° Comp. Retención)'
            elif not partner:
                linea.bloqueante = False
                linea.brecha = 'Proveedor nuevo — se creará'
            else:
                linea.bloqueante = False
                linea.brecha = False
